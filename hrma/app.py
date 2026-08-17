import os
import sys

# app.py DOĞRUDAN çalıştırılırsa (python hrma/app.py / python app.py) depo
# kökü sys.path'te olmaz → "ModuleNotFoundError: No module named 'hrma'"
# (2026-07-15 Windows geri dönütü). run.py/run_windows.py bunu zaten yapıyor;
# app.py'ye de eklendi ki hangi dosya çalıştırılırsa çalışsın 'hrma' bulunur.
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from flask import (Flask, render_template, request, jsonify, send_file,
                   send_from_directory)
import numpy as np
import json
import io
import math
import contextlib
import platform
import collections
import uuid

# Apply Windows fixes before importing other modules
if platform.system() == 'Windows':
    try:
        from hrma.utils.windows_compatibility import windows_compat, apply_windows_fixes
        windows_fixes = apply_windows_fixes()
        if windows_fixes:
            print(f"Windows compatibility fixes applied: {windows_fixes['fixes_applied']}")
    except ImportError:
        print("Windows compatibility module not found - continuing without fixes")

# Engines
from hrma.engines.hybrid_rocket_engine import HybridRocketEngine
from hrma.engines.solid_rocket_engine import SolidRocketEngine
from hrma.engines.liquid_rocket_engine import LiquidRocketEngine

# Utils
from hrma.utils.injector_design import (
    InjectorDesign,
    # v2.6.26: bu sabit app.py'de KULLANILIYOR ama import EDILMIYORDU.
    # Sonuc: impingement enjektorunde 'impingement_angle' bos/0 geldiginde
    # /calculate NameError ile HTTP 500 veriyordu (yayinlanmis surumde de
    # vardi). Arayuz alani doldurdugu icin gozden kacmisti; alani temizleyen
    # kullanici ya da API cagrisi cokertiyordu. Katman B sarsim bekcisi
    # bagimsiz olarak bu yolu deneyince yakalandi.
    IMPINGEMENT_HALF_ANGLE_DEG,
)
from hrma.utils.input_guard import safe_name, safe_arcname, is_safe_arcname
from hrma.utils.common_fixes import validation, calculations, graph_fixes, fuel_mixer, export_fixes
from hrma.utils.optimum_of_ratio import of_optimizer

# Validation
# v2.5.0 G1 (2026-07-17): experimental_validator emekli — sentetik kayitlar
# tests/fixtures'a tasindi, gercek deney DB'si hrma.validation.experiment_db
from hrma.validation.validation_system import validator
from hrma.validation.motor_validation import motor_validator

# Analysis
from hrma.analysis.regression_analysis import regression_analyzer
from hrma.analysis.launch_site import resolve_launch_site
from hrma.analysis.safety_analysis import SafetyAnalyzer
from hrma.analysis.structural_analysis import StructuralAnalyzer
from hrma.analysis.heat_transfer_analysis import HeatTransferAnalyzer
# v2.6.27: emekli cozuculerin (cfd_analysis, kinetic_analysis) acilis
# importlari kaldirildi — uclari 501 donuyor, govdeleri sokuldu (asagida).
from hrma.analysis.trajectory_analysis import TrajectoryAnalyzer

# Dalga 4A — hızlı gerçekçi modeller (sahte CFD/kinetik yerine):
# quasi-1D lüle akışı, kademeli kinetik verim, kullanıcı CSV doğrulaması,
# hafif iş kuyruğu (docs/ANALIZ_PLATFORM_PLANI.md)
from hrma.analysis.nozzle_flow_1d import NozzleFlow1D
from hrma.analysis.kinetic_efficiency import (
    kinetic_efficiency, VALID_FIDELITY_LEVELS as KINETIC_FIDELITY_LEVELS,
    CANTERA_AVAILABLE as KINETIC_CANTERA_AVAILABLE,
)
from hrma.validation.user_data_validation import (
    parse_thrust_csv, compare as compare_thrust_curves,
)
from hrma.utils.job_runner import job_runner

# Data
from hrma.data.propellant_database import propellant_db
from hrma.data.open_source_propellant_api import propellant_api
from hrma.data.chemical_database import chemical_db
from hrma.data.database_integrations import DatabaseManager

import threading
import traceback
import warnings

# Visualization
from hrma.visualization.visualization import (
    create_motor_plot, create_injector_plot, create_performance_plots,
    create_heat_transfer_plots, create_combustion_analysis_plots,
    create_structural_analysis_plots, create_real_time_dashboard,
    create_3d_motor_visualization, create_comparative_analysis_plot,
    create_chamber_pressure_mixture_ratio_3d_surface,
    create_nozzle_mach_area_ratio_contour,
    create_wall_heat_flux_waterfall_plot,
    create_improved_motor_cross_section,
    create_improved_injector_design,
    # SHOWERHEAD_PATTERNS: plaka yüz yerleşiminin TEK doğruluk kaynağı.
    # Desen adları burada kopyalanmaz — çizen taraf hangi desenleri
    # tanıyorsa istek sınırı da tam onları kabul eder.
    SHOWERHEAD_PATTERNS,
    # _fig_json: fig.to_json() yerine TEK JSON kapısı — plotly 6'nın bdata
    # çıktısını vendor plotly.js 1.58.5'in çizebileceği düz listeye açar
    # (boş grafik bugunun kökü). PALETTE: grafik serileri için ortak palet.
    _fig_json, PALETTE
)
from hrma.export.motor_geometry import (
    solid_results_to_motor_geometry,
    liquid_results_to_motor_geometry,
)
# NOT (göç, 16 Ağu 2026): NOZZLE_FRICTION_LOSS_FRACTION_DEFAULT buradan
# KALDIRILDI — /api/flow ucu artık sürtünme kesrini varsayılan olarak
# GEÇMİYOR (çözücü ölçülen sınır tabakası değerini yayımlıyor). Sabitin
# tek tanım yeri hrma.constants'tır ve orada yalnız yedek olarak durur.
from hrma.constants import G_0
from hrma.visualization.advanced_results import (
    create_cea_style_results, create_altitude_performance_plot,
    create_mass_fractions_plot, create_thrust_altitude_plot
)

# Export
from hrma.export.openrocket_integration import OpenRocketExporter
from hrma.export.cad_visualization import MotorCADDesigner

from datetime import datetime

app = Flask(__name__)

# İstek gövdesi üst sınırı (v2.6.2). Sınırsız bırakıldığında
# /api/validation/upload-csv gibi uçlara keyfi büyüklükte gövde gönderilip
# bellek tüketilebiliyordu. 32 MB, en büyük meşru girdiden (STEP/ork dosyası)
# fazlasıyla geniş; aşan istekler Flask tarafından 413 ile reddedilir.
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

#: 6-DOF entegrasyonu için üst zaman ufku [s]. Kaçış yörüngesinde araç yere
#: dönmediği için sınırsız t_max bitmeyen entegrasyona yol açıyordu.
_SIXDOF_T_MAX_LIMIT_S = 3600.0

#: 6-DOF çözücüsü için DUVAR SAATİ bütçesi [s] (Faz 5B / B5).
#:
#: ÖLÇÜM (HEAD 9d3728e, bu makine): NaN girdili tek bir istek 900 saniye
#: boyunca DÖNMEDİ — üç sonlandırma olayının üçü de karşılaştırmaya dayanıyor
#: ve NaN durumunda hiçbiri işaret değiştiremiyor. Kök neden Dalga 1'de
#: çözücü tarafında kapatıldı (sonlu-değer denetimi); bu bütçe İKİNCİ
#: savunma hattıdır.
#:
#: Sayı ölçümle seçildi. Geçerli ve fiziksel girdiyle ölçülen EN KÖTÜ süre
#: (t_max = 3600 s üst sınırında, burn_time = 3600 s) **10,1 s**; itki
#: 1e7 N + t_max 3600 s ile 8,1 s; sağlıklı çağrı 0,09 s. 60 s, ölçülen en
#: kötü hâlin ~6 katı — meşru hiçbir koşu kesilmez, ama bir gerileme
#: yeniden sonsuz döngü açarsa istek kilitlenmez.
#:
#: Uygulama daemon iş parçacığıdır: kesilemeyen bir çözücüyü öldüremeyiz, ama
#: ISTEK dönebilir ve daemon olduğu için süreç kapanışını da bloke etmez.
_SIXDOF_WALL_CLOCK_BUDGET_S = 60.0

# GÜVENLİK — v2.6.2: joker CORS kaldırıldı.
#
# Burada eskiden argümansız ``CORS(app)`` vardı; flask-cors varsayılanı TÜM
# rotalara ``Access-Control-Allow-Origin: *`` basar. Sunucu 127.0.0.1'e bağlı
# olsa bile bu yeterli koruma değildi: kullanıcı HRMA açıkken kötü niyetli bir
# siteye girdiğinde o sayfanın JS'i http://127.0.0.1:8080 uçlarına istek atıp
# YANITI OKUYABİLİYORDU (normalde same-origin politikası okumayı engeller).
# Bu, /download/stl yol-kaçışı açığıyla birleşince diskten dosya sızdırma
# zinciri oluşturuyordu.
#
# HRMA tek kullanıcılı bir masaüstü uygulaması: arayüz sayfaları uygulamanın
# KENDİ kökeninden servis edilir, yani çapraz köken erişimine hiç ihtiyaç yok.
# CORS tamamen kapatıldı ve ayrıca yabancı kökenli durum değiştiren istekler
# aşağıdaki süzgeçle reddediliyor.
#: Geri döngü (loopback) ana makine adları. Sayısal 127.0.0.0/8 adresleri
#: ayrıca ``ipaddress`` ile sınanır, bu küme yalnız adla gelenler içindir.
def _mm_to_m(value, default_m):
    """Arayuzden gelen milimetre degerini metreye cevirir.

    v2.6.25 — Birim sozlesmesi bu projede tekrar eden bir hata kaynagi oldu
    (termal panel mm degerini metre alanina basiyordu: 100 mm -> 100 m; tank
    STEP ihracati 1000 kati iki kez uyguluyordu). Donusum tek bir yerde
    yapilir; okunamayan ya da <=0 girdi sessizce 0 olmaz, VARSAYILANA doner —
    0 kalinlik termal modelde sonsuz iletkenlik demektir.
    """
    try:
        mm = float(value)
    except (TypeError, ValueError):
        return default_m
    if mm <= 0:
        return default_m
    return mm / 1000.0


def _mm_to_m_optional(value):
    """Opsiyonel milimetre girdisini metreye cevirir; yoksa None dondurur.

    _mm_to_m'den farki: VARSAYILAN YOKTUR. 'Bos birak = otomatik' anlamina
    gelen alanlar (or. kamara boyu ezmesi) icin dogru sozlesme budur —
    varsayilan enjekte etmek kullanicinin bos biraktigi alani bir TASARIM
    kararina cevirir. Okunamayan girdi de None doner; motor tarafi bunu
    kendi araligina gore uyararak reddeder.
    """
    if value is None or value == '':
        return None
    try:
        mm = float(value)
    except (TypeError, ValueError):
        return None
    if mm <= 0:
        return None
    return mm / 1000.0


def _positive_float(value):
    """Sonlu ve pozitif sayiya cevirir; olmuyorsa None."""
    if value is None or value == '':
        return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(out) or out <= 0:
        return None
    return out


def _resolve_vehicle_spec(data, motor_results):
    """Aracin TEK tanimi — yorunge ve OpenRocket dallari ayni araci gorsun.

    v2.6.26 olcumu: tek /calculate yanitinda UC farkli arac vardi. OpenRocket
    dali kendi sabit sozlugune dusuyordu (5 kg / 0,10 m / 1,5 m), yorunge dali
    var olmayan form alanlarindan okudugu icin kendi varsayilanlarina
    dusuyordu (50 kg / 0,15 m), motorun kendi yapisal kutlesi ise 89,3 kg idi.
    Govde capi motor kasasindan (0,184 m) ince, kuru kutle motorun kendi
    yapisindan 18 kat kucuktu; apoje 570.915 m'ye karsi 10.576 m cikiyordu ve
    mass_ratio/delta_v yanlis olanindan turuyordu.

    Sozlesme: sayi UYDURULMAZ. Kullanici vermediyse alan None kalir ve
    kaynagi 'not_supplied' diye beyan edilir; cagiran taraf ne yapacagina
    ona bakarak karar verir.
    """
    spec = {}
    sources = {}

    # Kuru kutle: sayfadaki 'final_mass' alani "yakit tukendikten sonraki
    # kutle (kuru kutle + faydali yuk)" diye tanimli — yorungenin kuru
    # kutlesi tam olarak budur.
    dry = _positive_float(data.get('vehicle_mass_dry'))
    if dry is not None:
        sources['dry_mass'] = 'request'
    else:
        # Alt sinir: motorun kendi atil kutlesi. Bu ARACIN kuru kutlesi
        # DEGILDIR (govde, kanat, faydali yuk haric), bu yuzden yalnizca
        # alt sinir olarak ve kaynagi yazilarak kullanilir.
        try:
            from hrma.export.openrocket_integration import OpenRocketExporter
            inert, inert_src = OpenRocketExporter.resolve_inert_mass(
                motor_results or {})
        except Exception:
            inert, inert_src = None, 'none'
        if inert is not None:
            dry = inert
            sources['dry_mass'] = f'motor_inert_lower_bound:{inert_src}'
        else:
            sources['dry_mass'] = 'not_supplied'
    spec['dry_mass'] = dry

    # Cap: sayfa referans ALANI (mm^2) soruyor, cap sormuyor. Donusum TEK
    # noktada, sunucuda yapilir; istemciye formul kopyalanmaz.
    area_m2 = _positive_float(data.get('vehicle_reference_area_m2'))
    diameter = _positive_float(data.get('vehicle_diameter'))
    if diameter is not None:
        sources['diameter'] = 'request'
    elif area_m2 is not None:
        diameter = math.sqrt(4.0 * area_m2 / math.pi)
        sources['diameter'] = 'reference_area'
    else:
        # Alt sinir: govde en azindan motor kasasini icermek zorunda. Bu da
        # olculmus bir sayidir (kasa ic capi + 2 x cidar), uydurma degil.
        # Alternatif — her dalin kendi varsayilanina dusmesi — tam olarak
        # duzeltmeye calistigimiz hataydi: openrocket 0,10 m, yorunge 0,15 m.
        try:
            from hrma.export.openrocket_integration import OpenRocketExporter
            # resolve_geometry ORNEK metodu (resolve_inert_mass ise statik);
            # statik cagrilirsa TypeError verip sessizce yedege duser.
            geo = OpenRocketExporter().resolve_geometry(motor_results or {})
            case_d = _positive_float((geo or {}).get('case_diameter'))
            case_src = (geo or {}).get('case_diameter_source') or 'unknown'
        except Exception:
            case_d, case_src = None, 'unknown'
        if case_d is not None:
            diameter = case_d
            sources['diameter'] = f'motor_case_lower_bound:{case_src}'
        else:
            sources['diameter'] = 'not_supplied'
    spec['diameter'] = diameter
    spec['reference_area_m2'] = area_m2

    cd = _positive_float(data.get('drag_coefficient'))
    sources['drag_coefficient'] = 'request' if cd is not None else 'not_supplied'
    spec['drag_coefficient'] = cd

    # Govde boyu icin sayfada alan YOK; uydurma bir boy uretmek yerine bos
    # birakilir (surukleme modeli boyu kullanmiyor, OpenRocket XML'i ise
    # kullaniciya birakilir).
    length = _positive_float(data.get('vehicle_length'))
    sources['length'] = 'request' if length is not None else 'not_modelled'
    spec['length'] = length

    spec['sources'] = sources
    return spec


#: Arayuzun enjektor malzemesi secenekleri -> materials_db kayit anahtarlari.
#: Arayuz "AISI 316 Stainless Steel" / "Titanium Grade 5" / "Brass (Low
#: Pressure)" yaziyor; materials_db kayitlari ss_316 / titanium_6al4v /
#: brass_c360. Esleme olmadan secim cozulemez (v2.6.25'te 'steel_304' ->
#: 'ss_304' esleme eksikligi ayni sinifta bir hataydi).
INJECTOR_MATERIAL_ALIASES = {
    'stainless_steel': 'ss_316',
    'stainless': 'ss_316',
    'titanium': 'titanium_6al4v',
    'brass': 'brass_c360',
}


def _injector_plate_report(material_name, data, motor_results,
                           injector_results):
    """Enjektor plakasinin secilen malzemeyle yapisal raporu.

    Girdiler cozucunun KENDI ciktilarindan alinir (plaka capi = yanma odasi
    capi, delik sayisi/capi enjektor cozumunden, basinc farki enjektor
    basinc dususunden). Eksik olan bir sey varsa uydurulmaz; rapor
    'not_analyzed' doner ve nedeni yazar.
    """
    from hrma.data.materials_db import get_material
    from hrma.utils.injector_design import injector_plate_structural

    raw = str(material_name or '').strip().lower().replace(' ', '_')
    key = INJECTOR_MATERIAL_ALIASES.get(raw, raw)
    try:
        props = get_material(key)
    except (KeyError, ValueError):
        return {'status': 'not_analyzed',
                'reason': (f"injector material '{material_name}' does not "
                           f"resolve to a materials_db record"),
                'requested': material_name}

    # Plaka capi = yanma odasi capi [m]; cozucunun kendi geometrisi.
    d_ch = motor_results.get('chamber_diameter') if isinstance(
        motor_results, dict) else None
    # Basinc farki: enjektor cozumunun kullandigi deger [bar].
    dp = injector_results.get('pressure_drop') if isinstance(
        injector_results, dict) else None
    if not dp:
        dp = data.get('pressure_drop')
    # Delik alani bilgisi yalniz delikli tiplerde vardir.
    n_holes = injector_results.get('n_holes') or 0
    d_hole_mm = (injector_results.get('hole_diameter')
                 or injector_results.get('orifice_diameter') or 0)
    # Plaka kalinligi: kullanicinin verdigi deger (showerhead/impingement).
    t_plate_mm = injector_results.get('plate_thickness')

    report = injector_plate_structural(
        delta_P_bar=dp,
        plate_diameter_m=d_ch,
        material_props=props,
        material_name=props.get('name', key),
        plate_thickness_m=(t_plate_mm / 1000.0) if t_plate_mm else None,
        n_holes=int(n_holes or 0),
        hole_diameter_m=(float(d_hole_mm) / 1000.0) if d_hole_mm else 0.0,
        required_sf=data.get('safety_factor'),
    )
    report['material_key'] = key
    report['material_requested'] = material_name
    return report


#: Parametrik taramada izin verilen en fazla adim. Her adim TAM bir motor
#: cozumudur (Cantera dengesi dahil); 50 adim saniyeler, 1000 adim dakikalar
#: surer. Sinir keyfi degil: arayuzun kendi max="50" beyaninin iki kati,
#: yani elle girilen makul bir tarama reddedilmez ama surec donmaz.
PARAMETRIC_MAX_STEPS = 100

#: Parametrik taramada KESİNLİKLE pozitif olması gereken süpürme
#: parametreleri. ÖLÇÜM (Faz 4, A9): ``param_type='of_ratio'`` ile
#: ``param_start=-2.0`` gönderildiğinde uç HTTP 200 ve Isp 204.77 s
#: döndürüyordu — negatif karışım oranı fiziksel olarak yok, ama çözücü
#: sayı üretiyor ve hiçbir katman itiraz etmiyordu. Kütle debisi, yoğunluk,
#: süre, basınç ve sıcaklık aynı sınıfta.
PARAMETRIC_POSITIVE_PARAMS = frozenset({
    'of_ratio', 'chamber_pressure', 'chamber_temperature', 'thrust',
    'burn_time', 'total_impulse', 'l_star', 'fuel_density', 'gas_constant',
    'throat_diameter', 'chamber_diameter_input', 'regression_a',
})

#: Sıfır "otomatik/serbest" anlamına geldiği için sıfıra izin verilir, ama
#: negatif değer geometrik olarak tanımsızdır.
PARAMETRIC_NON_NEGATIVE_PARAMS = frozenset({
    'expansion_ratio', 'thrust_coefficient', 'atmospheric_pressure',
})

#: Parametrik taramanın TOPLAM İŞ BÜTÇESİ [s] (Faz 5B / B12).
#:
#: ``PARAMETRIC_MAX_STEPS`` bir ADIM sınırıdır, SÜRE sınırı değil — ve süre
#: makineden makineye değişir. ÖLÇÜLDÜ (HEAD 9d3728e, bu makine):
#:   20 adım (varsayılan) -> 13,3 s ; 60 adım -> 40,4 s ; 100 adım -> 67,8 s
#: yani izin verilen en büyük tarama tek bir iş parçacığını ~68 saniye
#: tutuyordu. İş kuyruğu ya da iptal yok; waitress ``threads=8`` altında
#: bunlar doğrudan bloke istektir.
#:
#: Bütçe İKİ yerde uygulanır:
#:   1) ERKEN REDDETME — ilk nokta çözüldükten sonra ÖLÇÜLEN nokta maliyeti
#:      adım sayısıyla çarpılır; öngörü bütçeyi aşıyorsa istek ~1 saniyede
#:      422 ile döner (68 saniye beklemek yerine). Öngörü uydurma değil,
#:      kullanıcının kendi makinesinde ölçülen değerdir.
#:   2) SERT DURDURMA — döngü içinde geçen süre bütçeyi aşarsa tarama
#:      kesilir; yarım eğri "başarılı tam tarama" gibi yayımlanmaz.
#:
#: 30 s, arayüzün varsayılanı olan 10 adımın (~6,7 s) dört katından fazla
#: yer bırakır. Daha uzun taramayı isteyen ``time_budget_s`` gönderir —
#: yetenek kaybı yok, ama sessiz 68 saniyelik donma da yok.
PARAMETRIC_TIME_BUDGET_DEFAULT_S = 30.0

#: ``time_budget_s`` ile istenebilecek en büyük bütçe [s]. Bunun üstü
#: senkron bir HTTP isteğinde savunulamaz.
PARAMETRIC_TIME_BUDGET_MAX_S = 300.0


def _parametric_point_rejection(sweep_param, value):
    """Bu süpürme noktası fiziksel olarak geçerli mi?

    Geçerliyse ``None``, değilse makine-okur ret nedeni döner. Geçersiz nokta
    HESAPLANMAZ: sayı üretip "başarılı" demek, kullanıcıya var olmayan bir
    motorun performansını vermektir.
    """
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return 'not_a_number'
    if not math.isfinite(numeric):
        return 'not_finite'
    if sweep_param in PARAMETRIC_POSITIVE_PARAMS and numeric <= 0:
        return 'must_be_positive'
    if sweep_param in PARAMETRIC_NON_NEGATIVE_PARAMS and numeric < 0:
        return 'must_be_non_negative'
    if sweep_param == 'gamma' and numeric <= 1.0:
        return 'gamma_must_exceed_one'
    return None


def _request_trace_id():
    """Bu istek için kısa korelasyon kimliği — loga GÖVDE yerine bu gider.

    D7 (2026-08-02): ``/calculate_solid`` ve ``/calculate_liquid``
    ``print("... motor data received:", data)`` ile TAM istek gövdesini
    stdout'a basıyordu. Başlatıcı stdout'u ``Documents/HRMA/hrma_log.txt``
    dosyasına yönlendiriyor (``packaging/launcher.py:79``) ve destek paketi
    o dosyayı içine koyup kullanıcıya "bunu bize gönderin" diyor (:651).
    Yani kullanıcının motor tasarımının tamamı — itki, karışım oranı,
    geometri, malzeme, yakıt seçimi — bir destek dosyasıyla dışarı
    çıkabiliyordu. Gizli tasarım verisi hata ayıklama için gerekli değildir;
    gerekli olan, kullanıcının bildirdiği olayı log satırıyla
    EŞLEŞTİREBİLMEKTİR. Bu yüzden loga yalnız kararlı bir olay adı ve bu
    tahmin edilemez kimlik yazılır; aynı kimlik hata yanıtında da döner ki
    kullanıcı destek talebinde onu verebilsin.
    """
    return uuid.uuid4().hex[:12]


_LOOPBACK_HOSTNAMES = frozenset({'localhost', '::1'})

#: Sunucunun GERÇEKTEN bağlandığı port. Başlatıcı ``_pick_port()`` sonucunu
#: ``HRMA_SELF_PORT`` ortam değişkenine yazar (8080-8090 arasında değişebilir).
#: Köken kapısı bunu bildiğinde 'geri döngü olsun yeter' yerine 'benim portum
#: olsun' diyebilir. Ayarlı değilse (doğrudan ``python -m hrma.app``) kapı
#: gevşek davranır — geliştirme akışı bozulmasın.
try:
    app.config['HRMA_SELF_PORT'] = int(os.environ['HRMA_SELF_PORT'])
except (KeyError, ValueError, TypeError):
    app.config['HRMA_SELF_PORT'] = None


def _host_and_port(value, is_url=False):
    """('127.0.0.1', 8081) döndürür; ayrıştırılamazsa (None, None).

    ``is_url=False`` iken ``Host`` başlığı biçimi ('127.0.0.1:8081') beklenir;
    ``urlsplit`` bunu ancak '//' ön ekiyle netloc olarak ayrıştırır.
    """
    if not value:
        return (None, None)
    try:
        from urllib.parse import urlsplit
        parts = urlsplit(value if is_url else '//' + value)
        return (parts.hostname, parts.port)
    except (ValueError, TypeError):
        # Geçersiz port ('...:abc') .port erişiminde ValueError atar.
        return (None, None)


def _is_loopback(hostname):
    """Ana makine adı bu bilgisayarın kendisini mi gösteriyor?"""
    if not hostname:
        return False
    name = hostname.lower()
    if name in _LOOPBACK_HOSTNAMES:
        return True
    try:
        import ipaddress
        return ipaddress.ip_address(name).is_loopback
    except ValueError:
        return False


@app.before_request
def _reject_cross_origin():
    """Yabancı kökenden gelen durum değiştiren istekleri reddeder.

    Tarayıcı, çapraz kökenli isteklerde ``Origin`` başlığını kendisi ekler ve
    sayfa bunu değiştiremez; dolayısıyla bu denetim CSRF ve DNS-rebinding
    saldırılarını kapatır. ``Origin`` yoksa (curl, aynı köken GET, native
    webview) istek geçer — masaüstü kullanımını bozmamak için.

    v2.6.25 DÜZELTMESİ — SAHA HATASI (uygulama hiç hesap yapmıyordu):
    Burada eskiden sabit bir liste vardı: ``{127.0.0.1:8080, localhost:8080,
    127.0.0.1:5000, localhost:5000}``. Oysa masaüstü başlatıcısı
    (``packaging/launcher.py::_pick_port``) **8080-8090 arasında BOŞ port
    arar**: 8080 meşgulse uygulama 8081'e düşer. O durumda arayüz
    ``http://127.0.0.1:8081`` kökeninden servis edilir, tarayıcı her POST'a
    ``Origin: http://127.0.0.1:8081`` ekler, liste bunu tanımaz ve
    **uygulamanın kendi sayfası kendi API'sinden 403 alır**. Yani Hesapla
    düğmesi hiçbir motor tipinde çalışmaz. Sabit liste ile dinamik port seçimi
    iki ayrı dosyada durduğu, geliştirme makinesinde de 8080 hep boş olduğu
    için bu yerelde hiç görünmedi.

    Doğrusu portu sabitlemek değil, **kökenin geri döngü olmasını** şart
    koşmak: uzaktaki bir sayfanın (evil.example) kökeni asla geri döngü
    olamaz, DNS-rebinding denemesinde ise ``Host`` başlığı saldırganın alan
    adını taşır ve aşağıdaki ilk kapıya takılır.

    v2.6.26 DÜZELTMESİ — OKUMA İSTEKLERİ KAPININ DIŞINDAYDI:
    Burada eskiden ``if request.method in ('GET','HEAD','OPTIONS'): return None``
    satırı vardı ve **Host kapısından önce** geldiği için okuma isteklerinde
    hiçbir denetim çalışmıyordu. Ölçüldü: ``Host: evil.example:PORT`` başlıklı
    düz bir GET ile ``/api/projects`` proje listesini, ``/api/projects/load/<ad>``
    ise tam tasarım belgesini (yanma basıncı, alev sıcaklığı, kasa malzemesi,
    burn-rate katsayıları) döndürüyordu. DNS-rebinding tam olarak bu şekilde
    çalışır ve Host kapısı zaten bunun için yazılmıştı. Artık:
      * Host kapısı HER metodda çalışır (rebinding'e karşı tek savunma),
      * Origin kapısı yalnız durum değiştiren metodlarda (çapraz köken GET
        zaten CORS başlığı olmadığı için tarayıcıda okunamıyor).
    """
    def _reject(reason):
        return jsonify({
            'status': 'error',
            'error': 'Cross-origin request rejected',
            'reason': reason,
        }), 403

    # 1) DNS-rebinding kapısı: sunucu yalnız 127.0.0.1'e bağlı olduğu için
    #    meşru istekte Host mutlaka geri döngüdür. Saldırgan alan adını
    #    127.0.0.1'e çözümlerse Host 'evil.example:8081' olarak gelir.
    #    HER metodda çalışır: okuma uçları da tasarım verisi döndürüyor.
    self_host, _self_port = _host_and_port(request.host)
    if not _is_loopback(self_host):
        return _reject('host_not_loopback')

    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return None

    # 2) Köken kapısı (yalnız durum değiştiren metodlar): başlık varsa geri
    #    döngü olmalı. Port da bilindiği takdirde karşılaştırılır — aksi hâlde
    #    127.0.0.1'in BAŞKA bir portundaki sayfa (yerel geliştirme sunucusu,
    #    Jupyter, kötü niyetli yerel uygulama) bize CSRF yapabiliyordu.
    #    HRMA_SELF_PORT'u başlatıcı doldurur; doldurulmamışsa (doğrudan
    #    ``python -m hrma.app`` ile geliştirme) eski geri-döngü davranışına
    #    düşülür ki v2.6.25'teki "kendi sayfası 403 alıyor" hatası tekrarlanmasın.
    raw_origin = request.headers.get('Origin')
    if raw_origin:
        origin_host, origin_port = _host_and_port(raw_origin, is_url=True)
        if not _is_loopback(origin_host):
            # 'null' kökeni de buraya düşer (sandbox iframe, file://).
            return _reject('origin_not_loopback')
        self_port = app.config.get('HRMA_SELF_PORT')
        if self_port and origin_port and int(origin_port) != int(self_port):
            return _reject('origin_port_mismatch')
    return None


#: Gövde kapısının kendi okuma bütçesi. Bunun ÜSTÜNDEKİ istekler kapıya
#: uğramaz: ``/api/import/ork`` (20 MiB), ``/api/import/motor-file`` (2 MiB)
#: ve ``/api/projects/save`` kendi boyut kapılarını ``request.content_length``
#: ile ÖNCE çalıştırıyor; gövdeyi burada okumak o kapıları etkisiz bırakırdı.
#: Kapının kapattığı vakalar (bozuk JSON, skaler, dizi) küçük gövdelerdir.
_JSON_BODY_GATE_MAX_BYTES = 1024 * 1024


@app.before_request
def _reject_malformed_json_body():
    """JSON gövdesi bozuk / sözlük değilse 500 yerine 400 döndürür.

    Faz 5B / H3-B7 ÖLÇÜMÜ (HEAD 9d3728e; 65 uç tarandı):
    bozuk JSON ``{"a":`` 47 uçta, JSON skaler ``42`` 54 uçta, JSON dizi ``[]``
    23 uçta **HTTP 500** üretiyordu; boş gövde de aynı sınıftaydı. Tüm 500'lerin
    124 tanesi (261 içinde) yalnızca bu BİÇİMSEL nedendendi. İki kök neden vardı
    ve ikisi de tek tek uçlarda değil, ortak desende:

      1. ``request.json`` bozuk gövdede ``werkzeug.BadRequest`` atar; uçların
         ``except Exception`` bloğu bunu yakalayıp 500'e çevirir. Yanıt gövdesi
         komik biçimde 400 metnini taşıyordu ama HTTP kodu 500'dü.
      2. Gövde sözlük değilse ``data.get(...)`` →
         ``AttributeError: 'int' object has no attribute 'get'`` → 500.

    ``app.py`` içinde 62 ``request.json``/``get_json`` kullanımı var, yalnız
    7'si ``silent=True``. Kapıyı 55 yere tek tek yazmak yerine buraya, tek
    yere kondu: gövde ayrıştırılamıyorsa ya da JSON nesnesi (sözlük) değilse
    istek uca hiç ulaşmaz.

    Kapsam bilinçli olarak dar:

    * Yalnız gövde taşıyan metodlar (POST/PUT/PATCH).
    * Yalnız ``Content-Type`` JSON ise — ``multipart/form-data`` yükleme
      uçlarına (``/api/validation/upload-csv``, ``/api/import/*``,
      ``/api/step/import``) DOKUNULMAZ.
    * Yalnız ``_JSON_BODY_GATE_MAX_BYTES`` altındaki gövdeler; büyükler kendi
      boyut kapılarına gider.

    Ret kodları makine-okur: ``empty_json_body`` / ``malformed_json_body`` /
    ``body_not_an_object``. Boş gövde de burada kapanır: RFC 8259'a göre boş
    dize geçerli JSON DEĞİLDİR ve ölçümde 5 uçta 500 üretiyordu.
    """
    if request.method not in ('POST', 'PUT', 'PATCH'):
        return None
    if not request.is_json:
        return None
    # ``content_length`` yoksa (chunked gövde, ya da hiç gövde yok) kapı yine
    # çalışır: o durumda uçların KENDİ boyut kapıları da zaten devre dışıdır
    # (hepsi ``content_length is not None`` şartına bağlı), üst sınırı
    # ``MAX_CONTENT_LENGTH`` (32 MiB) verir. Bu dal olmadan "boş gövde" vakası
    # kapıdan kaçıyordu — ölçüldü: 5 uçta 500 kalmıştı.
    length = request.content_length
    if length is not None and length > _JSON_BODY_GATE_MAX_BYTES:
        return None

    raw = request.get_data(cache=True)
    if not raw.strip():
        return jsonify({
            'status': 'error',
            'error': 'empty_json_body',
            'message': ('The request declared Content-Type: application/json '
                        'but the body was empty. Send a JSON object.'),
        }), 400
    try:
        parsed = json.loads(raw)
    except (ValueError, UnicodeDecodeError):
        return jsonify({
            'status': 'error',
            'error': 'malformed_json_body',
            'message': ('The request body is not valid JSON and could not be '
                        'parsed. No analysis was attempted.'),
        }), 400
    if not isinstance(parsed, dict):
        return jsonify({
            'status': 'error',
            'error': 'body_not_an_object',
            'message': ('The request body must be a JSON object (a mapping of '
                        'field names to values); received '
                        + type(parsed).__name__ + '.'),
        }), 400
    return None


# --- Örnek proje tohumlama (ilk açılış) -------------------------------------
# Kurulu üründe kullanıcı proje dizini ilk açılışta BOŞTUR; kullanıcıyı boş
# formla bırakmamak için paketle gelen üç örnek .hrma projesi BİR KEZ
# kopyalanır. Mantık hrma/utils/projects.py::seed_examples içindedir: damga
# dosyası (.seeded_v1) sayesinde kullanıcı örnekleri silse de geri gelmezler,
# var olan dosya asla ezilmez. Gerçek giriş noktaları (hrma/run.py,
# packaging/launcher.py) yalnız 'hrma.app'i import ettiği için kanca burada,
# İLK İSTEK anında çalışır — modül import'unda DEĞİL: testler bu modülü
# import eder ve import yan etkisi gerçek kullanıcı dizinine yazardı.
_example_seed_state = {'done': False}
_EXAMPLE_SEED_LOCK = threading.Lock()


@app.before_request
def _seed_examples_once():
    """İlk istekte örnek projeleri kullanıcı proje dizinine tohumlar.

    pytest koşusunda atlanır: test istemcisi de before_request kancalarını
    çalıştırır ve HRMA_PROJECTS_DIR ayarlamayan onlarca test gerçek
    ~/Documents/HRMA/projects dizinine yazmış olurdu. Kancanın kendi bekçisi
    (tests/test_ornek_tohumlama.py) bu bayrağı kaldırıp tmp dizinle sınar.
    """
    if _example_seed_state['done']:
        return None
    if 'PYTEST_CURRENT_TEST' in os.environ:
        return None
    with _EXAMPLE_SEED_LOCK:
        if _example_seed_state['done']:
            return None
        # Bayrak, deneme BAŞARISIZ olsa da kalkar: her istekte diske vurmak
        # olmaz; bir sonraki uygulama açılışı zaten yeniden dener.
        _example_seed_state['done'] = True
        try:
            from hrma.utils import projects as _project_store
            info = _project_store.seed_examples()
            if info['status'] == 'seeded' and info['copied']:
                app.logger.info('Example projects seeded: %s',
                                ', '.join(info['copied']))
        except Exception as seed_error:  # tohumlama isteği ASLA kırmaz
            app.logger.warning('Example project seeding failed: %s', seed_error)
    return None


#: Hata gövdesindeki tek bir metin alanının üst sınırı (karakter).
#: ÖLÇÜLDÜ (ast taraması, ``hrma/app.py``): 300 karakteri geçen tek satırlık
#: tek bir ileti sabiti var — ``app.py:5438``, 343 karakter. 2000 bunun beş
#: katından fazla, yani meşru hiçbir ileti kırpılmaz.
_ERROR_TEXT_MAX_CHARS = 2000


def _clip_echo(value, limit=200):
    """Kullanıcı girdisini yanıta yankılarken kırpar; kırpmayı beyan eder.

    ``_clip_error_body`` yalnız HTTP >= 400 gövdelerini tarar; başarı
    gövdesindeki yankılar (ölçüldü: ``/api/get-fuel-properties`` → ``note``,
    100 049 karakter, HTTP 200) buradan geçer.
    """
    text = str(value)
    if len(text) <= limit:
        return text
    return (text[:limit]
            + f'... [truncated by HRMA: {limit} of {len(text)} characters]')


def _clip_long_strings(node, limit, state):
    """Hata gövdesindeki aşırı uzun metinleri kırpar; kırpmayı BEYAN eder."""
    if isinstance(node, dict):
        return {k: _clip_long_strings(v, limit, state) for k, v in node.items()}
    if isinstance(node, list):
        return [_clip_long_strings(v, limit, state) for v in node]
    if isinstance(node, str) and len(node) > limit:
        state['clipped'] = True
        return (node[:limit]
                + f'... [truncated by HRMA: {limit} of {len(node)} characters]')
    return node


@app.after_request
def _clip_error_body(response):
    """Hata yanıtındaki devasa metinleri kırpar (Faz 5B / H3-B13).

    ÖLÇÜM (HEAD 9d3728e): 100 000 karakterlik tek bir alan gönderildiğinde
    ``/calculate`` **200 109 baytlık** bir 400 gövdesi döndürüyordu — girdi
    hem ``details[0]`` (doğrulayıcı iletisi) hem ``motor_type`` yankısı olarak
    iki kez kopyalanmıştı. Aynı gövde Faz 4/D7'de ölçülen günlük ve destek
    paketi kanalına da gidiyor.

    Kırpma yalnız **hata** yanıtlarında (HTTP >= 400) çalışır: başarı
    gövdelerinde çizim/plot JSON'u meşru olarak on binlerce karakter
    olabiliyor (ölçüldü: ``/api/quick-geometry`` → ``plots.motor``, 27 238
    karakter) ve onlara dokunmak veri kaybı olurdu.

    Kırpılan metin sessizce kısaltılmaz; sonuna kaç karakterin atıldığını
    söyleyen açık bir beyan eklenir. Kırpma olmadıysa gövde bayt bayt
    DEĞİŞTİRİLMEZ.
    """
    if response.status_code < 400 or response.direct_passthrough:
        return response
    if not response.is_json:
        return response
    body = response.get_json(silent=True)
    if body is None:
        return response
    state = {'clipped': False}
    clipped = _clip_long_strings(body, _ERROR_TEXT_MAX_CHARS, state)
    if state['clipped']:
        response.set_data(json.dumps(clipped, sort_keys=True))
    return response


@app.after_request
def _add_security_headers(response):
    """Yerel arayüze uygulanabilir asgari güvenlik başlıkları.

    Abartılmadı: bu uygulama yalnız 127.0.0.1'e bağlı bir masaüstü penceresi
    olduğu için HSTS, COEP/COOP gibi başlıklar tören olurdu. Gerçekten iş
    gören iki tanesi burada:
      * ``X-Frame-Options`` / ``frame-ancestors``: uygulamanın sayfası başka
        bir sayfanın iframe'ine gömülüp tıklama-hırsızlığına alet edilmesin.
      * ``X-Content-Type-Options``: tarayıcı içerik tipini tahmin etmesin.
    """
    response.headers.setdefault('X-Frame-Options', 'DENY')
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('Referrer-Policy', 'no-referrer')
    response.headers.setdefault('Content-Security-Policy', "frame-ancestors 'none'")
    return response

# v2.5.5 modüler API'ler: dış format importu (.eng/.rse/.ork), STEP/CAD
# geometri analizi ve proje kaydet/yükle deposu Blueprint olarak yaşar
# (rotalar kendi dosyalarında tam yollarıyla tanımlı, çakışma yok).
from hrma.importers.api import importers_api
from hrma.importers.step_api import step_import_api
from hrma.utils.projects_api import projects_api
app.register_blueprint(importers_api)
app.register_blueprint(step_import_api)
app.register_blueprint(projects_api)

# Apply Windows-specific Flask configurations
if platform.system() == 'Windows':
    try:
        if 'windows_compat' in globals():
            windows_compat.fix_flask_configuration(app)
            print("Windows Flask configurations applied")
    except Exception as e:
        print(f"Could not apply Windows Flask fixes: {e}")

#: Bir istek boyunca kaç sonlu-olmayan değerin null'a çevrildiğini sayar.
#: Sunucu günlüğüne yazmak için; yanıt gövdesini etkilemez.
_non_finite_seen = {'count': 0}


def sanitize_json_values(obj):
    """JSON'a çevrilemeyen değerleri (NaN, Inf, NumPy dizileri) temizler.

    v2.6.2 düzeltmesi — SESSİZ VERİ BOZULMASI:
    Bu fonksiyon eskiden ``NaN → 0.0`` ve ``Inf → ±1e10`` dönüşümü yapıyordu.
    Neredeyse TÜM API yanıtlarının son filtresi olduğu için (50 çağrı yeri),
    hesabın içinde oluşan her sayısal hata kullanıcıya **geçerli bir ölçüm**
    gibi görünüyordu: sıfıra bölme, negatif karekök veya ıraksayan bir çözücü
    ekranda "0.00" olarak beliriyordu.

    Girdi tarafı 2026-07-23'te kapatılmıştı (``_reject_non_finite``, NaN/Inf
    girdi HTTP 400 alır) ama ÇIKTI tarafı açık kalmıştı — yani kullanıcının
    verdiği sayılar temizken bile hesabın ürettiği NaN sızıyordu.

    Bu, tek başına da bir zincirin son halkasıydı: dört üretim modülü
    ``warnings.filterwarnings('ignore')`` çağırıyor (argümansız çağrı SÜREÇ
    GENELİNDE catch-all filtre kurar), bu yüzden NaN'ı üreten numpy uyarısı da
    bastırılıyordu. Zincir: sayısal hata → uyarı yok → NaN → 0.0 → panelde
    gerçek sayı.

    Yeni davranış: sonlu olmayan değer ``None`` (JSON ``null``) döner.
    Ön yüzün sayı biçimleyicisi (``analysis_dock.js::fmt``) null ve sonlu
    olmayan değerleri zaten "—" olarak gösterir, yani kullanıcı eksik olanı
    eksik görür. Dönüşüm sunucu günlüğüne de sayılır.
    """
    if isinstance(obj, dict):
        sanitized = {}
        for k, v in obj.items():
            try:
                sanitized[str(k)] = sanitize_json_values(v)
            except Exception:
                sanitized[str(k)] = "serialization_error"
        return sanitized
    elif isinstance(obj, (list, tuple)):
        sanitized = []
        for item in obj:
            try:
                sanitized.append(sanitize_json_values(item))
            except Exception:
                sanitized.append("serialization_error")
        return sanitized
    elif isinstance(obj, np.ndarray):
        try:
            return sanitize_json_values(obj.tolist())  # Convert NumPy array to list
        except Exception:
            return "numpy_array_error"
    elif isinstance(obj, (np.integer, np.floating)):
        try:
            val = float(obj)  # Convert NumPy numbers to Python numbers
            if not math.isfinite(val):
                # NaN/Inf -> null. 0.0 döndürmek çözücü hatasını geçerli bir
                # ölçüm gibi gösterirdi (bkz. fonksiyon docstring'i).
                _non_finite_seen['count'] += 1
                return None
            return val
        except Exception:
            return None
    elif isinstance(obj, float):
        if not math.isfinite(obj):
            _non_finite_seen['count'] += 1
            return None
        return obj
    elif isinstance(obj, (int, bool, str, type(None))):
        return obj
    else:
        # Handle any other types by converting to string
        try:
            return str(obj)
        except Exception:
            return "unknown_type"

def _reject_non_finite(value, name):
    """NaN ve sonsuz değerleri reddeder.

    KARARLILIK DENETİMİ BULGUSU (2026-07-23): IEEE-754'te NaN ile yapılan HER
    karşılaştırma False döner, bu yüzden `value < min or value > max` kontrolü
    NaN'ı SESSİZCE geçiriyordu. Sonuç: /calculate_solid'e chamber_pressure=NaN
    gönderildiğinde yanıt HTTP 200 ve "successful" oluyor, ama 124 çıktı alanı
    0.0 dönüyor ve kesit grafiği sessizce kayboluyordu. Kullanıcı hesabın
    çöktüğünü anlamıyordu.
    """
    try:
        v = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{name} must be a number, given: {value!r}")
    if math.isnan(v) or math.isinf(v):
        raise ValueError(f"{name} must be a finite number, given: {value}")
    return v


def validate_input_range(value, min_val, max_val, name):
    """Validate input values within physical limits"""
    value = _reject_non_finite(value, name)
    if value < min_val or value > max_val:
        raise ValueError(f"{name} value must be between {min_val}-{max_val}, given: {value}")
    return True

def validate_positive(value, name):
    """Positive value check"""
    value = _reject_non_finite(value, name)
    if value <= 0:
        raise ValueError(f"{name} must be positive, given: {value}")
    return True


def _collect_unphysical_fields(data, positive=(), non_negative=(),
                               ranges=None, finite=()):
    """Fiziksel olarak imkânsız sayısal girdileri makine-okur biçimde toplar.

    Faz 5B ortak kapısı. Üç uç aynı sınıf hatayı üretiyordu ve üçünde de kök
    neden aynıydı: alan "verilmiş" sayılıyor, ama İŞARETİ/SONLULUĞU hiç
    denetlenmiyor. Ölçülen örnekler (HEAD ``9d3728e``):

    * ``/analyze_thermal_safety`` — bütün sayılar ``"NaN"`` iken ``risk_level``
      **HIGH'dan LOW'a** dönüyordu (IEEE-754'te NaN ile yapılan her
      karşılaştırma ``False``, bu yüzden hiçbir risk dalı girmiyor).
    * ``/analyze_structural_safety`` — ``chamber_length=0`` ve
      ``throat_diameter=0`` ile HTTP 200 + tam yapısal hüküm; dört zorunlu
      alanın hepsi ``-1`` iken de HTTP 200.
    * ``/api/six-dof-analysis`` — ``cd0=-1`` (negatif sürükleme) HTTP 200 ve
      4277 m/s tepe hız; ``body_diameter=0`` HTTP 500 (sıfıra bölme).

    Sözleşme, ``hrma/utils/input_guard.py``'nin ilkesiyle aynı: **``0`` ile
    "verilmedi" ASLA karıştırılmaz.** Bu yüzden ``None``/``''`` (verilmedi)
    burada hiç incelenmez — eksik alan kapısı ayrıdır; yalnızca GÖNDERİLMİŞ
    değerler denetlenir.

    Args:
        data: İstek sözlüğü.
        positive: Kesinlikle ``> 0`` olması gereken alan adları.
        non_negative: ``>= 0`` olması gereken alan adları (0 = "yok/serbest").
        ranges: ``{alan: (alt, ust)}`` kapalı aralık denetimi.
        finite: İşareti serbest ama sonlu olmak zorunda olan alan adları
            (ör. ``launch_altitude`` — Lut Gölü eksi rakımdadır).

    Returns:
        Ret listesi; her öğe ``{'field', 'reason', 'value'}``. Boş liste =
        gönderilen sayıların hepsi fiziksel olarak mümkün.
        ``reason`` kodları: ``not_a_number`` | ``not_finite`` |
        ``must_be_positive`` | ``must_be_non_negative`` | ``out_of_range``.
    """
    problems = []

    def _numeric(key):
        """(hata_sozlugu, deger) — hata varsa deger ``None``."""
        raw = data.get(key)
        if raw is None or raw == '':
            return None, None
        if isinstance(raw, bool):
            # ``True`` Python'da 1.0'a çevrilir; bir uzunluk/kütle alanında
            # bu sessiz bir tip hatasıdır, sayı değildir.
            return {'field': key, 'reason': 'not_a_number',
                    'value': _clip_echo(raw, 60)}, None
        try:
            value = float(raw)
        except (TypeError, ValueError):
            return {'field': key, 'reason': 'not_a_number',
                    'value': _clip_echo(raw, 60)}, None
        if not math.isfinite(value):
            return {'field': key, 'reason': 'not_finite',
                    'value': _clip_echo(raw, 60)}, None
        return None, value

    for key in positive:
        problem, value = _numeric(key)
        if problem is not None:
            problems.append(problem)
        elif value is not None and value <= 0:
            problems.append({'field': key, 'reason': 'must_be_positive',
                             'value': value})

    for key in non_negative:
        problem, value = _numeric(key)
        if problem is not None:
            problems.append(problem)
        elif value is not None and value < 0:
            problems.append({'field': key, 'reason': 'must_be_non_negative',
                             'value': value})

    for key, (low, high) in (ranges or {}).items():
        problem, value = _numeric(key)
        if problem is not None:
            problems.append(problem)
        elif value is not None and not (low <= value <= high):
            problems.append({'field': key, 'reason': 'out_of_range',
                             'value': value,
                             'allowed_range': [low, high]})

    for key in finite:
        problem, _value = _numeric(key)
        if problem is not None:
            problems.append(problem)

    return problems


def build_time_history(motor_results):
    """Gerçek zaman serilerinden dashboard time_history sözlüğü kurar.

    OPUS/keşif düzeltmesi: eski kod motor_results['time_history'] okuyordu —
    böyle bir anahtar hiç üretilmiyor, dashboard'un alt 3 paneli hep boş
    kalıyordu. Gerçek seriler port_history'de (Euler marşından, ~200 nokta).

    Dönen şema (create_real_time_dashboard beklentisi):
      {'time': [s], 'propellant_mass': [kg], 'burn_rate': [mm/s],
       'port_diameter': [mm]}
    Yakıt tüketimi D² oranıyla ölçeklenir (m_f·(D²−D0²)/(Df²−D0²)) —
    grain geometrisi kütle bütçesiyle aynı kaynaktan, ek anahtar gerekmez.
    """
    ph = (motor_results or {}).get('port_history') or {}
    t = ph.get('time')
    D = ph.get('port_diameter')
    if not t or not D or len(t) < 3 or len(t) != len(D):
        return None
    t = np.asarray(t, dtype=float)
    D = np.asarray(D, dtype=float)

    m_ox = float(motor_results.get('oxidizer_mass', 0.0) or 0.0)
    m_f = float(motor_results.get('fuel_mass', 0.0) or 0.0)
    t_b = float(motor_results.get('burn_time', t[-1]) or t[-1])
    D0, Df = D[0], D[-1]

    ox_consumed = m_ox * np.clip(t / max(t_b, 1e-9), 0.0, 1.0)
    denom = max(Df ** 2 - D0 ** 2, 1e-12)
    fuel_consumed = m_f * np.clip((D ** 2 - D0 ** 2) / denom, 0.0, 1.0)
    propellant_mass = (m_ox + m_f) - ox_consumed - fuel_consumed

    # Yanma hızı: r = (dD/dt)/2 [m/s] → mm/s
    burn_rate = np.gradient(D, t) / 2.0 * 1000.0

    return {
        'time': t.tolist(),
        'propellant_mass': np.maximum(propellant_mass, 0.0).tolist(),
        'burn_rate': np.maximum(burn_rate, 0.0).tolist(),
        'port_diameter': (D * 1000.0).tolist(),  # mm
    }

# Initialize database manager and trajectory analyzer
db_manager = DatabaseManager()
# v2.6.26 — BU TEKIL NESNE ARTIK ISTEK YOLUNDA KULLANILMIYOR.
# TrajectoryAnalyzer DURUM TASIR (arac ve kurtarma parametreleri).
# Modul duzeyinde paylasilinca bir istekte verilen paraşut bir sonraki
# istege siziyordu: birebir ayni istek %53 farkli inis hizi donduruyor
# ve `_assumed` bayragi uydurulmus degeri 'kullanici verdi' diye
# isaretliyordu (olculdu). Cozum semptomu yamamak degil: her istek
# kendi nesnesini kurar. Kurucu yalniz skaler atama yapar, maliyeti yok.
# Boylece `set_recovery_parameters` programatik sozlesmesi de bozulmaz
# (deger cagrilar arasi korunur — tek nesne icinde).
trajectory_analyzer = TrajectoryAnalyzer()  # yalniz geriye uyumluluk
openrocket_exporter = OpenRocketExporter()
cad_designer = MotorCADDesigner()

@app.context_processor
def inject_app_version():
    # Şablonlar sürümü tek kaynaktan gösterir (hrma/__init__.py)
    from hrma import __version__
    return {'app_version': __version__}

@app.route('/favicon.ico')
def favicon():
    """Tarayıcının kendiliğinden istediği /favicon.ico yolu.

    Dosya hrma/static/favicon.ico'da duruyordu ama kök yol tanımlı olmadığı
    için her sayfa yüklemesinde 404 üretiliyordu (uzun süredir bilinen
    kozmetik hata, 2026-07-23'te kapatıldı). Simge, uygulama ikonunun
    (packaging/icon_1024.png) çok boyutlu ICO türevidir.
    """
    return send_from_directory(
        os.path.join(app.root_path, 'static'), 'favicon.ico',
        mimetype='image/vnd.microsoft.icon')


# ---- Kullanma kılavuzu (uygulamayla birlikte gelen PDF) ----
# Kılavuz iki dilde hazırlanır ve hrma/static/docs altında paketlenir, böylece
# internet olmadan da açılır. Arayüz bu ucu çağırır; PDF sistemin kendi
# görüntüleyicisinde açılır (pywebview penceresi PDF göstermez).

USER_GUIDE_FILES = {
    'tr': 'HRMA-Kullanma-Kilavuzu-TR.pdf',
    'en': 'HRMA-User-Guide-EN.pdf',
}
USER_GUIDE_URL = 'https://github.com/berketez/HRMA/tree/main/docs/user_guide'


def _user_guide_path(lang):
    """Dile uygun kılavuz dosyasının yolu (yoksa None).

    Dosya adı istemciden GELMEZ: yalnızca yukarıdaki sabit eşlemeden seçilir,
    dolayısıyla dizin gezinme (path traversal) mümkün değildir.
    """
    name = USER_GUIDE_FILES.get('tr' if str(lang).lower().startswith('tr') else 'en')
    path = os.path.join(app.root_path, 'static', 'docs', name)
    return path if os.path.isfile(path) else None


@app.route('/api/user-guide/open', methods=['POST'])
def user_guide_open():
    lang = request.args.get('lang', 'en')
    path = _user_guide_path(lang)
    if not path:
        # Kılavuz paketlenmemişse (kaynaktan çalışma) çevrimiçi sürüme yönlendir
        return jsonify({'opened': False, 'url': USER_GUIDE_URL})
    try:
        import subprocess
        if sys.platform == 'darwin':
            subprocess.Popen(['open', path])
        elif os.name == 'nt':
            os.startfile(path)  # noqa: yalnızca Windows'ta var
        else:
            subprocess.Popen(['xdg-open', path])
        return jsonify({'opened': True, 'path': path})
    except Exception as exc:
        return jsonify({'opened': False, 'url': USER_GUIDE_URL,
                        'error': '%s: %s' % (type(exc).__name__, exc)})


@app.route('/api/user-guide/status')
def user_guide_status():
    """Arayüz, kılavuz paketlenmişse bağlantıyı gösterir."""
    return jsonify({
        'available': {lang: bool(_user_guide_path(lang))
                      for lang in USER_GUIDE_FILES},
        'url': USER_GUIDE_URL,
    })


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/hybrid')
def hybrid():
    return render_template('advanced.html')

@app.route('/solid')
def solid():
    return render_template('solid.html')

@app.route('/liquid')
def liquid():
    return render_template('liquid.html')

@app.route('/formulas')
def formulas():
    return render_template('formulas.html')

@app.route('/launch-site')
def launch_site_page():
    """İnteraktif 3B Dünya küresi: fırlatma sahası seçimi + uçuş yolu
    animasyonu (2026-07-23). Fizik hrma/analysis/launch_site.py'de."""
    return render_template('launch_site.html')

@app.route('/test')
def test():
    return jsonify({'status': 'ok', 'message': 'HRMA is running'})

# ---- Otomatik güncelleme (GitHub Releases) ----
# Arayüz açılışta /api/update/check'i çağırır; yeni sürüm varsa modal gösterir.
# İndirme URL'si istemciden alınmaz (bkz. hrma/utils/update_checker.py).

@app.route('/api/update/check')
def update_check():
    from hrma.utils.update_checker import check_for_update
    return jsonify(check_for_update())

@app.route('/api/update/download', methods=['POST'])
def update_download():
    from hrma.utils.update_checker import start_download
    return jsonify(start_download())

@app.route('/api/update/status')
def update_status():
    from hrma.utils.update_checker import download_status
    return jsonify(download_status())

@app.route('/api/update/install', methods=['POST'])
def update_install():
    # Sessiz otomatik kurulum: indirme bittikten sonra arayüz burayı çağırır.
    # "auto" modda uygulama kendini kapatır, yardımcı betik kurulumu yapıp
    # HRMA'yı yeniden başlatır; kurulamayan ortamlarda kurulum dosyası
    # açılır ("manual"). Dosya yolu istemciden alınmaz (self_install.py).
    from hrma.utils.update_checker import start_install
    return jsonify(start_install())

@app.route('/api/update/open-download', methods=['POST'])
def update_open_download():
    # Manuel/yedek indirme: uygulama içi indirme yavaş/başarısız olursa
    # (GitHub CDN yavaşlığı, ağ kısıtı) kullanıcı buradan sistem tarayıcısında
    # doğrudan asset URL'sini (yoksa Releases sayfasını) açar. Sunucu tarafı
    # webbrowser.open olduğu için pywebview/exe, Chromium ve her tarayıcıda
    # aynı çalışır — istemci ortamına bağımlı değil.
    from hrma.utils.update_checker import open_download_in_browser
    return jsonify(open_download_in_browser())

@app.route('/api/changelog')
def changelog():
    # Sürüm notları: paketle gelen hrma/data/changelog.json okunur
    # (GitHub Releases gövdelerinden derlenir; ağ gerektirmez, çevrimdışı
    # çalışır). Yol, hrma/data/offline_store.py ile aynı kalıpla paket
    # köküne göre çözülür — kaynak kurulumda da paketli kurulumda da aynı.
    # Dosya yoksa/bozuksa 500 atılmaz, boş liste döner (arayüz mesaj basar).
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        'data', 'changelog.json')
    try:
        with open(path, encoding='utf-8') as fh:
            data = json.load(fh)
        versions = data.get('versions', []) if isinstance(data, dict) else []
        if not isinstance(versions, list):
            versions = []
    except (OSError, ValueError):
        versions = []
    return jsonify({'versions': versions})

@app.route('/test-simple')
def test_simple():
    return '<h1>SIMPLE TEST</h1><p>If you see this, Flask is working!</p><a href="/">Home Page</a>'

#: Gönderilen ile kullanılan değeri "aynı" saymak için bağıl tolerans.
#: Yalnız yuvarlama/kayan nokta farkını yutar; 2500 K yerine 1681 K
#: kullanıldığında ayrım net biçimde görünür.
_INPUT_ECHO_REL_TOL = 1e-6


def _declare_overridden_inputs(data, motor_results, fields):
    """Gönderilmiş ama çözücünün ÜZERİNE YAZDIĞI girdileri beyan eder.

    B10 ÖLÇÜMÜ (HEAD ``9d3728e``) — ``/calculate`` (hibrit),
    ``chamber_temperature`` dışında her şey sabit:

    ======================  =====  ==================================
    gönderilen Tc           HTTP   sonucun sha256 imzası
    ======================  =====  ==================================
    (yok)                   200    ``d6fa7ad9d8df4f7d`` (referans)
    2500                    200    **bit-aynı**
    3500                    200    **bit-aynı**
    0                       200    **bit-aynı**
    -3000                   200    **bit-aynı**
    ======================  =====  ==================================

    Yanıt her durumda ``motor.chamber_temperature = 1681,73`` diyordu —
    yani kullanıcının verdiği sayı atılıyor ve yerine çözücünün bulduğu
    değer, "bu senin girdin değil" demeden geri veriliyordu. Üstelik aynı
    değer ``openrocket.flight_profile.motor_data`` ve
    ``trajectory.motor_data`` içinde de yankılanıyor.

    Sessizce yok sayılan girdi olmaz: alan modele bağlanamıyorsa
    KULLANILMADIĞI açıkça yazılır. Beyan VARSAYIMLA değil ÖLÇÜMLE üretilir —
    gönderilen sayı ile sonuçta fiilen duran sayı karşılaştırılır; eşitse
    hiçbir şey beyan edilmez.

    Returns:
        Beyan listesi; her öğe ``{'field', 'submitted', 'used_by_model',
        'reason', 'message'}``. Değer gerçekten kullanılmışsa boş liste.
    """
    declarations = []
    if not isinstance(motor_results, dict):
        return declarations
    for field in fields:
        raw = data.get(field)
        if raw is None or raw == '':
            continue
        try:
            submitted = float(raw)
        except (TypeError, ValueError):
            continue
        used = motor_results.get(field)
        try:
            used_value = float(used)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(submitted) or not math.isfinite(used_value):
            continue
        if math.isclose(submitted, used_value,
                        rel_tol=_INPUT_ECHO_REL_TOL, abs_tol=0.0):
            continue
        declarations.append({
            'field': field,
            'submitted': submitted,
            'used_by_model': used_value,
            'reason': 'solved_by_model',
            'message': (f"'{field}' was supplied as {submitted:g} but the "
                        f'solver computes this quantity itself; the analysis '
                        f'and every echo of this field in the response use '
                        f'{used_value:g}.'),
        })
    return declarations


#: ``/calculate`` içinde çözücünün üzerine yazdığı, bu yüzden beyan edilen
#: girdiler. Hibrit motorda hazne sıcaklığı yanma çözümünün SONUCUDUR
#: (``hybrid_rocket_engine.py`` içinde ``self.T_c`` denge/ampirik çözümle
#: yeniden atanır), girdisi değil.
#:
#: v2.6.27 (A3, Ayberk madde 2) — ``burn_time`` eklendi. ÖLÇÜLDÜ (bu dosya
#: değişmeden önce): ``{thrust: 500, burn_time: 20, total_impulse: 7500}``
#: gönderildiğinde yanıtta ``burn_time = 15,0 s`` dönüyordu, hiçbir uyarı
#: yoktu ve ``defaults_used`` boştu. Sebep ``hybrid_rocket_engine.py``
#: kurucusunda: toplam impuls VE itki verilmişse süre
#: ``t_b = I_total / F`` ile ÇÖZÜLÜR, kullanıcının girdiği süre hiç
#: okunmaz. Sayı doğru, sunum sessizdi: kullanıcının 20 s'i yok oluyordu.
#: Alan buraya girince mevcut ``_declare_overridden_inputs`` beyanı üretir;
#: süre gerçekten kullanıldığında (çakışma yoksa) hiçbir şey yazılmaz.
_CALCULATE_SOLVER_OWNED_FIELDS = ('chamber_temperature', 'burn_time')


@app.route('/calculate', methods=['POST'])
def calculate():
    try:
        data = request.json

        # Determine motor type (default to hybrid for this endpoint)
        motor_type = data.get('motor_type', 'hybrid')

        # Use comprehensive validator
        is_valid, validation_messages = motor_validator.validate_motor_data(data, motor_type)
        if not is_valid:
            return jsonify({
                'error': 'Validation failed',
                'details': validation_messages,
                'motor_type': motor_type,
                'status': 'validation_error'
            }), 400

        # Log warnings but continue
        if validation_messages:
            app.logger.info(f"Validation warnings: {validation_messages}")

        # P-6 (scratchpad/perf_audit_v262.md [P-6]): 10 Plotly figürünün
        # üretimi istek süresinin yarısından fazlasıydı (ölçülen: 674 ms'nin
        # ~%54'ü). include_plots=false gönderen istemci figür üretimini
        # TOPTAN atlar; alan verilmezse davranış birebir eskisidir (tam figür
        # seti). Bekçi + ölçüm: tests/test_istek_performansi.py.
        include_plots = bool(data.get('include_plots', True))
        # P-5 (perf_audit_v262 [P-5]): ?slim=1 sorgu parametresi ön yüzün hiç
        # okumadığı ağır alanları düşürür — sonuç sözlüğü kurulduktan sonra,
        # aşağıda uygulanır. Varsayılan (slim'siz) davranış değişmez.
        slim = request.args.get('slim') == '1'

        # --- v2.6.27 (A3 / Ayberk madde 3): EKSİK GİRDİ ≠ VARSAYILAN GİRDİ --
        # Ölçülen kırık zincir (bu satırlar değişmeden önce): uçtan gelen HER
        # istekte ``_mm_to_m`` varsayılanlı biçimiyle çağrılıp bir sayı
        # döndürüyordu (o fonksiyonun sözleşmesi gereği ASLA None dönmez), bu
        # yüzden motorda ``wall_thickness_user_supplied`` DAİMA True oluyor,
        # yapısal modül DOĞRULAMA moduna geçiyor ve rapor
        # "verified against user-supplied wall thickness" diyordu. Kullanıcı
        # hiçbir yapısal bilgi vermemişken sistem onun cidarına karşı
        # doğrulandığını söylüyordu — kazanılmamış hüküm.
        # Motorun kendi yorumu (hybrid_rocket_engine.py:450-458) zaten doğru
        # sözleşmeyi yazıyordu: "değer verilmediyse termal model yine 5 mm ile
        # çalışır ama yapısal modüle None geçilir". Eksik olan tek şey, uç
        # katmanının bu None'ı üretmesiydi.
        #
        # Malzemede de aynı desen vardı: ``or 'steel_4130'`` seçim yapılmamış
        # bir isteği sessizce çeliğe çeviriyordu. Motorun
        # ``_resolve_chamber_material`` işlevi None girdide zaten çeliğe düşer;
        # fark, artık bunun BEYAN EDİLMESİ (aşağıda structural_design_basis).
        chamber_material_input = data.get('chamber_material') or None
        wall_thickness_input_m = _mm_to_m_optional(data.get('wall_thickness'))

        # Create engine instance with support for total impulse
        # Only pass user-provided values, let the engine use fuel-specific defaults
        engine = HybridRocketEngine(
            # v2.6.27 parti 28: chug atalet kapıları — form kapıları
            # advanced.html'de, motor _feed_line_inertance_inputs ikiziyle
            # okur (uzunluk m, iç çap mm; bant dışı girdi _defaults_used
            # beyanıyla düşer). Geçilmezse chug çevrimi ataletsiz koşar ve
            # bunu beyan eder — sessiz varsayılan yok. Bekçiler:
            # tests/test_stability_hibrit_chug.py (skip-armed uçtan uca) +
            # tests/test_field_wiring_layer_b.py (ölü girdi taraması).
            feed_line_length_m=data.get('feed_line_length_m'),
            feed_line_inner_diameter_mm=data.get(
                'feed_line_inner_diameter_mm'),
            thrust=data.get('thrust'),
            burn_time=data.get('burn_time'),
            total_impulse=data.get('total_impulse'),
            of_ratio=data.get('of_ratio', 1.0),
            chamber_pressure=data.get('chamber_pressure', 20.0),
            atmospheric_pressure=data.get('atmospheric_pressure', 1.0),
            chamber_temperature=data.get('chamber_temperature'),  # None if not provided
            gamma=data.get('gamma', 1.25),
            gas_constant=data.get('gas_constant'),  # None if not provided
            l_star=data.get('l_star', 1.0),
            expansion_ratio=data.get('expansion_ratio', 0),
            nozzle_type=data.get('nozzle_type', 'conical'),
            thrust_coefficient=data.get('thrust_coefficient', 0),
            regression_a=data.get('regression_a'),  # None if not provided
            regression_n=data.get('regression_n'),  # None if not provided
            fuel_density=data.get('fuel_density'),  # None if not provided
            combustion_type=data.get('combustion_type', 'infinite'),
            chamber_diameter_input=data.get('chamber_diameter_input', 0),
            # v2.6.26: contraction_ratio arayuzden GONDERILIYOR ama buraya
            # hic gecirilmiyordu - motorda olcum sonucu sifir yaprak
            # degisimiydi, yani alan tamamen oluydu.
            # v2.6.27 (A8): fırlatma sahası motora GEÇMİYORDU — motor tarafı
            # hazır ve testliydi ama API her koşuda NOT_MODELLED dönüyordu.
            # Sözleşme: resolve_launch_site() çıktısı ya da en az
            # {'elevation_m'} / {'latitude_deg','longitude_deg'} taşıyan sözlük.
            launch_site=data.get('launch_site'),
            contraction_ratio=data.get('contraction_ratio', 0),
            # v2.6.26 (P2 devri): enjektor plaka kalinligi ve ortam sicakligi
            # formda VARDI ama motora hic gecmiyordu.
            #   - plate_thickness: orifis L/D'sini, dolayisiyla Cd'yi belirler
            #     (olculdu: 1 mm -> Cd 0,630 ; 20 mm -> 0,840 ; enjeksiyon
            #     alani %43 bant). Gecmedigi icin Cd sabit 0,78 kaliyordu.
            #   - ambient_temperature: isi ve yapisal moduller ayni kosuda
            #     iki farkli ortam sicakligi kullaniyordu (293,15 K / 300 K).
            plate_thickness=_mm_to_m_optional(data.get('plate_thickness')),
            ambient_temperature=data.get('ambient_temp'),
            fuel_type=data.get('fuel_type', 'htpb'),
            oxidizer_type=data.get('oxidizer_type', 'n2o'),
            injector_type=data.get('injector_type', 'showerhead'),
            initial_port_diameter=data.get('initial_port_diameter') or None,
            initial_gox=data.get('mass_flux_chamber') or None,
            # v2.6.25: termal sinir kosullari. Bu uc alan sayfada VARDI ve
            # sunucuya GELIYORDU, ama motora hic gecirilmiyordu; hibrit isi
            # transferi cagrisi malzeme/kalinlik/sogutmayi SABIT yaziyordu
            # (gerekce: hybrid_rocket_engine.py analyze_heat_transfer yorumu).
            # wall_thickness arayuzde mm, motorda m.
            # v2.6.27 (A3): ikisi de artik VERILMEDIYSE None gider — gerekce
            # yukarida, engine cagrisindan once.
            chamber_material=chamber_material_input,
            wall_thickness=wall_thickness_input_m,
            cooling_type=data.get('cooling_channels') or 'none',
            # v2.6.26: Katman A ile bulunan üç ölü girdi. Motorun
            # gövdesi ve çözücüleri 30 Temmuz'da yazılmıştı; imza ve
            # buradaki geçirme yarım kalmıştı (oturum çöktü).
            # chamber_length_override arayüzde MİLİMETRE, motorda METRE.
            safety_factor=data.get('safety_factor'),
            chamber_length_override=_mm_to_m_optional(data.get('chamber_length_override')),
            nozzle_material=data.get('nozzle_material'),
            tank_temperature=data.get('oxidizer_temp') or None,
            motor_name=data.get('motor_name', ''),
            motor_description=data.get('motor_description', '')
        )

        # Calculate motor geometry and performance
        motor_results = engine.calculate()

        # 2026-08-03 (Faz 6, T11 kök nedeni): kullanıcının girdiği tank basıncı
        # motor sonucuna HİÇ yazılmıyordu. Ölçüldü: tank 30 / 50 / 90 bar ile
        # /calculate çağrıldığında 90+ anahtarlı motor sözlüğünde 'tank' geçen
        # tek bir anahtar yoktu. Değer çözücüde VARDI (aynı koşuda kavitasyon
        # uyarısı K_c = -0,01 diyor, bu ancak P1 = 50 bar ile çıkar) ama
        # yayımlanmıyordu. Sonuç: basınç dağılımı çubuğu 'Tank' etiketiyle
        # Pc + ΔP geri düşüşünü gösteriyordu — kullanıcı 90 bar girse de
        # ekranda 24 bar duruyordu.
        # BURAYA konuyor, sonuç sözlüğünün kurulduğu yere DEĞİL: grafikler
        # aşağıda (satır ~1520 create_performance_plots) üretiliyor ve
        # motor_results'ı o an okuyor. Sonradan eklenen alan çubuğa yansımaz —
        # ilk denemede tam bu hata yapıldı ve ölçümle yakalandı.
        # Alan geldiği anda görselleştirme tarafında yapılacak bir şey yok:
        # _perf_panels_hybrid zaten motor_data.get('tank_pressure') okuyor ve
        # gerçek değer gelince çubuğu 'Tank' diye adlandırıyor; gelmezse
        # dürüst 'Inj. inlet' etiketine düşüyor.
        _tank_p = data.get('tank_pressure')
        if _tank_p is not None:
            try:
                motor_results['tank_pressure'] = float(_tank_p)
                motor_results['tank_pressure_source'] = 'user_input'
            except (TypeError, ValueError):
                pass  # geçersiz girdi: alan hiç konmaz, çubuk dürüst etikete düşer

        # --- v2.6.27 (A3): kazanılmamış yapısal hüküm bu yola da bağlandı ---
        # Kapı BURADA çağrılır, yanıt sözlüğü kurulurken değil: aşağıdaki CAD,
        # ihracat ve PDF dalları motor_results'ı bu noktadan SONRA okuyor;
        # geri çekilen hüküm onlara da gitsin. (Kapı yalnız hüküm/gösterge
        # alanlarını değiştirir — SF, gerilme ve kalınlık sayıları yerinde
        # kalır, o yüzden çizim/kütle zinciri etkilenmez.)
        structural_design_basis = _motor_structural_design_basis(
            motor_results, wall_thickness_input_m, chamber_material_input)

        # Design injector
        injector = InjectorDesign(
            mdot_ox=motor_results['mdot_ox'],
            chamber_pressure=data['chamber_pressure'],
            oxidizer_phase=data.get('oxidizer_phase', 'liquid'),
            oxidizer_density=data.get('oxidizer_density', 1220),
            oxidizer_viscosity=data.get('oxidizer_viscosity', 0.0002),
            oxidizer_temp=data.get('oxidizer_temp', 293),
            oxidizer_type=data.get('oxidizer_type', 'n2o'),
            tank_pressure=data.get('tank_pressure', 50.0),
            pressure_drop=data.get('pressure_drop', 0),
            discharge_coefficient=data.get('discharge_coefficient', 0.7),
            injector_type=data.get('injector_type', 'showerhead')
        )
        
        # Add type-specific parameters
        if data.get('injector_type', 'showerhead') == 'showerhead':
            injector.set_showerhead_params(
                target_velocity=data.get('target_velocity', 30),
                n_holes=data.get('n_holes', 0),
                hole_diameter_min=data.get('hole_diameter_min', 0.3),
                hole_diameter_max=data.get('hole_diameter_max', 2.0),
                plate_thickness=data.get('plate_thickness', 3.0)
            )
        elif data.get('injector_type', 'showerhead') == 'pintle':
            injector.set_pintle_params(
                outer_diameter=data.get('outer_diameter', 50),
                pintle_diameter=data.get('pintle_diameter', 25),
                # v2.6.27: 'Secondary Holes' seçimi çözücüye bağlandı.
                secondary_holes=data.get('secondary_holes', 'radial')
            )
        elif data.get('injector_type', 'showerhead') == 'swirl':
            # v2.6.26: 'swirl_chamber_diameter' ve 'swirl_angle' arayuzde
            # VARDI ama depoda hicbir yerde okunmuyordu (Katman A: sifir
            # yaprak). Ikisi de Giffen-Muraszew cozumune baglandi:
            #   swirl_chamber_diameter -> K = A_p/(D_s*d_o) icindeki D_s
            #   swirl_angle            -> hedef sprey yari acisi; ters cozucu
            #                             (swirl_K_from_theta) K'yi ondan bulur
            # Ters cozucu bu depoda ZATEN yaziliydi ve arayuzdeki alan ona
            # baglanmamisti.
            injector.set_swirl_params(
                n_slots=data.get('n_slots', 6),
                slot_width=data.get('slot_width', 0),
                slot_height=data.get('slot_height', 0),
                chamber_diameter=data.get('swirl_chamber_diameter', 0) or 0,
                target_half_angle=data.get('swirl_angle', 0) or 0
            )
        elif data.get('injector_type', 'showerhead') == 'impingement':
            # 2026-07-22 denetim bulgusu: form 6 alan gönderiyordu, çözücüye
            # HİÇBİRİ ulaşmıyordu. Eşleşenler bağlandı; eşleşmeyenler aşağıda
            # 'unused_inputs' ile AÇIKÇA bildirilir (sessiz yutma yasak).
            d_orif = data.get('orifice_diameter') or 0
            injector.set_impingement_params(
                n_pairs=data.get('element_pairs', 0) or 0,
                impingement_angle=data.get('impingement_angle')
                or 2 * IMPINGEMENT_HALF_ANGLE_DEG,
                # Kullanıcı delik çapını sabitlediyse arama bandı o değere
                # kilitlenir; vermediyse model kendi bandında çözer.
                hole_diameter_min=d_orif if d_orif > 0 else 0.3,
                hole_diameter_max=d_orif if d_orif > 0 else 2.0,
            )
        elif data.get('injector_type', 'showerhead') == 'coaxial':
            # İç jet debi payı, kullanıcının verdiği geometriden TÜRETİLİR:
            # f_inner = A_ic / (A_ic + A_anulus). Oran ölçekten bağımsız
            # olduğu için modelin debi-alan çözümüyle tutarlı kalır.
            d_in = data.get('inner_diameter') or 0
            d_out = data.get('outer_annulus_diameter') or 0
            f_inner = None
            if d_in > 0 and d_out > d_in:
                a_in = d_in ** 2
                a_ann = d_out ** 2 - d_in ** 2
                f_inner = a_in / (a_in + a_ann)
            if f_inner is not None:
                injector.set_coaxial_params(inner_flow_fraction=f_inner)

        injector_results = injector.calculate()

        # 2.6.27 — hole_pattern bağlaması ("kanal var kapı yok" kapanışı).
        # UI'daki desen seçimi ne sunucuya geliyor ne çizime yansıyordu; çizen
        # taraf (create_improved_injector_design) injector_data['hole_pattern']
        # alanını SHOWERHEAD_PATTERNS ile zaten doğrulayıp çiziyor. Desen yalnız
        # plaka/CAD yerleşimini etkiler — performans modeli YOK, advanced.html'in
        # 'no_model' beyanı doğru kalır. Tek doğruluk kaynağı çizicinin kayıt
        # defteri; burada kopya desen listesi tutulmaz.
        _hole_pattern = data.get('hole_pattern')
        if _hole_pattern is not None:
            _hp = str(_hole_pattern).lower()
            if _hp not in SHOWERHEAD_PATTERNS:
                return jsonify({
                    'status': 'error',
                    'error': 'invalid_hole_pattern',
                    'message': ('hole_pattern must be one of: '
                                + ', '.join(SHOWERHEAD_PATTERNS)
                                + '; received ' + repr(_hole_pattern) + '.'),
                }), 400
            injector_results['hole_pattern'] = _hp

        # --- Tüketilmeyen enjektör girdileri AÇIKÇA raporlanır -------------
        # Bu projede sessizce yutulan girdi yoktur: modelin kabul etmediği ya
        # da SONUÇ olarak hesapladığı alanlar kullanıcıya gerekçesiyle döner.
        _inj_type = data.get('injector_type', 'showerhead')
        _unused = []
        # i18n sözleşmesi (2026-08-04): backend EN üretir; TR karşılıkları
        # i18n_charts.js sözlüğünde/MSG_PATTERNS'ta durur (serverText çevirir).
        if _inj_type == 'impingement':
            if data.get('impingement_distance'):
                _unused.append(('impingement_distance',
                                'the model COMPUTES this distance from the '
                                'impingement angle and hole diameter; see the '
                                'value in the results'))
            if data.get('momentum_ratio'):
                _unused.append(('momentum_ratio',
                                'this path models like-on-like doublets; the '
                                'momentum-ratio criterion applies to unlike '
                                'impingement'))
            _pat = data.get('impingement_pattern')
            if _pat and _pat not in ('like_on_like', 'doublet'):
                _unused.append(('impingement_pattern',
                                f"the '{_pat}' pattern is not modelled on "
                                'this path; use the Injector Design panel '
                                'for a full solution '
                                '(doublet/triplet/like/coax-swirl)'))
        elif _inj_type == 'coaxial':
            if data.get('recess_length'):
                _unused.append(('recess_length',
                                'the model COMPUTES the recess from the inner '
                                'jet diameter; see the value in the results'))
            if data.get('n_elements'):
                _unused.append(('n_elements',
                                'this path sizes a single coaxial element; '
                                'use the Injector Design panel for a '
                                'multi-element array'))
        if _unused and isinstance(injector_results, dict):
            injector_results['unused_inputs'] = [
                {'field': f, 'reason': r} for f, r in _unused]
            injector_results.setdefault('warnings', []).extend(
                f"Input '{f}' was not consumed: {r}" for f, r in _unused)

        # --- Enjektor plakasi yapisal kontrolu (v2.6.26) --------------------
        # 'injector_material' arayuzde bu surume kadar hicbir hesaba
        # ulasmiyordu (yalniz CAD malzeme listesine yaziliyordu). Artik secilen
        # malzemenin akma dayanimi ve yogunlugu plaka egilme gerilmesi,
        # emniyet katsayisi, gereken kalinlik ve plaka kutlesini uretir.
        # Model: kenarindan tutturulmus dairesel plaka (Roark Tablo 11.2 durum
        # 10b) + ASME PG-52 ligament verimi. Ayrinti: injector_design.py
        # icindeki injector_plate_structural docstring'i.
        if isinstance(injector_results, dict):
            _inj_mat_name = data.get('injector_material')
            if _inj_mat_name:
                _plate = _injector_plate_report(
                    _inj_mat_name, data, motor_results, injector_results)
                injector_results['plate_structural'] = _plate
                for _pw in (_plate.get('warnings') or []):
                    injector_results.setdefault('warnings', []).append(_pw)

        # Fizik limiti doğrulaması (ValidationSystem, Sutton & Biblarz +
        # NASA SP-8089 aralıkları): rapor üretilemezse /calculate ASLA
        # kırılmaz — hata loglanır, 'validation' anahtarı yanıttan atlanır.
        validation_report = None
        try:
            combo = f"{data.get('oxidizer_type', 'n2o')}_{data.get('fuel_type', 'htpb')}"
            if combo not in validator.performance_limits['specific_impulse']:
                combo = 'n2o_htpb'
            validation_report = validator.comprehensive_validation(
                motor_results, injector_results, combo
            )
        except Exception as val_error:
            app.logger.warning(f"Validation report skipped: {val_error}")

        # Create visualizations - Use improved visuals
        # P-6: include_plots=false ise figür fonksiyonları HİÇ çağrılmaz;
        # 'plots' aşağıda null olarak beyan edilir (boş figür uydurulmaz).
        motor_plot = None
        injector_plot = None
        performance_plots = None
        if include_plots:
            try:
                # New improved motor cross-section
                motor_plot = create_improved_motor_cross_section(motor_results)
            except Exception:
                # Fallback to old version if new one fails
                motor_plot = create_motor_plot(motor_results)

            try:
                # New improved injector design
                injector_plot = create_improved_injector_design(injector_results)
            except Exception:
                # Fallback to old version if new one fails
                injector_plot = create_injector_plot(injector_results, data['injector_type'])
            performance_plots = create_performance_plots(motor_results, injector_results)

        # plots.injector = enjektör tip şeması, plots.performance = dashboard.
        # (Eski davranış şemayı dashboard ile eziyordu; ayrı div'lerde
        # gösterilir — advanced.html #injector_plot / #performance_plots.)
        
        # Create advanced analysis visualizations
        #
        # Dalga 0 (2026-07-14): Isı ve yapısal analiz burada İKİNCİ kez
        # hesaplanıp plot'lanıyordu (plots.heat_transfer /
        # plots.structural_analysis) ama HİÇBİR şablon bu plot'ları render
        # etmiyordu — ~251 KB ölü yük + çifte hesap. Kaldırıldı. Motor
        # İÇİNDEKİ sonuçlar (motor.heat_transfer_analysis,
        # motor.structural_analysis) KALIR: 3D ısı haritası
        # (motor_viz3d.js) ve analiz panelleri onları okur.
        combustion_analysis_plot = None
        real_time_dashboard_plot = None
        motor_3d_plot = None

        # Generate combustion analysis
        # (combustion_data yalnız figüre gider; include_plots=false iken
        # analizörü kurmak da gereksiz — P-6.)
        if include_plots and data.get('include_combustion_analysis', True):
            try:
                from hrma.engines.combustion_analysis import CombustionAnalyzer
                combustion_analyzer = CombustionAnalyzer()
                fuel_composition = {data.get('fuel_type', 'htpb'): 100.0}
                # v2.6.26: OKSITLEYICI SABIT 'N2O' YAZILIYDI. Kullanicinin
                # sectigi oksitleyici sayisal panellere gidiyordu ama
                # "Combustion Analysis" grafigine GITMIYORDU: LOX secili bir
                # motorda grafik N2O alev sicakligini, N2O tur dagilimini
                # (LOX/HTPB'de kimyasal olarak imkansiz olan %31 N2 dahil) ve
                # N2O Isp egrisini ciziyordu. Uc oksitleyicinin grafik JSON'u
                # bayt-ayni cikiyordu (olculdu). Motorun kendi cozucusu bu
                # hatayi 2026-07 denetiminde duzeltmisti; app.py'deki iki
                # cagri duzeltmenin disinda kalmis.
                combustion_data = combustion_analyzer.analyze_combustion(
                    fuel_composition,
                    data.get('oxidizer_type', 'n2o'),
                    data.get('of_ratio', 1.0),
                    data.get('chamber_pressure', 20.0)
                )
                combustion_analysis_plot = create_combustion_analysis_plots(combustion_data)
            except Exception as e:
                print(f"Combustion analysis error: {e}")
        
        # (Yapısal analiz çifte hesabı da kaldırıldı — bkz. yukarıdaki not.)

        # Generate real-time dashboard
        if include_plots and data.get('include_realtime_dashboard', True):
            try:
                # Gerçek port_history serilerinden kur (eski 'time_history'
                # anahtarı hiç üretilmiyordu — alt 3 panel hep boştu)
                time_data = build_time_history(motor_results)
                real_time_dashboard_plot = create_real_time_dashboard(motor_results, time_data)
            except Exception as e:
                print(f"Real-time dashboard error: {e}")
        
        # Generate 3D visualization
        motor_3d_plot = None
        if include_plots and data.get('include_3d_visualization', True):
            try:
                motor_3d_plot = create_3d_motor_visualization(motor_results)
            except Exception as viz_error:
                print(f"3D visualization error: {str(viz_error)}")
                motor_3d_plot = {'error': f'3D visualization failed: {str(viz_error)}'}
        
        # Create advanced analysis results
        cea_style_results = create_cea_style_results(motor_results)
        
        # Create additional plots if data is available
        altitude_performance_plot = None
        mass_fractions_plot = None
        thrust_altitude_plot = None
        
        if include_plots and 'altitude_performance' in motor_results:
            altitude_performance_plot = create_altitude_performance_plot(
                motor_results['altitude_performance']['altitude_performance']
            )

        if include_plots and 'mass_fractions' in motor_results:
            mass_fractions_plot = create_mass_fractions_plot(motor_results['mass_fractions'])

        if include_plots and 'thrust_altitude_analysis' in motor_results:
            thrust_altitude_plot = create_thrust_altitude_plot(
                motor_results['thrust_altitude_analysis']['thrust_altitude_data']
            )
        
        # Arac ve firlatma kosulu TEK yerde cozulur; yorunge dali da
        # OpenRocket dali da AYNI sozlugu kullanir. Oncesinde her dal kendi
        # varsayilanina dusuyordu ve tek yanitta uc farkli arac cikiyordu
        # (bkz. _resolve_vehicle_spec aciklamasi).
        vehicle_spec = _resolve_vehicle_spec(data, motor_results)
        launch_params = {
            'launch_angle': data.get('launch_angle'),
            'launch_altitude': data.get('launch_altitude'),
            'wind_speed': data.get('wind_speed'),
            'wind_direction': data.get('wind_direction'),
            # v2.6.26 — kurtarma parametreleri bu sozluge HIC konmuyordu.
            # calculate_trajectory ucunu de kabul ediyor (trajectory_analysis
            # .py:365-378) ve set_recovery_parameters acik bir API (:137), ama
            # /calculate yolu onlari hic tasimadigi icin inis hizi ARACTAN
            # BAGIMSIZ sabit 2,0 m^2 / Cd 1,4'ten turuyordu. Olculdu:
            # parachute_area=9.0 gonderildiginde yanittaki 152 yapragin 0'i
            # degisiyordu. landing_velocity bir GUVENLIK metrigi olarak
            # okunuyor; kullanicinin onu duzeltememesi kabul edilemez.
            'parachute_area': data.get('parachute_area'),
            'parachute_cd': data.get('parachute_cd'),
            'parachute_deploy_delay': data.get('parachute_deploy_delay'),
        }
        # Verilmeyen anahtar SOZLUKTEN CIKAR: disa aktarici "istekten geldi"
        # ile "verilmedi" ayrimini bu sayede yapip beyan edebiliyor.
        launch_params = {k: v for k, v in launch_params.items()
                         if v is not None and v != ''}

        rocket_params = {
            'dry_mass': vehicle_spec['dry_mass'],
            'diameter': vehicle_spec['diameter'],
            'length': vehicle_spec['length'],
            'drag_coefficient': vehicle_spec['drag_coefficient'],
            'sources': vehicle_spec['sources'],
        }

        # Generate OpenRocket export data
        openrocket_data = {
            'eng_file': openrocket_exporter.export_motor_file(motor_results),
            'motor_summary': openrocket_exporter.export_motor_summary(motor_results) if hasattr(openrocket_exporter, 'export_motor_summary') else {},
            'flight_profile': openrocket_exporter.create_flight_simulation_data(
                motor_results, rocket_params, launch_params)
        }
        
        # Generate 3D CAD design
        cad_data = None
        if data.get('generate_cad', True):
            try:
                # v2.6.26 — EKRAN ILE CAD FARKLI ENJEKTOR GOSTERIYORDU.
                # Hibritte iki bagimsiz enjektor cozucusu kosuyor:
                #   1) InjectorDesign (utils) -> yanittaki 'injector' blogu.
                #      Kullanicinin ENJEKTOR PANELI girdileri (hedef hiz, delik
                #      sayisi ezmesi, basinc dususu ezmesi, Cd) buraya gider.
                #      Tablo ve 2B sema bunu gosterir.
                #   2) design_injector (engines) -> motor icindeki
                #      'injector_design'. Devre modeli; N2O'da tank basincini
                #      degil doyma basincini kullanir.
                # CAD/cizim/STEP/DXF ve 3B model (2)'yi okuyordu. Olculdu:
                # ayni kosuda tablo 32 delik x 0,887 mm (dP 4,00 bar) derken
                # teknik cizim 3 delik x 2,10 mm (dP 30,37 bar) diyordu;
                # enjeksiyon alani 1,9 kat, hiz 3,5 kat farkliydi. Kullanici
                # hangisini delerse delsin diger tum analizler yanlis
                # enjektore aitti.
                #
                # _injector_spec zaten 'injector_results' anahtarini ONCELIKLI
                # okuyacak sekilde yazilmis, ama /calculate o anahtari HIC
                # doldurmuyordu; kopukluk tam oradaydi. Panel sonucu artik
                # CAD'e de veriliyor, boylece ekran ile cizim ayni enjektoru
                # anlatir. Panel sonucu yoksa eski davranis (motorun kendi
                # cozumu) surer.
                cad_input = motor_results
                if isinstance(injector_results, dict) and injector_results:
                    cad_input = dict(motor_results)
                    cad_input['injector_results'] = injector_results
                cad_data = cad_designer.generate_3d_motor_assembly(cad_input)
                
                # Export STL files if requested
                if data.get('export_stl', False):
                    if cad_data and 'assembly_meshes' in cad_data:
                        stl_files = cad_designer.export_stl_files(cad_data['assembly_meshes'])
                        cad_data['exported_stl_files'] = stl_files
            except Exception as cad_error:
                print(f"CAD generation error: {str(cad_error)}")
                cad_data = {'error': f'CAD generation failed: {str(cad_error)}'}
        
        # Calculate trajectory if requested
        trajectory_data = None
        # Istek basina TAZE analizor (bkz. modul duzeyindeki not): durum
        # tasiyan bir nesneyi istekler arasinda paylasmak sizinti demektir.
        trajectory_analyzer = TrajectoryAnalyzer()
        if data.get('calculate_trajectory', True):
            # Arac parametreleri: OpenRocket dalinin kullandigi AYNI sozluk.
            # set_vehicle_parameters None kabul etmiyor (cross_sectional_area
            # hesabi TypeError verir), bu yuzden cozulemeyen alanda metot HIC
            # cagrilmaz ve cozucunun kendi belgelenmis varsayilanlari gecerli
            # kalir — uydurma bir sayi enjekte edilmez.
            if (vehicle_spec['dry_mass'] is not None
                    and vehicle_spec['diameter'] is not None):
                kwargs = {
                    'mass_dry': vehicle_spec['dry_mass'],
                    'diameter': vehicle_spec['diameter'],
                }
                if vehicle_spec['drag_coefficient'] is not None:
                    kwargs['drag_coefficient'] = vehicle_spec['drag_coefficient']
                if vehicle_spec['length'] is not None:
                    kwargs['length'] = vehicle_spec['length']
                trajectory_analyzer.set_vehicle_parameters(**kwargs)

            # Calculate trajectory
            try:
                trajectory_data = trajectory_analyzer.calculate_trajectory(motor_results, launch_params)
                # P-6: yörünge VERİSİ her koşulda hesaplanır (sunum değil,
                # sonuç); yalnız figürü include_plots'a bağlıdır.
                trajectory_plot = (
                    trajectory_analyzer.create_trajectory_plots(trajectory_data)
                    if include_plots else None)
            except Exception as traj_error:
                print(f"Trajectory calculation error: {str(traj_error)}")
                trajectory_data = {'error': f'Trajectory calculation failed: {str(traj_error)}'}
                trajectory_plot = None
        else:
            trajectory_plot = None
        
        # Combine results
        results = {
            'motor': motor_results,
            'injector': injector_results,
            'trajectory': trajectory_data,
            'cea_results': cea_style_results,
            'openrocket': openrocket_data,
            'cad_design': cad_data,
            # Design outputs from engine calculation
            'design_summary': motor_results.get('design_summary', {}),
            'nozzle_angles': motor_results.get('nozzle_angles', {}),
            'grain_design': motor_results.get('grain_design', {}),
            'injector_design': motor_results.get('injector_design', {}),
            'plots': {
                'motor': motor_plot,
                'injector': injector_plot,
                'performance': performance_plots,
                'trajectory': trajectory_plot,
                'altitude_performance': altitude_performance_plot,
                'mass_fractions': mass_fractions_plot,
                'thrust_altitude': thrust_altitude_plot,
                # 'heat_transfer' ve 'structural_analysis' plot anahtarları
                # bilinçli olarak KALDIRILDI (Dalga 0): hiçbir şablon render
                # etmiyordu; analiz panelleri motor.* sonuçlarını okur.
                'combustion_analysis': combustion_analysis_plot,
                'realtime_dashboard': real_time_dashboard_plot,
                'motor_3d': motor_3d_plot
            # P-6: figürler istenmediyse 'plots' null — anahtarların None ile
            # tek tek doldurulması "figür var ama boş" izlenimi verirdi.
            } if include_plots else None
        }

        # ValidationSystem raporu (UI kontratı: results.validation)
        if validation_report is not None:
            results['validation'] = validation_report

        # v2.6.27 (A3): yapısal hükmün DAYANAĞI. Alan adı ve içeriği
        # /analyze_structural_safety'nin 'design_basis' bloğuyla aynı
        # sözleşmededir; iki yol aynı soruyu aynı dille cevaplasın diye.
        if structural_design_basis is not None:
            results['structural_design_basis'] = structural_design_basis

        # Faz 5B / B10 — sessizce yok sayılan girdi olmaz. Gönderilen değer
        # ile sonuçta fiilen duran değer karşılaştırılır; farklıysa alanın
        # KULLANILMADIĞI beyan edilir (bkz. _declare_overridden_inputs).
        inputs_not_used = _declare_overridden_inputs(
            data, motor_results, _CALCULATE_SOLVER_OWNED_FIELDS)
        if inputs_not_used:
            results['inputs_not_used'] = inputs_not_used

        # P-5 (perf_audit_v262 [P-5]) — ?slim=1: ön yüzün HİÇ okumadığı iki
        # ağır alan düşürülür (ölçülen: trajectory.trajectory 597,8 KB ham
        # zaman serisi + trajectory.motor_data 70,5 KB, üst seviye 'motor'
        # ile bit-aynı kopya; tüketici taraması perf_audit_v262 [P-5]).
        # Yörünge ÖZETİ (apogee, evreler, kurtarma, uyarılar) yerinde kalır;
        # düşürülen alanlar yanıtta AÇIKÇA beyan edilir — sessiz kesinti yok.
        # Bu dalga yalnız altyapı + şema kilidi
        # (tests/test_istek_performansi.py); ön yüz geçişi sonraki dalga.
        if slim:
            results['slim'] = True
            if isinstance(trajectory_data, dict) and 'error' not in trajectory_data:
                slim_trajectory = {
                    k: v for k, v in trajectory_data.items()
                    if k not in ('trajectory', 'motor_data')}
                slim_trajectory['omitted_fields'] = {
                    'trajectory': ('omitted by slim=1: raw time series; '
                                   'the front end reads plots.trajectory'),
                    'motor_data': ("omitted by slim=1: bit-identical copy "
                                   "of the top-level 'motor' block"),
                }
                results['trajectory'] = slim_trajectory

        # Sanitize results to handle NaN and Infinity values
        try:
            sanitized_results = sanitize_json_values(results)

            # Test JSON serialization before returning
            test_json = json.dumps(sanitized_results, indent=2)

            return jsonify(sanitized_results)
            
        except (TypeError, ValueError) as json_error:
            print(f"JSON Serialization Error: {str(json_error)}")
            
            # Return basic results without problematic data
            basic_results = {
                'motor': {
                    'thrust': motor_results.get('thrust', 0),
                    'specific_impulse': motor_results.get('specific_impulse', 0),
                    'chamber_pressure': motor_results.get('chamber_pressure', 0),
                    'burn_time': motor_results.get('burn_time', 0)
                },
                'cea_results': cea_style_results if isinstance(cea_style_results, str) else "Calculation completed",
                'error_info': f"Full results had serialization issues: {str(json_error)}"
            }
            
            return jsonify(sanitize_json_values(basic_results))
        
    except Exception as e:
        error_traceback = traceback.format_exc()
        print(f"Error in calculate: {str(e)}")
        print(f"Traceback: {error_traceback}")
        # v2.6.26 — TRACEBACK VE ISTEK GOVDESI ARTIK YANITTA DEGIL.
        # Yanit tam Python traceback'ini (gelistirinin mutlak dosya yollari,
        # ic modul yapisi, satir numaralari) ve istegin TAMAMINI geri
        # veriyordu. Tanilama degeri sunucu gunlugunde korunur; istemciye
        # yalnizca hatanin kendisi gider.
        return jsonify({
            'error': str(e),
            'error_type': type(e).__name__
        }), 400

@app.route('/api/burn-rate/resolve', methods=['POST'])
def api_burn_rate_resolve():
    """Merkezi burn_rate_db rejim fitinden tasarım basıncında (a, n) çözer.

    Girdi:  {propellant: 'kndx'|'knsb', pressure_bar: float}
    Çıktı:  motor konvansiyonunda a-n (r[m/s] = a·P[bar]^n) + rejim aralığı,
            geçerlilik bayrağı ve kaynak künyesi. Katı sayfasındaki burn-rate
            preset dropdown'ı bu endpoint'le a/n alanlarını doldurur — böylece
            tasarım yolu ile korelasyon/doğrulama yolu AYNI merkezi katsayıyı
            kullanır (CLAUDE.md kural 11).
    """
    try:
        from hrma.data import burn_rate_db
        data = request.json or {}
        prop = str(data.get('propellant', '')).lower()
        if not burn_rate_db.has_law(prop):
            return jsonify({'status': 'error',
                            'error': f"No published burn-rate law for "
                                     f"'{prop}'. Available: "
                                     f"{sorted(burn_rate_db.BURN_RATE_LAWS)}"
                            }), 400
        p_bar = float(data.get('pressure_bar', 0))
        if not (0 < p_bar <= 1000):
            return jsonify({'status': 'error',
                            'error': 'pressure_bar must be in (0, 1000]'}), 400
        result = burn_rate_db.resolve_engine_coeffs(prop, p_bar)
        result['status'] = 'success'
        result['propellant'] = prop
        result['pressure_bar'] = p_bar
        return jsonify(result)
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 400


@app.route('/api/quick-geometry', methods=['POST'])
def quick_geometry():
    """İnteraktif tasarım modu: yalnız motor çözücüsü + 2D kesit.

    /calculate'in ağır adımları (yörünge, CAD, OpenRocket, tüm grafikler)
    atlanır; 3D dijital ikiz ile kesitin slider'la canlı güncellenmesi için
    ~1 sn içinde geometri döndürür. Motor sözlüğü /calculate'teki
    results['motor'] ile aynı şemadadır (port_history ve
    heat_transfer_analysis dahil).
    """
    try:
        data = request.json or {}

        # GİRDİ DOĞRULAMASI (v2.6.2) — bu uç motor doğrulayıcısını HİÇ
        # çağırmıyordu, dolayısıyla 1e6 bar gibi fiziksel olmayan değerler
        # sessizce çözücüye giriyor ve "başarılı" bir geometri dönüyordu.
        # Verilmemiş (None) alanlara dokunulmaz; motor sınıfı kendi
        # varsayılanını kurar.
        if data.get('chamber_pressure') is not None:
            validate_input_range(data['chamber_pressure'], 1, 500,
                                 "Chamber pressure (bar)")
        if data.get('thrust') is not None:
            validate_input_range(data['thrust'], 1, 1e7, "Thrust (N)")
        if data.get('burn_time') is not None:
            validate_input_range(data['burn_time'], 0.1, 1000, "Burn time (s)")
        if data.get('of_ratio') is not None:
            validate_input_range(data['of_ratio'], 0.1, 50, "O/F ratio")
        if data.get('chamber_temperature') is not None:
            validate_input_range(data['chamber_temperature'], 300, 6000,
                                 "Chamber temperature (K)")

        engine = HybridRocketEngine(
            thrust=data.get('thrust'),
            burn_time=data.get('burn_time'),
            total_impulse=data.get('total_impulse'),
            of_ratio=data.get('of_ratio', 1.0),
            chamber_pressure=data.get('chamber_pressure', 20.0),
            atmospheric_pressure=data.get('atmospheric_pressure', 1.0),
            chamber_temperature=data.get('chamber_temperature'),
            gamma=data.get('gamma', 1.25),
            gas_constant=data.get('gas_constant'),
            l_star=data.get('l_star', 1.0),
            expansion_ratio=data.get('expansion_ratio', 0),
            nozzle_type=data.get('nozzle_type', 'conical'),
            thrust_coefficient=data.get('thrust_coefficient', 0),
            regression_a=data.get('regression_a'),
            regression_n=data.get('regression_n'),
            fuel_density=data.get('fuel_density'),
            combustion_type=data.get('combustion_type', 'infinite'),
            chamber_diameter_input=data.get('chamber_diameter_input', 0),
            # v2.6.26: contraction_ratio arayuzden GONDERILIYOR ama buraya
            # hic gecirilmiyordu - motorda olcum sonucu sifir yaprak
            # degisimiydi, yani alan tamamen oluydu.
            # v2.6.27 (A8): fırlatma sahası motora GEÇMİYORDU — motor tarafı
            # hazır ve testliydi ama API her koşuda NOT_MODELLED dönüyordu.
            # Sözleşme: resolve_launch_site() çıktısı ya da en az
            # {'elevation_m'} / {'latitude_deg','longitude_deg'} taşıyan sözlük.
            launch_site=data.get('launch_site'),
            contraction_ratio=data.get('contraction_ratio', 0),
            # v2.6.26 (P2 devri): enjektor plaka kalinligi ve ortam sicakligi
            # formda VARDI ama motora hic gecmiyordu.
            #   - plate_thickness: orifis L/D'sini, dolayisiyla Cd'yi belirler
            #     (olculdu: 1 mm -> Cd 0,630 ; 20 mm -> 0,840 ; enjeksiyon
            #     alani %43 bant). Gecmedigi icin Cd sabit 0,78 kaliyordu.
            #   - ambient_temperature: isi ve yapisal moduller ayni kosuda
            #     iki farkli ortam sicakligi kullaniyordu (293,15 K / 300 K).
            plate_thickness=_mm_to_m_optional(data.get('plate_thickness')),
            ambient_temperature=data.get('ambient_temp'),
            fuel_type=data.get('fuel_type', 'htpb'),
            oxidizer_type=data.get('oxidizer_type', 'n2o'),
            injector_type=data.get('injector_type', 'showerhead'),
            initial_port_diameter=data.get('initial_port_diameter') or None,
            initial_gox=data.get('mass_flux_chamber') or None,
            # v2.6.25: termal sinir kosullari. Bu uc alan sayfada VARDI ve
            # sunucuya GELIYORDU, ama motora hic gecirilmiyordu; hibrit isi
            # transferi cagrisi malzeme/kalinlik/sogutmayi SABIT yaziyordu
            # (gerekce: hybrid_rocket_engine.py analyze_heat_transfer yorumu).
            # wall_thickness arayuzde mm, motorda m.
            # v2.6.27 (A3): eksik girdi ≠ varsayılan girdi. Gerekçe
            # /calculate içindeki aynı ikilinin üstündeki yorumda; buradaki
            # kopya da aynı kırık zinciri besliyordu (uçtan gelen her istek
            # yapısal modülü DOĞRULAMA moduna sokuyordu).
            chamber_material=(data.get('chamber_material') or None),
            wall_thickness=_mm_to_m_optional(data.get('wall_thickness')),
            cooling_type=data.get('cooling_channels') or 'none',
            # v2.6.26: uc olu girdi daha baglandi. Ucu de sayfada VARDI ve
            # (chamber_length_override haric) sunucuya bile geliyordu, ama
            # motora hic gecirilmiyordu: Katman A taramasinda sifir yaprak
            # degisimi olcusmustu. safety_factor -> yapisal analizin tasarim
            # SF hedefi; chamber_length_override -> L* ile turetilen kamara
            # boyunu ezer (arayuz mm, motor m); nozzle_material -> bogaz
            # termal + erozyon degerlendirmesi.
            safety_factor=data.get('safety_factor'),
            chamber_length_override=_mm_to_m_optional(
                data.get('chamber_length_override')),
            nozzle_material=data.get('nozzle_material'),
            tank_temperature=data.get('oxidizer_temp') or None,
            motor_name=data.get('motor_name', ''),
            motor_description=data.get('motor_description', '')
        )
        motor_results = engine.calculate()

        # v2.6.27 (A3): bu uç de motor sözlüğünün TAMAMINI döndürüyor, yani
        # yapısal hüküm de içinde. Aynı kapı buraya da bağlanır; aksi hâlde
        # kullanıcı aynı kazanılmamış "doğrulandı" cümlesini interaktif
        # tasarım modunda görürdü.
        quick_design_basis = _motor_structural_design_basis(
            motor_results,
            _mm_to_m_optional(data.get('wall_thickness')),
            data.get('chamber_material') or None)

        try:
            motor_plot = create_improved_motor_cross_section(motor_results)
        except Exception as plot_err:
            print(f"Quick geometry cross-section error: {plot_err}")
            motor_plot = None

        payload = {
            'status': 'success',
            'motor': motor_results,
            'plots': {'motor': motor_plot}
        }
        if quick_design_basis is not None:
            payload['structural_design_basis'] = quick_design_basis
        return jsonify(sanitize_json_values(payload))
    except Exception as e:
        print(f"Quick geometry error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 400


@app.route('/api/transient-analysis', methods=['POST'])
def transient_analysis():
    """Zaman-çözümlü iç balistik: gerçek Pc(t) ve F(t) eğrileri.

    Girdi: /calculate ile aynı motor parametreleri + opsiyonel:
      feed_mode: 'regulated' (varsayılan) | 'blowdown'
      tank_temperature: blowdown tank başlangıç sıcaklığı [K, vars. 293.15]
      liquid_fill_fraction: tank sıvı doluluk oranı [vars. 0.85]

    Çıktı: time/thrust/chamber_pressure/of_ratio/port_diameter dizileri,
    blowdown'da tank basınç-sıcaklık geçmişi, durdurma olayı ve uyarılar.
    """
    try:
        data = request.json or {}
        engine = HybridRocketEngine(
            thrust=data.get('thrust'),
            burn_time=data.get('burn_time'),
            total_impulse=data.get('total_impulse'),
            of_ratio=data.get('of_ratio', 1.0),
            chamber_pressure=data.get('chamber_pressure', 20.0),
            atmospheric_pressure=data.get('atmospheric_pressure', 1.0),
            chamber_temperature=data.get('chamber_temperature'),
            gamma=data.get('gamma', 1.25),
            gas_constant=data.get('gas_constant'),
            l_star=data.get('l_star', 1.0),
            expansion_ratio=data.get('expansion_ratio', 0),
            nozzle_type=data.get('nozzle_type', 'conical'),
            thrust_coefficient=data.get('thrust_coefficient', 0),
            regression_a=data.get('regression_a'),
            regression_n=data.get('regression_n'),
            fuel_density=data.get('fuel_density'),
            combustion_type=data.get('combustion_type', 'infinite'),
            chamber_diameter_input=data.get('chamber_diameter_input', 0),
            # v2.6.26: contraction_ratio arayuzden GONDERILIYOR ama buraya
            # hic gecirilmiyordu - motorda olcum sonucu sifir yaprak
            # degisimiydi, yani alan tamamen oluydu.
            # v2.6.27 (A8): fırlatma sahası motora GEÇMİYORDU — motor tarafı
            # hazır ve testliydi ama API her koşuda NOT_MODELLED dönüyordu.
            # Sözleşme: resolve_launch_site() çıktısı ya da en az
            # {'elevation_m'} / {'latitude_deg','longitude_deg'} taşıyan sözlük.
            launch_site=data.get('launch_site'),
            contraction_ratio=data.get('contraction_ratio', 0),
            # v2.6.26 (P2 devri): enjektor plaka kalinligi ve ortam sicakligi
            # formda VARDI ama motora hic gecmiyordu.
            #   - plate_thickness: orifis L/D'sini, dolayisiyla Cd'yi belirler
            #     (olculdu: 1 mm -> Cd 0,630 ; 20 mm -> 0,840 ; enjeksiyon
            #     alani %43 bant). Gecmedigi icin Cd sabit 0,78 kaliyordu.
            #   - ambient_temperature: isi ve yapisal moduller ayni kosuda
            #     iki farkli ortam sicakligi kullaniyordu (293,15 K / 300 K).
            plate_thickness=_mm_to_m_optional(data.get('plate_thickness')),
            ambient_temperature=data.get('ambient_temp'),
            fuel_type=data.get('fuel_type', 'htpb'),
            oxidizer_type=data.get('oxidizer_type', 'n2o'),
            injector_type=data.get('injector_type', 'showerhead'),
            initial_port_diameter=data.get('initial_port_diameter') or None,
            initial_gox=data.get('mass_flux_chamber') or None,
            # v2.6.25: termal sinir kosullari. Bu uc alan sayfada VARDI ve
            # sunucuya GELIYORDU, ama motora hic gecirilmiyordu; hibrit isi
            # transferi cagrisi malzeme/kalinlik/sogutmayi SABIT yaziyordu
            # (gerekce: hybrid_rocket_engine.py analyze_heat_transfer yorumu).
            # wall_thickness arayuzde mm, motorda m.
            # v2.6.27 (A3): eksik girdi ≠ varsayılan girdi. Gerekçe
            # /calculate içindeki aynı ikilinin üstündeki yorumda; buradaki
            # kopya da aynı kırık zinciri besliyordu (uçtan gelen her istek
            # yapısal modülü DOĞRULAMA moduna sokuyordu).
            chamber_material=(data.get('chamber_material') or None),
            wall_thickness=_mm_to_m_optional(data.get('wall_thickness')),
            cooling_type=data.get('cooling_channels') or 'none',
            # v2.6.26: bkz. /calculate icindeki ayni ucluye dair yorum.
            safety_factor=data.get('safety_factor'),
            chamber_length_override=_mm_to_m_optional(
                data.get('chamber_length_override')),
            nozzle_material=data.get('nozzle_material'),
            tank_temperature=data.get('tank_temperature') or data.get('oxidizer_temp') or None,
        )
        engine.calculate()

        from hrma.analysis.transient_ballistics import (
            TransientBallistics, ThroatErosionModel)

        # Dalga 3 — opsiyonel boğaz erozyonu kuplajı (varsayılan KAPALI).
        # erosion_a_ref_mm_s: özel katsayı [mm/s @ 70 bar]; çelik/bakır gibi
        # 'not recommended' malzemelerde ZORUNLU (modelsiz ValueError → 400).
        erosion_model = None
        erosion_a_ref = data.get('erosion_a_ref_mm_s')
        if data.get('erosion_enabled') or erosion_a_ref is not None:
            erosion_model = ThroatErosionModel.for_material(
                data.get('throat_material', 'graphite'),
                a_ref_mm_s=(float(erosion_a_ref)
                            if erosion_a_ref is not None else None))

        solver = TransientBallistics(
            engine,
            feed_mode=data.get('feed_mode', 'regulated'),
            tank_temperature=float(data.get('tank_temperature', 293.15)),
            liquid_fill_fraction=float(data.get('liquid_fill_fraction', 0.85)),
            erosion_model=erosion_model,
        )
        tr = solver.solve()

        # Dizileri JSON'a uygun listelere çevir (sanitize NaN/Inf'i halleder)
        payload = {k: (v.tolist() if hasattr(v, 'tolist') else v)
                   for k, v in tr.items()}
        return jsonify(sanitize_json_values({
            'status': 'success',
            'transient': payload,
            'design_point': {
                'thrust': engine.F,
                'chamber_pressure_bar': engine.P_c,
                'burn_time': engine.t_b,
                'total_impulse_design': engine.F * engine.t_b,
            },
        }))
    except ValueError as e:
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/injector-design', methods=['POST'])
def injector_design_api():
    """Enjektör tasarımı — docs/10_Enjektor_ARGE.md bölüm C sözleşmesi.

    Girdi: spec B.1 alanları (motor_type, injector_type, mdot_ox, ...).
    Kolaylık: from_results=true + motor_results bloğu gönderilirse
    mdot_ox/mdot_fuel/Pc_bar/T_c_K/mw_gas oradan doldurulur (istekte
    açıkça verilen alan kazanır).

    Yanıt: 200 {'status':'success','design':{...}} | 400 doğrulama |
    500 beklenmeyen hata. Saf hesaptır, dosya yazmaz.
    """
    try:
        data = request.json or {}
        spec = {k: v for k, v in data.items()
                if k not in ('from_results', 'motor_results')}

        if data.get('from_results') and isinstance(
                data.get('motor_results'), dict):
            mr = data['motor_results']
            for key in ('mdot_ox', 'mdot_fuel', 'Pc_bar', 'T_c_K', 'mw_gas'):
                if spec.get(key) in (None, '', 0) and mr.get(key) is not None:
                    spec[key] = mr[key]

        from hrma.engines.injector_design import design_injector
        design = design_injector(spec)
        if isinstance(design, dict) and design.get('status') == 'error':
            return jsonify({'status': 'error',
                            'error': design.get('error',
                                                'injector design error')}), 400
        return jsonify(sanitize_json_values(
            {'status': 'success', 'design': design}))
    except ValueError as e:
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/six-dof-analysis', methods=['POST'])
def six_dof_analysis():
    """6-DOF rijit gövde uçuş analizi (Barrowman stabilite + weathercock).

    Girdi (JSON):
      Araç: body_diameter, body_length, nose_length [m], nose_type,
            dry_mass, propellant_mass [kg], cd0
      Kanat: fin_count, fin_root_chord, fin_tip_chord, fin_span,
             fin_sweep [m], fin_position (ops.)
      İtki: thrust [N] + burn_time [s] YA DA thrust_curve {time, thrust}
            (/api/transient-analysis çıktısı doğrudan verilebilir)
      Atış: launch_elevation_deg, launch_azimuth_deg, rail_length,
            wind_speed [m/s], wind_direction_deg (rüzgârın geldiği yön)
      CG:   x_cg_full, x_cg_empty [m, burundan] (ops.)

    Çıktı: apoje, maks hız/Mach/α, statik marj (dolu/boş), stabilite
    hükmü, yörünge zaman serileri (seyreltilmiş), ``solver_wall_time_s``.

    Faz 5B / B5 — FİZİKSEL GİRDİ KAPISI + DUVAR SAATİ BÜTÇESİ.
    """
    try:
        data = request.json or {}

        # --- FİZİKSEL GİRDİ KAPISI (Faz 5B / B5, H1-B5) ------------------
        # Bu uç bütün sayıları ``float(...)`` ile alıp doğrudan çözücüye
        # geçiriyordu; işaret/aralık denetimi hiçbir katmanda yoktu.
        # ÖLÇÜLDÜ (HEAD 9d3728e, HTTP 200 / status "success"):
        #   dry_mass = -5   -> apoje 3,95e14 m (2637 AU), tepe hız 9,96e11 m/s
        #                      (ışık hızının 3322 katı), stable = True
        #   cd0 = -1        -> apoje 1 060 260 m, tepe hız 4277 m/s (negatif
        #                      sürükleme aracı hızlandırıyor)
        #   thrust = -3000  -> HTTP 200, apoje 0,0, stable = True
        #   body_diameter=0 -> HTTP 500 'float division by zero'
        #                      (BarrowmanAero, ``(s/self.d)**2``)
        # ``stable: True`` hükmü tamamen geometrik statik marjdan geldiği
        # için ışık hızının 3322 katında uçan araca da "KARARLI" damgası
        # basılıyordu. Dalga 1 çözücü tarafında kütle/itki/sonluluk
        # denetimlerini kapattı; burası UÇ tarafıdır ve geri kalan geometri,
        # sürükleme ve atış açısı alanlarını da kapatır.
        # Desen ``/analyze_safety`` ile aynı: 422 + makine-okur alan listesi.
        invalid = _collect_unphysical_fields(
            data,
            positive=('body_diameter', 'body_length', 'nose_length',
                      'dry_mass', 'rail_length', 'thrust', 'burn_time',
                      'x_cg_full', 'x_cg_empty'),
            non_negative=('propellant_mass', 'fin_count', 'fin_root_chord',
                          'fin_tip_chord', 'fin_span', 'fin_sweep',
                          'fin_position', 'wind_speed'),
            # cd0 = 0 sürüklemesiz idealleştirmedir (meşru); negatifi enerji
            # üretir. Üst uç 5, kanatlı bir roket için fazlasıyla geniştir
            # (tipik cd0 0,3-0,8; küt gövdeli en kötü hâl ~2).
            ranges={'cd0': (0.0, 5.0),
                    'launch_elevation_deg': (0.0, 90.0),
                    'latitude_deg': (-90.0, 90.0)},
            finite=('launch_azimuth_deg', 'wind_direction_deg',
                    'launch_altitude'),
        )
        if invalid:
            return jsonify({
                'status': 'error',
                'error': 'invalid_six_dof_input',
                'message': ('A flight solution cannot be produced from '
                            'physically impossible inputs. Nothing was '
                            'integrated.'),
                'invalid_fields': invalid,
            }), 422

        from hrma.analysis.six_dof_trajectory import (
            BarrowmanAero, SixDOFTrajectory)
        aero = BarrowmanAero(
            body_diameter=float(data.get('body_diameter', 0.1)),
            nose_length=float(data.get('nose_length', 0.3)),
            body_length=float(data.get('body_length', 2.0)),
            nose_type=data.get('nose_type', 'ogive'),
            fin_count=int(data.get('fin_count', 4)),
            fin_root_chord=float(data.get('fin_root_chord', 0.15)),
            fin_tip_chord=float(data.get('fin_tip_chord', 0.075)),
            fin_span=float(data.get('fin_span', 0.1)),
            fin_sweep=float(data.get('fin_sweep', 0.05)),
            fin_position=data.get('fin_position'),
        )
        solver = SixDOFTrajectory(
            aero=aero,
            dry_mass=float(data.get('dry_mass', 20.0)),
            propellant_mass=float(data.get('propellant_mass', 10.0)),
            thrust_curve=data.get('thrust_curve'),
            thrust=data.get('thrust'),
            burn_time=data.get('burn_time'),
            x_cg_full=data.get('x_cg_full'),
            x_cg_empty=data.get('x_cg_empty'),
            cd0=float(data.get('cd0', 0.5)),
            wind_speed=float(data.get('wind_speed', 0.0)),
            wind_direction_deg=float(data.get('wind_direction_deg', 0.0)),
            launch_elevation_deg=float(data.get('launch_elevation_deg', 90.0)),
            launch_azimuth_deg=float(data.get('launch_azimuth_deg', 0.0)),
            rail_length=float(data.get('rail_length', 5.0)),
            # B1 — Coriolis: çözücü enlem parametreli Coriolis ivmesini
            # (−2·Ω×v) zaten destekliyordu ama bu uç değeri HİÇ geçirmiyordu,
            # yani düz-Dünya varsayımı fiilen yürürlükteydi. Fırlatma sahası
            # sayfası enlemi gönderiyor; enlem verilmezse çözücü kendi
            # varsayılanına düşer ve davranış eskisiyle aynı kalır.
            latitude_deg=(float(data['latitude_deg'])
                          if data.get('latitude_deg') is not None else None),
            launch_altitude=float(data.get('launch_altitude', 0.0)),
        )
        # t_max ÜST SINIRLI: sınırsız bırakıldığında kaçış yörüngesinde (araç
        # asla yere dönmez) entegrasyon bitmiyor ve istek süresiz asılıyordu.
        # 3600 s, atmosferik bir sounding roketi için fazlasıyla geniş.
        t_max = float(data.get('t_max', 400.0))
        if not (0.0 < t_max <= _SIXDOF_T_MAX_LIMIT_S):
            return jsonify({
                'status': 'error',
                'error': (f't_max must be in (0, {_SIXDOF_T_MAX_LIMIT_S:g}] s; '
                          'an unbounded horizon never terminates for escape '
                          'trajectories.'),
            }), 400

        # --- DUVAR SAATİ BÜTÇESİ (Faz 5B / B5) ---------------------------
        # ``solve_ivp``'ye duvar saati bütçesi verilemiyor ve çözücü
        # kesilemiyor; ama İSTEK bekletilmek zorunda değil. Çözüm daemon bir
        # iş parçacığında koşar: bütçe dolarsa uç 503 ile döner, iş parçacığı
        # kendi başına biter ve daemon olduğu için süreç kapanışını bloke
        # etmez. Ölçülen en kötü meşru süre 10,1 s (bkz.
        # ``_SIXDOF_WALL_CLOCK_BUDGET_S``), yani bu yol normalde HİÇ
        # tetiklenmez — bir gerileme olursa arayüz kilitlenmesin diye vardır.
        import time as _time
        _outcome = {}

        def _run_solver():
            try:
                _outcome['res'] = solver.solve(t_max=t_max)
            except BaseException as exc:      # noqa: BLE001 - yeniden atılır
                _outcome['exc'] = exc

        _t0 = _time.monotonic()
        _worker = threading.Thread(target=_run_solver, daemon=True,
                                   name='sixdof-solve')
        _worker.start()
        _worker.join(timeout=_SIXDOF_WALL_CLOCK_BUDGET_S)
        _elapsed = _time.monotonic() - _t0
        if _worker.is_alive():
            return jsonify({
                'status': 'error',
                'error': 'six_dof_time_budget_exceeded',
                'message': ('The 6-DOF integration did not finish within the '
                            f'{_SIXDOF_WALL_CLOCK_BUDGET_S:g} s budget. No '
                            'partial trajectory is reported: an unfinished '
                            'integration is not a flight.'),
                'budget_s': _SIXDOF_WALL_CLOCK_BUDGET_S,
                't_max_s': t_max,
            }), 503
        if 'exc' in _outcome:
            raise _outcome['exc']
        res = _outcome['res']

        # Zaman serilerini ~300 noktaya seyrelt
        import numpy as _np
        n = len(res['time'])
        idx = _np.linspace(0, n - 1, min(300, n)).astype(int)
        series = {
            'time': res['time'][idx].tolist(),
            'altitude': res['altitude'][idx].tolist(),
            'north': res['position'][0][idx].tolist(),
            'east': res['position'][1][idx].tolist(),
            'speed': res['speed'][idx].tolist(),
            'mach': res['mach'][idx].tolist(),
            'alpha_deg': res['alpha_deg'][idx].tolist(),
        }
        summary = {k: res[k] for k in (
            'apogee', 'apogee_time', 'max_speed', 'max_mach',
            'max_alpha_deg', 'static_margin_full', 'static_margin_empty',
            'stable', 'cn_alpha', 'x_cp', 'end_reason',
            'lateral_drift_at_end')}
        # ÖLÇÜLEN süre — uydurma değil, yukarıdaki monotonic farkı. Bütçenin
        # ne kadarının kullanıldığı böylece dışarıdan izlenebilir.
        return jsonify(sanitize_json_values({
            'status': 'success', 'summary': summary, 'series': series,
            'solver_wall_time_s': _elapsed}))
    except ValueError as e:
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/launch-site/resolve', methods=['POST'])
def launch_site_resolve():
    """Konumdan tam saha tanımı: rakım (DEM) + yerel g (WGS84) + yüzey atmosfer.

    Girdi (JSON): latitude, longitude [zorunlu]; elevation_m, temperature_k,
    pressure_pa [ops. elle datum]; use_online [ops. bool, Open-Meteo].

    KRİTİK: gravity_local_m_s2 enlem+rakımla değişir ama gravity_standard_m_s2
    her zaman 9.80665'tir (Isp/ideal-dV zinciri buna dokunmaz).
    """
    try:
        data = request.json or {}
        if data.get('latitude') is None or data.get('longitude') is None:
            return jsonify({'status': 'error',
                            'error': 'latitude and longitude are required'}), 400
        site = resolve_launch_site(
            float(data['latitude']), float(data['longitude']),
            elevation_m=(float(data['elevation_m'])
                         if data.get('elevation_m') not in (None, '') else None),
            temperature_k=(float(data['temperature_k'])
                           if data.get('temperature_k') not in (None, '') else None),
            pressure_pa=(float(data['pressure_pa'])
                         if data.get('pressure_pa') not in (None, '') else None),
            use_online=bool(data.get('use_online', False)),
        )
        return jsonify(sanitize_json_values({'status': 'success', 'site': site}))
    except (TypeError, ValueError) as e:
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:  # noqa: BLE001
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Dalga 3 — Analiz platformu endpoint'leri (docs/ANALIZ_PLATFORM_PLANI.md):
# basınçlı kap (ASME VIII / AIAA S-080 + Faupel burst), termal koruma
# (ablasyon Q* / heat-sink / radyasyon dengesi) ve cıvatalı bağlantı
# (Shigley). Modüller: hrma/analysis/{pressure_vessel, thermal_protection,
# bolted_joint}.py — burada yalnız girdi doğrulama + HTTP zarafeti var.
# ---------------------------------------------------------------------------

def _json_float(data, key):
    """JSON alanını float'a çevirir; yok / None / '' ise None döner.

    Sayıya çevrilemeyen değer ValueError yükseltir (endpoint 400'e çevirir).
    """
    v = data.get(key)
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        raise ValueError(f"'{key}' must be a number (got {v!r})")


def _json_bool(data, key, default):
    """JSON alanını bool'a çevirir ('true'/'false' string'leri dahil)."""
    v = data.get(key)
    if v is None or v == '':
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    if s in ('true', '1', 'yes', 'on'):
        return True
    if s in ('false', '0', 'no', 'off'):
        return False
    raise ValueError(f"'{key}' must be a boolean (got {v!r})")


@app.route('/api/pressure-vessel-analysis', methods=['POST'])
def pressure_vessel_analysis():
    """Basınçlı kap boyutlandırma + gerçek kopma (burst) basıncı.

    Girdi (JSON): meop_bar (ZORUNLU), inner_diameter_mm (ZORUNLU),
      material (vars. 'aluminum_6061'), wall_thickness_mm (None → otomatik
      boyutlandırma), temperature_K (vars. 293.15), weld_efficiency
      {0.70, 0.85, 1.00} (vars. 1.0), head_type (vars. 'ellipsoidal_2_1'),
      code_mode 'aiaa_s080' (vars.) | 'asme_viii'.

    Yanıt 200: PressureVesselAnalyzer.analyze() sözlüğü (status alanı
    PASS/MARGINAL/FAIL; actual_burst_pressure_bar "kaç barda patlar").
    Hata 400: {'error': mesaj} — tüm girdi hataları ValueError.
    """
    try:
        data = request.json or {}
        from hrma.analysis.pressure_vessel import PressureVesselAnalyzer

        meop_bar = _json_float(data, 'meop_bar')
        inner_diameter_mm = _json_float(data, 'inner_diameter_mm')
        if meop_bar is None:
            raise ValueError(
                "'meop_bar' is required (maximum expected operating "
                "pressure, bar)")
        if inner_diameter_mm is None:
            raise ValueError("'inner_diameter_mm' is required")

        temperature_K = _json_float(data, 'temperature_K')
        weld_efficiency = _json_float(data, 'weld_efficiency')

        result = PressureVesselAnalyzer().analyze(
            meop_bar=meop_bar,
            inner_diameter_mm=inner_diameter_mm,
            material=data.get('material', 'aluminum_6061'),
            wall_thickness_mm=_json_float(data, 'wall_thickness_mm'),
            temperature_K=(293.15 if temperature_K is None
                           else temperature_K),
            weld_efficiency=(1.0 if weld_efficiency is None
                             else weld_efficiency),
            head_type=data.get('head_type', 'ellipsoidal_2_1'),
            code_mode=data.get('code_mode', 'aiaa_s080'),
        )
        return jsonify(sanitize_json_values(result))
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# Termal koruma — mod bazlı parametre beyaz listeleri. Panel tek formdan
# tüm alanları gönderir; hedef modun kabul etmediği alanlar burada sessizce
# düşer (TypeError 500'e düşmesin). Anahtar adları modül imzalarıyla birebir
# (hrma/analysis/thermal_protection.py).
_TP_MODE_KEYS = {
    # v2.6.27 (B6-4): yeni-yol anahtarları eklendi — h_gas_W_m2K +
    # T_recovery_K çifti yüzey enerji dengesini açar (çekirdek yarım çifti
    # ValueError ile reddeder, sessiz yok sayma yok), gas_cp_J_kgK üfleme
    # blokajının B' çözümünü besler, station_radius_m geometrik kapıyı açar.
    'ablative': ('q_net_W_m2', 'burn_time_s', 'time_s', 'material',
                 'design_margin', 'density_kg_m3',
                 'h_gas_W_m2K', 'T_recovery_K', 'gas_cp_J_kgK',
                 'station_radius_m'),
    'heat_sink': ('h_gas_W_m2K', 'T_recovery_K', 'burn_time_s',
                  'wall_thickness_m', 'wall_material', 'T_initial_K',
                  'n_nodes', 'cfl_safety', 'store_history'),
    'radiation_equilibrium': ('h_gas_W_m2K', 'T_recovery_K', 'emissivity',
                              'material'),
}

#: Mod bazlı ZORUNLU alanlar. ``ThermalProtectionAnalyzer`` imzalarındaki
#: varsayılansız (positional) argümanların birebir karşılığı — ölçüldü:
#:   ablative_thickness    -> ['q_net_W_m2']
#:   heat_sink_transient   -> ['h_gas_W_m2K','T_recovery_K','burn_time_s',
#:                             'wall_thickness_m']
#:   radiation_equilibrium -> ['h_gas_W_m2K','T_recovery_K']
#: (``inspect.signature`` ile, ``hrma/analysis/thermal_protection.py``.)
_TP_MODE_REQUIRED = {
    'ablative': ('q_net_W_m2',),
    'heat_sink': ('h_gas_W_m2K', 'T_recovery_K', 'burn_time_s',
                  'wall_thickness_m'),
    'radiation_equilibrium': ('h_gas_W_m2K', 'T_recovery_K'),
}


@app.route('/api/thermal-protection', methods=['POST'])
@app.route('/api/analysis/thermal-protection', methods=['POST'])
def thermal_protection_analysis():
    """Termal koruma analizi (üç mod).

    Girdi (JSON): {'mode': 'ablative' | 'heat_sink' |
    'radiation_equilibrium', ...mod parametreleri} — şema için
    hrma/analysis/thermal_protection.py docstring'lerine bakınız.

    Kolaylıklar (panel sözleşmesi):
      * ablative: 'q_star_MJ_kg' kabul edilir → q_star_J_kg (x 1e6).
      * radiation_equilibrium: 'radiation_material' kabul edilir →
        'material' (panel, ablatif malzeme seçicisiyle çakışmasın diye
        ayrı alan adı kullanır).
      * heat_sink: store_history verilmezse True (panel T_w(t) grafiği).

    Yanıt 200: ThermalProtectionAnalyzer.analyze() sözlüğü (model_note
    alanı 'Simplified model' rozetine bağlanır). Hata 400: {'error': ...}.
    Eksik zorunlu alanda 422 + ``missing_fields`` (Faz 5B / B8).
    """
    try:
        data = request.json or {}
        mode = data.get('mode', 'ablative')
        if mode not in _TP_MODE_KEYS:
            raise ValueError(
                f"Unknown mode '{mode}'. "
                f"Available: {sorted(_TP_MODE_KEYS)}")

        params = {k: data[k] for k in _TP_MODE_KEYS[mode]
                  if data.get(k) not in (None, '')}

        # --- ZORUNLU ALAN KAPISI (Faz 5B / B8) ---------------------------
        # ÖLÇÜLDÜ (HEAD 9d3728e): panelde sayısal bir alanı BOŞALTMAK 500
        # üretiyordu, çünkü ham Python imza hatası dışarı sızıyordu:
        #   q_net_W_m2:''   -> 500 "ablative_thickness() missing 1 required
        #                          positional argument: 'q_net_W_m2'"
        #   h_gas_W_m2K:''  -> 500 (heat_sink_transient, aynı desen)
        #   T_recovery_K:'' -> 500 (radiation_equilibrium, aynı desen)
        #   {} (boş gövde)  -> 500 (mod varsayılanı 'ablative')
        # Zincir: ``analysis_dock.js`` boş alanı payload'a hiç koymuyor
        # (``if (Number.isFinite(v))``), yukarıdaki satır da ``''`` olanı
        # eliyor; zorunlu argüman böylece kayboluyordu. Panel ``data.error``
        # değerini ekrana bastığı için kullanıcı ham Python metnini görüyordu.
        # Bu bir istemci hatasıdır (500 değil) ve makine-okur olmalıdır.
        missing = [key for key in _TP_MODE_REQUIRED[mode]
                   if key not in params]
        if missing:
            return jsonify({
                'status': 'error',
                'error': 'incomplete_thermal_protection_input',
                'message': (f"Mode '{mode}' cannot be analysed without these "
                            'inputs; they have no default.'),
                'mode': mode,
                'missing_fields': missing,
            }), 422

        if mode == 'ablative' and data.get('q_star_MJ_kg') not in (None, ''):
            params['q_star_J_kg'] = float(data['q_star_MJ_kg']) * 1e6
        if (mode == 'radiation_equilibrium'
                and data.get('radiation_material') not in (None, '')):
            params['material'] = data['radiation_material']
        if mode == 'heat_sink':
            params['store_history'] = _json_bool(data, 'store_history', True)

        from hrma.analysis.thermal_protection import ThermalProtectionAnalyzer
        result = ThermalProtectionAnalyzer().analyze(mode, **params)
        return jsonify(sanitize_json_values(result))
    except (ValueError, KeyError) as e:
        # KeyError: bilinmeyen materials_db anahtarı (get_material)
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# Gimbal montajı — parametre beyaz listesi (C3 bağlaması, v2.6.27).
# Anahtar adları ``analyze_gimbal_mount`` imzasıyla BİREBİR aynıdır
# (hrma/analysis/gimbal_mount.py; tests/test_gimbal_baglama.py imzayla
# eşitliği inspect ile kilitler). Beyaz liste dışı anahtarlar termal-koruma
# ucundaki desenle SESSİZCE düşer (ölçüldü: /api/thermal-protection fazladan
# anahtarı 200 ile yok sayar; TypeError 500'e düşmesin diye).
_GIMBAL_KEYS = (
    'thrust_N', 'gimbal_angle_deg', 'actuator_arm_m', 'ring_offset_m',
    'thrust_offset_m', 'duct_torsional_stiffness_N_m_rad',
    'bearing_friction_moment_N_m', 'engine_inertia_kg_m2',
    'slew_acceleration_deg_s2', 'slew_rate_deg_s', 'slew_reversal_time_s',
    'actuators_per_axis', 'bolt_circle_diameter_m', 'bolt_count',
    'yaw_angle_deg',
)

#: ZORUNLU alanlar — ``analyze_gimbal_mount`` imzasındaki varsayılansız
#: argümanların birebir karşılığı. Ölçüldü (``inspect.signature`` ile):
#:   analyze_gimbal_mount -> ['thrust_N', 'gimbal_angle_deg',
#:                            'actuator_arm_m']
_GIMBAL_REQUIRED = ('thrust_N', 'gimbal_angle_deg', 'actuator_arm_m')


@app.route('/api/gimbal-mount', methods=['POST'])
def gimbal_mount_analysis():
    """Gimbal montaj yük zinciri: itki bileşenleri + aktüatör + halka + cıvata.

    Girdi (JSON): ``_GIMBAL_KEYS`` beyaz listesindeki alanlar — şema için
    hrma/analysis/gimbal_mount.py docstring'lerine bakınız. Boş dize ve
    None gönderilen alan YOK sayılır (panel boş alanı payload'a koymaz;
    burada ikinci savunma hattı).

    Yanıt 200: ``analyze_gimbal_mount`` sözlüğü (validity/warnings/
    not_modelled beyanları aynen taşınır — panel bunları basar, sayı
    uydurmaz). Eksik zorunlu alanda 422 + ``missing_fields`` (termal-koruma
    ucunun sözleşmesi). Modülün ValueError'ı (ör. 45 derece üstü açı,
    3'ten az cıvata) 400 + makine-okur mesajla döner; geçerlilik beyanının
    İÇERİĞİ korunur (metin aynen iletilir, yumuşatılmaz).
    """
    try:
        data = request.json or {}
        params = {k: data[k] for k in _GIMBAL_KEYS
                  if data.get(k) not in (None, '')}

        # --- ZORUNLU ALAN KAPISI (termal-koruma deseni) -------------------
        # Eksik zorunlu argüman ham TypeError olarak 500'e sızmasın; bu bir
        # istemci hatasıdır (422) ve makine-okur olmalıdır.
        missing = [key for key in _GIMBAL_REQUIRED if key not in params]
        if missing:
            return jsonify({
                'status': 'error',
                'error': 'incomplete_gimbal_mount_input',
                'message': ('Gimbal mount loads cannot be analysed without '
                            'these inputs; they have no default.'),
                'missing_fields': missing,
            }), 422

        from hrma.analysis.gimbal_mount import analyze_gimbal_mount
        result = analyze_gimbal_mount(**params)
        return jsonify(sanitize_json_values(result))
    except ValueError as e:
        # Modülün geçerlilik beyanı (açı bandı, cıvata sayısı, işaret) —
        # metin aynen iletilir ki kullanıcı NEDENİ görsün.
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/bolted-joint', methods=['POST'])
def bolted_joint_analysis():
    """Cıvatalı bağlantı (flanş/kapak) analizi — Shigley yöntemi.

    Girdi (JSON): pressure_bar + seal_diameter_mm VEYA
    external_axial_load_n (en az biri zorunlu); bolt_count (ZORUNLU);
    size 'M4'..'M24' (vars. 'M8'); property_class '8.8'|'10.9'|'12.9'|
    'A2-70' (vars. '8.8'); grip_length_mm (vars. 20); member_material
    (materials_db anahtarı, vars. 'aluminum_6061'); lubricated (vars.
    false); reusable (vars. true).

    Yanıt 200: {'status': 'success', 'joint': {...}} — tork önerisi
    ±%25 ön-yük saçılım bandıyla, ayrılma marjı ve emniyet faktörleri.
    Hata 400: {'status': 'error', 'error': mesaj}.
    """
    try:
        data = request.json or {}
        from hrma.analysis.bolted_joint import analyze_bolted_joint

        pressure_bar = _json_float(data, 'pressure_bar')
        seal_diameter_mm = _json_float(data, 'seal_diameter_mm')
        external_load_n = _json_float(data, 'external_axial_load_n')
        if pressure_bar is None and external_load_n is None:
            raise ValueError(
                "Provide either 'pressure_bar' (with 'seal_diameter_mm') "
                "or 'external_axial_load_n'")

        bolt_count = data.get('bolt_count')
        if bolt_count in (None, ''):
            raise ValueError("'bolt_count' is required (integer >= 1)")
        try:
            bolt_count = int(bolt_count)
        except (TypeError, ValueError):
            raise ValueError(
                f"'bolt_count' must be an integer >= 1 (got {bolt_count!r})")

        grip_length_mm = _json_float(data, 'grip_length_mm')

        joint = analyze_bolted_joint(
            pressure_bar=pressure_bar,
            seal_diameter_mm=seal_diameter_mm,
            bolt_count=bolt_count,
            size=data.get('size', 'M8'),
            property_class=data.get('property_class', '8.8'),
            grip_length_mm=(20.0 if grip_length_mm is None
                            else grip_length_mm),
            member_material=data.get('member_material', 'aluminum_6061'),
            lubricated=_json_bool(data, 'lubricated', False),
            reusable=_json_bool(data, 'reusable', True),
            external_axial_load_n=external_load_n,
        )
        return jsonify(sanitize_json_values(
            {'status': 'success', 'joint': joint}))
    except (ValueError, KeyError) as e:
        # KeyError: bilinmeyen materials_db anahtarı (get_material)
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Yanma kararlılığı ucu (F2 API bağlaması, parti 28 / A2) — desen:
# thermal_protection_analysis İKİZİ (mode kapısı + zorunlu alan kapısı +
# 422 ``missing_fields`` + ValueError → 400). Çekirdek:
# hrma/stability/{chug,damping}.py — bu katman HİÇBİR fizik hesabı yapmaz ve
# kendi eklediği anahtarlarda hüküm TAŞIMAZ: ``assessment`` içindeki
# ``verdict`` çekirdekten kapsam etiketiyle gelir (make_verdict), uç katmanın
# kabuğunda 'verdict' bulunmadığı tests/test_api_combustion_stability.py'de
# ``forbid_verdict_key`` bekçisiyle kilitlidir.
# ---------------------------------------------------------------------------

#: Mod bazlı ZORUNLU alanlar (uç sözleşmesi). Çekirdek imzalarının
#: varsayılansız karşılığı — ölçüldü (``inspect.signature``):
#:   assess_chug                 -> ['dp_ratio_j', 'tau_s', 'tau_c_s']
#:   nozzle_damping_quasi_steady -> ['sound_speed_m_s', 'chamber_length_m',
#:                                   'gamma', 'nozzle_entrance_mach']
#: (``hrma/stability/chug.py`` ve ``hrma/stability/damping.py``.)
#: Termal-koruma ucundan bilinçli fark: mode'un VARSAYILANI YOKTUR — iki mod
#: tamamen ayrı fizik sorularıdır, sessizce birini seçmek uydurma varsayılan
#: olurdu; mode'suz istek 400 alır.
_CS_MODE_REQUIRED = {
    'chug': ('dp_ratio_j', 'tau_s', 'tau_c_s'),
    'damping': ('sound_speed_m_s', 'chamber_length_m', 'gamma',
                'nozzle_entrance_mach'),
}

#: Nötr eğri örnekleme ızgarası: J ∈ [0,02, 0,48], 93 nokta (adım 0,005 —
#: sözleşmenin "en az 60 örnek" şartının üstünde). Bant uçları çekirdeğin
#: kendi tanım bandından: J > 0 (``_positive`` kapısı) ve
#: J < CHUG_GAIN_J_MAX = 0,5 (nötr çözümün varlık şartı, ``chug.py``);
#: 0,02/0,48 uçları iki tekilliğe (J → 0⁺'da τ/τ_c → 0, J → 0,5⁻'de +∞)
#: yaklaşmadan durur. Yeni fizik eşiği DEĞİL, çizim bandıdır.
_CS_NEUTRAL_J_MIN = 0.02
_CS_NEUTRAL_J_MAX = 0.48
_CS_NEUTRAL_N = 93

#: Kök yeri taraması: işletme J'si çevresinde ±0,10 pencere, 41 nokta.
#: Pencere merkezi [0,12, 0,38] bandına kıstırılır ki tarama daima nötr eğri
#: bandının ([0,02, 0,48]) İÇİNDE kalsın — sözleşmedeki "[0,02, 0,48]
#: kesişimi" böyle uygulanır (pencere bandın dışına taşacaksa kaydırılır,
#: nokta atılmaz).
_CS_LOCUS_HALF_WIDTH = 0.10
_CS_LOCUS_N = 41


def _combustion_stability_feed_tau(feed_line):
    """``feed_line`` sözlüğünden τ_f + yankı beyanı türetir.

    Uç sözleşmesinin alan adları: ``length_m``, ``area_m2`` |
    ``diameter_mm``, ``mass_flow_kg_s`` (+ yalnız-yankı ``density_kg_m3``).
    τ_f = ℓ·ṁ/(2·A·ΔP_inj) formülü ΔP_inj'siz KAPANMAZ
    (``hrma.stability.chug.feed_inertance_time_constant`` imzası dördünü de
    ister) ve ΔP burada başka hiçbir girdiden türetilemez: uç J = ΔP/P_c
    oranını alır ama P_c'yi almaz, yoğunluk da birinci mertebe atalet formuna
    girmez. Uydurma varsayılan yasağı gereği ``dp_injector_Pa`` bu yüzden
    feed_line içinde ZORUNLU alınır (sözleşme listesinden gerekçeli sapma —
    raporda beyanlı). Dairesel kesit çevirisi motor tarafıyla birebir aynı
    formüldür (``liquid_rocket_engine._feed_line_inertance_inputs``:
    A = π·(d/1000)²/4).

    Returns:
        (tau_f_s, echo, missing): eksik alan varsa (None, None, [adlar]).
    """
    if not isinstance(feed_line, dict):
        raise ValueError(
            "'feed_line' must be a JSON object with the line geometry "
            "(length_m, area_m2 or diameter_mm, mass_flow_kg_s, "
            "dp_injector_Pa).")
    missing = []
    length = _json_float(feed_line, 'length_m')
    if length is None:
        missing.append('feed_line.length_m')
    area_given = _json_float(feed_line, 'area_m2')
    d_mm = _json_float(feed_line, 'diameter_mm')
    area = area_given
    if area is None:
        if d_mm is not None:
            area = math.pi * (d_mm / 1000.0) ** 2 / 4.0
        else:
            missing.append('feed_line.area_m2 | feed_line.diameter_mm')
    mdot = _json_float(feed_line, 'mass_flow_kg_s')
    if mdot is None:
        missing.append('feed_line.mass_flow_kg_s')
    dp_pa = _json_float(feed_line, 'dp_injector_Pa')
    if dp_pa is None:
        missing.append('feed_line.dp_injector_Pa')
    if missing:
        return None, None, missing

    from hrma.stability.chug import feed_inertance_time_constant
    tau_f = feed_inertance_time_constant(
        line_length_m=length, line_area_m2=area,
        mass_flow_kg_s=mdot, dp_injector_Pa=dp_pa)
    echo = {
        'line_length_m': length,
        'line_area_m2': area,
        'mass_flow_kg_s': mdot,
        'dp_injector_Pa': dp_pa,
        '_basis': (
            'Feed line geometry supplied by the API caller; tau_f = '
            'l*mdot/(2*A*dP_inj) via '
            'hrma.stability.chug.feed_inertance_time_constant. Where area_m2 '
            'was not given, the flow area was derived from the circular '
            'inner diameter (A = pi*d^2/4, same conversion as the liquid '
            'engine). density_kg_m3, if echoed, is NOT used by this '
            'first-order inertance form; it is echoed for the record only.'),
    }
    if area_given is None and d_mm is not None:
        echo['line_diameter_mm'] = d_mm
    density = _json_float(feed_line, 'density_kg_m3')
    if density is not None:
        echo['density_kg_m3'] = density   # yalnız yankı — hesaba girmez
    return tau_f, echo, []


def _combustion_stability_chug(data):
    """chug modu gövdesi: assessment AYNEN + nötr eğri + kök yeri taraması."""
    from hrma.stability.chug import (
        assess_chug,
        chug_neutral_tau_ratio,
        chug_rightmost_root,
    )

    dp_ratio_j = _json_float(data, 'dp_ratio_j')
    tau_s = _json_float(data, 'tau_s')
    tau_c_s = _json_float(data, 'tau_c_s')

    tau_f_s = _json_float(data, 'tau_f_s')
    feed_line_raw = data.get('feed_line')
    feed_echo = None
    if tau_f_s is not None:
        # τ_f doğrudan verildi; feed_line (varsa) türetimde KULLANILMAZ ve
        # bu, yankının kendi beyanında açıkça yazılır.
        if feed_line_raw is not None:
            if not isinstance(feed_line_raw, dict):
                raise ValueError("'feed_line' must be a JSON object.")
            feed_echo = dict(feed_line_raw)
            feed_echo['_basis'] = (
                'Echo only: tau_f_s was supplied directly by the caller, so '
                'this feed_line block was NOT used to derive it.')
    elif feed_line_raw is not None:
        tau_f_s, feed_echo, missing = _combustion_stability_feed_tau(
            feed_line_raw)
        if missing:
            return jsonify({
                'status': 'error',
                'error': 'incomplete_combustion_stability_input',
                'message': ("Mode 'chug' cannot derive the feed line "
                            'inertance time constant without these inputs; '
                            'they have no default.'),
                'mode': 'chug',
                'missing_fields': missing,
            }), 422

    assessment = assess_chug(dp_ratio_j=dp_ratio_j, tau_s=tau_s,
                             tau_c_s=tau_c_s, tau_f_s=tau_f_s,
                             feed_line=feed_echo)

    # Nötr eğri — çekirdek fonksiyonun bire bir örneklenmesi. Ara katman
    # hesabı yok: her nokta chug_neutral_tau_ratio(J) çağrısının kendisidir
    # (bit-aynılık bekçisi tests/test_api_combustion_stability.py'de).
    step = (_CS_NEUTRAL_J_MAX - _CS_NEUTRAL_J_MIN) / (_CS_NEUTRAL_N - 1)
    curve_j = [_CS_NEUTRAL_J_MIN + i * step for i in range(_CS_NEUTRAL_N)]
    neutral_curve = {
        'dp_ratio_j': curve_j,
        'tau_over_tau_c': [chug_neutral_tau_ratio(j) for j in curve_j],
    }

    # Kök yeri: işletme J'si çevresinde tarama — τ, τ_c, τ_f işletme
    # noktasının DEĞERLERİYLE sabit tutulur, yalnız J taranır. Çekirdeğin
    # reddettiği (ValueError) ya da kalıntı denetimini geçemeyen (None)
    # nokta ATLANIR ve skipped_points'te adıyla beyan edilir; uydurma sayı
    # yayımlanmaz.
    j_op = assessment['dp_ratio_j']
    tau = assessment['tau_s']
    tau_c = assessment['tau_c_s']
    tau_f = assessment['tau_f_s'] if assessment['tau_f_s'] else 0.0
    center = min(max(j_op, _CS_NEUTRAL_J_MIN + _CS_LOCUS_HALF_WIDTH),
                 _CS_NEUTRAL_J_MAX - _CS_LOCUS_HALF_WIDTH)
    locus_step = 2.0 * _CS_LOCUS_HALF_WIDTH / (_CS_LOCUS_N - 1)
    locus_j, locus_sigma, locus_freq, skipped = [], [], [], []
    for i in range(_CS_LOCUS_N):
        j = center - _CS_LOCUS_HALF_WIDTH + i * locus_step
        try:
            root = chug_rightmost_root(j, tau, tau_c, tau_f)
        except ValueError as exc:
            skipped.append({'dp_ratio_j': j, 'reason': str(exc)})
            continue
        if root is None:
            skipped.append({'dp_ratio_j': j, 'reason': (
                'the continued root did not meet the residual check '
                '(chug_rightmost_root returned None); no number is '
                'fabricated for this point.')})
            continue
        locus_j.append(j)
        locus_sigma.append(float(root.real))
        locus_freq.append(abs(float(root.imag)) / (2.0 * math.pi))

    return jsonify(sanitize_json_values({
        'status': 'ok',
        'mode': 'chug',
        'assessment': assessment,
        'neutral_curve': neutral_curve,
        'root_locus': {
            'dp_ratio_j': locus_j,
            'sigma_1_s': locus_sigma,
            'frequency_hz': locus_freq,
            'skipped_points': skipped,
            '_basis': (
                'Rightmost characteristic root traced over a J window of '
                f'+/-{_CS_LOCUS_HALF_WIDTH:g} around the operating point '
                f'(window centre clamped into [{_CS_NEUTRAL_J_MIN + _CS_LOCUS_HALF_WIDTH:g}, '
                f'{_CS_NEUTRAL_J_MAX - _CS_LOCUS_HALF_WIDTH:g}] so the sweep '
                f'stays inside the neutral-curve band '
                f'[{_CS_NEUTRAL_J_MIN:g}, {_CS_NEUTRAL_J_MAX:g}]), with tau, '
                'tau_c and tau_f held at their operating values. Points the '
                'core rejects or cannot resolve are listed in '
                'skipped_points, never fabricated.'),
        },
        'operating_point': {
            'dp_ratio_j': assessment['dp_ratio_j'],
            'tau_over_tau_c': assessment['tau_over_tau_c'],
        },
    }))


def _combustion_stability_damping(data):
    """damping modu gövdesi: lüle terimi + bütçe, çekirdek sözlükleri AYNEN."""
    from hrma.stability.damping import (
        damping_budget,
        nozzle_damping_quasi_steady,
    )
    nozzle = nozzle_damping_quasi_steady(
        sound_speed_m_s=_json_float(data, 'sound_speed_m_s'),
        chamber_length_m=_json_float(data, 'chamber_length_m'),
        gamma=_json_float(data, 'gamma'),
        nozzle_entrance_mach=_json_float(data, 'nozzle_entrance_mach'))
    budget = damping_budget([nozzle])
    return jsonify(sanitize_json_values({
        'status': 'ok',
        'mode': 'damping',
        'nozzle': nozzle,
        'budget': budget,
    }))


@app.route('/api/combustion-stability', methods=['POST'])
@app.route('/api/analysis/combustion-stability', methods=['POST'])
def combustion_stability_analysis():
    """Yanma kararlılığı analizi (iki mod: chug + damping).

    Girdi (JSON): {'mode': 'chug' | 'damping', ...mod parametreleri}.

    mode='chug' (ZORUNLU: dp_ratio_j, tau_s, tau_c_s):
      Opsiyonel tau_f_s VEYA feed_line {length_m, area_m2 | diameter_mm,
      mass_flow_kg_s, dp_injector_Pa (+ yalnız-yankı density_kg_m3)} —
      τ_f, hrma.stability.chug.feed_inertance_time_constant ile türetilir
      ve yankısı beyanlı taşınır.
      Yanıt 200: {status:'ok', mode:'chug',
        assessment: assess_chug() sözlüğü AYNEN (yeniden adlandırma yok;
        içindeki verdict çekirdekten kapsam etiketli gelir, bu uç kendi
        hüküm alanı EKLEMEZ),
        neutral_curve: {dp_ratio_j, tau_over_tau_c} (93 örnek,
        chug_neutral_tau_ratio bire bir),
        root_locus: {dp_ratio_j, sigma_1_s, frequency_hz, skipped_points}
        (chug_rightmost_root, işletme J'si çevresinde tarama),
        operating_point: {dp_ratio_j, tau_over_tau_c}}.

    mode='damping' (ZORUNLU: sound_speed_m_s, chamber_length_m, gamma,
    nozzle_entrance_mach):
      Yanıt 200: {status:'ok', mode:'damping',
        nozzle: nozzle_damping_quasi_steady() sözlüğü AYNEN,
        budget: damping_budget([lüle terimi]) sözlüğü AYNEN}.

    Eksik zorunlu alanda 422 + ``missing_fields`` (termal-koruma ucunun
    sözleşmesi); bilinmeyen/eksik mode ve çekirdeğin her ValueError'ı 400.
    Mode'un varsayılanı BİLEREK yoktur (iki ayrı fizik sorusu; sessiz seçim
    uydurma varsayılan olurdu).
    """
    try:
        data = request.json or {}
        mode = data.get('mode')
        if mode not in _CS_MODE_REQUIRED:
            raise ValueError(
                f"Unknown mode {mode!r}. "
                f"Available: {sorted(_CS_MODE_REQUIRED)}")

        # --- ZORUNLU ALAN KAPISI (termal-koruma deseni) -------------------
        # Eksik zorunlu girdi ham çekirdek ValueError'ı olarak 400'e
        # düşmesin; bu makine-okur bir istemci hatasıdır (422) ve panel
        # missing_fields listesini alanlara eşler.
        missing = [key for key in _CS_MODE_REQUIRED[mode]
                   if data.get(key) in (None, '')]
        if missing:
            return jsonify({
                'status': 'error',
                'error': 'incomplete_combustion_stability_input',
                'message': (f"Mode '{mode}' cannot be analysed without these "
                            'inputs; they have no default.'),
                'mode': mode,
                'missing_fields': missing,
            }), 422

        if mode == 'chug':
            return _combustion_stability_chug(data)
        return _combustion_stability_damping(data)
    except ValueError as e:
        # Çekirdeğin geçerlilik beyanı (J/τ/γ kapıları, sayısal menzil
        # kapısı) — metin aynen iletilir ki kullanıcı NEDENİ görsün.
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Dalga 4B — Soğutma & besleme sistemi endpoint'leri
# (docs/ANALIZ_PLATFORM_PLANI.md). Modüller literatür referanslı, HTTP katmanı
# burada yalnız girdi doğrulama + birim çevirisi + zarafet yapar:
#   /api/regen-cooling      -> hrma.analysis.regen_cooling.RegenCooling
#   /api/slosh-analysis     -> hrma.analysis.slosh_analysis.analyze_slosh
#   /api/pressurant-sizing  -> hrma.analysis.pressurant_sizing.analyze_pressurant
#   /api/water-hammer       -> hrma.analysis.water_hammer.WaterHammerAnalyzer
# Tüm modüller geçersiz girdide ValueError yükseltir -> endpoint 400'e çevirir.
# Basınçlar arayüzde bar; modül SI (Pa) beklediği için burada çevrilir.
# ---------------------------------------------------------------------------


def _bar_to_pa(data, key):
    """'<key>' alanını bar'dan Pa'ya çevirir; yok/boş ise None döner."""
    v = _json_float(data, key)
    return None if v is None else v * 1e5


@app.route('/api/regen-cooling', methods=['POST'])
def regen_cooling_analysis():
    """Rejeneratif soğutma 1D istasyon marşı (analiz modu — otomatik
    boyutlandırma YOK, kanal geometrisi kullanıcı girdisidir).

    Girdi (JSON): chamber_pressure (bar, ZORUNLU), chamber_temperature (K,
      ZORUNLU), throat_diameter (m, ZORUNLU); nozul için expansion_ratio (>1)
      VEYA exit_diameter (m) — en az biri; gamma (vars. 1.2),
      molecular_weight (g/mol, vars. 24); coolant 'water'|'rp1' (vars.
      water); coolant_mdot (kg/s, vars. 1); coolant_inlet_temp (K, vars.
      300); coolant_inlet_pressure (bar, vars. 30); n_channels (vars. 64);
      channel_width, channel_height, wall_thickness (MM — burada m'ye
      çevrilir); wall_material (materials_db anahtarı, vars. copper);
      flow_direction 'counterflow'|'coflow' (vars. counterflow); n_stations
      (20-50, vars. 40).

    Yanıt 200: {'status':'success', 'cooling': RegenCooling.solve()} —
      istasyon dizileri (x_mm, T_wall_hot_K, T_wall_cold_K, T_coolant_K,
      P_coolant_bar, q_MW_m2, velocity_m_s), summary (peak wall T, malzeme
      limiti, dP, çıkış T, koklaşma durumu, uyarılar) ve model_note.
    Hata 400: {'status':'error','error': mesaj} — tüm girdi hataları ValueError.
    """
    try:
        data = request.json or {}
        from hrma.analysis.regen_cooling import RegenCooling

        chamber_pressure_bar = _json_float(data, 'chamber_pressure')
        chamber_temperature = _json_float(data, 'chamber_temperature')
        throat_diameter = _json_float(data, 'throat_diameter')
        if chamber_pressure_bar is None:
            raise ValueError("'chamber_pressure' is required (bar)")
        if chamber_temperature is None:
            raise ValueError("'chamber_temperature' is required (K)")
        if throat_diameter is None:
            raise ValueError("'throat_diameter' is required (m)")

        gamma = _json_float(data, 'gamma')
        molecular_weight = _json_float(data, 'molecular_weight')
        coolant_mdot = _json_float(data, 'coolant_mdot')
        coolant_inlet_temp = _json_float(data, 'coolant_inlet_temp')
        coolant_inlet_pressure_bar = _json_float(data, 'coolant_inlet_pressure')
        n_channels = _json_float(data, 'n_channels')
        n_stations = _json_float(data, 'n_stations')
        expansion_ratio = _json_float(data, 'expansion_ratio')
        exit_diameter = _json_float(data, 'exit_diameter')

        # Kanal geometrisi arayüzde mm; modül SI (m) bekler.
        channel_width_mm = _json_float(data, 'channel_width')
        channel_height_mm = _json_float(data, 'channel_height')
        wall_thickness_mm = _json_float(data, 'wall_thickness')

        kwargs = dict(
            chamber_pressure=chamber_pressure_bar * 1e5,
            chamber_temperature=chamber_temperature,
            throat_diameter=throat_diameter,
            coolant=data.get('coolant', 'water'),
            wall_material=data.get('wall_material', 'copper'),
            flow_direction=data.get('flow_direction', 'counterflow'),
        )
        if gamma is not None:
            kwargs['gamma'] = gamma
        if molecular_weight is not None:
            kwargs['molecular_weight'] = molecular_weight
        if coolant_mdot is not None:
            kwargs['coolant_mdot'] = coolant_mdot
        if coolant_inlet_temp is not None:
            kwargs['coolant_inlet_temp'] = coolant_inlet_temp
        if coolant_inlet_pressure_bar is not None:
            kwargs['coolant_inlet_pressure'] = coolant_inlet_pressure_bar * 1e5
        if n_channels is not None:
            kwargs['n_channels'] = int(n_channels)
        if n_stations is not None:
            kwargs['n_stations'] = int(n_stations)
        if expansion_ratio is not None:
            kwargs['expansion_ratio'] = expansion_ratio
        if exit_diameter is not None:
            kwargs['exit_diameter'] = exit_diameter
        if channel_width_mm is not None:
            kwargs['channel_width'] = channel_width_mm / 1e3
        if channel_height_mm is not None:
            kwargs['channel_height'] = channel_height_mm / 1e3
        if wall_thickness_mm is not None:
            kwargs['wall_thickness'] = wall_thickness_mm / 1e3

        result = RegenCooling(**kwargs).solve()
        return jsonify(sanitize_json_values(
            {'status': 'success', 'cooling': result}))
    except (ValueError, KeyError) as e:
        # KeyError: bilinmeyen materials_db anahtarı (get_material)
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/slosh-analysis', methods=['POST'])
def slosh_analysis_api():
    """Yakıt çalkalanması (slosh) analizi — dik silindirik tank, doğrusal
    serbest yüzey teorisi (NASA SP-106 / Dodge 2000).

    Girdi (JSON): radius (m, ZORUNLU), fill_height (m, ZORUNLU), g_eff
      (m/s^2, vars. 9.80665), fluid_density (kg/m^3, ops.), liquid_mass
      (kg, ops.), baffle_width_ratio (w/R, ops. — verilmezse hedef sönümleme
      önerisi döner), baffle_depth_ratio (d_s/R, vars. 0.10),
      control_frequencies / structural_frequencies (Hz listeleri, ops.),
      coincidence_margin (vars. 0.20).

    Yanıt 200: {'status':'success', 'slosh': analyze_slosh()} — f1_hz,
      slosh_mass_ratio, pendulum_length, modes, fill_sweep (doluluk eğrisi
      dizileri), baffle (sönümleme), coincidence_warnings, model_note.
    Hata 400: {'status':'error','error': mesaj}.
    """
    try:
        data = request.json or {}
        from hrma.analysis.slosh_analysis import analyze_slosh

        radius = _json_float(data, 'radius')
        fill_height = _json_float(data, 'fill_height')
        if radius is None:
            raise ValueError("'radius' is required (m)")
        if fill_height is None:
            raise ValueError("'fill_height' is required (m)")

        g_eff = _json_float(data, 'g_eff')
        fluid_density = _json_float(data, 'fluid_density')
        liquid_mass = _json_float(data, 'liquid_mass')
        baffle_width_ratio = _json_float(data, 'baffle_width_ratio')
        baffle_depth_ratio = _json_float(data, 'baffle_depth_ratio')
        coincidence_margin = _json_float(data, 'coincidence_margin')

        def _freq_list(key):
            v = data.get(key)
            if v in (None, ''):
                return None
            if isinstance(v, (list, tuple)):
                out = [float(x) for x in v if x not in (None, '')]
                return out or None
            # virgülle ayrılmış string de kabul et (panel kolaylığı)
            try:
                out = [float(x) for x in str(v).split(',') if x.strip()]
                return out or None
            except (TypeError, ValueError):
                raise ValueError(f"'{key}' must be a list of frequencies [Hz]")

        analyze_kwargs = dict(
            radius=radius,
            fill_height=fill_height,
            g_eff=(9.80665 if g_eff is None else g_eff),
            fluid_density=fluid_density,
            liquid_mass=liquid_mass,
            baffle_depth_ratio=(0.10 if baffle_depth_ratio is None
                                else baffle_depth_ratio),
            control_frequencies=_freq_list('control_frequencies'),
            structural_frequencies=_freq_list('structural_frequencies'),
        )
        if baffle_width_ratio is not None:
            analyze_kwargs['baffle_width_ratio'] = baffle_width_ratio
        if coincidence_margin is not None:
            analyze_kwargs['coincidence_margin'] = coincidence_margin

        result = analyze_slosh(**analyze_kwargs)
        return jsonify(sanitize_json_values(
            {'status': 'success', 'slosh': result}))
    except (ValueError, KeyError) as e:
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/pressurant-sizing', methods=['POST'])
def pressurant_sizing_api():
    """Basınçlandırıcı gaz (helyum/azot) boyutlandırma — regüleli veya
    blowdown besleme mimarisi (Sutton 9. baskı Böl. 6; Huzel & Huang Böl. 5).

    Girdi (JSON): mode 'regulated' (vars.) | 'blowdown'.
      Ortak: propellant_volume (m^3, ZORUNLU); gas 'helium'|'nitrogen';
        initial_temperature (K, vars. 293.15).
      regulated: tank_pressure (bar, ZORUNLU); storage_pressure (bar, vars.
        200); regulator_margin (vars. 0.10); collapse_factor (vars. 1.0).
      blowdown: initial_ullage_volume (m^3, ZORUNLU); initial_pressure (bar,
        ZORUNLU); polytropic_n (vars. 1.2).

    Yanıt 200: {'status':'success','mode':mode,'pressurant': result}.
    Hata 400: {'status':'error','error': mesaj}.
    """
    try:
        data = request.json or {}
        from hrma.analysis.pressurant_sizing import analyze_pressurant

        mode = str(data.get('mode', 'regulated')).strip().lower()
        if mode not in ('regulated', 'blowdown'):
            raise ValueError("'mode' must be 'regulated' or 'blowdown'")

        propellant_volume = _json_float(data, 'propellant_volume')
        if propellant_volume is None:
            raise ValueError("'propellant_volume' is required (m^3)")
        initial_temperature = _json_float(data, 'initial_temperature')
        gas = data.get('gas', 'helium' if mode == 'regulated' else 'nitrogen')

        if mode == 'regulated':
            tank_pressure_pa = _bar_to_pa(data, 'tank_pressure')
            if tank_pressure_pa is None:
                raise ValueError("'tank_pressure' is required (bar)")
            storage_pressure_pa = _bar_to_pa(data, 'storage_pressure')
            regulator_margin = _json_float(data, 'regulator_margin')
            collapse_factor = _json_float(data, 'collapse_factor')
            kwargs = dict(
                mode='regulated',
                propellant_volume=propellant_volume,
                tank_pressure=tank_pressure_pa,
                gas=gas,
            )
            if initial_temperature is not None:
                kwargs['initial_temperature'] = initial_temperature
            if storage_pressure_pa is not None:
                kwargs['storage_pressure'] = storage_pressure_pa
            if regulator_margin is not None:
                kwargs['regulator_margin'] = regulator_margin
            if collapse_factor is not None:
                kwargs['collapse_factor'] = collapse_factor
        else:  # blowdown
            initial_ullage_volume = _json_float(data, 'initial_ullage_volume')
            initial_pressure_pa = _bar_to_pa(data, 'initial_pressure')
            if initial_ullage_volume is None:
                raise ValueError("'initial_ullage_volume' is required (m^3)")
            if initial_pressure_pa is None:
                raise ValueError("'initial_pressure' is required (bar)")
            polytropic_n = _json_float(data, 'polytropic_n')
            kwargs = dict(
                mode='blowdown',
                propellant_volume=propellant_volume,
                initial_ullage_volume=initial_ullage_volume,
                initial_pressure=initial_pressure_pa,
                gas=gas,
            )
            if initial_temperature is not None:
                kwargs['initial_temperature'] = initial_temperature
            if polytropic_n is not None:
                kwargs['polytropic_n'] = polytropic_n

        result = analyze_pressurant(**kwargs)
        return jsonify(sanitize_json_values(
            {'status': 'success', 'mode': mode, 'pressurant': result}))
    except (ValueError, KeyError) as e:
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/water-hammer', methods=['POST'])
def water_hammer_api():
    """Su koçu (water hammer) besleme hattı geçici basınç analizi
    (Joukowsky/Allievi + ince-cidar hoop basınç sınıfı kıyası).

    Girdi (JSON): fluid 'water'|'n2o'|'rp1'|'lox' (özel sıvı için
      bulk_modulus_Pa + density_kg_m3 birlikte); line_length_m (ZORUNLU),
      line_id_mm (ZORUNLU), wall_thickness_mm (ZORUNLU), working_pressure_bar
      (ZORUNLU); mdot_kg_s VEYA flow_velocity_m_s (en az biri); pipe_material
      (materials_db anahtarı, vars. ss_304); valve_closure_time_ms (ops. —
      None ani kapanma); pipe_mawp_bar (ops.); delta_v_m_s (ops.).

    Yanıt 200: {'status':'success','water_hammer': WaterHammerAnalyzer.
      analyze()} — wave_speed, critical_closure_time, joukowsky/applied
      pressure rise, peak_pressure, pipe MAWP/akma/kopma, status
      (SAFE/MARGINAL/UNSAFE), recommendation, recommended_closure_time_ms.
    Hata 400: {'status':'error','error': mesaj}.
    """
    try:
        data = request.json or {}
        from hrma.analysis.water_hammer import WaterHammerAnalyzer

        fluid = data.get('fluid', 'water')
        line_length_m = _json_float(data, 'line_length_m')
        line_id_mm = _json_float(data, 'line_id_mm')
        wall_thickness_mm = _json_float(data, 'wall_thickness_mm')
        working_pressure_bar = _json_float(data, 'working_pressure_bar')
        if line_length_m is None:
            raise ValueError("'line_length_m' is required (m)")
        if line_id_mm is None:
            raise ValueError("'line_id_mm' is required (mm)")
        if wall_thickness_mm is None:
            raise ValueError("'wall_thickness_mm' is required (mm)")
        if working_pressure_bar is None:
            raise ValueError("'working_pressure_bar' is required (bar)")

        result = WaterHammerAnalyzer().analyze(
            fluid=fluid,
            line_length_m=line_length_m,
            line_id_mm=line_id_mm,
            wall_thickness_mm=wall_thickness_mm,
            working_pressure_bar=working_pressure_bar,
            mdot_kg_s=_json_float(data, 'mdot_kg_s'),
            flow_velocity_m_s=_json_float(data, 'flow_velocity_m_s'),
            valve_closure_time_ms=_json_float(data, 'valve_closure_time_ms'),
            pipe_material=data.get('pipe_material', 'ss_304'),
            pipe_mawp_bar=_json_float(data, 'pipe_mawp_bar'),
            bulk_modulus_Pa=_json_float(data, 'bulk_modulus_Pa'),
            density_kg_m3=_json_float(data, 'density_kg_m3'),
            delta_v_m_s=_json_float(data, 'delta_v_m_s'),
        )
        return jsonify(sanitize_json_values(
            {'status': 'success', 'water_hammer': result}))
    except (ValueError, KeyError) as e:
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


class EngineContractViolation(TypeError):
    """Motor ``calculate_performance()`` sözleşme ihlali (tip kapısı).

    v2.6.27 — sessiz 200 kusuru (denetçi kanıtı, 2026-08-04): motor dict
    yerine ``None`` döndürdüğünde ardındaki geniş çizim ``try/except``leri
    ``None.setdefault`` hatasını yutuyor, uç ``success`` loglayıp HTTP 200
    ``null`` gövdesi dönüyordu. Sözlük dışı sonuç sunucu tarafı bir sözleşme
    ihlalidir — istemci hatası (400) değil; bu sınıf hata dalında 500'e
    eşlenir. ``TypeError`` alt sınıfı olması bilinçli: kapı bir tip kapısıdır
    ve gövdedeki ``error_type`` alanı da bunu söyler.
    """


def _require_dict_result(results, engine_name, route_name):
    """Tip kapısı: motor sonucu dict değilse EngineContractViolation yükseltir."""
    if not isinstance(results, dict):
        raise EngineContractViolation(
            f'{engine_name}.calculate_performance() must return a dict, got '
            f'{type(results).__name__}; refusing to report {route_name} '
            f'success for a result that was never computed.')
    return results


def _build_solid_engine(data, **base_kwargs):
    """Katı motoru kurar ve yanma hızı katsayılarının KAYNAĞINI beyan eder.

    v2.6.27 (A3, Ayberk madde 3) — ÖLÇÜLEN KUSUR: uç katmanı
    ``burn_rate_a`` için 0,005, ``burn_rate_n`` için 0,35 varsayılanı
    enjekte ediyordu. Bu çift, merkezî katalogun apcp kaydından
    (a = 0,0022334, n = 0,35) 2,24 KAT sapar — referans basınçta
    18,18 mm/s'ye karşı 8,12 mm/s. Motor bunu KULLANICI girdisi sanıyor ve
    ``warn.solid.burn_rate_off_catalog`` uyarısını ateşliyordu; yani
    kullanıcı, kendi girmediği bir sayı yüzünden uyarı alıyordu. Sapmanın
    bedeli gözle görünmez: toplam impuls doğru kalır, yanma süresi yarılanır,
    ortalama itki iki katına çıkar.

    Sözleşme:
      * Alan geldiyse motora AYNEN geçer, kaynağı ``'request'``.
      * Gelmediyse motora HİÇ geçmez ve değer SEÇİLEN yakıtın merkezî
        katalog kaydından çözülür (CLAUDE.md kural 11: sayı tek kaynaktan).
        Kaynak ``'central_catalog:<yakıt>'`` diye yazılır.
      * Yakıt katalogda yoksa motorun kurucu varsayılanı kalır — ama o
        varsayılan APCP'nin katsayısıdır ve seçilen yakıtın DEĞİLDİR, bu
        yüzden ``'engine_constructor_default:apcp_catalog'`` diye adıyla
        beyan edilir.

    Yakıt kimliğini uç katmanı yeniden çözmez: motorun kendi
    ``_resolve_propellant_type`` işlevi (öncelik: ``propellant_type`` >
    kurucu argümanı > ``burn_rate_preset`` > ``propellant_name``) zaten
    koşmuştur, sonucu ``motor.propellant_type``. Katsayı değişmesi
    gerekiyorsa motor o kimlikle YENİDEN kurulur; kurulum maliyeti ölçüldü
    (0,09 ms) ve bu, motorun iç durumunu dışarıdan elle değiştirmekten
    güvenlidir.

    Args:
        data: İstek gövdesi (``overrides`` olarak da motora geçer).
        **base_kwargs: ``SolidRocketEngine`` kurucusunun katsayı DIŞINDAKİ
            argümanları.

    Returns:
        ``(motor, burn_rate_sources)`` — ikinci öğe yanıta konacak beyan.
    """
    burn_rate_a = data.get('burn_rate_a')
    burn_rate_n = data.get('burn_rate_n')
    motor_kwargs = dict(base_kwargs)
    motor_kwargs['overrides'] = data
    if burn_rate_a not in (None, ''):
        motor_kwargs['burn_rate_a'] = float(burn_rate_a)
    if burn_rate_n not in (None, ''):
        motor_kwargs['burn_rate_n'] = float(burn_rate_n)
    motor = SolidRocketEngine(**motor_kwargs)

    sources = {
        'burn_rate_a_source': ('request' if burn_rate_a not in (None, '')
                               else None),
        'burn_rate_n_source': ('request' if burn_rate_n not in (None, '')
                               else None),
    }
    if None in sources.values():
        from hrma.engines.solid_rocket_engine import _catalog_burn_rate
        a_cat, n_cat = _catalog_burn_rate(motor.propellant_type)
        katalog_etiketi = f'central_catalog:{motor.propellant_type}'
        yedek_etiket = 'engine_constructor_default:apcp_catalog'
        degisti = False
        if sources['burn_rate_a_source'] is None:
            if a_cat is not None:
                motor_kwargs['burn_rate_a'] = float(a_cat)
                sources['burn_rate_a_source'] = katalog_etiketi
                degisti = True
            else:
                sources['burn_rate_a_source'] = yedek_etiket
        if sources['burn_rate_n_source'] is None:
            if n_cat is not None:
                motor_kwargs['burn_rate_n'] = float(n_cat)
                sources['burn_rate_n_source'] = katalog_etiketi
                degisti = True
            else:
                sources['burn_rate_n_source'] = yedek_etiket
        if degisti:
            motor = SolidRocketEngine(**motor_kwargs)

    sources['burn_rate_a_used'] = float(motor.a)
    sources['burn_rate_n_used'] = float(motor.n)
    sources['_basis'] = (
        'Saint-Robert coefficients are never invented by the API layer. '
        "A value marked 'request' came from the request body; "
        "'central_catalog:<propellant>' means the field was not supplied and "
        'the value was resolved from the central propellant catalog for the '
        'propellant the solver resolved. Whether the solver actually uses '
        "this pair is a separate question - see 'burn_rate_basis' (a "
        'piecewise regime table overrides it).')
    return motor, sources


@app.route('/calculate_solid', methods=['POST'])
def calculate_solid():
    trace_id = _request_trace_id()
    try:
        data = request.json or {}
        # D7: istek GÖVDESİ loglanmaz (gerekçe: _request_trace_id docstring'i).
        # Loga yalnız kararlı olay adı + korelasyon kimliği + alan SAYISI
        # gider; hiçbir alan adı ve hiçbir değer yazılmaz.
        print(f"[{trace_id}] calculate_solid.request_accepted "
              f"fields={len(data)}")

        # --- Faz 4B, bulgu 57.3 (katı eşleniği): eksik kritik girdi
        # sessizce dolmaz. ÖLÇÜM (2 Ağustos 2026): `POST /calculate_solid
        # -d '{}'` HTTP 200 ve `status: CALCULATED` ile eksiksiz bir
        # Ø100 / 500 mm / APCP / 40 bar tasarımı döndürüyordu. Sıvı ucuyla
        # AYNI sözleşme uygulanır; iki uç arasında farklı davranmak
        # kullanıcının hangisine güveneceğini bilememesi demektir.
        # burn_rate_a/n zorunlu DEĞİL: verilmezse yakıt veritabanından
        # türetilir, yani kaynağı bellidir.
        #
        # YAKIT KİMLİĞİ TEK BİR ANAHTARDA DEĞİL. İlk denemede kritik alan
        # listesine `propellant_type` konmuştu ve 45 test kırıldı: gerçek
        # kayıtlı projeler (examples/*.hrma) yakıtı `propellant_name` +
        # `burn_rate_preset` ile taşıyor, `propellant_type` hiç yok. Kapı,
        # veri modelinin kendisine uydurulur — tersi değil. Kimlik şu
        # yollardan HERHANGİ biriyle karşılanır; hiçbiri yoksa reddedilir.
        # UCUN İKİ GEÇERLİ GİRDİ KİPİ VAR ve kapı ikisini de tanımak
        # zorundadır:
        #   (1) Geometri kipi  — çap/boy/çekirdek verilir, performans çıkar.
        #   (2) Tasarım noktası kipi — itki + yanma süresi verilir, geometri
        #       BOYUTLANDIRILIR (bkz. tests/test_solid_design_point.py).
        # İlk denemede yalnız (1) zorunlu tutulmuştu; (2) kipini kullanan
        # meşru istekler 422 alıyordu. Ortak zorunluluk: oda basıncı + yakıt
        # kimliği; buna ek olarak iki kipten EN AZ BİRİ eksiksiz olmalı.
        GEOMETRI_ALANLARI = ('chamber_diameter', 'grain_length', 'core_diameter')
        TASARIM_NOKTASI_ALANLARI = ('thrust', 'burn_time')
        YAKIT_KIMLIK_ANAHTARLARI = ('propellant_type', 'propellant_name',
                                    'burn_rate_preset')
        OGRETICI_VARSAYILAN = {'chamber_diameter': 100, 'grain_length': 500,
                               'core_diameter': 30, 'chamber_pressure': 40,
                               'propellant_type': 'apcp'}

        def _dolu(ad):
            return data.get(ad) not in (None, '')

        ogretici_mod = bool(data.get('use_tutorial_defaults'))
        eksik_girdiler = []
        if not _dolu('chamber_pressure'):
            eksik_girdiler.append('chamber_pressure')
        if not (any(_dolu(ad) for ad in YAKIT_KIMLIK_ANAHTARLARI)
                or (_dolu('burn_rate_a') and _dolu('burn_rate_n'))):
            eksik_girdiler.append('propellant_type')
        geometri_tam = all(_dolu(ad) for ad in GEOMETRI_ALANLARI)
        tasarim_noktasi_tam = all(_dolu(ad) for ad in TASARIM_NOKTASI_ALANLARI)
        if not (geometri_tam or tasarim_noktasi_tam):
            eksik_girdiler.extend(
                ad for ad in GEOMETRI_ALANLARI if not _dolu(ad))
        if eksik_girdiler and not ogretici_mod:
            print(f"[{trace_id}] calculate_solid.incomplete_input "
                  f"missing={len(eksik_girdiler)}")
            return jsonify({
                'status': 'incomplete_input',
                # i18n: backend EN üretir; TR karşılığı i18n_charts.js sözlüğünde.
                'error': ('Required inputs for the solid motor calculation '
                          'are missing; defaults were not applied and no '
                          'design was produced.'),
                'missing_fields': eksik_girdiler,
                'required_fields': {
                    'always': ['chamber_pressure', 'propellant_type'],
                    'either_geometry': list(GEOMETRI_ALANLARI),
                    'or_design_point': list(TASARIM_NOKTASI_ALANLARI),
                },
                'hint': ('Geometry mode takes diameter/length/core, '
                         'design-point mode takes thrust + burn time. For '
                         'the tutorial scenario send '
                         '"use_tutorial_defaults": true; the result declares '
                         'which inputs came from defaults in the '
                         '"defaults_applied" field.'),
            }), 422
        uygulanan_varsayilanlar = []
        if ogretici_mod and eksik_girdiler:
            for ad in eksik_girdiler:
                if ad in OGRETICI_VARSAYILAN:
                    data[ad] = OGRETICI_VARSAYILAN[ad]
                    uygulanan_varsayilanlar.append(ad)

        # Solid motor input validation
        chamber_diameter = data.get('chamber_diameter', 100)
        validate_input_range(chamber_diameter, 10, 2000, "Chamber diameter (mm)")
        
        grain_length = data.get('grain_length', 500)
        validate_input_range(grain_length, 50, 5000, "Grain length (mm)")
        
        core_diameter = data.get('core_diameter', 30)
        validate_input_range(core_diameter, 5, chamber_diameter-5, "Core diameter (mm)")
        
        chamber_pressure = data.get('chamber_pressure', 40)
        validate_input_range(chamber_pressure, 5, 200, "Chamber pressure (bar)")
        
        # --- v2.6.27 (A3 / Ayberk madde 3): a ve n ARTIK UÇTA UYDURULMUYOR --
        # ÖLÇÜLDÜ (bu blok değişmeden önce): a/n GÖNDERİLMEDEN yapılan
        # ``POST /calculate_solid`` (apcp) çağrısında uç katmanı
        # ``a = 0,005`` ve ``n = 0,35`` enjekte ediyordu. Bu çift, merkezî
        # katalogun apcp kaydından (a = 0,0022334, n = 0,35) 2,24 KAT sapar:
        # referans basınçta 18,18 mm/s'ye karşı 8,12 mm/s. Motor bu değeri
        # KULLANICI GİRDİSİ sanıp ``warn.solid.burn_rate_off_catalog``
        # uyarısını ateşliyordu — yani kullanıcı, kendi girmediği bir sayı
        # yüzünden "kataloğun dışındasın" uyarısı alıyordu.
        # Sapmanın bedeli gözle görünmez: toplam impuls doğru kalır, yanma
        # süresi yarılanır, ortalama itki iki katına çıkar
        # (solid_rocket_engine.py ``_check_burn_rate_coefficients``).
        #
        # Yeni sözleşme: alan gelmediyse motora HİÇ geçilmez; değer SEÇİLEN
        # yakıtın merkezî katalog kaydından çözülür (CLAUDE.md kural 11 —
        # sayı tek kaynaktan gelir) ve kaynağı yanıtta beyan edilir.
        # NOT: motorun kurucu varsayılanı ``DEFAULT_BURN_RATE_A/N``
        # APCP'nin katsayısıdır ve SEÇİLEN yakıttan bağımsızdır; KNDX seçen
        # kullanıcıya APCP'nin yanma hızını vermemek için katalog çözümü
        # yakıt çözüldükten SONRA yapılır (aşağıda).
        burn_rate_a = data.get('burn_rate_a')
        burn_rate_n = data.get('burn_rate_n')
        if burn_rate_a not in (None, ''):
            validate_input_range(burn_rate_a, 0.0001, 0.1,
                                 "Burn rate coefficient")

        # Alt sınır -0.5: KN-şeker plateau/mesa rejimlerinde n NEGATİFTİR
        # (Nakka 1999 KNDX n=-0.148, KNSB n=-0.314 — bkz. burn_rate_db).
        # Eski [0.1, 1.0] aralığı merkezi db preset'lerini reddediyordu.
        if burn_rate_n not in (None, ''):
            validate_input_range(burn_rate_n, -0.5, 1.0, "Burn rate exponent")

        # Create solid motor instance
        # overrides=data: formun yoğunluk/C*/gama/segman/star/sıcaklık gibi
        # alanları motora işlensin (2026-07-13 — girdi-backend kopukluğu fixi;
        # motor yalnız tanıdığı ve fiziksel aralıktaki anahtarları uygular)
        motor, burn_rate_sources = _build_solid_engine(
            data,
            grain_type=data.get('grain_type', 'bates'),
            propellant_type=data.get('propellant_type', 'apcp'),
            chamber_diameter=chamber_diameter,
            grain_length=grain_length,
            core_diameter=core_diameter,
            chamber_pressure=chamber_pressure,
        )

        # Calculate motor performance
        results = motor.calculate_performance()

        # v2.6.27 tip kapısı — sözlük dışı sonuç burada durdurulur; aşağıdaki
        # çizim try/except'leri artık DARALTILDIĞI için bu hatayı yutamaz
        # (eski akış: None.setdefault yutulur, 'success' loglanır, HTTP 200
        # 'null' döner — bkz. EngineContractViolation).
        results = _require_dict_result(
            results, 'SolidRocketEngine', 'calculate_solid')

        # Sanitize results
        sanitized_results = sanitize_json_values(results)

        # v2.6.27 (A3): yanma hızı katsayılarının KAYNAĞI yanıtta durur.
        sanitized_results['burn_rate_inputs'] = burn_rate_sources

        # v2.6.27 (A3): kazanılmamış kap onayı geri çekilir. Ölçüldü:
        # kullanıcı hiç kasa kalınlığı vermediği hâlde dört ayrı hazne
        # basıncında (20/40/70/100 bar) ``pressure_safety.vessel_status``
        # 'PASS' dönüyordu. Sebep totoloji: cidarı HRMA'nın kendisi
        # ``_case_design()`` içinde Barlow ile SF'yi sağlayacak şekilde
        # boyutlandırıyor, sonra aynı cidarı sınayıp geçtiğini söylüyor.
        _withhold_unearned_vessel_verdict(
            sanitized_results,
            case_thickness_supplied=(data.get('case_thickness')
                                     not in (None, '')))

        # Hibrit paritesi: motor kesiti + ortak geometri (2026-07-13)
        # v2.6.27: geometri dönüşümü try DIŞINA alındı — geniş kapsam,
        # motor sonucunun bozukluğunu da yutuyordu; koruma yalnız kesit
        # ÇİZİMİNİN kendisi içindir.
        geo = solid_results_to_motor_geometry(sanitized_results)
        sanitized_results['motor_geometry'] = sanitize_json_values(geo)
        try:
            sanitized_results.setdefault('plots', {})['motor'] = \
                create_improved_motor_cross_section(geo, motor_type='solid')
        except Exception:
            traceback.print_exc()  # kesit çizimi hesabı düşürmesin

        # Hibrit paritesi (v2.5.2): performans panosu artık motor tipini
        # sonuç sözlüğünden kendisi tespit ediyor, tek argümanla çağrılır.
        try:
            sanitized_results.setdefault('plots', {})['performance'] = \
                create_performance_plots(sanitized_results)
        except Exception:
            traceback.print_exc()  # pano hesabı düşürmesin

        # Öğretici modda sonuç, girdisinin nereden geldiğini TAŞIR
        # (Faz 4B, bulgu 57.3 — sıvı ucuyla aynı sözleşme).
        if uygulanan_varsayilanlar:
            sanitized_results['defaults_applied'] = uygulanan_varsayilanlar
            sanitized_results['input_source'] = 'tutorial_defaults'

        print(f"[{trace_id}] calculate_solid.success")
        return jsonify(sanitized_results)

    except Exception as e:
        # D7: hata dalında da gövde yazılmaz. Yığın izi değişken DEĞERLERİ
        # taşımaz (yalnız dosya/satır/kaynak metni), bu yüzden kalabilir.
        error_traceback = traceback.format_exc()
        print(f"[{trace_id}] calculate_solid.error "
              f"type={type(e).__name__}")
        print(f"[{trace_id}] Traceback: {error_traceback}")
        # v2.6.26: traceback yanittan cikarildi (bkz. /calculate notu).
        # v2.6.27 tip kapısı: motor sözleşme ihlali SUNUCU hatasıdır — 500.
        # Diğer istisnalar (girdi doğrulama vb.) 400 olarak kalır.
        return jsonify({
            'error': str(e),
            'error_type': type(e).__name__,
            # Kullanıcı destek talebinde bu kimliği verir; log satırı
            # gövdeyi saklamadan olayla eşleşir.
            'trace_id': trace_id
        }), (500 if isinstance(e, EngineContractViolation) else 400)

# --- Motor stdout tekrar bastırma (konsol gürültüsü) ------------------------
# ÖLÇÜLDÜ (2026-08-04, scratchpad noise_before): TEK /calculate_liquid çağrısı
# stdout'a 301 satır basıyor, bunların yalnız 105'i benzersiz — "NASA
# Validation" bloğu 27, "Effective C* set" satırları 4'er kez aynen
# tekrarlıyor (kaynak: hrma/engines/liquid_rocket_engine.py; motor dosyası bu
# bakım dalgasının kapsamı DIŞINDA, o yüzden süzgeç burada, çağıran katmanda).
# Kural: bilgi SİLİNMEZ — her benzersiz satır İLK görüldüğü anda aynen basılır,
# yalnız birebir aynı tekrarlar bastırılır ve kapanışta tek satırlık sayım
# beyan edilir (total/unique/suppressed). Bekçi + ölçüm:
# tests/test_istek_performansi.py.

class _ThreadLocalLineDedup(io.TextIOBase):
    """sys.stdout üstüne BİR kez takılan, iş parçacığı-yerel süzgeç.

    Kapsam (begin_scope/end_scope) yalnız kendi iş parçacığında etkilidir;
    kapsam yokken her yazma olduğu gibi alttaki akışa geçer. Böylece eşzamanlı
    bir başka istek ne süzülür ne bozulur — contextlib.redirect_stdout'un
    süreç-genel (ve iç içe çıkışta yanlış akım bırakabilen) davranışından
    bilinçli olarak kaçınıldı.
    """

    def __init__(self, target):
        self._target = target
        self._scopes = threading.local()

    @property
    def target(self):
        return self._target

    def writable(self):
        return True

    def write(self, s):
        state = getattr(self._scopes, 'state', None)
        if state is None:
            return self._target.write(s)
        state['buffer'] += s
        while '\n' in state['buffer']:
            line, state['buffer'] = state['buffer'].split('\n', 1)
            state['total'] += 1
            seen = state['seen']
            count = seen.get(line, 0)
            seen[line] = count + 1
            if count == 0:  # ilk görülüş aynen geçer, tekrarlar bastırılır
                self._target.write(line + '\n')
        return len(s)

    def flush(self):
        self._target.flush()

    def begin_scope(self):
        state = {'buffer': '', 'seen': {}, 'total': 0}
        self._scopes.state = state
        return state

    def end_scope(self):
        state = getattr(self._scopes, 'state', None)
        self._scopes.state = None
        if state and state['buffer']:
            # Satır sonu gelmemiş kuyruk parçası kaybolmaz, aynen basılır.
            self._target.write(state['buffer'])
        return state


_STDOUT_DEDUP_LOCK = threading.Lock()


def _install_stdout_dedup():
    """Süzgeci sys.stdout üstüne (bir kez) tak ve döndür.

    pytest her testte sys.stdout'u kendi yakalayıcısıyla değiştirir; o yüzden
    "takılı mı" denetimi her çağrıda yeniden yapılır — eski sarmalayıcı
    değiştirilen akışla birlikte doğal olarak devreden çıkar.
    """
    with _STDOUT_DEDUP_LOCK:
        if not isinstance(sys.stdout, _ThreadLocalLineDedup):
            sys.stdout = _ThreadLocalLineDedup(sys.stdout)
        return sys.stdout


@contextlib.contextmanager
def _dedup_engine_stdout(trace_id, route_name):
    """Motor hesabı boyunca birebir tekrar eden stdout satırlarını bastır.

    Çıkışta tek satırlık sayım basar (yalnız tekrar VARSA): kaç satır geldi,
    kaçı benzersizdi, kaçı bastırıldı — bilgi silinmez, tekrar kesilir.
    """
    stream = _install_stdout_dedup()
    stream.begin_scope()
    try:
        yield
    finally:
        state = stream.end_scope()
        if state and state['total'] > len(state['seen']):
            suppressed = state['total'] - len(state['seen'])
            print(f"[{trace_id}] {route_name}.stdout_dedup "
                  f"total={state['total']} unique={len(state['seen'])} "
                  f"suppressed={suppressed}")


@app.route('/calculate_liquid', methods=['POST'])
def calculate_liquid():
    trace_id = _request_trace_id()
    try:
        data = request.json or {}
        # D7: istek GÖVDESİ loglanmaz (gerekçe: _request_trace_id docstring'i).
        print(f"[{trace_id}] calculate_liquid.request_accepted "
              f"fields={len(data)}")

        # --- Faz 4B, bulgu 57.3: eksik kritik girdi sessizce dolmaz ---
        # ÖLÇÜM (2 Ağustos 2026): `POST /calculate_liquid -d '{}'` HTTP 200 ve
        # eksiksiz bir 10 kN / RP1-LOX / 100 bar tasarımı döndürüyordu. Çağıran,
        # bu sayıların KENDİ girdisinden mi yoksa gövdedeki `data.get(..., X)`
        # varsayılanlarından mı geldiğini yanıta bakarak ayırt edemiyordu.
        # Beş kritik alan artık zorunlu; hiçbiri için sessiz varsayılan yok.
        # Öğretici/tanıtım senaryosu kaybolmasın diye AÇIK bir katılım anahtarı
        # bırakıldı (`use_tutorial_defaults`); o yolda sonuç, hangi alanların
        # varsayılandan geldiğini `defaults_applied` ile beyan eder.
        # Arayüz kırılmaz: liquid.html `collectAllParameters()` beş alanın
        # hepsini gönderiyor (thrust, chamber_pressure, mixture_ratio,
        # fuel_type, oxidizer_type).
        KRITIK_GIRDILER = ('thrust', 'chamber_pressure', 'mixture_ratio',
                           'fuel_type', 'oxidizer_type')
        OGRETICI_VARSAYILAN = {'thrust': 10000, 'chamber_pressure': 100,
                               'mixture_ratio': 2.5, 'fuel_type': 'rp1',
                               'oxidizer_type': 'lox'}
        ogretici_mod = bool(data.get('use_tutorial_defaults'))
        eksik_girdiler = [ad for ad in KRITIK_GIRDILER
                          if data.get(ad) in (None, '')]
        if eksik_girdiler and not ogretici_mod:
            print(f"[{trace_id}] calculate_liquid.incomplete_input "
                  f"missing={len(eksik_girdiler)}")
            return jsonify({
                'status': 'incomplete_input',
                # i18n: backend EN üretir; TR karşılığı i18n_charts.js sözlüğünde.
                'error': ('Required inputs for the liquid motor calculation '
                          'are missing; defaults were not applied and no '
                          'design was produced.'),
                'missing_fields': eksik_girdiler,
                'required_fields': list(KRITIK_GIRDILER),
                'hint': ('For the tutorial/demo scenario send '
                         '"use_tutorial_defaults": true; the result declares '
                         'which inputs came from defaults in the '
                         '"defaults_applied" field.'),
            }), 422
        uygulanan_varsayilanlar = []
        if ogretici_mod and eksik_girdiler:
            for ad in eksik_girdiler:
                data[ad] = OGRETICI_VARSAYILAN[ad]
                uygulanan_varsayilanlar.append(ad)

        # Liquid motor input validation
        thrust = data.get('thrust', 10000)
        validate_positive(thrust, "Thrust")
        validate_input_range(thrust, 100, 1e7, "Thrust (N)")
        
        chamber_pressure = data.get('chamber_pressure', 100)
        validate_input_range(chamber_pressure, 10, 500, "Chamber pressure (bar)")
        
        mixture_ratio = data.get('mixture_ratio', 2.5)
        validate_input_range(mixture_ratio, 0.5, 20, "Mixture ratio")
        
        # Validate tank pressure (Issue #6)
        tank_pressure = data.get('tank_pressure', chamber_pressure * 1.5)
        is_valid, msg = validation.validate_pressure_consistency(tank_pressure, chamber_pressure)
        if not is_valid:
            raise ValueError(msg)
        
        # Create liquid motor instance
        # v2.5.2: formdaki ~55 sayısal girdinin HİÇBİRİ motora ulaşmıyordu
        # (kurucu 7 parametre alıyordu, katı motordaki overrides bağlantısı
        # sıvıda hiç kurulmamıştı). Kullanıcı genişleme oranı, L*, soğutma
        # kanalı, enjektör ΔP gibi onlarca alanı doldurup sonucun değişmediğini
        # göremiyordu. Motor artık aralık doğrulamalı `overrides` kabul ediyor;
        # bağlanamayan alanlar sonuçta `unwired_inputs`, aralık dışı değerler
        # `input_warnings` ile AÇIKÇA beyan ediliyor (sessiz yutma yok).
        # Konsol gürültüsü: motor iç döngüleri aynı bilgi satırlarını yüzlerce
        # kez basıyor (ölçüm ve gerekçe _ThreadLocalLineDedup üstündeki notta).
        # Süzgeç yalnız motor hesabı + figür üretimini sarar; rotanın kendi
        # [trace_id] satırları kapsam dışında kalır.
        with _dedup_engine_stdout(trace_id, 'calculate_liquid'):
            engine = LiquidRocketEngine(
                thrust=thrust,
                chamber_pressure=chamber_pressure,
                mixture_ratio=mixture_ratio,
                fuel_type=data.get('fuel_type', 'rp1'),
                oxidizer_type=data.get('oxidizer_type', 'lox'),
                cooling_type=data.get('cooling_type', 'regenerative'),
                injector_type=data.get('injector_type', 'impinging'),
                overrides=data
            )

            # Calculate engine performance
            results = engine.calculate_performance()

            # v2.6.27 tip kapısı — bkz. calculate_solid'deki eş kapı ve
            # EngineContractViolation gerekçesi (sessiz 200 kusuru).
            results = _require_dict_result(
                results, 'LiquidRocketEngine', 'calculate_liquid')

            # Sanitize results
            sanitized_results = sanitize_json_values(results)

            # Hibrit paritesi: motor kesiti + ortak geometri (2026-07-13)
            # v2.6.27: geometri dönüşümü try DIŞINA alındı — geniş kapsam,
            # motor sonucunun bozukluğunu da yutuyordu; koruma yalnız kesit
            # ÇİZİMİNİN kendisi içindir.
            geo = liquid_results_to_motor_geometry(sanitized_results)
            sanitized_results['motor_geometry'] = sanitize_json_values(geo)
            try:
                sanitized_results.setdefault('plots', {})['motor'] = \
                    create_improved_motor_cross_section(geo, motor_type='liquid')
            except Exception:
                traceback.print_exc()  # kesit çizimi hesabı düşürmesin

            # Hibrit paritesi (v2.5.2): performans panosu artık motor tipini
            # sonuç sözlüğünden kendisi tespit ediyor, tek argümanla çağrılır.
            try:
                sanitized_results.setdefault('plots', {})['performance'] = \
                    create_performance_plots(sanitized_results)
            except Exception:
                traceback.print_exc()  # pano hesabı düşürmesin

        # Öğretici modda üretilen sonuç, kendi girdisinin nereden geldiğini
        # TAŞIR: sayıya bakan biri bunun kullanıcı tasarımı değil tanıtım
        # senaryosu olduğunu yanıttan görebilmeli (Faz 4B, bulgu 57.3).
        if uygulanan_varsayilanlar:
            sanitized_results['defaults_applied'] = uygulanan_varsayilanlar
            sanitized_results['input_source'] = 'tutorial_defaults'

        print(f"[{trace_id}] calculate_liquid.success")
        return jsonify(sanitized_results)

    except Exception as e:
        # D7: hata dalında da gövde yazılmaz (bkz. /calculate_solid notu).
        error_traceback = traceback.format_exc()
        print(f"[{trace_id}] calculate_liquid.error "
              f"type={type(e).__name__}")
        print(f"[{trace_id}] Traceback: {error_traceback}")
        # v2.6.26: traceback yanittan cikarildi (bkz. /calculate notu).
        # v2.6.27 tip kapısı: motor sözleşme ihlali SUNUCU hatasıdır — 500.
        return jsonify({
            'error': str(e),
            'error_type': type(e).__name__,
            'trace_id': trace_id
        }), (500 if isinstance(e, EngineContractViolation) else 400)

@app.route('/api/solid-monte-carlo', methods=['POST'])
def solid_monte_carlo():
    """Katı motor Monte Carlo analizi — üretim toleransı belirsizlikleri.

    Girdi: /calculate_solid ile aynı form alanları + opsiyonel n_samples.
    Çıktı: başarı oranı, itki/Isp/yanma süresi/tepe basıncı istatistikleri
    ve histogram verileri (frontend çizer).
    """
    try:
        data = request.json or {}
        # v2.6.27 (A3): katsayı enjeksiyonu burada da vardı ve etkisi daha
        # büyüktü — Monte Carlo, sapmalı bir NOMİNAL etrafında saçılım
        # üretiyordu (a katalogtan 2,24 kat uzakken tüm dağılım kayar).
        # Kurulum /calculate_solid ile AYNI yoldan yapılır.
        motor, burn_rate_sources = _build_solid_engine(
            data,
            grain_type=data.get('grain_type', 'bates'),
            propellant_type=data.get('propellant_type', 'apcp'),
            chamber_diameter=data.get('chamber_diameter', 100),
            grain_length=data.get('grain_length', 500),
            core_diameter=data.get('core_diameter', 30),
            chamber_pressure=data.get('chamber_pressure', 40),
        )
        mc = motor.run_monte_carlo(n_samples=int(data.get('n_samples', 300)))
        if mc.get('error'):
            return jsonify({'status': 'error', 'error': mc['error']}), 400
        return jsonify({'status': 'success',
                        'burn_rate_inputs': burn_rate_sources,
                        **sanitize_json_values(mc)})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/export_tank_cad', methods=['POST'])
def export_tank_cad():
    """Export tank CAD files (STEP, STL, drawings)"""
    try:
        data = request.get_json()
        tank_data = data.get('tank_data')
        
        if not tank_data:
            return jsonify({'error': 'Tank data not found'}), 400
        
        # Import CAD generator
        from hrma.export.cad_export import cad_generator
        
        # Generate CAD files
        print("Generating tank CAD files...")
        zip_file_path = cad_generator.generate_tank_cad(tank_data)
        
        print(f"CAD files generated: {zip_file_path}")
        
        # Return zip file
        return send_file(
            zip_file_path,
            as_attachment=True,
            download_name=f'propellant_tanks_cad_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip',
            mimetype='application/zip'
        )
        
    except Exception as e:
        traceback.print_exc()
        print(f"CAD export error: {str(e)}")
        return jsonify({'error': f'CAD export error: {str(e)}'}), 500

def _motor_cad_zip_response(geo, motor_type, default_name):
    """Ortak CAD paket üreticisi: STEP (varsa) + STL'leri ZIP'ler.

    STEP build123d yoksa paket STL-only iner; MANIFEST durumu açıkça yazar
    (sessiz eksilme yasak). Enjektör katısı katıda, grain katısı sıvıda üretilmez.
    """
    name = safe_name(geo.get('motor_name'), default=default_name)
    arc = {}
    # i18n: README/hata metinleri EN üretilir (backend sözleşmesi); hata
    # gövdesinin TR karşılığı i18n_charts.js MSG_PATTERNS'ta.
    manifest = [f'HRMA CAD package — {name} ({motor_type})', '']

    # STEP (gerçek parametrik katılar)
    try:
        from hrma.export.step_export import generate_step_assembly
        step_files = generate_step_assembly(geo, motor_type=motor_type)
        for key, path in step_files.items():
            arc[f'step/{name}_{key}.step'] = path
        manifest.append(f'STEP: {len(step_files)} files (AP214, mm)')
    except Exception as e:
        manifest.append(f'STEP: FAILED ({e})')

    # STL (mesh'ler) — motor tipine uymayan bileşenler filtrelenir
    try:
        cad_data = cad_designer.generate_3d_motor_assembly(geo)
        meshes = cad_data.get('assembly_meshes') or []
        skip = {'solid': {'Injector'}, 'liquid': {'Fuel Grain'}}.get(motor_type, set())
        meshes = [(n, m) for n, m in meshes if n not in skip]
        stl_files = cad_designer.export_stl_files(meshes) if meshes else []
        for p in stl_files:
            arc[safe_arcname('stl', os.path.basename(p))] = p
        manifest.append(f'STL: {len(stl_files)} files (mm, 3D printing/CAM)')
    except Exception as e:
        manifest.append(f'STL: FAILED ({e})')

    if not arc:
        return jsonify({'status': 'error',
                        'error': ('CAD generation failed: '
                                  + ' | '.join(manifest))}), 500

    buf = _zip_files(arc, readme_text='\n'.join(manifest) + '\n')
    return send_file(buf, as_attachment=True,
                     download_name=f'{name}_CAD_package.zip',
                     mimetype='application/zip')


@app.route('/export_solid_motor_cad', methods=['POST'])
def export_solid_motor_cad():
    """Katı motor CAD paketi: STEP + STL (kamara, nozul, grain)."""
    try:
        data = request.get_json() or {}
        # JS ya ham /calculate_solid sonucunu ('results') ya da eski sözleşmeyle
        # 'motor_data' gönderir — ikisi de katı sonuç sözlüğü kabul edilir
        results = data.get('results') or data.get('motor_data')
        if not results:
            return jsonify({'error': 'Motor data not found'}), 400
        geo = solid_results_to_motor_geometry(results)
        return _motor_cad_zip_response(geo, 'solid', 'UZAYTEK_SOLID')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'CAD export error: {str(e)}'}), 500


@app.route('/export_liquid_cad', methods=['POST'])
def export_liquid_cad():
    """Sıvı motor CAD paketi: STEP + STL (kamara, nozul, enjektör)."""
    try:
        data = request.get_json() or {}
        results = data.get('results') or data.get('motor_data')
        if not results:
            return jsonify({'error': 'Motor data not found'}), 400
        geo = liquid_results_to_motor_geometry(results)
        return _motor_cad_zip_response(geo, 'liquid', 'UZAYTEK_LIQUID')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'CAD export error: {str(e)}'}), 500

@app.route('/parametric-analysis', methods=['POST'])
def parametric_analysis():
    """Parametric analysis for motor design optimization"""
    try:
        data = request.json
        
        # Get base parameters (all form data except sweep parameters)
        base_params = {k: v for k, v in data.items() if k not in ['param_type', 'param_start', 'param_end', 'param_steps']}
        
        # Get sweep parameters from the request
        sweep_param = data.get('param_type', 'of_ratio')
        param_start = data.get('param_start', 0.5)
        param_end = data.get('param_end', 3.0)
        # v2.6.26 — IS BUTCESI. Bu deger dogrudan np.linspace'e gidiyordu;
        # tip denetimi ve ust sinir yoktu. OLCULDU: param_steps=1e9 ile
        # istek iki dakikadan uzun surdu ve DONMEDI (8 GB linspace).
        # Arayuzde max="50" yaziyor ama HTML max'i zorlamaz ve kayitli
        # proje / API cagrisi onu tamamen atlar. Sunucu tarafinda sinir
        # koymak tek gercek koruma.
        try:
            sweep_points = int(data.get('param_steps', 20))
        except (TypeError, ValueError):
            return jsonify({
                'error': "'param_steps' must be a whole number",
                'status': 'invalid_input'}), 422
        if not (2 <= sweep_points <= PARAMETRIC_MAX_STEPS):
            return jsonify({
                'error': (f"'param_steps' must be between 2 and "
                          f"{PARAMETRIC_MAX_STEPS}; each step runs a full "
                          f"motor solution."),
                'requested': sweep_points,
                'status': 'invalid_input'}), 422
        
        # TOPLAM İŞ BÜTÇESİ (Faz 5B / B12). Bkz.
        # PARAMETRIC_TIME_BUDGET_DEFAULT_S.
        # ``or`` ile varsayılana düşülmez: ``time_budget_s = 0`` yanlışsıl
        # (falsy) bir değerdir ve sessizce 30 saniyeye çevrilirse kullanıcı
        # sıfır bütçe istediğini sanır. Yalnız GERÇEKTEN verilmemiş alan
        # varsayılana düşer.
        _budget_raw = data.get('time_budget_s')
        if _budget_raw is None or _budget_raw == '':
            _budget_raw = PARAMETRIC_TIME_BUDGET_DEFAULT_S
        try:
            time_budget_s = float(_budget_raw)
        except (TypeError, ValueError):
            return jsonify({
                'error': "'time_budget_s' must be a number (seconds)",
                'status': 'invalid_input'}), 422
        if not (math.isfinite(time_budget_s)
                and 0.0 < time_budget_s <= PARAMETRIC_TIME_BUDGET_MAX_S):
            return jsonify({
                'error': (f"'time_budget_s' must be in (0, "
                          f"{PARAMETRIC_TIME_BUDGET_MAX_S:g}] seconds; a "
                          f"synchronous request cannot block longer."),
                'requested': time_budget_s,
                'status': 'invalid_input'}), 422

        sweep_range = [param_start, param_end]

        # Generate sweep values
        sweep_values = np.linspace(sweep_range[0], sweep_range[1], sweep_points)

        import time as _time
        _sweep_t0 = _time.monotonic()
        _budget_stop = None

        results = []
        # A9 (2026-08-02): başarısız noktalar artık SESSİZCE ATLANMIYOR.
        # ÖLÇÜLDÜ: 5 nokta istendiğinde 4 dönüyor, yanıt yine
        # 'status':'success' ve hiçbir başarısızlık alanı yok — kullanıcı
        # eğrinin bir parçasının hiç hesaplanmadığını göremiyordu.
        failed_points = []

        for _point_index, value in enumerate(sweep_values):
            # SERT DURDURMA: bütçe dolduysa yeni nokta BAŞLATILMAZ.
            _elapsed = _time.monotonic() - _sweep_t0
            if _elapsed >= time_budget_s:
                _budget_stop = {
                    'reason': 'time_budget_exhausted',
                    'elapsed_s': round(_elapsed, 3),
                    'budget_s': time_budget_s,
                    'points_started': _point_index,
                }
                break

            rejection = _parametric_point_rejection(sweep_param, value)
            if rejection is not None:
                failed_points.append({
                    'sweep_value': float(value)
                    if math.isfinite(float(value)) else None,
                    'sweep_parameter': sweep_param,
                    'reason': rejection,
                    'stage': 'input_validation',
                })
                continue
            try:
                # Update sweep parameter
                current_params = base_params.copy()
                current_params[sweep_param] = value
                
                # Create engine with current parameters
                engine = HybridRocketEngine(
                    thrust=current_params.get('thrust'),
                    burn_time=current_params.get('burn_time'),
                    total_impulse=current_params.get('total_impulse'),
                    of_ratio=current_params.get('of_ratio', 1.0),
                    chamber_pressure=current_params.get('chamber_pressure', 20.0),
                    atmospheric_pressure=current_params.get('atmospheric_pressure', 1.0),
                    chamber_temperature=current_params.get('chamber_temperature'),  # None if not provided
                    gamma=current_params.get('gamma', 1.25),
                    gas_constant=current_params.get('gas_constant'),  # None if not provided
                    l_star=current_params.get('l_star', 1.0),
                    expansion_ratio=current_params.get('expansion_ratio', 0),
                    nozzle_type=current_params.get('nozzle_type', 'conical'),
                    thrust_coefficient=current_params.get('thrust_coefficient', 0),
                    regression_a=current_params.get('regression_a'),  # None if not provided
                    regression_n=current_params.get('regression_n'),  # None if not provided
                    fuel_density=current_params.get('fuel_density'),  # None if not provided
                    combustion_type=current_params.get('combustion_type', 'infinite'),
                    chamber_diameter_input=current_params.get('chamber_diameter_input', 0),
                    fuel_type=current_params.get('fuel_type', 'htpb')
                )
                
                # Calculate results
                motor_results = engine.calculate()
                
                # Store key results
                point_result = {
                    'sweep_value': value,
                    'isp': motor_results['isp'],
                    'thrust': motor_results['thrust'],
                    'total_impulse': motor_results['total_impulse'],
                    'chamber_pressure': motor_results['chamber_pressure'],
                    'propellant_mass_total': motor_results['propellant_mass_total'],
                    'throat_diameter': motor_results['throat_diameter'] * 1000,  # Convert to mm
                    'expansion_ratio': motor_results['expansion_ratio'],
                    'c_star': motor_results['c_star'],
                    'cf': motor_results['cf']
                }
                
                # Calculate trajectory if requested
                if data.get('include_trajectory', False):
                    trajectory_analyzer.set_vehicle_parameters(
                        mass_dry=data.get('vehicle_mass_dry', 50),
                        diameter=data.get('vehicle_diameter', 0.15)
                    )
                    
                    launch_params = {
                        'launch_angle': data.get('launch_angle', 85),
                        'launch_altitude': data.get('launch_altitude', 0)
                    }
                    
                    trajectory_data = trajectory_analyzer.calculate_trajectory(motor_results, launch_params)
                    point_result['max_altitude'] = trajectory_data['performance']['trajectory_metrics']['max_altitude']
                    point_result['max_velocity'] = trajectory_data['performance']['trajectory_metrics']['max_velocity']
                    point_result['total_flight_time'] = trajectory_data['performance']['trajectory_metrics']['total_flight_time']
                
                results.append(point_result)

                # ERKEN REDDETME: ilk çözülen noktanın ÖLÇÜLEN maliyetiyle
                # tüm taramanın süresi öngörülür. Öngörü bütçeyi aşıyorsa
                # kullanıcıyı 68 saniye bekletmek yerine hemen söylenir ve
                # bu makinede kaç adımın sığdığı BİLDİRİLİR (ölçüme dayalı,
                # uydurma değil).
                if _point_index == 0 and len(results) == 1:
                    _per_point_s = _time.monotonic() - _sweep_t0
                    _projected_s = _per_point_s * sweep_points
                    if _projected_s > time_budget_s:
                        _fits = int(time_budget_s / _per_point_s) \
                            if _per_point_s > 0 else sweep_points
                        return jsonify({
                            'status': 'invalid_input',
                            'error': 'parametric_time_budget_exceeded',
                            'message': (
                                f'This sweep is projected to take '
                                f'{_projected_s:.1f} s, over the '
                                f'{time_budget_s:g} s budget. Reduce '
                                f"'param_steps' or raise 'time_budget_s' "
                                f'(max {PARAMETRIC_TIME_BUDGET_MAX_S:g} s).'),
                            'measured_seconds_per_point': round(
                                _per_point_s, 3),
                            'projected_seconds': round(_projected_s, 1),
                            'budget_s': time_budget_s,
                            'points_requested': int(sweep_points),
                            'points_that_fit_budget': max(2, _fits),
                        }), 422

            except Exception as e:
                # Nokta çözülemedi: girdisi ve nedeni yanıta girer.
                print(f"Failed calculation for {sweep_param}={value}: {str(e)}")
                failed_points.append({
                    'sweep_value': float(value),
                    'sweep_parameter': sweep_param,
                    'reason': f'{type(e).__name__}: {e}'[:200],
                    'stage': 'solver',
                })
                continue

        # Create parametric analysis plot
        parametric_plot = create_parametric_plot(results, sweep_param)

        response = {
            'sweep_parameter': sweep_param,
            'sweep_range': sweep_range,
            'results': results,
            'plot': parametric_plot,
            'plot_data': parametric_plot,  # Add plot_data field for compatibility
            'points_requested': int(sweep_points),
            'points_succeeded': len(results),
            'points_failed': len(failed_points),
            'failed_points': failed_points,
            # ÖLÇÜLEN toplam süre (monotonic farkı), uydurma değil.
            'sweep_wall_time_s': round(_time.monotonic() - _sweep_t0, 3),
        }
        # Bütçe yüzünden kesilen tarama TAM tarama gibi yayımlanmaz: kaç
        # noktanın hiç BAŞLATILMADIĞI açıkça yazılır.
        if _budget_stop is not None:
            _not_started = int(sweep_points) - _budget_stop['points_started']
            response['truncated_by_time_budget'] = dict(
                _budget_stop, points_not_attempted=_not_started)
        if results:
            # Kısmi tarama hâlâ çizilebilir bir eğridir, bu yüzden
            # 'success' sözleşmesi korunur (advanced.html:3571 yalnız
            # status==='success' dalında grafik çiziyor). Eksiklik gizlenmez:
            # points_failed / failed_points / warning alanları yanıtta.
            response['status'] = 'success'
            if failed_points:
                response['warning'] = (
                    f'{len(failed_points)} of {sweep_points} sweep points '
                    f'could not be computed; see failed_points.')
            if _budget_stop is not None:
                response['warning'] = (
                    f'The sweep was cut short after '
                    f"{response['truncated_by_time_budget']['elapsed_s']} s "
                    f'(budget {time_budget_s:g} s); '
                    f"{response['truncated_by_time_budget']['points_not_attempted']}"
                    f' of {sweep_points} points were never attempted. The '
                    f'curve is partial; see truncated_by_time_budget.')
        else:
            # Hiçbir nokta hesaplanamadı — bu bir başarı değildir.
            response['status'] = 'error'
            response['error'] = (
                f'No sweep point could be computed for {sweep_param}; '
                f'see failed_points for the reason of each.')
            return jsonify(response), 422
        return jsonify(response)

    except Exception as e:
        return jsonify({'error': str(e)}), 400

def create_parametric_plot(results, sweep_param):
    """Create parametric analysis visualization"""
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    
    if not results:
        return None
    
    # Extract data
    sweep_values = [r['sweep_value'] for r in results]
    isp_values = [r['isp'] for r in results]
    thrust_values = [r['thrust'] for r in results]
    mass_values = [r['propellant_mass_total'] for r in results]
    throat_diameter_values = [r['throat_diameter'] for r in results]
    
    # Create subplots
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=(
            f'Specific Impulse vs {sweep_param.replace("_", " ").title()}',
            f'Thrust vs {sweep_param.replace("_", " ").title()}',
            f'Propellant Mass vs {sweep_param.replace("_", " ").title()}',
            f'Throat Diameter vs {sweep_param.replace("_", " ").title()}'
        )
    )
    
    # Süpürülen parametrenin insan-okur adı (hover ve eksen başlıkları için)
    sweep_label = sweep_param.replace('_', ' ').title()

    # Seri renkleri merkezi paletten atanır (v2.5.5): kırmızı/mavi/yeşil
    # CSS adları koyu temada tutarsız ve okunaksızdı.
    # Isp plot
    fig.add_trace(
        go.Scatter(
            x=sweep_values,
            y=isp_values,
            mode='lines+markers',
            name='Specific Impulse',
            line=dict(color=PALETTE[0], width=3),
            marker=dict(size=6),
            hovertemplate=(sweep_label + ': %{x:.4g}<br>'
                           'Isp: %{y:.1f} s<extra></extra>')
        ),
        row=1, col=1
    )

    # Thrust plot
    fig.add_trace(
        go.Scatter(
            x=sweep_values,
            y=thrust_values,
            mode='lines+markers',
            name='Thrust',
            line=dict(color=PALETTE[1], width=3),
            marker=dict(size=6),
            hovertemplate=(sweep_label + ': %{x:.4g}<br>'
                           'Thrust: %{y:.1f} N<extra></extra>')
        ),
        row=1, col=2
    )

    # Mass plot
    fig.add_trace(
        go.Scatter(
            x=sweep_values,
            y=mass_values,
            mode='lines+markers',
            name='Propellant Mass',
            line=dict(color=PALETTE[2], width=3),
            marker=dict(size=6),
            hovertemplate=(sweep_label + ': %{x:.4g}<br>'
                           'Propellant Mass: %{y:.2f} kg<extra></extra>')
        ),
        row=2, col=1
    )

    # Throat diameter plot
    fig.add_trace(
        go.Scatter(
            x=sweep_values,
            y=throat_diameter_values,
            mode='lines+markers',
            name='Throat Diameter',
            line=dict(color=PALETTE[3], width=3),
            marker=dict(size=6),
            hovertemplate=(sweep_label + ': %{x:.4g}<br>'
                           'Throat Diameter: %{y:.2f} mm<extra></extra>')
        ),
        row=2, col=2
    )

    # Add trajectory data if available
    if 'max_altitude' in results[0]:
        altitude_values = [r['max_altitude'] / 1000 for r in results]  # Convert to km
        fig.add_trace(
            go.Scatter(
                x=sweep_values,
                y=altitude_values,
                mode='lines+markers',
                name='Max Altitude (km)',
                line=dict(color=PALETTE[4], width=3),
                marker=dict(size=6),
                yaxis='y5',
                hovertemplate=(sweep_label + ': %{x:.4g}<br>'
                               'Max Altitude: %{y:.2f} km<extra></extra>')
            ),
            row=1, col=1
        )

    # Update layout — sabit width KALDIRILDI (v2.5.5): genişlik konteynere
    # uyar (autosize), yükseklik panel sabitlemesi için korunur.
    fig.update_layout(
        title=dict(
            text=f'Parametric Analysis: {sweep_label} Sweep',
            x=0.5,
            font=dict(size=16, family='Arial')
        ),
        showlegend=False,
        height=600,
        autosize=True
    )
    
    # Update axis labels
    fig.update_xaxes(title_text=sweep_label, row=1, col=1)
    fig.update_yaxes(title_text='Isp (s)', row=1, col=1)
    fig.update_xaxes(title_text=sweep_label, row=1, col=2)
    fig.update_yaxes(title_text='Thrust (N)', row=1, col=2)
    fig.update_xaxes(title_text=sweep_label, row=2, col=1)
    fig.update_yaxes(title_text='Mass (kg)', row=2, col=1)
    fig.update_xaxes(title_text=sweep_label, row=2, col=2)
    fig.update_yaxes(title_text='Throat Diameter (mm)', row=2, col=2)

    # _fig_json: bdata'sız düz JSON (vendor plotly.js 1.58.5 uyumu, v2.5.5)
    return _fig_json(fig)

@app.route('/api/comparative-analysis', methods=['POST'])
def comparative_analysis():
    """Create comparative analysis between multiple motor configurations.

    Dalga 4A onarımı (2026-07-14): eski kod eksik metrik anahtarlarında
    (thrust/isp/total_impulse/total_mass) KeyError -> 500 veriyordu.
    Şema doğrulaması artık onarılmış create_comparative_analysis_plot
    içinde yapılır (ValueError -> net 400 mesajı); "en iyi" sıralamaları
    yalnız ilgili metriği taşıyan konfigürasyonlar üzerinden hesaplanır.
    """
    try:
        data = request.get_json(silent=True) or {}
        motor_configs = data.get('motor_configs', {})

        if not isinstance(motor_configs, dict):
            return jsonify({
                'status': 'error',
                'error': ("motor_configs must be an object of "
                          "{config_name: {metric: value}} entries."),
            }), 400
        if len(motor_configs) < 2:
            return jsonify({
                'status': 'error',
                'error': ('At least 2 motor configurations are required '
                          'for comparison.'),
            }), 400

        # Onarılmış plot fonksiyonu: eksik anahtar tolere edilir, yapısal
        # bozukluk ValueError ile net mesaj verir (visualization.py).
        try:
            comparative_plot = create_comparative_analysis_plot(motor_configs)
        except ValueError as exc:
            return jsonify({'status': 'error', 'error': str(exc)}), 400

        def _numeric(value):
            return isinstance(value, (int, float)) and np.isfinite(value)

        def _best_by(metric_fn):
            # metric_fn(cfg) -> sayısal skor veya None; skoru olmayan
            # konfigürasyon sıralamaya girmez (eski kodun KeyError tuzağı)
            scored = {}
            for name, cfg in motor_configs.items():
                if not isinstance(cfg, dict):
                    continue
                score = metric_fn(cfg)
                if score is not None:
                    scored[name] = score
            if not scored:
                return None
            return max(scored, key=scored.get)

        best_thrust = _best_by(
            lambda c: c['thrust'] if _numeric(c.get('thrust')) else None)
        best_isp = _best_by(
            lambda c: c['isp'] if _numeric(c.get('isp')) else None)
        best_efficiency = _best_by(
            lambda c: (c['isp'] / c['total_mass'])
            if (_numeric(c.get('isp')) and _numeric(c.get('total_mass'))
                and c['total_mass'] > 0) else None)

        return jsonify({
            'status': 'success',
            'plot': comparative_plot,
            'analysis': {
                'best_thrust': best_thrust,
                'best_isp': best_isp,
                'best_efficiency': best_efficiency,
                'total_configs': len(motor_configs)
            }
        })

    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/advanced-analysis', methods=['POST'])
def advanced_analysis():
    """Generate comprehensive advanced analysis plots"""
    try:
        data = request.json
        motor_data = data.get('motor_data', {})
        analysis_types = data.get('analysis_types', [])
        
        results = {}
        
        # Heat transfer analysis
        # Dalga 0 düzeltmesi (2026-07-14): analyze_chamber_thermal diye bir
        # metot HİÇ olmadı — bu dal her çağrıda AttributeError -> 500
        # veriyordu. Gerçek API analyze_heat_transfer'dir; girdi sözlüğü
        # onun beklediği anahtarlarla kurulur.
        if 'heat_transfer' in analysis_types:
            heat_analyzer = HeatTransferAnalyzer()
            ht_input = {
                'chamber_pressure': float(motor_data.get('chamber_pressure', 20.0)),   # bar
                'chamber_temperature': float(motor_data.get('chamber_temperature', 3000.0)),  # K
                'chamber_diameter': float(motor_data.get('chamber_diameter', 0.1)),    # m
                'chamber_length': float(motor_data.get('chamber_length', 0.5)),        # m
                'burn_time': float(motor_data.get('burn_time', 10.0)),                 # s
                'mdot_total': float(motor_data.get('mdot_total', 1.0)),                # kg/s
            }
            # Varsa gerçek gaz/boğaz değerlerini geçir (Bartz fallback'i yerine)
            for key in ('gamma', 'molecular_weight', 'gas_constant',
                        'throat_diameter', 'c_star'):
                if motor_data.get(key) is not None:
                    try:
                        ht_input[key] = float(motor_data[key])
                    except (TypeError, ValueError):
                        pass
            heat_data = heat_analyzer.analyze_heat_transfer(
                ht_input,
                material=data.get('material_type', 'steel'),
                wall_thickness=float(data.get('wall_thickness', 0.005)),
                cooling_type=data.get('cooling_type', 'natural')
            )
            # Plot fonksiyonunun beklediği zones/effectiveness alanlarını ekle
            if 'cooling_analysis' in heat_data and 'zones' not in heat_data['cooling_analysis']:
                ca = heat_data['cooling_analysis']
                ca['zones'] = ['Chamber', 'Throat', 'Nozzle']
                ca['effectiveness'] = [ca.get('cooling_efficiency', 0.8)] * 3
            results['heat_transfer_plot'] = create_heat_transfer_plots(heat_data)
            results['heat_analysis'] = heat_data
        
        # Combustion analysis
        if 'combustion' in analysis_types:
            from hrma.engines.combustion_analysis import CombustionAnalyzer
            combustion_analyzer = CombustionAnalyzer()
            fuel_composition = {data.get('fuel_type', 'htpb'): 100.0}
            # v2.6.26: bkz. /calculate icindeki ayni sabit-oksitleyici notu.
            combustion_data = combustion_analyzer.analyze_combustion(
                fuel_composition,
                data.get('oxidizer_type', 'n2o'),
                data.get('of_ratio', 1.0),
                data.get('chamber_pressure', 20.0)
            )
            results['combustion_plot'] = create_combustion_analysis_plots(combustion_data)
            results['combustion_analysis'] = combustion_data
        
        # Structural analysis
        if 'structural' in analysis_types:
            structural_analyzer = StructuralAnalyzer()
            structural_data = structural_analyzer.analyze_structure(
                motor_data, material=data.get('material_type', 'steel_4130')
            )
            results['structural_plot'] = create_structural_analysis_plots(structural_data)
            results['structural_analysis'] = structural_data
        
        # 3D visualization
        if '3d_visualization' in analysis_types:
            results['motor_3d_plot'] = create_3d_motor_visualization(motor_data)
        
        # Real-time dashboard
        if 'realtime_dashboard' in analysis_types:
            time_data = build_time_history(motor_data)
            results['dashboard_plot'] = create_real_time_dashboard(motor_data, time_data)
        
        return jsonify({
            'status': 'success',
            'results': results
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/oxidizer-properties', methods=['POST'])
def get_live_oxidizer_properties():
    """Get oxidizer properties with proper data for different oxidizers"""
    try:
        data = request.json
        oxidizer_type = data.get('oxidizer_type', 'n2o')
        temperature = data.get('temperature', 293.15)
        
        print(f"OXIDIZER REQUEST: {oxidizer_type} at {temperature}K")
        
        # Define comprehensive oxidizer properties
        oxidizer_properties = {
            'n2o': {
                'density': get_oxidizer_density('n2o', temperature),
                'viscosity': 2.8e-4,
                'formula': 'N2O',
                'molecular_weight': 44.013,
                'boiling_point': 184.67,
                'vapor_pressure_20c': 5.17e6,  # Pa
                'enthalpy_formation': -82.05,  # kJ/mol
                'name': 'Nitrous Oxide',
                'phase_at_stp': 'gas',
                'storage_pressure': 5.17e6  # Pa, self-pressurizing
            },
            'lox': {
                'density': get_oxidizer_density('lox', temperature),
                'viscosity': 1.95e-4,
                'formula': 'O2',
                'molecular_weight': 31.998,
                'boiling_point': 90.15,
                'vapor_pressure_20c': 0,  # Cryogenic
                'enthalpy_formation': 0.0,
                'name': 'Liquid Oxygen',
                'phase_at_stp': 'liquid',
                'storage_pressure': 3.5e5  # Pa, typical tank pressure
            },
            'h2o2': {
                'density': 1450 - 1.5 * (temperature - 293.15),  # Temperature dependent
                'viscosity': 1.2e-3,
                'formula': 'H2O2',
                'molecular_weight': 34.015,
                'boiling_point': 423.35,
                'vapor_pressure_20c': 200,  # Pa
                'enthalpy_formation': -187.78,  # kJ/mol
                'name': 'Hydrogen Peroxide',
                'phase_at_stp': 'liquid',
                'storage_pressure': 1.5e5  # Pa
            },
            'air': {
                'density': 1.225 * (293.15 / temperature) * (101325 / 101325),  # Ideal gas
                'viscosity': 1.8e-5,
                'formula': 'Air',
                'molecular_weight': 28.97,
                'boiling_point': 78.8,  # N2 dominant
                'vapor_pressure_20c': 101325,  # Pa
                'enthalpy_formation': 0.0,
                'name': 'Compressed Air',
                'phase_at_stp': 'gas',
                'storage_pressure': 2.0e7  # Pa, high pressure
            }
        }
        
        if oxidizer_type in oxidizer_properties:
            properties = oxidizer_properties[oxidizer_type]
            
            print(f"OXIDIZER RESPONSE: {oxidizer_type} - density: {properties['density']:.1f} kg/m³")

            # Faz 5B / H3-B9: ``temperature`` yankısı ve sıcaklığa bağlı
            # yoğunluk sonlu olmayan değer taşıyabiliyordu. Ölçüldü:
            # ``temperature: Infinity`` gönderilince gövde
            # ``"temperature":Infinity`` çıkıyordu — Python ``json.loads``
            # kabul eder, tarayıcının ``JSON.parse``'ı ETMEZ, panel gerçek
            # sorun yerine ayrıştırma hatası gösterir.
            return jsonify(sanitize_json_values({
                'status': 'success',
                'properties': properties,
                'source': 'HRMA Oxidizer Database',
                'temperature': temperature
            }))
        else:
            return jsonify({
                'status': 'error', 
                'error': f'Unknown oxidizer type: {oxidizer_type}'
            })
        
    except Exception as e:
        print(f"Oxidizer properties error: {str(e)}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/validate-fuel', methods=['POST'])
def validate_fuel_composition():
    """Validate fuel composition with NASA CEA"""
    try:
        data = request.json
        composition = data.get('composition', [])
        
        # Convert composition to required format
        composition_tuples = [(comp['formula'], comp['percentage']) for comp in composition]
        
        result = db_manager.validate_fuel_composition(composition_tuples)
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

#: /api/database-status sonucu için kısa ömürlü önbellek (değer, zaman damgası).
_db_status_cache = {'value': None, 'ts': 0.0}
#: Önbellek ömrü [s]. Bağlantı durumu saniyeler içinde değişen bir şey değil.
_DB_STATUS_TTL_S = 120.0


@app.route('/api/database-status', methods=['GET'])
def check_database_status():
    """Veri tabanı bağlantı durumu — önbellekli.

    v2.6.2 performans düzeltmesi: bu uç her çağrıda CANLI NIST isteği yapıyor
    ve istek ``timeout=10`` taşıyor. ``advanced.html`` sayfayı açarken bunu
    tetiklediği için, ağ kesikken veya NIST yavaşken sayfa açılışı tam 10
    saniye asılıyordu — üstelik her sekme değişiminde yeniden.

    Durum bilgisi saniyeler içinde değişen bir şey olmadığından sonuç kısa
    süreli önbelleğe alınır. ``?refresh=1`` ile zorla tazelenebilir.
    """
    import time as _time
    force = request.args.get('refresh') in ('1', 'true', 'yes')
    now = _time.monotonic()
    if (not force and _db_status_cache['value'] is not None
            and now - _db_status_cache['ts'] < _DB_STATUS_TTL_S):
        cached = dict(_db_status_cache['value'])
        cached['cached'] = True
        cached['cache_age_s'] = round(now - _db_status_cache['ts'], 1)
        return jsonify(cached)
    try:
        status = db_manager.test_connections()
        _db_status_cache['value'] = status
        _db_status_cache['ts'] = now
        return jsonify(status)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------------------------------------------------------------------------
# A2 — 3B Dünya uydu karoları (NASA GIBS) + kalıcı önbellek
#
# Mantığın tamamı hrma/analysis/tile_cache.py içinde ve Flask'tan bağımsız;
# burası yalnız HTTP çevirisi yapar. Güvenlik kapıları (katman allowlist,
# z/x/y aralığı, tarih regex'i, realpath önek denetimi) modülde uygulanır —
# çıkış hostu sabit GIBS olduğu için SSRF yüzeyi yoktur.
# ---------------------------------------------------------------------------

@app.route('/api/tile/<layer_key>/<int:z>/<int:x>/<int:y>', methods=['GET'])
def tile_proxy(layer_key, z, x, y):
    """Tek bir uydu karosunu sunar (önbellekten ya da GIBS'ten çekerek).

    Çevrimdışıyken sahte doku ÜRETİLMEZ: 503 döner ve ön yüz o karoyu atlar,
    altındaki Blue Marble taban dokusu görünür kalır (kısmi yükleme kabul).
    """
    from hrma.analysis import tile_cache

    res = tile_cache.resolve_tile(layer_key, z, x, y,
                                  date=request.args.get('date'))
    if not res.get('ok'):
        payload = {k: v for k, v in res.items() if k != 'code'}
        return jsonify(payload), res.get('code', 500)

    resp = send_file(res['path'], mimetype=res['mimetype'])
    # Karolar içerik-adresli (layer/date/z/x/y) — değişmezler.
    resp.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    resp.headers['X-Tile-Cached'] = '1' if res.get('cached') else '0'
    # 2026-08-03 (Faz 6, T54-ek): burada bir ``X-Tile-Attribution`` başlığı vardı
    # ve atıf metnini olduğu gibi yazıyordu. Metin em-dash içeriyor
    # ("NASA GIBS — Blue Marble ...") ; HTTP başlıkları latin-1 kodlanır ve
    # U+2014 latin-1'de YOKTUR. Sonuç sessiz değil ama görünmez bir arıza:
    # werkzeug durum satırını yazdıktan SONRA send_header'da
    # ``UnicodeEncodeError`` atıyordu, yani günlüğe "200" düşüyor ama istemciye
    # tek bayt gitmiyordu. Ölçüldü: ``curl --max-time 15 /api/tile/...`` ->
    # exit 28, 0 bayt, 15 s asılı; karo diske YAZILIYOR (önbellek büyüyor) ama
    # ekrana gelmiyor. Tarayıcıda daha kötüsü: Chrome host başına 6 bağlantı
    # tutar, altı karo isteği asılınca sayfanın TÜM ağ trafiği duruyordu.
    # Başlık zaten hiçbir yerde okunmuyordu; atıf ``/api/tile/cache/status``
    # gövdesinde (UTF-8 JSON) yayımlanıyor ve arayüz #ls-tile-attr'ı oradan
    # dolduruyor. Düz yazı HTTP başlığına KONMAZ — geri eklenirse
    # tests/test_faz6_baslik_kodlama.py düşer.
    return resp


@app.route('/api/tile/cache/status', methods=['GET'])
def tile_cache_status():
    """Karo önbelleğinin disk kullanımı ve konumu."""
    from hrma.analysis import tile_cache
    return jsonify({'status': 'success', **tile_cache.cache_status(),
                    'layers': tile_cache.list_layers()})


@app.route('/api/tile/cache/clear', methods=['POST'])
def tile_cache_clear():
    """Karo önbelleğini boşaltır (kullanıcı isteğiyle)."""
    from hrma.analysis import tile_cache
    return jsonify({'status': 'success', **tile_cache.cache_clear()})


# ---------------------------------------------------------------------------
# A1 — Uçurulacak araç köprüsü
#
# Motor sayfaları ile /launch-site AYRI sayfalardır: launch-site'ta motor formu
# ya da currentResults yoktur. Bu uç, üç motor tipinin BİRBİRİNDEN FARKLI alan
# adlarını (hybrid/solid/liquid) tek şemaya indirger — normalize mantığı
# hrma/analysis/flight_vehicle.py içinde tektir (parametre tutarlılığı kuralı).
#
# İki kaynak: 'results' (oturum köprüsü, hesap sonucu doğrudan gelir) ve
# 'project' (.hrma dosyası; proje yalnız girdi + özet sakladığı için itki eğrisi
# ve propelan kütlesi YENİDEN HESAPLANIR).
# ---------------------------------------------------------------------------

@app.route('/api/flight-vehicle', methods=['POST'])
def flight_vehicle():
    """Motor sonucunu ya da kayıtlı projeyi tek araç şemasına çevirir."""
    from hrma.analysis import flight_vehicle as fv

    data = request.get_json(silent=True) or {}
    source = str(data.get('source') or 'results').lower()
    try:
        if source == 'results':
            motor_type = str(data.get('motor_type') or '').lower()
            results = data.get('results') or {}
            if not motor_type or not results:
                return jsonify({
                    'status': 'error',
                    'error': "source='results' requires 'motor_type' and 'results'",
                }), 400
            vehicle = fv.normalize(motor_type, results)

        elif source == 'project':
            name = data.get('name')
            if not name:
                return jsonify({'status': 'error',
                                'error': "source='project' requires 'name'"}), 400
            from hrma.utils import projects
            # load_project -> (doc, warnings) çifti döndürür
            doc, load_warnings = projects.load_project(name)
            # Faz 6 / T30: ``recompute_from_project`` (motor_type, RESULTS)
            # döndürür — normalize edilmiş araç DEĞİL. Buradaki eski kod ikinci
            # değeri doğrudan ``vehicle`` sanıyordu; uca ham motor sonucu
            # (ÖLÇÜLDÜ: 50 anahtar, thrust=None, motor_name=None, source=None)
            # gidiyordu. launch_site.html ``num_(veh.thrust) || 6500`` yazdığı
            # için kullanıcı kendi projesini seçtiğinde çözücüye ÖRNEK aracın
            # 6500 N'u gidiyor, itki eğrisi ise projeden geliyordu: karışık
            # kökenli araç. 'results' kolu zaten normalize ediyor; bu kol da
            # aynı tek şemaya indirgemek zorunda (flight_vehicle.py modül
            # başındaki şema; tests/test_flight_vehicle.py:362 aynı deseni
            # kullanıyor).
            motor_type, results = fv.recompute_from_project(doc)
            vehicle = fv.normalize(
                motor_type, results,
                motor_name=(doc.get('name') if isinstance(doc, dict) else None) or name,
                source='project')
            if load_warnings:
                vehicle['load_warnings'] = list(load_warnings)
            # Proje airframe taşıyorsa geri ver — panel alanları doldurulsun.
            airframe = ((doc.get('inputs') or {}).get('airframe')
                        if isinstance(doc, dict) else None)
            if airframe:
                vehicle['airframe'] = airframe

        else:
            return jsonify({'status': 'error',
                            'error': f"unknown source '{source}'"}), 400

        return jsonify(sanitize_json_values(
            {'status': 'success', 'vehicle': vehicle}))

    except (ValueError, KeyError) as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 400
    except FileNotFoundError as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 404
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(exc)}), 500


@app.route('/api/altitude-to-pressure', methods=['POST'])
def altitude_to_pressure():
    """Convert altitude to atmospheric pressure"""
    try:
        data = request.json
        altitude = data.get('altitude', 0)
        
        # Standard atmosphere calculation
        P0 = 1.01325  # Sea level pressure in bar
        T0 = 288.15   # Sea level temperature in K
        L = 0.0065    # Temperature lapse rate in K/m
        g = 9.80665   # Gravitational acceleration
        M = 0.0289644 # Molar mass of air
        R = 8.31432   # Universal gas constant
        
        if altitude < 11000:
            # Troposphere
            T = T0 - L * altitude
            pressure = P0 * (T / T0) ** ((g * M) / (R * L))
        else:
            # Simplified stratosphere
            T11 = T0 - L * 11000
            P11 = P0 * (T11 / T0) ** ((g * M) / (R * L))
            pressure = P11 * np.exp((-g * M * (altitude - 11000)) / (R * T11))
        
        # Faz 5B / H3-B9: ölçüldü — ``altitude: -Infinity`` gönderilince gövde
        # ``{"altitude":-Infinity,"pressure":Infinity,"temperature":Infinity}``
        # çıkıyordu. Bu RFC 8259 dışıdır; ``advanced.html:4404``'teki
        # ``await response.json()`` çağrısı patlar ve kullanıcı gerçek sorun
        # yerine bir ayrıştırma hatası görür. Süzgeç sonlu olmayan değeri
        # ``null`` yapar — uydurma sayı KOYMAZ.
        return jsonify(sanitize_json_values({
            'altitude': altitude,
            'pressure': pressure,
            'temperature': T if altitude < 11000 else T11
        }))

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Note: Removed duplicate /api/find-optimum-of endpoint - using the newer version below

@app.route('/api/export-eng', methods=['POST'])
def export_eng_file():
    """Export motor data as .eng file for OpenRocket"""
    try:
        data = request.json
        motor_data = data.get('motor_data', {})
        
        # Generate .eng file content
        eng_content = openrocket_exporter.export_eng_file(motor_data)
        
        # Generate filename
        motor_name = motor_data.get('motor_name', 'UZAYTEK-HRM-001')
        filename = f"{motor_name.replace(' ', '_')}.eng"
        
        return jsonify({
            'status': 'success',
            'filename': filename,
            'content': eng_content,
            'motor_summary': openrocket_exporter.export_motor_summary(motor_data)
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500


# ---------------------------------------------------------------------------
# GERÇEK export uçları (2026-07-13): DXF / çizim PDF'i / STEP / STL zip /
# komple paket. Eski popup butonları alert'ten ibaretti; artık her buton
# gerçek dosya indirir. Üreticiler: hrma/export/drawing_generator.py (kaleido
# + ezdxf + reportlab) ve hrma/export/step_export.py (build123d/OCC).
# ---------------------------------------------------------------------------

#: "Bu anahtar sözlükte hiç yok" ile "var ama None" ayrımı için nöbetçi.
_ABSENT = object()

#: Export üreticilerinin (``step_export.py``, ``drawing_generator.py``,
#: ``cad_visualization.py``) doğrudan okuduğu geometri alanları. Bir alan
#: istekte VARSA ama sonlu bir sayıya çözülmüyorsa katı cisim üretilmez.
_EXPORT_GEOMETRY_FIELDS = (
    'chamber_diameter', 'chamber_length', 'throat_diameter', 'exit_diameter',
    'grain_length', 'port_diameter_initial', 'nozzle_length',
    'expansion_ratio', 'wall_thickness',
)

#: Bunlardan EN AZ BİRİ sonlu ve pozitif olmalı. Hiçbiri yoksa ortada
#: kullanıcı geometrisi yoktur ve üretici baştan sona kendi varsayılanlarını
#: kullanır. ÖLÇÜM (Faz 4, A4): tüm geometri alanları NaN olan bir istek
#: /api/export-step, -dxf, -drawings-pdf ve -complete-zip uçlarından HTTP 200
#: aldı; üretilen STEP OCC ile geri okununca 308 x 109 mm, 5.179e5 mm3 katı
#: cisim çıktı. Yani kullanıcı hiç vermediği bir motorun imalat dosyasını
#: indiriyordu. ``step_export._num`` sonlu olmayan değeri sessizce kendi
#: yedeğine çeviriyor (step_export.py:38-43) — bu yüzden kapı ROTA
#: katmanında olmak zorunda.
_EXPORT_GEOMETRY_PRIMARY = ('chamber_diameter', 'chamber_length',
                            'throat_diameter')


def _geometry_field_value(motor_data, key):
    """Alanı üst düzeyde, yoksa ``motor_geometry`` alt sözlüğünde arar."""
    if key in motor_data:
        return motor_data[key]
    geo = motor_data.get('motor_geometry')
    if isinstance(geo, dict) and key in geo:
        return geo[key]
    return _ABSENT


def _export_geometry_problem(motor_data):
    """Geometri export'a uygun mu? Uygunsa ``None``, değilse yanıt gövdesi.

    İki ret nedeni var ve ikisi de FAIL-CLOSED:

    * ``invalid_export_geometry`` — alan verilmiş ama sonlu sayı değil
      (NaN / Inf / sayıya çevrilemeyen metin). Sessizce varsayılana düşmek
      kullanıcının vermediği bir motoru imalata göndermek demektir.
    * ``missing_export_geometry`` — birincil alanların hiçbiri sonlu ve
      pozitif değil. Bu durumda üretilecek katı cisim tamamen üreticinin
      kendi yedek sayılarından oluşur; "hesaplanmadı" beyanı yerine sahte
      bir çizim vermek yasak.

    Biçim ``/analyze_safety``'nin 422 kapısıyla aynı (app.py:4426 civarı):
    ``status`` + makine-okur ``error`` kodu + insan-okur ``message`` + hangi
    alanın sorunlu olduğunu söyleyen liste.

    Açıkça ``None``/boş dize gelen alan "verilmedi" sayılır (``input_guard``
    modülünün belgelenmiş ilkesi): tek başına ret nedeni değildir, ama
    birincil alanların hepsi böyleyse ikinci kapıya takılır.
    """
    if not isinstance(motor_data, dict):
        return {
            'status': 'error',
            'error': 'invalid_export_geometry',
            'message': "'motor_data' must be an object with geometry fields.",
            'invalid_fields': [],
        }

    invalid = []
    finite_primary = []
    for key in _EXPORT_GEOMETRY_FIELDS:
        raw = _geometry_field_value(motor_data, key)
        if raw is _ABSENT or raw is None:
            continue
        if isinstance(raw, str) and raw.strip() == '':
            continue
        if isinstance(raw, bool):
            invalid.append({'field': key, 'reason': 'not_a_number',
                            'value': repr(raw)})
            continue
        try:
            value = float(raw)
        except (TypeError, ValueError):
            invalid.append({'field': key, 'reason': 'not_a_number',
                            'value': str(raw)[:64]})
            continue
        if not math.isfinite(value):
            invalid.append({'field': key, 'reason': 'not_finite',
                            'value': str(raw)[:64]})
            continue
        if key in _EXPORT_GEOMETRY_PRIMARY and value > 0:
            finite_primary.append(key)

    if invalid:
        return {
            'status': 'error',
            'error': 'invalid_export_geometry',
            'message': ('Manufacturing files cannot be generated from '
                        'non-finite geometry. Fix the listed fields and '
                        'run the analysis again.'),
            'invalid_fields': invalid,
        }
    if not finite_primary:
        return {
            'status': 'error',
            'error': 'missing_export_geometry',
            'message': ('No usable motor geometry was supplied; the export '
                        'would consist entirely of generator defaults. '
                        'Run an analysis first.'),
            'required_any_of': list(_EXPORT_GEOMETRY_PRIMARY),
        }
    return None


def _reject_unexportable_geometry(motor_data):
    """Rota kapısı: sorun varsa ``(yanıt, 422)``, yoksa ``None``."""
    problem = _export_geometry_problem(motor_data)
    if problem is None:
        return None
    return jsonify(problem), 422


def _step_length_units(paths):
    """Üretilen STEP dosyalarının başlığındaki uzunluk birimini OKUR.

    A1 düzeltmesi: README'ye sabit 'Units: millimetres.' yazılıyordu. Birim
    iddiası, üretilen dosyadan okunmadıkça iddia edilmez. AP214 başlığında
    birim ``SI_UNIT(.MILLI.,.METRE.)`` biçiminde geçer; inç kullanan
    dosyalarda ``CONVERSION_BASED_UNIT('INCH', ...)`` görünür.

    Dönüş: okunabilen birim adları kümesi. Boş küme = okunamadı → çağıran
    metinde birim İDDİA ETMEZ.
    """
    units = set()
    for path in paths:
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as handle:
                head = handle.read(400_000)
        except OSError:
            continue
        # Boşluk/satır sonu STEP'te anlamsızdır; kalıp eşleşmesi için silinir.
        compact = ''.join(head.upper().split())
        if 'SI_UNIT(.MILLI.,.METRE.)' in compact:
            units.add('millimetres')
        elif 'SI_UNIT($,.METRE.)' in compact:
            units.add('metres')
        elif "CONVERSION_BASED_UNIT('INCH'" in compact:
            units.add('inches')
    return units


def _step_readme_text(paths):
    """STEP paketinin README'si — birim satırı ölçülen değerden gelir."""
    lines = [
        'HRMA STEP export (AP214)',
        'Solver-generated parametric solids: chamber, nozzle (true contour),',
        'fuel grain, injector plate (drilled orifices) + assembly.',
    ]
    units = _step_length_units(paths)
    if len(units) == 1:
        lines.append('Length unit read from the STEP header of the files in '
                     'this archive: ' + units.pop() + '.')
    elif units:
        lines.append('WARNING: the files in this archive do not share one '
                     'length unit; headers report: '
                     + ', '.join(sorted(units)) + '.')
    else:
        lines.append('Length unit could not be read from the generated '
                     'files; check the UNIT block in the STEP header before '
                     'importing.')
    return '\n'.join(lines) + '\n'


def _stl_readme_text(paths):
    """STL paketinin README'si — birim İDDİA ETMEZ, ölçülen değeri yazar.

    A1 ölçümü: metin 'Units: millimetres.' ve 'Watertight closed-profile
    revolve solids' diyordu; aynı ZIP'te STEP sınırlayıcı kutusu
    [1069.62 163.50 163.50] mm iken STL kutusu [0.1635 0.1635 1.0696]
    çıkıyordu (yani metre, 1000x hata) ve ``motor_assembly.stl``
    ``is_watertight=False`` idi.

    STL biçiminde birim beyanı YOKTUR — dosyadan okunamaz, dolayısıyla
    iddia da edilemez. Bunun yerine ölçülen sınırlayıcı kutu ve gerçek
    su-sızdırmazlık durumu yazılır; kullanıcı büyüklüğe bakıp hangi birimde
    olduğunu görebilir. Metreden mm'ye çevirmek üretici tarafın
    (``hrma/export/cad_visualization.py``) işidir; orası mm'ye geçtiğinde bu
    metin kendiliğinden doğru sayıyı gösterir, burada değişiklik gerekmez.
    """
    lines = [
        'HRMA STL export',
        'Revolve solids generated from solver geometry.',
        'motor_assembly.stl = combined single-file model.',
        'The STL format carries no unit declaration, so none is claimed '
        'here.',
        'Measured properties of the files in this archive:',
    ]
    measured = False
    try:
        import trimesh
    except Exception:
        trimesh = None
    if trimesh is not None:
        for path in paths:
            try:
                mesh = trimesh.load_mesh(path)
                extents = getattr(mesh, 'extents', None)
                watertight = bool(getattr(mesh, 'is_watertight', False))
            except Exception:
                continue
            if extents is None:
                continue
            measured = True
            lines.append(
                '  {name}: bounding box {x:.4g} x {y:.4g} x {z:.4g} '
                '(file units), watertight={wt}'.format(
                    name=os.path.basename(path),
                    x=float(extents[0]), y=float(extents[1]),
                    z=float(extents[2]),
                    wt='yes' if watertight else 'no'))
    if not measured:
        lines.append('  not measured (mesh library unavailable on this '
                     'machine); no geometric claim is made.')
    return '\n'.join(lines) + '\n'


def _zip_files(file_map, readme_text=None, text_map=None):
    """{arşiv_adı: dosya_yolu} sözlüğünü bellekte ZIP'ler; BytesIO döner.

    ``text_map`` ({arşiv_adı: metin}) doğrudan içerik yazmak içindir; aynı
    kapıdan geçer.

    v2.6.26 BEKÇİSİ — ZIP SLIP: Girdi adları ``is_safe_arcname`` beyaz
    listesinden geçmezse istisna atılır. Eskiden kullanıcının ``motor_name``
    değeri buraya ham geliyordu ve ``motor_name='../../EVIL'`` ile arşivde
    ``../EVIL_chamber.step`` gibi girdiler üretilebiliyordu (ölçüldü).
    Çağıranlar adı zaten ``safe_name``'den geçiriyor; bu bekçi ileride
    eklenecek yeni bir çağıranın aynı hataya düşmesini MEKANİK olarak
    engeller — dikkat değil, kod garanti eder.
    """
    import zipfile
    from hrma.utils.input_guard import is_safe_arcname

    def _check(arcname):
        if not is_safe_arcname(arcname):
            raise ValueError(f'unsafe archive entry name: {arcname!r}')
        return arcname

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for arcname, path in file_map.items():
            if path and os.path.exists(path):
                zf.write(path, _check(arcname))
        for arcname, content in (text_map or {}).items():
            if content:
                zf.writestr(_check(arcname), content)
        if readme_text:
            zf.writestr('README.txt', readme_text)
    buf.seek(0)
    return buf


#: Çizim uçlarının BEYAN EDİLEN birim sözleşmesi. Aşağıdaki
#: ``_declare_drawing_units`` her çağrıda gelen geometriyi bu birime indirger
#: ve sözlüğe ``length_units`` damgasını basar.
DRAWING_ENDPOINT_LENGTH_UNIT = 'm'


def _declare_drawing_units(motor_data):
    """Çizim uçlarına giren geometrinin birimini ÇÖZER ve BEYAN EDER.

    Faz 5B / H4-2 ÖLÇÜMÜ (HEAD 9d3728e): ``/api/export-dxf`` girdinin birimi
    hakkında hiçbir şey söylemiyordu; üretici (``drawing_generator._dims_mm``)
    girdiyi KOŞULSUZ metre kabul edip 1000 ile çarpıyordu. Hibrit yanıtı SI
    olduğu için tarayıcıdan görünmüyordu, ama ``/calculate_solid`` yanıtı
    doğrudan uca verilince çizim 1000× büyük çıkıyordu::

        Ø_throat  = 47927.25 mm   (gerçek 47.93 mm)
        Ø_chamber = 100000.0 mm   (gerçek 100.0 mm)

    Sıvıda daha da kötüsü: kamara 1000× büyük, boğaz/çıkış DOĞRU — tek çizimde
    iki farklı ölçek. Kök neden, sözleşmenin hiçbir tarafta yazılı olmamasıydı;
    depoda aynı sorunun BEŞ ayrı, birbirinden habersiz çözümü vardı
    (``openrocket_integration``, ``performance_panel.js``, ``analysis_dock.js``,
    ``pdf_generator``, ``drawing_generator``).

    Bu fonksiyon uç tarafındaki payı kapatır ve iki şey yapar:

      1. Dönüşümü TEK yerden geçirir — ortak
         ``hrma.export.motor_geometry.normalise_export_geometry``. Kendi eşiğini
         ya da kendi çarpanını KURMAZ.
      2. Sonucu ``length_units='m'`` damgasıyla beyan eder; bundan sonra bu
         sözlüğü okuyan her üretici birimi tahmin etmek zorunda değildir.

    Ortak normalize edici idempotenttir (damgalı sözlüğü tekrar verirseniz
    ``declared:m`` yoluyla aynı değerleri döndürür), bu yüzden üreticinin
    kendisi de aynı işlevi çağırıyorsa çift dönüşüm OLMAZ.

    Çözülemeyen alanlara DOKUNULMAZ (uydurma değer yazılmaz); hangi alanın
    hangi yoldan çözüldüğü dönen sözlükteki ``geometry_unit_resolution``
    raporunda durur.
    """
    if not isinstance(motor_data, dict):
        return motor_data
    from hrma.export.motor_geometry import normalise_export_geometry
    normalised, _report = normalise_export_geometry(motor_data)
    normalised['length_units'] = DRAWING_ENDPOINT_LENGTH_UNIT
    return normalised


@app.route('/api/export-dxf', methods=['POST'])
def export_dxf():
    """2D imalat çizimi (DXF): iç akış konturu + kamara + grain profili.

    BİRİM SÖZLEŞMESİ: bu uç girdinin birimini VARSAYMAZ. Gelen geometri
    ``_declare_drawing_units`` ile metreye indirgenir ve ``length_units='m'``
    damgasıyla üreticiye verilir (gerekçe: bkz. o fonksiyonun açıklaması).
    """
    try:
        from hrma.export.drawing_generator import generate_dxf
        motor_data = (request.json or {}).get('motor_data', {})
        # A4 kapısı: sonlu olmayan geometriden imalat çizimi üretilmez.
        # Kapı HAM girdide çalışır — birim çözümü sonlu olmayanı gizlemesin.
        rejected = _reject_unexportable_geometry(motor_data)
        if rejected:
            return rejected
        # H4-2: birim burada, TEK yerde çözülür ve beyan edilir.
        motor_data = _declare_drawing_units(motor_data)
        path = generate_dxf(motor_data)
        name = safe_name(motor_data.get('motor_name'))
        return send_file(path, as_attachment=True,
                         download_name=f'{name}_profile.dxf',
                         mimetype='application/dxf')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/export-drawings-pdf', methods=['POST'])
def export_drawings_pdf():
    """Antetli çok sayfalı teknik çizim PDF'i (kesit + enjektör + tablo).

    BİRİM SÖZLEŞMESİ ``/api/export-dxf`` ile aynıdır: girdi varsayılmaz,
    ``_declare_drawing_units`` ile metreye indirgenip beyan edilir. Bu uç ve
    DXF ucu aynı ``_dims_mm`` yolundan geçtiği için sözleşme de ortaktır —
    ölçülen kusur (katı/sıvı sonucunda 1000× büyük ölçü tablosu) ikisinde de
    aynıydı.
    """
    try:
        from hrma.export.drawing_generator import generate_drawing_pdf
        motor_data = (request.json or {}).get('motor_data', {})
        # A4 kapısı: sonlu olmayan geometriden çizim paketi üretilmez.
        rejected = _reject_unexportable_geometry(motor_data)
        if rejected:
            return rejected
        # H4-2: birim burada, TEK yerde çözülür ve beyan edilir.
        motor_data = _declare_drawing_units(motor_data)
        path = generate_drawing_pdf(motor_data)
        name = safe_name(motor_data.get('motor_name'))
        return send_file(path, as_attachment=True,
                         download_name=f'{name}_technical_drawings.pdf',
                         mimetype='application/pdf')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/export-step', methods=['POST'])
def export_step_files():
    """Gerçek STEP katıları (build123d): bileşenler + assembly, ZIP olarak.

    Faz 5B / H3-B14 ÖLÇÜMÜ (HEAD 9d3728e): boş gövdeyle (``{}``) bu uç DOĞRU
    biçimde 422 dönüyordu — ama önce +433 MB KALICI bellek alıyordu:

        app import sonrasi tepe RSS : 216.4 MB
        POST /api/export-step {}    : 422, 3.41 s, tepe RSS 649.4 MB

    Sebep sıra hatasıydı: ``from hrma.export.step_export import ...`` satırı
    kapıdan ÖNCE geliyordu, yani reddedilecek bir istek bile build123d/OCC
    yığınını sürece yüklüyordu. ``sys.modules`` yüklü kaldığı için maliyet
    kalıcıdır; masaüstü uygulamasının ayak izi TEK bir başarısız export
    denemesiyle 220 MB'tan ~650 MB'a çıkıyordu.

    Düzeltme: ithalat kapının ARKASINA alındı. Kabul edilen istekte maliyet
    aynıdır (katı üretmek için o yığın gerçekten gerekli); reddedilen istekte
    hiç ödenmez. Bellek başka bir yerde tutulmuyor — tek tutucu modül
    nesnelerinin kendisi, dolayısıyla "yükleme" tek serbest bırakma noktası.
    """
    try:
        motor_data = (request.json or {}).get('motor_data', {})
        # A4 kapısı: sonlu olmayan geometriden katı cisim üretilmez.
        rejected = _reject_unexportable_geometry(motor_data)
        if rejected:
            return rejected
        # H3-B14: ağır ithalat KAPIDAN SONRA — reddedilen istek ödemez.
        from hrma.export.step_export import generate_step_assembly
        files = generate_step_assembly(motor_data)
        name = safe_name(motor_data.get('motor_name'))
        arc = {f'{name}_{k}.step': p for k, p in files.items()}
        # A1: birim satırı artık üretilen dosyanın başlığından OKUNUYOR.
        buf = _zip_files(arc, readme_text=_step_readme_text(files.values()))
        return send_file(buf, as_attachment=True,
                         download_name=f'{name}_STEP_package.zip',
                         mimetype='application/zip')
    except RuntimeError as e:
        # build123d yok — kullanıcıya açık mesaj (sessiz düşüş yasak)
        return jsonify({'status': 'error', 'error': str(e)}), 501
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/export-stl-zip', methods=['POST'])
def export_stl_zip():
    """Tüm bileşen STL'leri + birleşik assembly tek ZIP'te.

    Faz 5B / H3-B4 ÖLÇÜMÜ (HEAD 9d3728e): bu uç KARDEŞLERİNDEN farklı olarak
    hiçbir geometri kapısından geçmiyordu. Aynı ``motor_data`` üç uca
    gönderildiğinde:

        motor_data   /api/export-stl   /api/export-complete-zip  BURASI
        {}           400               422 missing_export_geometry  200, 120 426 baytlık ZIP
        hepsi 0      422               422                          500
        negatif      422               422                          500
        NaN          422               422                          500

    Boş ``motor_data`` ile dönen ZIP'in içi ölçüldü: ``motor_assembly.stl``
    sınırlayıcı kutusu ``109 x 109 x 489.7`` — bu kullanıcının motoru değil,
    ``cad_visualization.py:507-509``'daki ŞABLON (0.1 / 0.02 / 0.04 m). Yani
    hiç hesap yapmamış bir kullanıcı "başarıyla indirilmiş" bir imalat dosyası
    alıyordu. Kısmi veri daha sinsiydi: tek alan eksik olduğunda uydurulan
    ölçü ZIP'te hiçbir yerde beyan edilmiyordu.

    Düzeltme: kardeş EXPORT uçlarının kapısı AYNI BİÇİMDE uygulanır —
    ``_reject_unexportable_geometry`` (sonlu olmayan alan, ya da hiçbir
    birincil ölçü sonlu-pozitif değil → 422). ``/api/export-dxf``,
    ``/api/export-drawings-pdf``, ``/api/export-step`` ve
    ``/api/export-complete-zip`` tam olarak bu kapıyı kullanıyor; eksik olan
    tek uç burasıydı. Kapı üretimden ÖNCE koşar: reddedilen istekte CAD
    montajı hiç kurulmaz.

    AÇIK KALAN ASİMETRİ (bilinçli, karar gerektirir): ``/api/export-stl``
    ayrıca ``_STL_REQUIRED_FIELDS`` sözleşmesini uyguluyor ve hibritte
    ``port_diameter`` istiyor; bu uç istemiyor. Aynı sözleşmeyi buraya da
    koymak, ``tests/test_faz4_app_export.py::TestPackageReadmeHonesty``
    fikstürünü (``GOOD_GEOMETRY``, ``port_diameter`` taşımıyor) reddeder ve
    üç README bekçisini kırar. Sözleşmeyi sıkılaştırmak fikstürün de
    değişmesini gerektirdiği için ayrı bir karardır; burada ölçülen kusur
    (kapı YOK) kapatıldı, sözleşme genişletilmedi.
    """
    try:
        motor_data = (request.json or {}).get('motor_data', {})
        rejected = _reject_unexportable_geometry(motor_data)
        if rejected:
            return rejected
        cad_data = cad_designer.generate_3d_motor_assembly(motor_data)
        if not cad_data or 'assembly_meshes' not in cad_data:
            # i18n: EN üretilir; TR karşılığı i18n_charts.js sözlüğünde.
            return jsonify({'status': 'error',
                            'error': 'CAD assembly could not be generated'}), 500
        stl_files = cad_designer.export_stl_files(cad_data['assembly_meshes'])
        if not stl_files:
            return jsonify({'status': 'error',
                            'error': 'STL generation failed'}), 500
        name = safe_name(motor_data.get('motor_name'))
        arc = {os.path.basename(p): p for p in stl_files}
        # A1: birim ve su-sızdırmazlık iddiası kaldırıldı; metin artık
        # üretilen mesh'in ÖLÇÜLEN sınırlayıcı kutusunu ve gerçek
        # ``is_watertight`` değerini yazıyor.
        buf = _zip_files(arc, readme_text=_stl_readme_text(stl_files))
        return send_file(buf, as_attachment=True,
                         download_name=f'{name}_STL_package.zip',
                         mimetype='application/zip')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/export-complete-zip', methods=['POST'])
def export_complete_zip():
    """Komple tasarım paketi: STL + DXF + çizim PDF'i + STEP + .eng + geometri.

    Her alt üretici bağımsız denenir; başaramayanlar MANIFEST'te 'FAILED'
    olarak raporlanır (paket sessizce eksilmez).
    """
    try:
        motor_data = (request.json or {}).get('motor_data', {})
        # A4 kapısı: paketin İÇİNDEKİ her üretici aynı geometriyi okuyor;
        # kapı en dışta bir kez kurulur.
        rejected = _reject_unexportable_geometry(motor_data)
        if rejected:
            return rejected
        name = safe_name(motor_data.get('motor_name'))
        arc = {}
        manifest = []

        def attempt(label, fn):
            try:
                fn()
                manifest.append(f'[OK]     {label}')
            except Exception as exc:
                manifest.append(f'[FAILED] {label}: {exc}')

        def add_stl():
            cad_data = cad_designer.generate_3d_motor_assembly(motor_data)
            for p in cad_designer.export_stl_files(cad_data['assembly_meshes']):
                arc[safe_arcname('stl', os.path.basename(p))] = p

        def add_dxf():
            from hrma.export.drawing_generator import generate_dxf
            arc[f'drawings/{name}_profile.dxf'] = generate_dxf(motor_data)

        def add_drawpdf():
            from hrma.export.drawing_generator import generate_drawing_pdf
            arc[f'drawings/{name}_technical_drawings.pdf'] = \
                generate_drawing_pdf(motor_data)

        def add_step():
            from hrma.export.step_export import generate_step_assembly
            for k, p in generate_step_assembly(motor_data).items():
                arc[f'step/{name}_{k}.step'] = p

        eng_holder = {}

        def add_eng():
            eng_holder['content'] = openrocket_exporter.export_motor_file(motor_data)

        attempt('STL solids', add_stl)
        attempt('DXF manufacturing profile', add_dxf)
        attempt('Technical drawing PDF', add_drawpdf)
        attempt('STEP solids (build123d)', add_step)
        attempt('OpenRocket .eng (real thrust curve if transient present)', add_eng)

        # v2.6.26: burası eskiden kendi zipfile.ZipFile'ını açıyordu ve
        # _zip_files bekçisinin DIŞINDA kalıyordu. Üstelik zipfile.writestr,
        # zf.write'ın uyguladığı normpath temizliğini de yapmadığı için
        # 'openrocket/{name}.eng' girdisi traversal dizesini bir kademe DAHA
        # derin taşıyordu (ölçüldü). Artık tek kapı: _zip_files.
        text_map = {
            'geometry/motor_geometry.json':
                json.dumps(sanitize_json_values(motor_data), indent=2),
            'MANIFEST.txt':
                'HRMA COMPLETE DESIGN PACKAGE\n'
                + datetime.now().strftime('%Y-%m-%d %H:%M') + '\n\n'
                + '\n'.join(manifest) + '\n',
        }
        if eng_holder.get('content'):
            text_map[safe_arcname('openrocket', f'{name}.eng')] = \
                eng_holder['content']
        buf = _zip_files(arc, text_map=text_map)
        return send_file(buf, as_attachment=True,
                         download_name=f'{name}_complete_package.zip',
                         mimetype='application/zip')
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/export-cad', methods=['POST'])
def export_cad_files():
    """Export CAD files (STL, technical drawings, etc.)"""
    try:
        data = request.json
        motor_data = data.get('motor_data', {})
        export_formats = data.get('formats', ['stl', 'technical_drawings'])
        
        results = {}
        
        # Generate CAD assembly
        cad_data = cad_designer.generate_3d_motor_assembly(motor_data)
        
        # Export STL files if requested
        if 'stl' in export_formats:
            stl_files = cad_designer.export_stl_files(cad_data['assembly_meshes'])
            results['stl_files'] = stl_files
            # A10 (2026-08-02): indirme bağlantısı artık ÜRETİCİNİN DÖNDÜRDÜĞÜ
            # yola bağlanıyor. Eskiden link yalnız dosya adını taşıyordu ve
            # /download/stl ucu adı ``cwd/cad_exports`` altında arıyordu;
            # ``export_stl_files`` ise v2.6.26'dan beri her çağrıda kendi
            # ``mkdtemp`` dizinine yazıyor. ÖLÇÜLDÜ: sunulan dosyanın sha256'sı
            # o istekte üretilenden farklıydı, mtime 31 Temmuz — kullanıcı iki
            # gün önceki başka bir motorun katısını indiriyordu.
            results['stl_download_links'] = [
                '/download/stl/' + _register_stl_download(path)
                for path in stl_files
            ]
        
        # Technical drawings
        if 'technical_drawings' in export_formats:
            results['technical_drawings'] = cad_data['technical_drawings']
        
        # Material specifications
        if 'materials' in export_formats:
            results['material_specs'] = cad_data['material_specifications']
            results['manufacturing_notes'] = cad_data['manufacturing_notes']
        
        # 3D visualization
        if '3d_plot' in export_formats:
            results['plotly_3d'] = cad_data['plotly_visualization']
        
        # Performance summary
        results['performance_summary'] = cad_data['performance_summary']
        
        return jsonify({
            'status': 'success',
            'cad_exports': results,
            'motor_name': motor_data.get('motor_name', 'UZAYTEK-HRM-001')
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/export-openrocket', methods=['POST'])
def export_openrocket_files():
    """Export OpenRocket compatible files"""
    try:
        data = request.json
        motor_data = data.get('motor_data', {})
        rocket_params = data.get('rocket_params', None)
        
        # Generate motor file (.eng format)
        eng_content = openrocket_exporter.export_motor_file(motor_data)
        
        # Generate flight simulation data
        flight_data = openrocket_exporter.create_flight_simulation_data(motor_data, rocket_params)
        
        # Generate motor designation
        # v2.5.2 (Codex bulgusu): burada kendi kopyası kuruluyordu ve
        # throat_diameter'ı METRE varsayıp 1000 ile çarpıyordu. Katı motorda
        # o alan zaten mm olduğu için isimlendirme "N47927-..." çıkıyordu.
        # Tek doğruluk kaynağı dışa aktarıcının kendi çözücüsüdür (normalize
        # motor_geometry varsa ondan, yoksa büyüklük çıkarımıyla).
        motor_designation = openrocket_exporter._designation(motor_data)
        
        return jsonify({
            'status': 'success',
            'motor_designation': motor_designation,
            'eng_file_content': eng_content,
            'flight_simulation': flight_data,
            'download_filename': f"{motor_designation}.eng",
            'openrocket_instructions': [
                "1. Save the .eng file to OpenRocket's motor directory",
                "2. In OpenRocket, go to Edit → Preferences → Motors",
                "3. Add the motor directory path",
                "4. Select your motor in the motor selection dialog",
                "5. Run simulation with your rocket design"
            ]
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/generate-complete-package', methods=['POST'])
def generate_complete_design_package():
    """Generate complete motor design package with all files"""
    try:
        data = request.json
        motor_data = data.get('motor_data', {})
        package_options = data.get('package_options', {
            'include_cad': True,
            'include_openrocket': True,
            'include_analysis': True,
            'include_manufacturing': True
        })
        
        complete_package = {}
        # Alt bölümler isteğe bağlı üretiliyor; özet ve imalat blokları bu iki
        # sözlüğü okuduğu için burada tanımlanırlar. Üretilmemiş bölüm boş
        # kalır ve özet "analiz yok" hükmüne düşer.
        structural_pkg = {}
        safety_section = {}

        # CAD files and drawings
        if package_options.get('include_cad', True):
            cad_data = cad_designer.generate_3d_motor_assembly(motor_data)
            stl_files = cad_designer.export_stl_files(cad_data['assembly_meshes'])
            
            complete_package['cad'] = {
                'stl_files': stl_files,
                'technical_drawings': cad_data['technical_drawings'],
                'material_specifications': cad_data['material_specifications'],
                'plotly_3d_model': cad_data['plotly_visualization'],
                'performance_summary': cad_data['performance_summary']
            }
        
        # OpenRocket integration
        if package_options.get('include_openrocket', True):
            eng_content = openrocket_exporter.export_motor_file(motor_data)
            flight_data = openrocket_exporter.create_flight_simulation_data(motor_data)
            
            complete_package['openrocket'] = {
                'eng_file': eng_content,
                'flight_simulation': flight_data,
                'motor_class': openrocket_exporter._get_motor_class(motor_data.get('total_impulse', 10000))
            }
        
        # Analysis reports
        if package_options.get('include_analysis', True):
            # Dalga 2 (2026-07-14): Eski sabit 'safety_factor: 4.0' ve ondan
            # türetilen uydurma burst_pressure/material_limits kaldırıldı.
            # Gerçek yapısal analiz sonucu varsa o raporlanır; yoksa alan
            # 'NOT ANALYZED' olarak işaretlenir — değer UYDURULMAZ.
            structural_pkg = motor_data.get('structural_analysis') or {}
            safety_sub_pkg = structural_pkg.get('safety_analysis') or {}
            safety_section.update({
                'chamber_pressure': motor_data.get('chamber_pressure', 0),
            })
            if structural_pkg.get('safety_factor') is not None:
                safety_section.update({
                    'safety_factor': structural_pkg.get('safety_factor'),
                    'safety_factor_pressure': structural_pkg.get('safety_factor_pressure'),
                    'safety_factor_total': structural_pkg.get('safety_factor_total'),
                    'status': safety_sub_pkg.get('status', 'UNKNOWN'),
                    'risk_level': safety_sub_pkg.get('risk_level', 'UNKNOWN'),
                })
            else:
                safety_section.update({
                    'safety_factor': None,
                    'status': 'NOT ANALYZED',
                    'note': 'Run the structural analysis to obtain real safety factors.',
                })
            complete_package['analysis'] = {
                'motor_performance': motor_data,
                'safety_analysis': safety_section,
                'weight_breakdown': {
                    'chamber_mass': cad_data['performance_summary']['mass_breakdown']['chamber_mass'] if 'cad_data' in locals() else 'N/A',
                    'nozzle_mass': cad_data['performance_summary']['mass_breakdown']['nozzle_mass'] if 'cad_data' in locals() else 'N/A',
                    'total_dry_mass': cad_data['performance_summary']['mass_breakdown']['total_dry_mass'] if 'cad_data' in locals() else 'N/A'
                }
            }
        
        # Manufacturing package
        #
        # v2.6.26 DÜRÜSTLÜK DÜZELTMESİ: burada eskiden hesapla hiç ilgisi olmayan
        # sabit bir malzeme listesi vardı (AISI 304 kamara, ATJ lüle, AISI 316
        # enjektör, 3 Viton O-ring, 8 adet M8x30 cıvata) ve yanında ±0.1 mm,
        # Ra 3.2, 1.5x basınç testi gibi tolerans hükümleri. Ölçüldü: 500 N'lik
        # motor da 50 kN'lik motor da AYNI listeyi alıyordu. Bağlayıcı sayı
        # (cıvata sayısı, sızdırmazlık eleman adedi, tolerans) HRMA tarafından
        # hesaplanmıyor; imalatçıya verilen bir pakette bunları sanki
        # hesaplanmış gibi göstermek bu projedeki en ağır hata sınıfıdır.
        # Artık: analizden gelen malzeme kullanılır, gelmiyorsa NOT_DEFINED
        # yazılır; boyutlandırılmayan bağlantı elemanları listeye HİÇ girmez.
        if package_options.get('include_manufacturing', True):
            structural_params = structural_pkg.get('design_parameters') or {}

            def _material_of(*keys):
                for key in keys:
                    value = motor_data.get(key) or structural_params.get(key)
                    if value:
                        return {'material': str(value), 'source': 'analysis input'}
                return {'material': 'NOT_DEFINED',
                        'source': 'not specified in the analysis'}

            bom = [
                dict({'part': 'Combustion Chamber', 'quantity': 1},
                     **_material_of('chamber_material', 'material', 'case_material')),
                dict({'part': 'Nozzle', 'quantity': 1},
                     **_material_of('nozzle_material')),
            ]
            if motor_data.get('injector_type') or motor_data.get('injector'):
                bom.append(dict({'part': 'Injector', 'quantity': 1},
                                **_material_of('injector_material')))

            design_pressure_bar = structural_params.get('design_pressure')
            chamber_pressure_bar = motor_data.get('chamber_pressure')
            if design_pressure_bar:
                proof_note = (f"Proof test to the structural design pressure "
                              f"({float(design_pressure_bar):.1f} bar)")
            elif chamber_pressure_bar:
                proof_note = (f"Proof test pressure is NOT_DEFINED; operating "
                              f"pressure is {float(chamber_pressure_bar):.1f} bar "
                              f"(run the structural analysis for a design pressure)")
            else:
                proof_note = 'Proof test pressure is NOT_DEFINED'

            complete_package['manufacturing'] = {
                'basis': ('Generic workshop template. HRMA does not size fasteners, '
                          'seals, tolerances or surface finishes; those must come '
                          'from your own detail design and applicable standards.'),
                'bill_of_materials': bom,
                'bill_of_materials_note': (
                    'Fasteners, seals and gaskets are not sized by HRMA and are '
                    'deliberately omitted from this list.'),
                'manufacturing_notes': cad_data['manufacturing_notes'] if 'cad_data' in locals() else [],
                'assembly_instructions': [
                    "1. Machine all components per the technical drawings",
                    f"2. {proof_note}",
                    "3. Install the fuel grain with proper centering",
                    "4. Mount the nozzle with a high-temperature sealant",
                    "5. Attach the injector with its seals",
                    "6. Perform a final leak test before use",
                ],
                'quality_control': [
                    "Visual inspection of all welds",
                    "Dimensional verification against the technical drawings "
                    "(tolerances are NOT_DEFINED by HRMA)",
                    "Surface finish per your detail design (NOT_DEFINED by HRMA)",
                    "Proof test certification",
                ],
            }

        # Generate summary report
        motor_name = safe_name(motor_data.get('motor_name'))

        # v2.6.26: 'Ready for manufacturing' KOŞULSUZ basılıyordu. Ölçüldü:
        # paketin KENDİ güvenlik bölümü UNSAFE (SF=1.28) derken de, boş {}
        # gönderildiğinde de, thrust=-5000 gibi saçma girdide de aynı ifade
        # çıkıyordu. Artık durum güvenlik bölümünden türetilir ve hiçbir
        # koşulda sertifikasyon iması taşımaz — bağımsız inceleme yerine
        # geçecek bir damga üretmiyoruz.
        # include_analysis kapalıysa safety_section hiç kurulmamış olur; o
        # durumda da "hazır" demeyiz — analiz yoksa hüküm de yoktur.
        safety_status = str(safety_section.get('status', 'NOT ANALYZED')).upper()
        safety_factor = safety_section.get('safety_factor')
        if safety_status in ('NOT ANALYZED', 'UNKNOWN'):
            design_status = ('NOT ANALYZED - run the structural analysis before '
                             'manufacturing')
        elif safety_status == 'UNSAFE' or str(
                safety_section.get('risk_level', '')).upper() == 'HIGH':
            detail = (f" (SF={float(safety_factor):.2f})"
                      if isinstance(safety_factor, (int, float)) else '')
            design_status = f'NOT READY - structural safety inadequate{detail}'
        else:
            design_status = ('Design outputs complete - independent review '
                             'required before manufacturing')

        def _fmt(key, spec, unit):
            value = motor_data.get(key)
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                return 'NOT_DEFINED'
            if not math.isfinite(value):
                return 'NOT_COMPUTABLE'
            return f"{value:{spec}} {unit}"

        complete_package['summary'] = {
            'motor_designation': motor_name,
            'total_impulse': _fmt('total_impulse', '.0f', 'N.s'),
            'thrust': _fmt('thrust', '.0f', 'N'),
            'burn_time': _fmt('burn_time', '.1f', 's'),
            'isp': _fmt('isp', '.1f', 's'),
            'chamber_pressure': _fmt('chamber_pressure', '.1f', 'bar'),
            'design_status': design_status,
            'safety_factor': safety_factor,
        }
        # 'estimated_cost' ve 'development_time' KALDIRILDI: '$500-800 USD' ve
        # '2-4 weeks' koddaki iki literal satırdı, hiçbir tedarikçi verisine
        # veya hesaba dayanmıyordu ve motor 500 N de olsa 50 kN de olsa aynıydı.

        return jsonify({
            'status': 'success',
            'complete_package': complete_package,
            'package_info': {
                'motor_name': motor_name,
                'generation_date': datetime.now().isoformat(),
                'package_version': '1.0',
                'files_included': len([k for k, v in package_options.items() if v])
            }
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

#: A10 — üretilen STL dosyalarının jeton -> GERÇEK YOL kaydı.
#:
#: ``export_stl_files`` her çağrıda ``tempfile.mkdtemp`` ile kendi dizinine
#: yazar (v2.6.26 eşzamanlılık düzeltmesi). ``/download/stl`` ucu ise sabit
#: bir dizini (``cwd/cad_exports``) tarıyordu; ikisi buluşmuyordu ve uç, adı
#: tutan ESKİ bir dosyayı sunuyordu. Kayıt defteri bu varsayımı kaldırır:
#: bağlantı, o istekte üretilen dosyanın kendisini gösterir.
#:
#: Sınırlı boyutta (LRU): eski girdiler düşer, dosyalar silinmez (geçici
#: dizin temizliği işletim sisteminde kalır).
_STL_DOWNLOAD_REGISTRY = collections.OrderedDict()
_STL_REGISTRY_LOCK = threading.Lock()
_STL_REGISTRY_MAX = 256


def _register_stl_download(path):
    """Üretilen bir STL yolunu kaydeder, indirme jetonunu döndürür.

    Jeton ``<16 hex>_<güvenli taban ad>`` biçimindedir; ``/download/stl``
    ucunun beyaz listesine (``[A-Za-z0-9._-]{1,128}\\.stl``) uyar. Jeton
    tahmin edilemez olduğu için başka bir sekmenin dosyasına rastlanmaz.
    """
    token = uuid.uuid4().hex[:16] + '_' + safe_name(os.path.basename(path),
                                                    default='motor.stl')
    if not token.lower().endswith('.stl'):
        token += '.stl'
    with _STL_REGISTRY_LOCK:
        _STL_DOWNLOAD_REGISTRY[token] = os.path.abspath(path)
        while len(_STL_DOWNLOAD_REGISTRY) > _STL_REGISTRY_MAX:
            _STL_DOWNLOAD_REGISTRY.popitem(last=False)
    return token


@app.route('/download/stl/<filename>')
def download_stl_file(filename):
    """Üretilen bir STL dosyasını indirir (kayıt defteri, yoksa cad_exports/).

    GÜVENLİK — v2.6.2 düzeltmesi (rastgele dosya okuma):
    Burası eskiden ``send_file(f"./cad_exports/{filename}")`` yapıyordu; adı
    hiç doğrulamıyordu. Flask'ın varsayılan ``string`` dönüştürücüsü ``/``
    geçirmez ama TERS BÖLÜ geçirir; Windows'ta ``\\`` da yol ayracı olduğundan
    ``/download/stl/..\\..\\..\\Windows\\win.ini`` cad_exports dizininin dışına
    çıkıyordu. HRMA Windows'ta exe dağıttığı için bu gerçek bir açıktı ve
    ``CORS(app)`` joker kuralıyla birleşince yanıt herhangi bir web sayfasından
    okunabiliyordu: kullanıcı HRMA açıkken kötü niyetli bir siteye girdiğinde
    o sayfa diskten dosya okuyup dışarı gönderebilirdi.

    Üç katmanlı savunma: (1) ad yalnız güvenli karakterlerden oluşmalı ve
    ``.stl`` ile bitmeli, (2) ``basename`` ile her türlü dizin bileşeni atılır,
    (3) çözümlenmiş mutlak yolun gerçekten export dizini altında kaldığı
    ``os.path.commonpath`` ile doğrulanır (sembolik bağlantı dahil).

    A10 (2026-08-02): ad önce KAYIT DEFTERİNDE aranır. ``/api/export-cad``
    ürettiği her dosyayı oradaki gerçek yoluyla kaydeder, yani bağlantı sabit
    bir dizin tahminine değil üreticinin döndürdüğü yola bağlıdır. Defterde
    olmayan adlar için eski ``cad_exports/`` davranışı korunur (kullanıcı
    ``output_dir`` vererek oraya toplu dışa aktarım yapabiliyor).
    """
    import os
    import re

    # (1) Beyaz liste: harf/rakam/nokta/tire/alt çizgi + .stl uzantısı.
    if not re.fullmatch(r'[A-Za-z0-9._-]{1,128}\.stl', filename or '',
                        flags=re.IGNORECASE):
        return jsonify({'error': 'Invalid filename'}), 400
    # ".." bileşeni beyaz listeden geçebilir (nokta izinli) — açıkça reddet.
    if '..' in filename:
        return jsonify({'error': 'Invalid filename'}), 400

    # (0) Bu istekte üretilmiş bir dosya mı? Yol defterden gelir, kullanıcı
    # girdisinden birleştirilmez — dolayısıyla yol kaçışı imkânsızdır.
    with _STL_REGISTRY_LOCK:
        registered = _STL_DOWNLOAD_REGISTRY.get(filename)
    if registered:
        if os.path.isfile(registered):
            return send_file(registered, as_attachment=True)
        # Geçici dizin temizlenmiş: ESKİ bir dosyayı ikame etmek yerine
        # açıkça "yok" denir (yanlış katıyı sunmak bu bulgunun kendisiydi).
        return jsonify({'error': 'File no longer available'}), 404

    export_dir = os.path.realpath(os.path.join(os.getcwd(), 'cad_exports'))
    # (2) Kalan her dizin bileşenini at.
    candidate = os.path.realpath(os.path.join(export_dir,
                                              os.path.basename(filename)))
    # (3) Çözümlenen yol gerçekten export dizininin altında mı?
    try:
        if os.path.commonpath([export_dir, candidate]) != export_dir:
            return jsonify({'error': 'Invalid filename'}), 400
    except ValueError:
        # Farklı sürücü harfleri (Windows) -> commonpath ValueError verir.
        return jsonify({'error': 'Invalid filename'}), 400

    if not os.path.isfile(candidate):
        return jsonify({'error': 'File not found'}), 404
    return send_file(candidate, as_attachment=True)

@app.route('/api/export-simulation', methods=['POST'])
def export_simulation_file():
    """Export complete simulation data for OpenRocket"""
    try:
        data = request.json
        motor_data = data.get('motor_data', {})
        rocket_data = data.get('rocket_data', None)
        
        # Generate simulation file
        simulation_content = openrocket_exporter.create_simulation_file(motor_data, rocket_data)
        flight_profile = openrocket_exporter.generate_flight_profile(motor_data, rocket_data)
        
        motor_name = motor_data.get('motor_name', 'UZAYTEK-HRM-001')
        filename = f"{motor_name.replace(' ', '_')}_simulation.json"
        
        return jsonify({
            'status': 'success',
            'filename': filename,
            'simulation_content': simulation_content,
            'flight_profile': flight_profile
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/generate-3d', methods=['POST'])
def generate_3d():
    """Generate 3D visualization for motor"""
    try:
        data = request.json
        motor_data = data.get('motor_data', {})
        injector_data = data.get('injector_data', {})
        
        # Generate 3D visualization safely
        try:
            from hrma.visualization.visualization import create_3d_motor_visualization
            motor_3d_plot = create_3d_motor_visualization(motor_data)
        except Exception as viz_error:
            # Fallback: Create simple 3D plot
            import plotly.graph_objects as go
            fig = go.Figure()
            
            # Simple 3D cylinder representation
            theta = np.linspace(0, 2*np.pi, 20)
            z = np.linspace(0, 100, 20)
            theta_mesh, z_mesh = np.meshgrid(theta, z)
            x = 50 * np.cos(theta_mesh)
            y = 50 * np.sin(theta_mesh)
            
            fig.add_trace(go.Surface(
                x=x, y=y, z=z_mesh,
                colorscale='Viridis',
                name='Motor Chamber'
            ))
            
            fig.update_layout(
                title='3D Motor Visualization',
                scene=dict(
                    xaxis_title='X (mm)',
                    yaxis_title='Y (mm)',
                    zaxis_title='Z (mm)'
                ),
                width=800,
                height=600
            )

            # _fig_json: bdata'sız düz JSON (vendor plotly.js 1.58.5 uyumu)
            motor_3d_plot = _fig_json(fig)
        
        return jsonify({
            'status': 'success',
            'plot_data': motor_3d_plot
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

# /api/export-stl için zorunlu geometri alanları (motor tipine göre).
# Bu alanlar OLMADAN üretilecek katı, kullanıcının tasarladığı motor değildir;
# eksikse istek reddedilir (aşağıdaki fail-closed gerekçesine bakınız).
_STL_REQUIRED_FIELDS = {
    'hybrid': ('chamber_diameter', 'chamber_length', 'port_diameter'),
    'solid':  ('chamber_diameter', 'chamber_length'),
    'liquid': ('chamber_diameter', 'chamber_length'),
}


def _reject_incomplete_stl_geometry(motor_data):
    """STL sözleşmesi kapısı: eksikse ``(yanıt, 422)``, tamsa ``None``.

    Faz 5B / H3-B4: bu denetim ``/api/export-stl`` gövdesine gömülüydü, yani
    adı yoktu ve tek başına sınanamıyordu. Artık adlandırılmış tek bir yerde
    durur; sözleşmeyi paylaşmak isteyen bir uç eklendiğinde denetimi ikinci
    kez yazmak zorunda kalmaz (kusurun sınıfı tam olarak buydu). Yanıt gövdesi
    bilerek değiştirilmedi — istemciler ve mevcut bekçiler aynı
    ``status``/``missing_fields`` alanlarını okuyor.

    ``/api/export-stl-zip`` bu kapıyı ŞU AN kullanmıyor; gerekçe o ucun
    açıklamasında (fikstür sahipliği kaynaklı, ayrı karar).

    ``0`` burada "verilmedi" ile aynı kefeye konur (mevcut davranış korundu):
    sıfır çaplı ya da sıfır boylu bir katı imal edilemez, dolayısıyla ret
    nedeni her iki okumada da aynıdır.
    """
    if not isinstance(motor_data, dict):
        motor_data = {}
    motor_type = motor_data.get('motor_type', 'hybrid')
    required = _STL_REQUIRED_FIELDS.get(motor_type,
                                        _STL_REQUIRED_FIELDS['hybrid'])
    missing = [f for f in required if motor_data.get(f) in (None, '', 0)]
    if not missing:
        return None
    return jsonify({
        'status': 'incomplete_geometry',
        'error': ('Cannot export STL: required geometry is missing. '
                  'These values define the exported solid and are not '
                  'assumed on your behalf.'),
        'missing_fields': missing,
        'motor_type': motor_type,
    }), 422


@app.route('/api/export-stl', methods=['POST'])
def export_stl():
    """Motor tasarımını STL olarak dışa aktarır — FAIL-CLOSED.

    v2.6.2 düzeltmesi (sessiz veri bozulması):
    Bu uç eskiden DÖRT ayrı yedek yola sahipti ve hepsi HTTP 200 dönüyordu:
    CAD üretimi çökerse basitleştirilmiş geometri, STL yazımı çökerse
    ``generate_basic_stl_content`` (toplam 6 üçgen: iki düzlemde birer
    çeyrek-daire yelpazesi — kapalı katı değil, nozul yok, port yok, gövde
    yok), o da çökerse TEK üçgenlik 10 mm'lik bir dosya. Ön yüz her durumda
    "STL exported successfully" yazıyordu. Yani başarısız bir dışa aktarım,
    başarılı bir indirme gibi görünüyordu.

    Ayrıca eksik alanlar motor tipine göre sessizce dolduruluyordu
    (hybrid→HTPB/N2O, solid→APCP/BATES, liquid→RP-1/LOX ve itkiden tahmini
    port çapı ``0.02·√(F/1000)`` — kaynaksız, üstelik aynı kavram için kod
    tabanında üç ayrı sihirli sayı vardı). Kullanıcıya bunun bir tahmin
    olduğu hiçbir yerde söylenmiyordu.

    Yeni davranış: eksik zorunlu geometri → 422; CAD/STL üretimi çökerse →
    500 + yapılandırılmış hata. Görselleştirme amaçlı basit geometri artık
    imalata gidebilecek bir dosya olarak dönmüyor.
    """
    try:
        data = request.json or {}
        motor_data = data.get('motor_data', {})
        motor_type = motor_data.get('motor_type', 'hybrid')

        is_valid, validation_msg = motor_validator.validate_export_request(data, 'stl')
        if not is_valid:
            return jsonify({'error': validation_msg, 'status': 'failed'}), 400

        motor_data = motor_validator.sanitize_export_data(motor_data)

        # Zorunlu geometri denetimi — "iyi niyetli varsayılan" YOK.
        # Faz 5B / H3-B4: kapı buradan ``_reject_incomplete_stl_geometry``
        # içine taşındı; ``/api/export-stl-zip`` de aynı yerden geçiyor.
        # ``motor_type`` sanitize sonrası sözlükten yeniden okunur.
        incomplete = _reject_incomplete_stl_geometry(motor_data)
        if incomplete:
            return incomplete

        cad_data = cad_designer.generate_3d_motor_assembly(motor_data)

        if not cad_data or 'assembly_meshes' not in cad_data:
            return jsonify({
                'status': 'cad_failed',
                'error': ('CAD assembly could not be generated for this motor. '
                          'No placeholder geometry is returned — a simplified '
                          'shape would not represent the analysed design.'),
            }), 500

        stl_files = cad_designer.export_stl_files(cad_data['assembly_meshes'])
        if not stl_files:
            return jsonify({
                'status': 'stl_write_failed',
                'error': 'CAD assembly succeeded but no STL file was written.',
            }), 500

        main_stl_path = next(
            (p for p in stl_files
             if 'motor_assembly' in p.lower() or 'complete' in p.lower()),
            stl_files[0])

        import os
        if not os.path.exists(main_stl_path):
            return jsonify({
                'status': 'stl_missing',
                'error': 'STL export reported success but the file is absent.',
            }), 500

        with open(main_stl_path, 'rb') as f:
            stl_content = f.read()
        # v2.6.26: export artik istek basina gecici dizine yaziyor
        # (cad_visualization.export_stl_files). Icerik okundugu an dizin
        # silinir; birikmesini beklemek disk sizintisi olurdu.
        try:
            import shutil, tempfile
            # GUVENLIK KAPISI (31 Tem 2026): burada korumasiz bir
            # `rmtree(os.path.dirname(main_stl_path))` vardi. main_stl_path
            # goreli bir yol oldugunda dirname "." doner ve rmtree CALISMA
            # DIZININI siler — depo kokunun tamami. Uc kez yasandi (test
            # kosarken agac ucdu). Artik yalniz bu modulun kendi urettigi
            # gecici dizin silinir: hem tempdir altinda olmali hem de
            # mkdtemp onekini tasimali.
            _stl_dir = os.path.dirname(os.path.abspath(main_stl_path))
            _tmp_kok = os.path.realpath(tempfile.gettempdir())
            if (os.path.realpath(_stl_dir).startswith(_tmp_kok)
                    and os.path.basename(_stl_dir).startswith('hrma_stl_')):
                shutil.rmtree(_stl_dir, ignore_errors=True)
        except Exception:
            pass
        if not stl_content.strip():
            return jsonify({
                'status': 'stl_empty',
                'error': 'Generated STL file is empty.',
            }), 500

        # Dosya adı kullanıcı girdisinden geliyor -> başlık enjeksiyonuna karşı
        # temizle (safe_name yalnız [A-Za-z0-9._-] bırakır).
        from hrma.utils.input_guard import safe_name
        name = safe_name(motor_data.get('motor_name')
                         or f'UZAYTEK_{motor_type.upper()}_Motor')
        filename = f"{name}_{motor_type}.stl"

        from flask import Response
        return Response(
            stl_content,
            mimetype='application/sla',
            headers={'Content-Disposition': f'attachment;filename="{filename}"'}
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'status': 'failed',
            'error': f'STL export failed: {e}',
        }), 500

# v2.6.26 - OLU YEDEK GEOMETRI URETICILERI KALDIRILDI.
#
# generate_basic_stl_content ve generate_fallback_cad_geometry: CAD
# uretimi basarisiz oldugunda basitlestirilmis bir govde (silindir +
# koni) uretip gecerli export gibi donduren yedek yollardi. v2.6.2'de
# fail-closed davranisa gecildiginde CAGRILARI kaldirilmisti ama
# fonksiyonlar dosyada kaldi. Cagrilmayan kod tehlikelidir: iyi
# niyetli bir refaktor onu diriltir ve uydurma geometri sessizce
# imalata gider. Bekciler: test_stl_export_fail_closed.py,
# test_v262_release_gate.py.



# ---------------------------------------------------------------------------
# Faz 6 / T14 — depolama durumu düzeltmesi (gaz fazı sızıntısı)
#
# ``open_source_propellant_api.DEFAULT_STORAGE_STATE`` her akışkanı GERÇEK
# saklama koşulunda sorar (LOX 90,19 K doymuş sıvı, LH2 20,28 K, N2O 293,15 K
# kendinden basınçlı). Ama ``get_comprehensive_properties`` imzası
# ``temperature=298.15, pressure=101325`` VARSAYILANIYLA geliyor ve bu değerleri
# açıkça geçiriyor; ``temperature is None`` dalı hiç çalışmıyor, yani depolama
# durumu mantığı ETKİSİZ kalıyor.
#
# ÖLÇÜLDÜ (2026-08-03, /api/get-propellant-properties):
#     lox : density 1,3088 kg/m³  (doğrusu 1141,16 →  872x)
#     n2o : density 1,8089 kg/m³  (doğrusu  785,10 →  434x)
#     lh2 : density 0,0823 kg/m³  (doğrusu   70,95 →  862x)
#     lox : viscosity 2,055e-5 Pa·s (doğrusu 1,947e-4 → 9,5x)
# Bu sayılar /liquid Panel 1'deki "Oxidizer Density" alanına YAZILIYOR
# (liquid.html:2075-2081) ve yanına yeşil "Real-time Data" rozeti konuyor.
# Çözücünün kendisi 1141,7 kullandığı için form ile hesap 872 kat ayrışıyordu.
#
# DÜZELTMENİN İLKESİ — kanıtlı değiştirme: bir alan ancak değeri YANLIŞ
# DURUMDAKİ CoolProp sorgusunun (298,15 K / 1 atm) döndürdüğü sayıyla birebir
# eşleşiyorsa değiştirilir. Böylece yerel tablodan gelen değerlere DOKUNULMAZ.
# Bu ayrım şart: yerel tablo ``specific_heat``'i kJ/(kg·K) tutuyor (LH2 14,3;
# RP-1 2,1), CoolProp ise J/(kg·K) veriyor (LH2 9722,9). Körü körüne üzerine
# yazmak 1000x'lik sessiz bir birim hatası doğururdu.
# ---------------------------------------------------------------------------

#: Duruma (T, P, faz) bağlı olan ve bu yüzden yanlış durumda sorulduğunda
#: fiziksel olarak anlamsız çıkan alanlar. Hepsi iki yolda da SI: kg/m³,
#: Pa·s, W/(m·K), J/(kg·K).
_STATE_DEPENDENT_PROPS = ('density', 'viscosity', 'thermal_conductivity',
                          'specific_heat')

#: CoolProp'un depolama durumunda döndürdüğü durum künyesi alanları.
_STORAGE_STATE_META = ('state_temperature_K', 'state_pressure_Pa', 'phase')


def _coolprop_state(propellant_name, temperature=None, pressure=None):
    """CoolProp özelliklerini getir; başarısızlıkta boş sözlük.

    ``temperature=None`` -> akışkanın ``DEFAULT_STORAGE_STATE`` girdisi
    (kriyojenlerde doymuş sıvı). Açık sıcaklık verilirse o kullanılır.
    CoolProp kurulu değilse ya da akışkanı tanımıyorsa termofiziksel anahtar
    HİÇ gelmez — bu durumda düzeltme yapılmaz (uydurma değer yok).
    """
    try:
        props = propellant_api.get_coolprop_properties(
            propellant_name, temperature, pressure)
    except Exception:
        return {}
    return props if isinstance(props, dict) else {}


def _same_number(a, b):
    """İki değer aynı CoolProp sorgusundan mı geldi? (bit-eş ya da 1e-12 bağıl)"""
    if not isinstance(a, (int, float)) or not isinstance(b, (int, float)):
        return False
    if isinstance(a, bool) or isinstance(b, bool):
        return False
    if not (math.isfinite(a) and math.isfinite(b)):
        return False
    return math.isclose(a, b, rel_tol=1e-12, abs_tol=0.0)


def _correct_storage_state(propellant_name, merged_props):
    """Yanlış durumdan (298 K / 1 atm) sızmış alanları depolama durumuyla değiştir.

    Dönüş: ``(merged_props, storage_state | None, coolprop_katkisi_var_mi)``.

    * Bir alan yalnız değeri ORTAM sorgusunun sonucuyla eşleşiyorsa değişir —
      yani sızıntı olduğu KANITLANMIŞSA. Yerel tablodan gelen değer korunur.
    * Depolama durumunda karşılığı olmayan sızmış alan SİLİNİR; yerine tahmin
      konmaz (sıvı yoğunluğu + gaz viskozitesi karışımı üretmemek için).
    """
    storage = _coolprop_state(propellant_name)
    if not any(k in storage for k in _STATE_DEPENDENT_PROPS):
        # CoolProp bu akışkanı tanımıyor (RP-1, HTPB, kerosen) ya da kurulu
        # değil: düzeltilecek bir şey yok, uydurma da yapılmaz.
        return merged_props, None, False

    ambient = _coolprop_state(propellant_name, 298.15, 101325)
    for key in _STATE_DEPENDENT_PROPS:
        if key not in merged_props:
            continue
        if not _same_number(merged_props.get(key), ambient.get(key)):
            continue                      # yerel tablodan gelmiş — dokunma
        if key in storage:
            merged_props[key] = storage[key]
        else:
            merged_props.pop(key, None)   # doğrusu yok -> alan kaldırılır

    state = {k: storage.get(k) for k in _STORAGE_STATE_META if k in storage}
    if storage.get('state'):
        state['description'] = storage['state']
    return merged_props, (state or None), True


@app.route('/api/get-propellant-properties', methods=['POST'])
def get_propellant_properties():
    """Get propellant properties from open-source databases"""
    try:
        data = request.json
        propellant_type = data.get('propellant_type', 'hybrid_fuel')
        propellant_name = data.get('propellant_name', 'htpb')

        # First try local database
        local_props = propellant_db.get_propellant_properties(propellant_name)

        # Then fetch from open-source APIs
        api_props = propellant_api.get_propellant_for_ui(propellant_type, propellant_name)

        # Merge properties (API data takes precedence for real-time accuracy)
        if local_props:
            merged_props = {**local_props, **api_props}
        else:
            merged_props = dict(api_props)

        # T14: gaz fazı sızıntısını gerçek depolama durumuyla değiştir.
        merged_props, storage_state, coolprop_used = _correct_storage_state(
            propellant_name, merged_props)
        if storage_state:
            merged_props['storage_state'] = storage_state

        # T14 (künye dürüstlüğü): ``get_coolprop_properties`` akışkanı hiç
        # tanımasa bile ``source`` alanını koşulsuz 'CoolProp (NIST
        # REFPROP-based)' yazıyor (open_source_propellant_api.py:378). RP-1 ve
        # HTPB'de CoolProp TEK BİR sayı bile üretmiyor; künye o hâlde yalanmış
        # oluyor. Katkı yoksa gerçek kaynağı (yerel tablo) bildir.
        source = api_props.get('data_source', 'Combined sources')
        if not coolprop_used and str(source).startswith('CoolProp'):
            source = ((local_props or {}).get('source')
                      or 'Built-in propellant table')
            merged_props['data_source'] = source

        return jsonify({
            'status': 'success',
            'properties': merged_props,
            'source': source
        })

    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/find-optimum-of', methods=['POST'])
def find_optimum_of_ratio():
    """Find optimum O/F ratio for maximum ISP.

    Tek doğruluk kaynağı: CombustionAnalyzer'ın gerçek denge taraması.
    Desteklenmeyen yakıt/oksitleyici çiftinde sessiz 7.0 varsayılanı
    dönmek yerine 400 + açıklayıcı hata döner (kullanıcı şikayeti:
    propellant seçilmeden 'optimum' üretiliyordu).
    """
    try:
        data = request.json
        motor_type = data.get('motor_type', 'hybrid')
        oxidizer = data.get('oxidizer', 'n2o')
        fuel = data.get('fuel', 'htpb')
        chamber_pressure = data.get('chamber_pressure', 20.0)

        if fuel in ('custom', 'mixture') or oxidizer == 'custom':
            return jsonify({
                'status': 'error',
                'error': ('Optimum O/F requires a defined propellant pair. '
                          'Select a specific fuel and oxidizer first; custom or mixture '
                          'compositions need a full combustion analysis run.')
            }), 400

        from hrma.engines.combustion_analysis import CombustionAnalyzer
        analyzer = CombustionAnalyzer()
        fuel_composition = {fuel: 100.0}
        opt = analyzer.find_optimum_of_ratio(
            fuel_composition, oxidizer, chamber_pressure
        )
        max_isp = float(opt.get('maximum_isp', 0) or 0)
        optimum_of = float(opt.get('optimum_of_ratio', 0) or 0)
        # minimize_scalar başarısız noktalara -1000 cezası verir; tüm
        # noktalar başarısızsa max_isp fiziksel bandın dışında kalır.
        if not (50.0 < max_isp < 600.0) or optimum_of <= 0:
            return jsonify({
                'status': 'error',
                'error': (f'No reliable combustion data for {oxidizer.upper()}/{fuel.upper()}. '
                          'Optimum O/F cannot be determined for this pair — verify the '
                          'propellant selection.')
            }), 400

        # Isp-O/F eğrisi (UI performance_curve bekliyor)
        performance_curve = None
        try:
            import numpy as _np
            of_scan = _np.linspace(max(0.5, optimum_of * 0.4), optimum_of * 1.8, 15)
            isp_vals = []
            for _of in of_scan:
                try:
                    r = analyzer.analyze_combustion(
                        fuel_composition, oxidizer, float(_of), chamber_pressure)
                    isp_vals.append(float(r['performance']['isp']))
                except Exception:
                    isp_vals.append(None)
            if any(v is not None for v in isp_vals):
                performance_curve = {
                    'of_ratios': [float(v) for v in of_scan],
                    'isp_values': isp_vals,
                }
        except Exception as curve_err:
            app.logger.info(f"Optimum O/F curve skipped: {curve_err}")

        recommendation = None
        try:
            recommendation = of_optimizer.get_recommendation(motor_type, oxidizer, fuel)
        except Exception:
            pass

        return jsonify({
            'status': 'success',
            'optimum_of_ratio': optimum_of,
            'max_isp': max_isp,
            'method': 'combustion equilibrium scan (CombustionAnalyzer)',
            'performance_curve': performance_curve,
            'recommendation': recommendation,
        })

    except (ValueError, KeyError) as e:
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/regression-analysis', methods=['POST'])
def regression_analysis():
    """Perform regression rate analysis for hybrid motors"""
    try:
        data = request.json
        motor_data = data.get('motor_data', {})
        
        # Perform regression analysis
        regression_data = regression_analyzer.analyze_regression_vs_time(motor_data)
        
        # Create regression plot
        regression_plot = regression_analyzer.create_regression_plot(regression_data)
        
        # Fuel comparison if requested
        comparison_plot = None
        if data.get('compare_fuels', False):
            comparison_plot = regression_analyzer.compare_fuel_types(motor_data)
        
        # Faz 5B / H3-B9: ölçüldü — ``motor_data`` içinde sonlu olmayan bir
        # değer varken yanıt 198 adet ham ``Infinity`` ve 1 ham ``NaN``
        # taşıyordu (``regression_data.port_diameter`` dizisi baştan sona).
        # Tarayıcı bu gövdeyi ayrıştıramaz. Süzgeç null'a çevirir; sayı
        # UYDURULMAZ, dizi kısaltılmaz.
        return jsonify(sanitize_json_values({
            'status': 'success',
            'regression_data': regression_data,
            'regression_plot': regression_plot,
            'comparison_plot': comparison_plot
        }))
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/trajectory-analysis', methods=['POST'])
def trajectory_analysis():
    """Perform trajectory analysis"""
    try:
        data = request.json
        
        # Extract trajectory parameters
        initial_mass = float(data.get('initial_mass', 50))
        final_mass = float(data.get('final_mass', 25))
        drag_coefficient = float(data.get('drag_coefficient', 0.5))
        reference_area = float(data.get('reference_area', 0.1))
        
        # Genel itki kaynağı (2026-07-13): istek doğrudan thrust + burn_time
        # veriyorsa (katı/sıvı sayfaları) hibrit motor kurulmaz — mevcut
        # hesap sonuçları kullanılır. Verilmemişse eski hibrit yolu çalışır.
        direct_thrust = data.get('thrust')
        direct_burn_time = data.get('burn_time')
        if not (direct_thrust and direct_burn_time):
            # v2.6.26 — burada eskiden thrust=1000 N / burn_time=10 s ile
            # UYDURMA bir hibrit motor kuruluyordu ve yorunge o hayali
            # motordan cikiyordu. Kullanicinin gordugu apoje, delta-v ve
            # ucus suresi kendi motoruna ait degildi. Motor verisi yoksa
            # yorunge hesaplanamaz; sessizce sayi uretmek yerine acikca
            # eksik alan bildirilir.
            return jsonify({
                'status': 'error',
                'error': 'missing_fields',
                'missing_fields': [f for f in ('thrust', 'burn_time')
                                   if not data.get(f)],
                'detail': ('Trajectory analysis needs the motor thrust and '
                           'burn time. Run a motor calculation first, or '
                           'send thrust and burn_time explicitly.'),
            }), 422

        # Create trajectory analyzer
        trajectory_analyzer = TrajectoryAnalyzer()
        
        # Set vehicle parameters
        trajectory_analyzer.set_vehicle_parameters(
            mass_dry=final_mass,
            diameter=np.sqrt(4 * reference_area / np.pi),  # Calculate diameter from reference area
            drag_coefficient=drag_coefficient
        )
        
        # Prepare motor data for trajectory analysis
        thrust = float(direct_thrust)
        burn_time = float(direct_burn_time)
        isp = float(data.get('isp', 200.0))
        motor_data = {
            'thrust': thrust,
            'burn_time': burn_time,
            'total_impulse': float(data.get('total_impulse',
                                            thrust * burn_time)),
            'isp': isp,
            'mass_flow_rate': thrust / (isp * G_0) if isp > 0 else 0.0,
            'propellant_mass_total': initial_mass - final_mass
        }

        # Firlatma kosulu ISTEKTEN gelir.
        #
        # v2.6.26 — bu sozluk tamamen sabitti: 85 derece, 0 m rakim, 40 derece
        # enlem, 0 m/s ruzgar. Oysa istek launch_angle ve wind_speed'i ZATEN
        # tasiyordu (app.js) ve rakim icin trajectory_start_altitude geliyordu
        # ama okunmuyordu. Ayrica sabit 40 derece enlem, kullanicinin hic
        # vermedigi bir sahaya gore yerel yercekimi uyguluyordu; ayni motorun
        # /calculate yolunda ise saha yok ve g0 kullaniliyordu — tek uygulamada
        # iki farkli yercekimi.
        #
        # Verilmeyen anahtar SOZLUGE KONMAZ: cozucunun kendi belgelenmis
        # varsayilani gecerli kalir, uydurma bir saha enjekte edilmez.
        launch_params = {
            'initial_mass': initial_mass,
            'final_mass': final_mass,
        }
        for anahtar, kaynak in (
            ('launch_angle', 'launch_angle'),
            ('launch_altitude', 'trajectory_start_altitude'),
            ('wind_speed', 'wind_speed'),
            ('wind_direction', 'wind_direction'),
            ('launch_latitude', 'launch_latitude'),
            ('launch_longitude', 'launch_longitude'),
        ):
            deger = _positive_float(data.get(kaynak))
            if deger is None:
                # 0 gecerli bir deger (rakim, ruzgar, boylam); _positive_float
                # onu elediginden sifiri ayrica kabul et.
                ham = data.get(kaynak)
                if ham not in (None, '') and float(ham) == 0.0:
                    deger = 0.0
            if deger is not None:
                launch_params[anahtar] = deger
        
        # Calculate trajectory with error tracking
        try:
            print("About to call calculate_trajectory...")
            print(f"Motor data keys: {motor_data.keys()}")
            print(f"Launch params keys: {launch_params.keys()}")
            results = trajectory_analyzer.calculate_trajectory(motor_data, launch_params)
            print("calculate_trajectory completed successfully")
        except Exception as calc_error:
            print(f"calculate_trajectory failed: {calc_error}")
            print(f"Error type: {type(calc_error)}")
            print("Calculate trajectory traceback:")
            traceback.print_exc()
            raise calc_error
        
        # Debug: Print result structure
        print("Trajectory results keys:", results.keys() if isinstance(results, dict) else type(results))
        
        # Create trajectory plot with detailed error tracking
        try:
            print("About to call create_trajectory_plots...")
            trajectory_plot = trajectory_analyzer.create_trajectory_plots(results)
            print("create_trajectory_plots completed successfully")
            
        except Exception as plot_error:
            print(f"create_trajectory_plots failed: {plot_error}")
            print(f"Error type: {type(plot_error)}")
            print(f"Error args: {plot_error.args}")
            print("Full traceback:")
            traceback.print_exc()
            
            # UYDURMA GRAFİK SÖKÜLDÜ (2026-07-23 kararlılık denetimi):
            # burada 0 s'de 0 m, 10 s'de 1000 m gösteren SAHTE bir yörünge
            # çiziliyordu ve yanıt 'status': 'success' dönüyordu. Kullanıcı,
            # çizimin çöktüğünü anlamadan uydurma bir yörüngeye bakıyordu.
            # Artık grafik yerine hata bilgisi döner; çözüm verisi
            # (trajectory_data) hesaplandığı için yanıt yine gönderilir ama
            # grafiğin ÜRETİLEMEDİĞİ açıkça bildirilir.
            trajectory_plot = None
            # i18n: EN üretilir; TR karşılığı i18n_charts.js MSG_PATTERNS'ta.
            plot_error_message = (
                'Trajectory plot could not be generated: %s' % plot_error)

        return jsonify({
            'status': 'success',
            'plot_error': locals().get('plot_error_message'),
            'trajectory_data': sanitize_json_values(results),
            'plot_data': trajectory_plot,
            'engine_data': {
                'thrust': motor_data['thrust'],
                'isp': motor_data['isp'],
                'burn_time': motor_data['burn_time'],
                'total_impulse': motor_data['total_impulse']
            }
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/analyze_safety', methods=['POST'])
def analyze_safety():
    """Comprehensive safety analysis endpoint"""
    try:
        data = request.json or {}

        # v2.6.26 FAIL-CLOSED: bu uç eskiden BOŞ istekten ({}) varsayılan bir
        # motor kurup TAM güvenlik hükmü üretiyordu — tahliye mesafeleri,
        # muayene aralığı, 'ACCEPTABLE' kabul kararı, hatta tıbbi müdahale
        # bölümü. Ölçüldü: `POST /analyze_safety -d '{}'` -> HTTP 200 ve
        # 11100 baytlık analiz, hepsi kullanıcının hiç vermediği
        # varsayılanlardan. Güvenlik hükmü, verisi olmayan bir motor için
        # üretilemez; eksik girdide 422 döner.
        REQUIRED_SAFETY_FIELDS = (
            'chamber_pressure', 'propellant_mass', 'chamber_diameter',
            'wall_thickness',
        )
        missing = [key for key in REQUIRED_SAFETY_FIELDS
                   if data.get(key) in (None, '')]
        if missing:
            return jsonify({
                'status': 'error',
                'error': 'incomplete_safety_input',
                'message': ('A safety assessment cannot be produced from '
                            'defaults. Provide the missing values.'),
                'missing_fields': missing,
            }), 422

        # Extract motor parameters
        motor_type = data.get('motor_type', 'hybrid')
        chamber_pressure = float(data.get('chamber_pressure'))  # bar
        chamber_temperature = float(data.get('chamber_temperature', 3000))  # K
        thrust = float(data.get('thrust', 1000))  # N
        burn_time = float(data.get('burn_time', 10))  # s
        propellant_mass = float(data.get('propellant_mass'))  # kg
        propellant_type = data.get('propellant_type', 'composite')
        facility_type = data.get('facility_type', 'test_stand')

        # Zorunlu olmayan alanlarda varsayılan kullanıldıysa bunu SAKLAMA:
        # kullanıcı hangi sayının kendi verisi olmadığını görmeli.
        defaults_applied = [key for key, default in (
            ('chamber_temperature', 3000), ('thrust', 1000),
            ('burn_time', 10), ('propellant_type', 'composite'),
            ('facility_type', 'test_stand'), ('material', 'steel_4130'),
        ) if data.get(key) in (None, '')]

        # Prepare motor data dictionary
        motor_data = {
            'chamber_pressure': chamber_pressure,
            'chamber_temperature': chamber_temperature,
            'thrust': thrust,
            'burn_time': burn_time,
            'chamber_diameter': float(data.get('chamber_diameter')),
            'wall_thickness': float(data.get('wall_thickness')),
        }

        # Initialize safety analyzer
        safety_analyzer = SafetyAnalyzer()
        
        # Perform comprehensive safety analysis
        # Dalga 0 (2026-07-14): malzeme artık istekten geçer — yapısal
        # emniyet merkezi materials_db dayanımlarıyla hesaplanır (eski
        # sabit 250/400 MPa jenerik çelik kalktı).
        # v2.6.26: motor_type OKUNUYOR ama analize HİÇ geçirilmiyordu — yanıtta
        # daima 'not supplied' görünüyordu ve MOTOR_TYPE_TO_EXPLOSIVE_CLASS
        # eşlemesi bu uçtan asla tetiklenemiyordu (patlayıcı sınıfı yalnız
        # propellant_type'tan çözülüyordu).
        safety_kwargs = dict(
            motor_data=motor_data,
            propellant_mass=propellant_mass,
            propellant_type=propellant_type,
            facility_type=facility_type,
            material=data.get('material', 'steel_4130'),
        )
        import inspect
        if 'motor_type' in inspect.signature(
                safety_analyzer.analyze_comprehensive_safety).parameters:
            safety_kwargs['motor_type'] = motor_type
        safety_results = safety_analyzer.analyze_comprehensive_safety(**safety_kwargs)

        return jsonify({
            'status': 'success',
            'defaults_applied': defaults_applied,
            'safety_analysis': sanitize_json_values(safety_results)
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

#: Yapısal hükmün ASGARİ girdisi. Hiçbirinin varsayılanı yoktur: bunlar
#: olmadan ortada değerlendirilecek bir basınçlı kap yoktur.
REQUIRED_STRUCTURAL_FIELDS = (
    'chamber_pressure', 'chamber_diameter', 'chamber_length',
    'throat_diameter',
)

#: Kazanılmamış hükmün yerine yazılan değer. 'SAFE'/'ACCEPTABLE' gibi bir
#: sıraya girmez; ``structural_panel.js::statusKind`` bilinmeyen değeri
#: nötr 'info' rozetiyle çizer, yani ekranda yeşil onay görünmez.
STRUCTURAL_VERDICT_NOT_EVALUATED = 'NOT_EVALUATED'


def _withhold_unearned_structural_verdict(structural_results,
                                          thermal_input_supplied):
    """Kazanılmamış yapısal hükmü yanıttan çeker, yerine gerekçe koyar.

    B3 ÖLÇÜMÜ (HEAD ``a7ff1e7``, 2026-08-02):
    ``POST /analyze_structural_safety -d '{}'`` AYNI yanıtta hem
    ``safety_factor_is_tautological: true`` hem ``status: 'ACCEPTABLE'``,
    ``risk_level: 'LOW'`` ve ``peak_wall_temperature_K: 300.0``
    döndürüyordu. Bayrak yazılıyor, kimse okumuyordu. İki ayrı kazanılmamış
    çıktı var:

    * **Hüküm.** Boyutlandırma (``design_mode == 'size'``) modunda emniyet
      katsayısı bağımsız bir doğrulama değil, hedefin geri okunmasıdır:
      cidar zaten ``SF_hedef`` sağlayacak şekilde hesaplanır, sonra
      imalat payıyla çarpılır — çelikte 4.0 x 1.2 = 4.8. Bu sayıdan
      "ACCEPTABLE / LOW risk" çıkarmak, kendi cevabını kendine sormaktır.
      Modül bunu ``safety_factor_is_tautological`` ile zaten söylüyordu.
    * **Sıcaklık.** Gaz sıcaklığı ve cidar sıcaklıkları verilmediğinde
      ``structural_analysis._estimate_wall_delta_T`` termal etkiyi KAPATIR
      ve iç/dış cidarı ortam sıcaklığına eşitler (o dosyada:
      ``T_chamber is None -> return T_ambient, T_ambient``). Yanıttaki
      300.0 K bu yüzden "hesaplanmış tepe cidar sıcaklığı" değil, "termal
      analiz koşmadı" demektir. Sahte gösterge yasağı gereği hesaplanmayan
      değer yayımlanmaz.

    Anahtarlar SİLİNMEZ — panel ve sözleşme testleri onları okuyor
    (``tests/test_analysis_dock_contract.py::test_safety_analysis_keys``);
    hüküm alanları ``NOT_EVALUATED``'e, hesaplanmamış sıcaklıklar
    ``None``'a çekilir ve nedeni makine-okur kodla yazılır.

    Döner: yanıtın üst seviyesine konacak beyan sözlüğü.
    """
    chamber = structural_results.get('chamber_analysis') or {}
    safety = structural_results.get('safety_analysis')
    design_mode = chamber.get('design_mode')
    tautological = bool(chamber.get('safety_factor_is_tautological'))

    declaration = {
        'design_mode': design_mode,
        'is_verification': bool(design_mode == 'verify'),
        'thermal_input': ('user_supplied' if thermal_input_supplied
                          else 'not_supplied'),
        'verdict': 'issued',
        # Hükmün NEDEN verilmediği (yalnız verdict == 'withheld' iken dolu).
        'verdict_withheld_reasons': [],
        # Hüküm verilmiş olsa bile KOŞMAYAN değerlendirmeler.
        'not_evaluated': [],
    }

    if not isinstance(safety, dict):
        return declaration

    if tautological:
        declaration['verdict'] = 'withheld'
        declaration['verdict_withheld_reasons'].append(
            'safety_factor_is_tautological')
        # ONAY geri çekilir, UYARI geri çekilmez. B3'ün ifadesi "status EN
        # FAZLA NOT_EVALUATED olsun": NOT_EVALUATED bir TAVANDIR. 'SAFE' ve
        # 'ACCEPTABLE' kazanılmamış onaylardır, tavana çekilir. 'MARGINAL' /
        # 'UNSAFE' ise tehlike bildirimidir — bunu "değerlendirilmedi"ye
        # çevirmek, kazanılmamış bir onaydan daha kötüdür: gerçek bir uyarıyı
        # susturur. (Bu uçta uyarı iki yerden gelebilir: kullanıcının kendi
        # düşük SF hedefinin geri okunması ve termal marj — ikincisi cidar
        # kalınlığından bağımsızdır, yani totolojik değildir.)
        if str(safety.get('status', '')).upper() in ('SAFE', 'ACCEPTABLE'):
            safety['status'] = STRUCTURAL_VERDICT_NOT_EVALUATED
            safety['risk_level'] = STRUCTURAL_VERDICT_NOT_EVALUATED
        safety['verdict_basis'] = (
            'No acceptance is issued: the wall was sized by HRMA to meet the '
            'target safety factor, so the reported safety factor is that '
            'target read back (target x manufacturing allowance), not an '
            'independent verification. Supply the real wall_thickness (m) '
            'to obtain a verification. A remaining MARGINAL/UNSAFE status is '
            'a warning and is never suppressed.')
        # Sayılar KALIR (gerçekten hesaplandılar) ama ne oldukları yazılır.
        safety['minimum_safety_factor_is_tautological'] = True
    else:
        safety['verdict_basis'] = (
            'Verified against the user-supplied wall thickness.')
        safety['minimum_safety_factor_is_tautological'] = False

    if not thermal_input_supplied:
        declaration['not_evaluated'].append('wall_temperature')
        # Hesaplanmamış sıcaklık YAYIMLANMAZ (300.0 K = ortam varsayılanı).
        for key in ('peak_wall_temperature_K', 'derating_wall_temperature_K',
                    'thermal_margin_ratio'):
            if key in safety:
                safety[key] = None
        safety['thermal_assessment'] = 'not_evaluated'
        safety['thermal_assessment_basis'] = (
            'No gas temperature (chamber_temperature) and no wall '
            'temperatures were supplied, so the thermal path did not run. '
            'The values previously reported here were the ambient default '
            'that marks the thermal model as OFF, not computed wall '
            'temperatures. Safety factors above are pressure-only.')
        thermal = structural_results.get('thermal_analysis')
        if isinstance(thermal, dict):
            for key in ('wall_temperature_inner_K', 'wall_temperature_outer_K',
                        'wall_delta_T_K'):
                if key in thermal:
                    thermal[key] = None
            thermal['status'] = 'NOT_MODELLED'
            thermal['basis'] = (
                'Thermal path not run: no gas or wall temperature supplied.')
    else:
        safety['thermal_assessment'] = 'evaluated'

    return declaration


#: ``vessel_status`` alanında ONAY sayılan değerler. Yalnız bunlar geri
#: çekilir; 'FAIL' / 'MARGINAL' bir TEHLİKE BİLDİRİMİDİR ve asla susturulmaz
#: (aynı ayrım ``_withhold_unearned_structural_verdict`` içinde de var).
_VESSEL_APPROVAL_VERDICTS = ('PASS', 'SAFE', 'ACCEPTABLE', 'OK')


def _withhold_unearned_vessel_verdict(solid_results, case_thickness_supplied):
    """Katı motorda kazanılmamış basınçlı kap onayını geri çeker.

    v2.6.27 (A3, Ayberk madde 3 sınıfı) — ÖLÇÜLDÜ: kullanıcı hiçbir kasa
    kalınlığı vermeden ``POST /calculate_solid`` çağırdığında yanıtta
    ``safety_analysis.pressure_safety.vessel_status = 'PASS'`` dönüyordu ve
    bu, denenen dört hazne basıncının hepsinde aynıydı. Sebep totolojidir:
    ``solid_rocket_engine._case_design()`` cidarı verilmediğinde Barlow ile
    ``t = P*r/(sigma_y/SF)`` diye BOYUTLANDIRIR; sonra
    ``_calculate_safety_analysis`` aynı cidarı ``PressureVesselAnalyzer``a
    verip "geçti" der. Kap, geçmek üzere tasarlanmıştır — bu bir doğrulama
    değil, hedefin geri okunmasıdır.

    Sözleşme, hibrit yapısal kapıyla aynı: ONAY geri çekilir, UYARI
    çekilmez; sayılar (kopma basıncı, marj, kalınlık) SİLİNMEZ — gerçekten
    hesaplandılar — ama ne oldukları yazılır.

    Args:
        solid_results: ``/calculate_solid`` sonuç sözlüğü (yerinde güncellenir).
        case_thickness_supplied: Kullanıcı ``case_thickness`` verdi mi?

    Returns:
        Uygulanan beyan sözlüğü ya da kap bloğu yoksa None.
    """
    if not isinstance(solid_results, dict):
        return None
    safety = solid_results.get('safety_analysis')
    if not isinstance(safety, dict):
        return None
    press = safety.get('pressure_safety')
    if not isinstance(press, dict):
        return None

    press['wall_thickness_source'] = ('user_supplied' if case_thickness_supplied
                                      else 'sized_by_hrma')
    press['vessel_status_is_tautological'] = bool(not case_thickness_supplied)
    if case_thickness_supplied:
        press['vessel_status_basis'] = (
            'The case wall thickness was supplied by the user, so the burst '
            'margin and the vessel status are an evaluation of that wall.')
        return press

    if str(press.get('vessel_status', '')).upper() in _VESSEL_APPROVAL_VERDICTS:
        press['vessel_status'] = STRUCTURAL_VERDICT_NOT_EVALUATED
    press['vessel_status_basis'] = (
        'No acceptance is issued: no case thickness was supplied, so HRMA '
        'sized the wall itself from the hoop stress to meet the design '
        'safety factor. Checking that wall and reporting a pass is the '
        'design target read back, not an independent verification. Supply '
        "'case_thickness' (mm) to obtain one. A remaining MARGINAL/FAIL "
        'status is a warning and is never suppressed.')
    return press


def _motor_structural_design_basis(motor_results, wall_thickness_m,
                                   chamber_material):
    """Motor çözücüsünün yapısal hükmüne aynı dürüstlük kapısını uygular.

    v2.6.27 (A3, Ayberk madde 3) — ``_withhold_unearned_structural_verdict``
    yalnız ``/analyze_structural_safety`` ucuna bağlıydı. Oysa kullanıcının
    fiilen kullandığı yol ``/calculate``: hibrit sayfası hesabı oradan alır
    ve yapısal hüküm ``motor.structural_analysis`` içinde döner. ÖLÇÜLDÜ
    (bu işlev yazılmadan önce, cidar ve malzeme GÖNDERİLMEDEN):

        design_mode          : verify
        safety_factor_basis  : verified against user-supplied wall thickness
        wall_thickness_used  : 5,0 mm

    Yani uç katmanının kendi enjekte ettiği 5 mm, kullanıcının cidarı gibi
    "doğrulanmış" sayılıyordu. Girdi tarafı yukarıda (motor kurulumunda)
    düzeltildi; bu işlev de aynı kapının hüküm tarafını bu yola bağlar.

    Args:
        motor_results: Motor çözücüsünün sonuç sözlüğü (yerinde güncellenir).
        wall_thickness_m: İSTEKTEN gelen cidar [m] ya da None.
        chamber_material: İSTEKTEN gelen malzeme adı ya da None.

    Returns:
        Yanıtın üst seviyesine konacak beyan sözlüğü; yapısal sonuç yoksa
        None (uydurma bir beyan üretilmez).
    """
    if not isinstance(motor_results, dict):
        return None
    structural_results = motor_results.get('structural_analysis')
    if not isinstance(structural_results, dict):
        return None

    # Termal yol GERÇEKTEN koştu mu? Tahmin edilmez, yapısal modülün kendi
    # beyanından (v2.6.27'de eklenen ``wall_temperature_source``) okunur.
    thermal = structural_results.get('thermal_analysis')
    wall_temp_source = None
    if isinstance(thermal, dict):
        wall_temp_source = thermal.get('wall_temperature_source')
    thermal_input_supplied = bool(
        wall_temp_source not in (None, 'not_evaluated'))

    basis = _withhold_unearned_structural_verdict(
        structural_results, thermal_input_supplied)
    basis['wall_temperature_source'] = wall_temp_source
    # Malzeme seçilmediyse çelik VARSAYILDIĞI yazılır; sessiz enjeksiyon yok.
    basis['chamber_material_source'] = ('user_supplied' if chamber_material
                                        else 'default:steel_4130')

    # --- Kullanıcının cidarı GERÇEKTEN değerlendirildi mi? ---------------
    # ÖLÇÜLDÜ: aralık dışı bir kalınlık (0,05 mm ya da 500 mm) gönderilince
    # motor uyarıp 5 mm'ye düşüyor, ama "kullanıcı verdi" bayrağını
    # düşüşten ÖNCE hesapladığı için rapor yine "verified against
    # user-supplied wall thickness" diyor ve değerlendirilen 5 mm'yi
    # kullanıcının cidarı gibi gösteriyordu. Sayı doğru, İDDİA yanlış.
    # Karşılaştırma DEĞERLENDİRİLEN kalınlıkla yapılır: iddia ancak
    # istekteki cidar gerçekten hesaba girdiyse kurulur.
    chamber = structural_results.get('chamber_analysis') or {}
    try:
        evaluated_m = float(chamber.get('wall_thickness_used_mm')) / 1000.0
    except (TypeError, ValueError):
        evaluated_m = None
    istek_uygulandi = (
        wall_thickness_m is not None and evaluated_m is not None
        and math.isclose(wall_thickness_m, evaluated_m,
                         rel_tol=1e-6, abs_tol=1e-12))

    if wall_thickness_m is None:
        basis['wall_thickness_source'] = 'sized_by_hrma'
        basis['message'] = (
            'No wall thickness was supplied, so HRMA sized the wall itself. '
            'This result is a DESIGN PROPOSAL, not a verification of a wall '
            'you built.')
    elif istek_uygulandi:
        basis['wall_thickness_source'] = 'user_supplied'
        basis['message'] = (
            'The supplied wall thickness was evaluated against the design '
            'pressure; this result is a verification.')
    else:
        # Girdi geldi ama hesaba GİRMEDİ (aralık dışı ya da okunamadı).
        basis['wall_thickness_source'] = 'user_value_not_applied'
        basis['is_verification'] = False
        basis['submitted_wall_thickness_mm'] = wall_thickness_m * 1000.0
        basis['evaluated_wall_thickness_mm'] = (
            evaluated_m * 1000.0 if evaluated_m is not None else None)
        if basis['verdict'] != 'withheld':
            basis['verdict'] = 'withheld'
        if 'user_wall_thickness_not_applied' not in basis[
                'verdict_withheld_reasons']:
            basis['verdict_withheld_reasons'].append(
                'user_wall_thickness_not_applied')
        basis['message'] = (
            'The wall thickness you submitted was NOT the one evaluated: '
            'the solver rejected it (out of range or unreadable) and fell '
            'back to its own value. The numbers below describe that other '
            'wall, so they are not a verification of your design.')
        # Kazanılmamış ONAY geri çekilir; UYARI dokunulmaz (aynı ayrım
        # _withhold_unearned_structural_verdict içinde de uygulanıyor).
        safety = structural_results.get('safety_analysis')
        if isinstance(safety, dict) and str(
                safety.get('status', '')).upper() in ('SAFE', 'ACCEPTABLE'):
            safety['status'] = STRUCTURAL_VERDICT_NOT_EVALUATED
            safety['risk_level'] = STRUCTURAL_VERDICT_NOT_EVALUATED
            safety['verdict_basis'] = basis['message']
    return basis


@app.route('/analyze_structural_safety', methods=['POST'])
def analyze_structural_safety():
    """Detailed structural safety analysis endpoint.

    v2.6.26 (Faz 4B) — ÜÇ KAPI EKLENDİ:

    **1. Fail-closed girdi kapısı (B3).** Bu uç eskiden BOŞ istekten
    (``{}``) tam bir yapısal hüküm üretiyordu: 20 bar / 100 mm / 500 mm /
    20 mm varsayılanlarıyla kurulmuş, kullanıcının hiç vermediği bir
    motorun gerilmeleri, emniyet katsayıları ve "ACCEPTABLE" kararı.
    Aynı depodaki ``/analyze_safety`` kapısı (bu dosyada, biraz yukarıda)
    buraya uyarlandı: eksik girdide 422 + ``missing_fields``.

    **2. Doğrulama modu bağlandı (B4).** ``StructuralAnalyzer.
    analyze_structure`` iki argüman kabul ediyor — ``actual_wall_thickness``
    ve ``design_safety_factor`` — ama bu uç İKİSİNİ DE geçirmiyordu.
    ÖLÇÜLDÜ: ``wall_thickness=0.001`` (1 mm) ve ``safety_factor=2.0``
    gönderildiğinde modül kendi boyutlandırdığı 5.887 mm'yi kullanıp
    SF 4.8 (= 4.0 hedef x 1.2 imalat payı) ile 'SAFE' diyordu; kullanıcının
    1 mm'lik cidarı hiç sınanmadı. Doğru desen aynı depoda vardı:
    ``hrma/engines/hybrid_rocket_engine.py:1380-1386`` ikisini de geçiriyor.

    **3. Kazanılmamış hüküm geri çekildi (B3).** Bkz.
    ``_withhold_unearned_structural_verdict``.
    """
    try:
        data = request.json or {}

        # --- KAPI 1: eksik girdide hüküm üretilmez -----------------------
        missing = [key for key in REQUIRED_STRUCTURAL_FIELDS
                   if data.get(key) in (None, '')]
        if missing:
            return jsonify({
                'status': 'error',
                'error': 'incomplete_structural_input',
                'message': ('A structural assessment cannot be produced from '
                            'defaults. Provide the missing values.'),
                'missing_fields': missing,
            }), 422

        # --- KAPI 1b: 0 "verildi" SAYILMAZ (Faz 5B / B6) -----------------
        # KAPI 1 yalnız ``None``/``''`` bakıyordu, yani 0 gerçek bir değer
        # kabul ediliyordu. ÖLÇÜLDÜ (HEAD 9d3728e), dört zorunlu alan tek tek
        # 0 yapıldığında:
        #   chamber_pressure=0 -> HTTP 500 'float division by zero'
        #   chamber_diameter=0 -> HTTP 500 'float division by zero'
        #   chamber_length=0   -> HTTP 200 + TAM yapısal hüküm
        #   throat_diameter=0  -> HTTP 200 + TAM yapısal hüküm
        # Yani sıfır boyda / sıfır çapta bir oda için emniyet hükmü
        # çıkarılıyordu. Negatifi daha da kötüydü: dördünde de -1 ile
        # HTTP 200 + tam hüküm. Ayrıca ``chamber_pressure=[1,2]`` (liste)
        # HTTP 500 "float() argument ... not 'list'" veriyordu.
        # ``hrma/utils/input_guard.py:9-12`` projenin kendi ilkesini zaten
        # yazıyor: "0 ile 'verilmedi' ASLA karıştırılmaz".
        invalid = _collect_unphysical_fields(
            data, positive=REQUIRED_STRUCTURAL_FIELDS + ('burn_time',))
        if invalid:
            return jsonify({
                'status': 'error',
                'error': 'invalid_structural_input',
                'message': ('A structural assessment cannot be produced for a '
                            'chamber with non-finite or non-positive '
                            'dimensions. No verdict was computed.'),
                'invalid_fields': invalid,
            }), 422

        # Extract parameters
        chamber_pressure = float(data.get('chamber_pressure'))  # bar
        chamber_diameter = float(data.get('chamber_diameter'))  # m
        chamber_length = float(data.get('chamber_length'))  # m
        throat_diameter = float(data.get('throat_diameter'))  # m
        burn_time = float(data.get('burn_time', 10))  # s
        material = data.get('material', 'steel_4130')

        # Zorunlu olmayan alanlarda varsayılan kullanıldıysa SAKLANMAZ —
        # kullanıcı hangi sayının kendi verisi olmadığını görmeli
        # (``/analyze_safety`` ile aynı desen).
        defaults_applied = [key for key in ('burn_time', 'material')
                            if data.get(key) in (None, '')]

        # --- KAPI 2: doğrulama modu girdileri ----------------------------
        # ``wall_thickness`` METRE (bu uçtaki bütün uzunluklar gibi;
        # ``/analyze_thermal_safety`` de aynı birimi kullanıyor).
        actual_wall_thickness = _json_float(data, 'wall_thickness')
        design_safety_factor = _json_float(data, 'safety_factor')

        if actual_wall_thickness is not None and actual_wall_thickness <= 0:
            # 0 / negatif = "cidarı sen boyutlandır" (thrust alanındaki
            # 0 = "atla" sözleşmesiyle aynı); doğrulama modu açılmaz.
            actual_wall_thickness = None
        if actual_wall_thickness is not None:
            radius = chamber_diameter / 2.0
            if not math.isfinite(actual_wall_thickness):
                return jsonify({
                    'status': 'error', 'error': 'invalid_wall_thickness',
                    'message': "'wall_thickness' must be a finite number (m).",
                }), 422
            if actual_wall_thickness >= radius:
                # Neredeyse her zaman birim hatasıdır (mm yerine m). Sessizce
                # kabul edip dolu silindir gibi hesaplamak yerine söylenir.
                return jsonify({
                    'status': 'error', 'error': 'invalid_wall_thickness',
                    'message': ("'wall_thickness' is given in METRES and must "
                                "be smaller than the chamber radius "
                                f"({radius} m); received "
                                f"{actual_wall_thickness}."),
                }), 422
        if design_safety_factor is not None and (
                not math.isfinite(design_safety_factor)
                or design_safety_factor <= 0):
            return jsonify({
                'status': 'error', 'error': 'invalid_safety_factor',
                'message': ("'safety_factor' must be a finite number greater "
                            "than zero."),
            }), 422

        motor_data = {
            'chamber_pressure': chamber_pressure,
            'chamber_diameter': chamber_diameter,
            'chamber_length': chamber_length,
            'throat_diameter': throat_diameter,
            'burn_time': burn_time
        }
        # İTKİ (v2.6.2, fizik denetimi F075): yapısal modül eksenel burkulmayı
        # motorun kendi itkisinden gelen BASMA yüküyle hesaplar
        # (structural_analysis:483 `motor_data.get('thrust', 0)`). Hiçbir çağıran
        # bu anahtarı geçirmediği için eksenel kuvvet DAİMA 0 kalıyor, burkulma
        # emniyet katsayısı sonsuz çıkıyor ve kontrol her zaman "SAFE" diyordu —
        # yani burkulma kontrolü fiilen ölü koddu.
        # Verilmezse eski davranış (saf iç basınç) korunur.
        if data.get('thrust'):
            motor_data['thrust'] = float(data['thrust'])
        # Termal senaryo bu uçta pasif kalıyordu: gaz sıcaklığı geçilmeyince
        # yapısal modül termal gerilmeyi hiç değerlendirmiyordu (2026-07-14).
        # İstemci gönderirse geçir; 0/boş "termal analizi atla" demektir.
        if data.get('chamber_temperature'):
            motor_data['chamber_temperature'] = float(data['chamber_temperature'])

        # Isı transfer analizinden gelen gerçek cidar sıcaklıkları varsa
        # geçir — yapısal modül termal gradyanı tahmin etmek yerine bunları
        # kullanır (v2.5.2 sözleşmesi, structural_analysis._estimate_wall_delta_T).
        for wall_key in ('wall_temperature_hot', 'wall_temperature_cold'):
            if data.get(wall_key):
                motor_data[wall_key] = float(data[wall_key])

        # Termal yol GERÇEKTEN koştu mu? (Bu üç anahtardan biri yoksa
        # structural_analysis._estimate_wall_delta_T iç/dış cidarı ortam
        # sıcaklığına eşitler, yani termal model KAPALIDIR.)
        thermal_input_supplied = any(
            key in motor_data for key in
            ('chamber_temperature', 'wall_temperature_hot',
             'wall_temperature_cold'))

        # Initialize structural analyzer
        structural_analyzer = StructuralAnalyzer()

        # Perform structural analysis
        # B4: kullanıcının GERÇEK cidarı ve tasarım emniyet katsayısı artık
        # geçiyor. Verilmezse ikisi de None kalır ve modül eski boyutlandırma
        # davranışını sürdürür — ama sonuç 'verify' sayılmaz (aşağıdaki kapı).
        structural_results = structural_analyzer.analyze_structure(
            motor_data=motor_data,
            material=material,
            design_pressure_factor=1.5,
            design_safety_factor=design_safety_factor,
            actual_wall_thickness=actual_wall_thickness,
        )

        # --- KAPI 3: kazanılmamış hüküm yayımlanmaz ----------------------
        design_basis = _withhold_unearned_structural_verdict(
            structural_results, thermal_input_supplied)
        design_basis['wall_thickness_source'] = (
            'user_supplied' if actual_wall_thickness is not None
            else 'sized_by_hrma')
        design_basis['design_safety_factor_source'] = (
            'user_supplied' if design_safety_factor is not None
            else 'materials_database_default')
        if actual_wall_thickness is None:
            design_basis['message'] = (
                'No wall thickness was supplied, so HRMA sized the wall '
                'itself. This result is a DESIGN PROPOSAL, not a '
                'verification of a wall you built.')
        else:
            design_basis['message'] = (
                'The supplied wall thickness was evaluated against the '
                'design pressure; this result is a verification.')

        return jsonify({
            'status': 'success',
            'defaults_applied': defaults_applied,
            'design_basis': design_basis,
            'structural_analysis': sanitize_json_values(structural_results)
        })

    except ValueError as e:
        # _json_float / float() sayıya çeviremedi — bu istemci hatasıdır.
        return jsonify({'status': 'error', 'error': 'invalid_input',
                        'message': str(e)}), 400
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

#: ``/analyze_thermal_safety`` girdilerinden fiziksel olarak POZİTİF olmak
#: zorunda olanlar. Hepsi mutlak ölçekte bir büyüklük: mutlak sıcaklık,
#: mutlak basınç, uzunluk, süre, kütle debisi. Sıfır veya negatifi olan yok.
_THERMAL_SAFETY_POSITIVE_FIELDS = (
    'chamber_pressure', 'chamber_temperature', 'chamber_diameter',
    'chamber_length', 'burn_time', 'mdot_total', 'wall_thickness',
)

#: Kazanılmamış termal hükmün yerine yazılan değer. ``'LOW'`` sırasına
#: GİRMEZ; ``STRUCTURAL_VERDICT_NOT_EVALUATED`` ile aynı sözleşme.
THERMAL_VERDICT_NOT_EVALUATED = 'NOT_EVALUATED'


def _withhold_unevaluated_thermal_verdict(thermal_results):
    """Sonlu olmayan emniyet katsayılarından çıkarılan risk hükmünü geri çeker.

    B2 ÖLÇÜMÜ (HEAD ``9d3728e``) — aynı motor, üç girdi:

    ======================  =====  ==========  ==========  ============
    girdi                   HTTP   risk_level  melting_sf  stress_sf
    ======================  =====  ==========  ==========  ============
    normal                  200    **HIGH**    0,5967      22,15
    bütün sayılar ``NaN``   200    **LOW**     ``null``    **1000000,0**
    bütün sayılar ``-1``    200    **LOW**     8,073       **1000000,0**
    ======================  =====  ==========  ==========  ============

    Yani bozuk girdi güvenlik hükmünü TERS çeviriyordu: kullanıcı "risk
    düşük" görüp devam ediyordu. Kök neden
    ``heat_transfer_analysis.py:_analyze_thermal_safety`` içinde IEEE-754
    NaN karşılaştırma tuzağı: ``nan < 1.5`` ve ``nan < 2.0`` ikisi de
    ``False``, ``nan > 0`` da ``False``. Hiçbir dal girmediği için
    ``risk_level`` başlangıç değeri ``'LOW'``da kalıyor ve
    ``stress_safety_factor`` sentinel ``1e6`` olarak yayımlanıyordu.

    Asıl koruma girdi kapısıdır (aşağıda, ``_collect_unphysical_fields``);
    bu işlev İKİNCİ savunma hattıdır: hesap içinde başka bir yoldan NaN
    doğarsa hüküm yine ``NOT_EVALUATED`` olur — bozuk girdi risk seviyesini
    ASLA DÜŞÜREMEZ.

    Yerinde değiştirir ve geri çekilen hüküm sayısını döner.
    """
    if not isinstance(thermal_results, dict):
        return 0
    safety = thermal_results.get('safety_analysis')
    if not isinstance(safety, dict):
        return 0

    # Hükmün DAYANDIĞI üç sayı. Biri bile sonlu değilse hüküm kazanılmamıştır.
    unevaluated = []
    for key in ('temperature_safety_factor', 'melting_safety_factor',
                'stress_safety_factor'):
        value = safety.get(key)
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            unevaluated.append(key)
            continue
        if not math.isfinite(numeric):
            unevaluated.append(key)

    if not unevaluated:
        return 0

    safety['risk_level'] = THERMAL_VERDICT_NOT_EVALUATED
    safety['risk_level_withheld_because'] = (
        'The risk level is derived from safety factors that could not be '
        'computed as finite numbers ('
        + ', '.join(unevaluated)
        + '). A missing computation is not a low risk, so no risk level is '
          'published.')
    safety['unevaluated_safety_factors'] = unevaluated
    return 1


@app.route('/analyze_thermal_safety', methods=['POST'])
def analyze_thermal_safety():
    """Detailed thermal safety analysis endpoint.

    Faz 5B / B2 — İKİ KAPI EKLENDİ:

    **1. Fiziksel girdi kapısı.** Gönderilen sayıların sonlu ve pozitif
    olması zorunlu (``_THERMAL_SAFETY_POSITIVE_FIELDS``); değilse 422 +
    ``invalid_fields``, hiçbir hüküm üretilmez.

    **2. Kazanılmamış hüküm geri çekilir.** Bkz.
    ``_withhold_unevaluated_thermal_verdict``.
    """
    try:
        data = request.json or {}

        # --- KAPI 1: bozuk girdi risk seviyesini DÜŞÜREMEZ ---------------
        invalid = _collect_unphysical_fields(
            data, positive=_THERMAL_SAFETY_POSITIVE_FIELDS)
        if invalid:
            return jsonify({
                'status': 'error',
                'error': 'invalid_thermal_input',
                'message': ('A thermal safety verdict cannot be produced from '
                            'non-finite or non-positive values. No risk level '
                            'was computed.'),
                'invalid_fields': invalid,
            }), 422

        # Extract parameters
        chamber_pressure = float(data.get('chamber_pressure', 20))  # bar
        chamber_temperature = float(data.get('chamber_temperature', 3000))  # K
        chamber_diameter = float(data.get('chamber_diameter', 0.1))  # m
        chamber_length = float(data.get('chamber_length', 0.5))  # m
        burn_time = float(data.get('burn_time', 10))  # s
        mdot_total = float(data.get('mdot_total', 1.0))  # kg/s
        material = data.get('material', 'steel')
        wall_thickness = float(data.get('wall_thickness', 0.005))  # m
        cooling_type = data.get('cooling_type', 'natural')
        
        motor_data = {
            'chamber_pressure': chamber_pressure,
            'chamber_temperature': chamber_temperature,
            'chamber_diameter': chamber_diameter,
            'chamber_length': chamber_length,
            'burn_time': burn_time,
            'mdot_total': mdot_total
        }
        
        # Initialize heat transfer analyzer
        thermal_analyzer = HeatTransferAnalyzer()
        
        # Perform thermal analysis
        thermal_results = thermal_analyzer.analyze_heat_transfer(
            motor_data=motor_data,
            material=material,
            wall_thickness=wall_thickness,
            ambient_temp=293.15,
            cooling_type=cooling_type
        )

        # --- KAPI 2: kazanılmamış hüküm yayımlanmaz ----------------------
        # sanitize_json_values NaN'ı ``null``a çevirir, yani geri çekme
        # ONDAN ÖNCE yapılmalı — sonrasında sonlu-olmama bilgisi kaybolur.
        _withhold_unevaluated_thermal_verdict(thermal_results)

        return jsonify({
            'status': 'success',
            'thermal_analysis': sanitize_json_values(thermal_results)
        })

    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/analysis/wall-profile', methods=['POST'])
def analyze_wall_profile():
    """Eksenel cidar ısı profili (Dalga 2).

    Nozul konturu boyunca (hazne -> boğaz -> çıkış) A(x)/A_t, izantropik
    M(x), Bartz h_g(x), tasarım ısı akısı q(x) ve denge cidar sıcaklığı
    T_wall_eq(x) dizilerini döner. Grafik ÇİZMEZ — frontend çizer.

    Girdi şeması /analyze_thermal_safety ile aynı çekirdek alanları kullanır
    (chamber_pressure, chamber_temperature, chamber_diameter, chamber_length,
    burn_time, mdot_total, material, wall_thickness, cooling_type) + isteğe
    bağlı geometri/gaz alanları (throat_diameter, exit_diameter,
    expansion_ratio, nozzle_type, gamma, molecular_weight, n_stations...).
    """
    try:
        data = request.json or {}

        # Çekirdek alanlar — /analyze_thermal_safety ile bire bir aynı
        motor_data = {
            'chamber_pressure': float(data.get('chamber_pressure', 20)),   # bar
            'chamber_temperature': float(data.get('chamber_temperature', 3000)),  # K
            'chamber_diameter': float(data.get('chamber_diameter', 0.1)),  # m
            'chamber_length': float(data.get('chamber_length', 0.5)),      # m
            'burn_time': float(data.get('burn_time', 10)),                 # s
            'mdot_total': float(data.get('mdot_total', 1.0)),              # kg/s
        }
        material = data.get('material', 'steel')
        wall_thickness = float(data.get('wall_thickness', 0.005))  # m
        cooling_type = data.get('cooling_type', 'natural')

        # İsteğe bağlı sayısal alanlar: verilirse geçir (0/boş = "kullanma")
        optional_numeric = (
            'gamma', 'molecular_weight', 'gas_constant', 'c_star',
            'throat_diameter', 'exit_diameter', 'expansion_ratio',
            'throat_radius_curvature', 'coolant_side_coefficient',
        )
        for key in optional_numeric:
            value = data.get(key)
            if value in (None, '', 0, '0'):
                continue
            try:
                motor_data[key] = float(value)
            except (TypeError, ValueError):
                pass
        if data.get('nozzle_type'):
            # sample_nozzle_inner_contour konik/bell ayrımını buradan okur
            motor_data['nozzle_angles'] = {'nozzle_type': str(data['nozzle_type'])}

        # Parti 25 (D2 bell/parabolik dirilişi): motorun YAYIMLADIĞI kontur
        # gövdeyle gelirse aynen geçirilir — analyze_axial_profile artık
        # yayımlanmış konturu tercih ediyor ve eksen, FEA köprüsünün okuduğu
        # diziyle TANIM GEREĞİ örtüşüyor. Ölçüldü: yalnız nozzle_type geçirmek
        # 38,1 mm, tam nozzle_angles bile 9,7 mm sapıyordu; kontur geçirmek
        # sapmayı 0,000 mm yapıyor (tests/test_wall_profile_ekseni.py).
        # Şekil denetimi asgari: points listesi taşıyan sözlük; dürüstlük
        # denetimlerinin kendisi analyze_axial_profile içindedir.
        kontur = data.get('nozzle_contour')
        if (isinstance(kontur, dict)
                and isinstance(kontur.get('points'), list)
                and len(kontur['points']) >= 2):
            motor_data['nozzle_contour'] = kontur

        n_stations = int(data.get('n_stations', 40))

        thermal_analyzer = HeatTransferAnalyzer()
        profile = thermal_analyzer.analyze_axial_profile(
            motor_data,
            n_stations=n_stations,
            material=material,
            wall_thickness=wall_thickness,
            ambient_temp=float(data.get('ambient_temp', 293.15)),
            cooling_type=cooling_type,
        )

        return jsonify({
            'status': 'success',
            'wall_profile': sanitize_json_values(profile)
        })

    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

# ===========================================================================
# D5 — YAPISAL FEA UCU (v2.7 analiz modülünün kullanıcı yüzü)
# ---------------------------------------------------------------------------
# hrma/fea/ çözücüleri (D1 yapısal + mesh + D4 köprü) depoda hazırdı ama
# HİÇBİR kullanıcı yüzü yoktu: kullanıcı ne gerilme alanını ne de mesh'i
# görebiliyordu. Bu uç, köprüyü (bridge.run_structural_from_motor) motor
# sonuç sözlüğüyle koşturur ve ÇİZİLEBİLİR ham veriyi döner — grafik
# üretmez, karar vermez (çizim fea_panel.js'in işidir).
#
# Sahte veri yasağı: girdi eksikse köprünün NOT_MODELLED sonucu AYNEN
# (200 ile) geçirilir; uydurma gerilme/SF alanı hiçbir dalda üretilmez.
# Çözücü çökerse 500 döner — "boş ama başarılı" yanıt üretilmez.
# ===========================================================================

#: Tek koşuda üretilecek eleman sayısı üst sınırı. Ardışık inceltme her
#: turda eksenel ve radyal bölümü REFINE_FACTOR ile çarptığı için eleman
#: sayısı tur başına F² kat artar; sınır aşılacaksa TUR SAYISI kısılır
#: (mesh sessizce bozulmaz, kısıtlama yanıtta 'limits' ile beyan edilir).
#: Masaüstü tek-worker sunucuda uzun koşu riskine karşı konmuştur.
#:
#: 2026-08-15 (parti 24) 20000 → 80000: dış yüzey eğrilik tabanı sonrası
#: sıvı varsayılanında tepe vM SINIRLI ama 16384 elemanda son tur değişimi
#: %2,37 ölçüldü. Parti 25: yön-ayrışık inceltme (structural_axisym
#: refine_policy="directional") geldi — canlı sıvı hükmü artık 32768
#: elemanda yeşil (%0,65, 14,8 sn); tavan yine gerekli (bütçe formülü
#: F^(2·rounds) tavanı aynı biçimde sınırlar) ve 5. tura alan bırakır.
FEA_MAX_ELEMS = 80000

#: Cidar (yapısal) ucunun varsayılan inceltme turu — çözücünün ortak
#: DEFAULT_MAX_REFINE_ROUNDS'undan (4) BİLİNÇLİ ayrı. Parti 25 canlı
#: ölçümü (yön-ayrışık inceltme politikasıyla, gerçek sıvı motoru):
#: rounds=4 → YAKINSAMADI (%1,75, 16384 el.); rounds=5 → YAKINSADI
#: (%0,65, 32768 el., 14,8 sn — birlikte-katlamanın 65536'sının yarısı).
#: Tane ucu 4'te kalır — kabul ölçütü (port gerinimi) 2. turda oturuyor;
#: erken duran koşular (hibrit, toleransta durur) etkilenmez. Not: örnekleyici
#: kaynaklı DETERMİNİSTİK test vakası bu turda da yakınsamaz (poligon köşe
#: konsantrasyonu, bulgu defterinde ayrı kalem) — o vaka hüküm beyanını
#: dürüstçe YAKINSAMADI taşır, varsayılanın gerekçesi CANLI vakadır.
FEA_STRUCTURAL_DEFAULT_ROUNDS = 5

#: Eleman kalite eşikleri — TEK TANIM YERİ. Panel bu değerleri yanıttan
#: okur (kaynakta ikinci kez yazılmaz, parametre tutarlılığı kuralı).
#: Kaynak: Verdict / CUBIT dörtgen eleman ölçüt tablosu — 'Aspect Ratio'
#: kabul aralığı 1..4 (Robinson 1987), 'Scaled Jacobian' kabul aralığı
#: 0.5..1 (Knupp 2000); C. J. Stimpson, C. D. Ernst, P. Knupp, P. P. Pebay,
#: D. Thompson, "The Verdict Geometric Quality Library", Sandia National
#: Laboratories, SAND2007-1751, 2007.
FEA_QUALITY_ASPECT_MAX = 4.0
FEA_QUALITY_SCALED_JACOBIAN_MIN = 0.5


def _fea_pick_motor_results(data):
    """İstek gövdesinden FEA'ya verilecek motor sonuç sözlüğünü seçer.

    Hibrit sayfada sonuç ``{'motor': {...}}`` sarmalı içinde, sıvı/katı
    sayfalarda üst düzey sözlüktedir. Seçim TAHMİNLE değil, köprünün kendi
    motor-tipi imzasıyla (``detect_engine_layout``) yapılır: imzayı taşıyan
    ilk aday kullanılır. Hiçbir aday imza taşımıyorsa ilk sözlük adayı
    döner — köprü onun için dürüst NOT_MODELLED üretir (uydurma yok).

    Dönüş: (motor_results | None, hangi alandan alındığı).
    """
    from hrma.fea.bridge import detect_engine_layout

    adaylar = []
    for alan in ('motor_results', 'motor', 'results'):
        deger = data.get(alan)
        if isinstance(deger, dict):
            adaylar.append((deger, alan))
            ic = deger.get('motor')
            if isinstance(ic, dict):
                adaylar.append((ic, alan + '.motor'))
    if isinstance(data, dict) and data:
        adaylar.append((data, 'body'))

    for aday, alan in adaylar:
        if detect_engine_layout(aday) is not None:
            return aday, alan
    for aday, alan in adaylar:
        if alan != 'body':
            return aday, alan
    return (data if isinstance(data, dict) and data else None), 'body'


def _fea_quad_quality(nodes, elems):
    """Dörtgen eleman kalite ölçütleri: en-boy oranı + ölçekli Jacobian.

    En-boy oranı (quad merkezinde): karşılıklı kenarların ortalaması olan
    iki asal eksen X1 = (P1-P0) + (P2-P3), X2 = (P2-P1) + (P3-P0) alınır ve
    ``max(|X1|, |X2|) / min(|X1|, |X2|)`` döner (oran olduğu için 1/2
    normalizasyonu sadeleşir). Kare elemanda 1, uzayan elemanda büyür.

    Ölçekli Jacobian: her köşede köşeye gelen ve köşeden çıkan kenar
    vektörlerinin çapraz çarpımı kendi boylarının çarpımına bölünür (köşe
    açısının sinüsü); eleman değeri dört köşenin EN KÜÇÜĞÜDÜR. Kare
    elemanda 1, dejenere/ters elemanda <= 0. Mesh üreticisinin köşe
    Jacobian denetimiyle (mesh_axisym._corner_jacobian_check) aynı çapraz
    çarpımların normalize hâlidir — yani burada ölçülen, mesh'in kabul
    ölçütünün sürekli karşılığıdır.

    Kaynak: Verdict / CUBIT dörtgen ölçüt tablosu (SAND2007-1751; Aspect
    Ratio "maximum edge length ratios at quad center", Robinson 1987;
    Scaled Jacobian "minimum Jacobian divided by the lengths of the 2 edge
    vectors", Knupp 2000).
    """
    P = np.asarray(nodes, dtype=float)[np.asarray(elems, dtype=int)]  # (M,4,2)
    e_next = np.roll(P, -1, axis=1) - P          # köşeden çıkan kenar
    e_in = np.roll(e_next, 1, axis=1)            # köşeye gelen kenar
    l_next = np.hypot(e_next[:, :, 0], e_next[:, :, 1])
    l_in = np.hypot(e_in[:, :, 0], e_in[:, :, 1])
    cross = (e_in[:, :, 0] * e_next[:, :, 1]
             - e_in[:, :, 1] * e_next[:, :, 0])
    payda = l_in * l_next
    with np.errstate(divide='ignore', invalid='ignore'):
        corner_ratio = np.where(payda > 0.0, cross / payda, 0.0)
    scaled_jacobian = np.min(corner_ratio, axis=1)

    X1 = (P[:, 1] - P[:, 0]) + (P[:, 2] - P[:, 3])
    X2 = (P[:, 2] - P[:, 1]) + (P[:, 3] - P[:, 0])
    l1 = np.hypot(X1[:, 0], X1[:, 1])
    l2 = np.hypot(X2[:, 0], X2[:, 1])
    buyuk = np.maximum(l1, l2)
    kucuk = np.minimum(l1, l2)
    with np.errstate(divide='ignore', invalid='ignore'):
        aspect = np.where(kucuk > 0.0, buyuk / kucuk, np.inf)
    return aspect, scaled_jacobian


@app.route('/api/fea/structural', methods=['POST'])
def api_fea_structural():
    """Motor sonucundan uçtan uca yapısal FEA koşusu (D5 kullanıcı yüzü).

    Girdi (payload): ``motor_results`` / ``motor`` / ``results`` alanlarından
    biri motor çözücüsünün sonuç sözlüğünü taşır (hibrit sayfada
    ``{'motor': {...}}``, sıvı/katı sayfalarda üst düzey sözlük). Alan
    seçimi köprünün motor-tipi imzasıyla yapılır, tahminle değil.

    Çıktı: mesh (düğüm/eleman/indeks ızgarası), von Mises ve emniyet
    katsayısı düğüm alanları, eleman kalite ölçütleri + eşikleri,
    yakınsama geçmişi ve köprünün beyan zinciri. GRAFİK ÜRETİLMEZ —
    çizim istemcinin (fea_panel.js) işidir.

    Dürüstlük sözleşmesi:
      * Girdi eksik  -> köprünün NOT_MODELLED payloadsi AYNEN, HTTP 200
        (uydurma alan yok; 'missing' eksikleri adlandırır).
      * Çözücü hatası -> HTTP 500 (boş ama 'başarılı' yanıt üretilmez).
      * Uzun koşu riski -> eleman sayısı FEA_MAX_ELEMS ile sınırlanır ve
        kısıtlamanın kendisi yanıtın 'limits' bloğunda beyan edilir.
    """
    try:
        data = request.get_json(silent=True) or {}
    except Exception:
        data = {}
    if not isinstance(data, dict):
        data = {}

    motor_results, kaynak_alan = _fea_pick_motor_results(data)
    if not isinstance(motor_results, dict) or not motor_results:
        return jsonify({
            'status': 'error',
            'error': ('Structural FEA needs a motor result object in the '
                      'request body (motor_results / motor / results).'),
        }), 400

    from hrma.fea import bridge as fea_bridge
    from hrma.fea.structural_axisym import (
        DEFAULT_ELEMS_THROUGH_WALL,
        DEFAULT_MAX_REFINE_ROUNDS,
        DEFAULT_N_AXIAL0,
        DEFAULT_REFINE_TOL,
        REFINE_FACTOR,
    )

    def _pozitif_int(alan, varsayilan, alt, ust):
        deger = data.get(alan)
        try:
            sayi = int(deger)
        except (TypeError, ValueError):
            return int(varsayilan)
        return int(min(max(sayi, alt), ust))

    n_axial0 = _pozitif_int('n_axial0', DEFAULT_N_AXIAL0, 4, 256)
    n_radial0 = _pozitif_int('n_radial0', DEFAULT_ELEMS_THROUGH_WALL, 1, 32)
    rounds_istenen = _pozitif_int('max_rounds', FEA_STRUCTURAL_DEFAULT_ROUNDS,
                                  0, 8)

    # Tur kısıtlaması: n_elems(k) = n_axial0 * n_radial0 * F^(2k) <= sınır.
    rounds_izinli = 0
    while rounds_izinli < rounds_istenen:
        sonraki = (n_axial0 * n_radial0
                   * REFINE_FACTOR ** (2 * (rounds_izinli + 1)))
        if sonraki > FEA_MAX_ELEMS:
            break
        rounds_izinli += 1

    try:
        sonuc = fea_bridge.run_structural_from_motor(
            motor_results,
            tol=DEFAULT_REFINE_TOL,
            n_axial0=n_axial0,
            n_radial0=n_radial0,
            max_rounds=rounds_izinli,
        )
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': 'Structural FEA solver failed.',
            'detail': str(e),
        }), 500

    limits = {
        'max_elems': FEA_MAX_ELEMS,
        'rounds_requested': rounds_istenen,
        'rounds_allowed': rounds_izinli,
        'refine_factor': REFINE_FACTOR,
        'n_axial0': n_axial0,
        'n_radial0': n_radial0,
        'clamped': rounds_izinli < rounds_istenen,
        '_basis': ('refinement rounds are capped so that the final element '
                   'count stays within max_elems; the mesh is never silently '
                   'degraded, the cap itself is reported here'),
    }

    if sonuc.get('status') != fea_bridge.BRIDGE_STATUS_OK:
        # Köprünün redli sonucu AYNEN geçer (missing / reason / notes /
        # warning). Hiçbir alan eklenmez, hiçbir sayı uydurulmaz.
        payload = sanitize_json_values(sonuc)
        payload['input_field'] = kaynak_alan
        payload['limits'] = limits
        return jsonify({'status': 'success', 'fea': payload}), 200

    mesh = sonuc['mesh']
    ref = sonuc['refinement']
    inputs = sonuc['inputs']
    malzeme = inputs.get('material')

    aspect, scaled_jacobian = _fea_quad_quality(mesh['nodes'], mesh['elems'])
    aspect_bayrak = ~(aspect <= FEA_QUALITY_ASPECT_MAX)
    jacobian_bayrak = ~(scaled_jacobian >= FEA_QUALITY_SCALED_JACOBIAN_MIN)

    quality = {
        '_source': 'hrma.app._fea_quad_quality',
        '_basis': ('Verdict / CUBIT quadrilateral metric definitions: aspect '
                   'ratio from the two principal axes at the quad centre '
                   '(acceptable range 1..4, Robinson 1987) and scaled '
                   'Jacobian as the minimum corner cross product normalised '
                   'by the two adjacent edge lengths (acceptable range '
                   '0.5..1, Knupp 2000); C. J. Stimpson et al., "The Verdict '
                   'Geometric Quality Library", Sandia SAND2007-1751, 2007. '
                   'Elements outside these ranges are flagged for review; '
                   'the flag is a mesh-quality warning, not a failed result. '
                   'A thin wall swept along a long contour produces elongated '
                   'elements by construction, so a high aspect-ratio count is '
                   'expected here; the scaled Jacobian is the metric that '
                   'reports actual element distortion, and the convergence '
                   'history is what decides whether the mesh is fine enough.'),
        'aspect_ratio': aspect,
        'scaled_jacobian': scaled_jacobian,
        'thresholds': {
            'aspect_ratio_max': FEA_QUALITY_ASPECT_MAX,
            'scaled_jacobian_min': FEA_QUALITY_SCALED_JACOBIAN_MIN,
        },
        'counts': {
            'n_elems': int(mesh['n_elems']),
            'aspect_ratio_flagged': int(np.count_nonzero(aspect_bayrak)),
            'scaled_jacobian_flagged': int(np.count_nonzero(jacobian_bayrak)),
            'flagged': int(np.count_nonzero(aspect_bayrak | jacobian_bayrak)),
        },
        'worst': {
            'aspect_ratio_max': float(np.max(aspect)) if aspect.size else None,
            'scaled_jacobian_min': (float(np.min(scaled_jacobian))
                                    if scaled_jacobian.size else None),
        },
    }

    fea = {
        'status': fea_bridge.BRIDGE_STATUS_OK,
        'engine_layout': sonuc.get('engine_layout'),
        'input_field': kaynak_alan,
        'mesh': {
            'nodes': mesh['nodes'],
            'elems': mesh['elems'],
            # Yapısal ızgara haritası çözücüden AYNEN gelir; istemci düğüm
            # sırasını kendi varsayımıyla yeniden kurmaz.
            'node_index_grid': ref.mesh.node_index_grid,
            'n_nodes': int(mesh['n_nodes']),
            'n_elems': int(mesh['n_elems']),
            'n_axial': int(mesh['n_axial']),
            'n_radial': int(mesh['n_radial']),
            'meta': mesh.get('meta'),
            'coordinate_units': 'm',
            'node_order': ('node index = node_index_grid[i][j]; i is the '
                           'axial station (0..n_axial), j is the through-wall '
                           'layer (0 inner surface, n_radial outer surface)'),
        },
        'fields': {
            'von_mises_pa': sonuc['von_mises_nodal'],
            'safety_factor': sonuc['safety_factor_nodal'],
            '_basis': ('nodal fields come from the solver as returned '
                       '(stress recovery method is declared in meta.cozucu); '
                       'safety factor is null when the material record '
                       'carries no yield strength - no substitute value is '
                       'invented'),
        },
        'scalars': {
            'von_mises_gauss_max_pa': sonuc['von_mises_gauss_max'],
            'min_safety_factor': sonuc['min_safety_factor'],
            'yield_strength_pa': getattr(malzeme, 'yield_strength', None),
            'material_key': inputs.get('material_key'),
            'wall_thickness_m': inputs.get('thickness_m'),
            'inner_pressure_pa': inputs.get('inner_pressure_pa'),
        },
        'quality': quality,
        'convergence': sonuc['convergence'],
        'limits': limits,
        'meta': sonuc['meta'],
    }
    return jsonify({'status': 'success', 'fea': sanitize_json_values(fea)})


# ===========================================================================
# D2 — TERMAL FEA UCU (geçici ısı iletimi çözücüsünün kullanıcı yüzü)
# ---------------------------------------------------------------------------
# hrma/fea/thermal_axisym.py (D2) doğrulanmış hâlde depodaydı ama hiçbir
# çağıranı yoktu: köprünün termal yolu "girdiler hazır, çözücü çağrılmadı"
# durumunda duruyordu. Bu uç zinciri kapatır — köprü mesh'i kurar, sınır
# koşullarını bağlar ve çözücüyü sürer; uç yalnız SINIRLARI koyar ve ham
# veriyi döner (grafik üretmez, karar vermez).
#
# GAZ TARAFI PROFİLİ İSTEK GÖVDESİNDEN GELİR. Motor sonuç sözlükleri
# Bartz h(z) dizisini YAYIMLAMAZ (kod okundu: hibrit motor
# HeatTransferAnalyzer.analyze_axial_profile'i içeride çağırır ama yalnız
# boğaz skalerlerini nozzle_material_analysis.throat_thermal'a koyar; sıvı
# ve katı eksenel profili hiç üretmez). Diziyi üreten uç zaten vardır:
# POST /api/analysis/wall-profile → {'x_mm', 'h_g', 'T_recovery', ...}.
# Beklenen çağrı sırası bu yüzden: /calculate* → /api/analysis/wall-profile
# → /api/fea/thermal. Motor sözlüğünden profil TÜRETİLMEZ; eksikse köprünün
# NOT_MODELLED yanıtı aynen geçer.
# ===========================================================================

#: Zaman adımı bütçesi. D2 sürücüsü adım sayısını ikiye katlayarak yakınsar;
#: üst sınır aşılacaksa TUR SAYISI kısılır (adım sessizce bozulmaz, kısıtlama
#: yanıtta 'limits' ile beyan edilir ve yakınsamama D2 beyanında görünür).
FEA_THERMAL_MAX_STEPS = 4096

#: Sıcaklık geçmişi bellek tavanı: (n_steps + 1) * n_nodes değer. 8 bayt/değer
#: üzerinden 2e6 değer ≈ 16 MB'lık tek koşu; otomatik sürücü aynı anda iki
#: koşu tuttuğu için tepe ≈ 32 MB. Masaüstü tek-worker sunucuda bellek
#: patlamasını önler; eleman bütçesiyle birlikte zaman yakınsamasının ne
#: kadar derine inebileceğini belirler.
FEA_THERMAL_MAX_HISTORY_VALUES = 2_000_000


def _fea_thermal_int(data, alan, varsayilan, alt, ust):
    """İstek gövdesinden tam sayı parametre okur ve [alt, ust] aralığına çeker.

    (Yapısal ucun kendi içindeki eşdeğeri ayrı durur; iki uç birbirinin
    gövdesini değiştirmesin diye burada bağımsız tanımlıdır.)
    """
    try:
        sayi = int(data.get(alan))
    except (TypeError, ValueError):
        return int(varsayilan)
    return int(min(max(sayi, alt), ust))


@app.route('/api/fea/thermal', methods=['POST'])
def api_fea_thermal():
    """Motor sonucu + eksenel gaz profilinden geçici termal FEA koşusu.

    Girdi (payload):
      * ``motor_results`` / ``motor`` / ``results`` — motor çözücüsünün
        sonuç sözlüğü (alan seçimi köprünün motor-tipi imzasıyla yapılır).
      * ``axial_profile`` — ``{'x_mm', 'h_g', 'T_recovery'}``;
        ``/api/analysis/wall-profile`` yanıtının ``wall_profile`` bloğu
        doğrudan verilebilir. ZORUNLUDUR: motor sözlüğünde h(z) yoktur ve
        boğaz skalerinden türetilmez.
      * ``ambient_temperature_K`` — başlangıç (tekdüze) ortam sıcaklığı.
        ZORUNLUDUR: hiçbir motor çözücüsü yayımlamaz, uydurulmaz.
      * ``outer_ambient`` (isteğe bağlı) — ``{'h_W_m2K', 'T_ambient_K',
        'emissivity'}``. Verilmezse dış yüzey ADYABATİK alınır: uydurma
        katsayı konmaz, ısı kaçışı sıfır sayılır (tepe cidar sıcaklığı için
        konservatif üst sınır) ve bu beyan edilir.
      * ``n_axial`` / ``n_radial`` (isteğe bağlı) — uzamsal bölüm sayıları.

    Çıktı: mesh, son andaki düğüm sıcaklık alanı, iç yüzey sıcaklık dizisi,
    tepe cidar sıcaklığı geçmişi, enerji bütçesi, zaman adımı beyanı,
    malzeme sıcaklık sınırı karşılaştırması ve köprünün beyan zinciri.
    Tüm zaman adımlarının alan geçmişi (T_history) yanıta KONMAZ — boyutu
    adım sayısı × düğüm sayısıdır; yanıt son alan + tepe geçmişiyle sınırlı
    tutulur ve bu 'limits' içinde beyan edilir.

    Dürüstlük sözleşmesi:
      * Girdi eksik  -> köprünün NOT_MODELLED payloadsi AYNEN, HTTP 200
        ('missing' eksiği adlandırır; hiçbir sıcaklık alanı üretilmez).
      * Çözücü modülü yok/API eksik -> köprünün NOT_AVAILABLE /
        INPUTS_READY payloadsi, HTTP 200.
      * Mesh veya çözücü hatası -> HTTP 500 (boş ama 'başarılı' yanıt yok).
      * Uzun koşu/bellek riski -> eleman ve zaman adımı bütçeleri
        uygulanır ve kısıtlamanın kendisi 'limits' içinde beyan edilir.
    """
    try:
        data = request.get_json(silent=True) or {}
    except Exception:
        data = {}
    if not isinstance(data, dict):
        data = {}

    motor_results, kaynak_alan = _fea_pick_motor_results(data)
    if not isinstance(motor_results, dict) or not motor_results:
        return jsonify({
            'status': 'error',
            'error': ('Thermal FEA needs a motor result object in the '
                      'request body (motor_results / motor / results).'),
        }), 400

    from hrma.fea import bridge as fea_bridge
    from hrma.fea.mesh_axisym import (
        DEFAULT_ELEMS_THROUGH_WALL,
        MIN_ELEMS_THROUGH_WALL,
    )
    from hrma.fea.structural_axisym import DEFAULT_N_AXIAL0

    n_radial = _fea_thermal_int(data, 'n_radial', DEFAULT_ELEMS_THROUGH_WALL,
                                MIN_ELEMS_THROUGH_WALL, 32)
    n_axial_istenen = _fea_thermal_int(data, 'n_axial', DEFAULT_N_AXIAL0,
                                       4, 512)
    # Eleman bütçesi: mesh üreticisi katman sayısını alt sınıra ÇEKEBİLİR,
    # o yüzden bütçe fiilen kullanılacak katman sayısıyla hesaplanır.
    n_radial_fiili = max(n_radial, MIN_ELEMS_THROUGH_WALL)
    n_axial = min(n_axial_istenen,
                  max(4, FEA_MAX_ELEMS // max(n_radial_fiili, 1)))
    n_nodes_tahmin = (n_axial + 1) * (n_radial_fiili + 1)

    # Zaman adımı bütçesi: D2 sürücüsü n_steps0 * 2^k adım kullanır.
    # D2 depoda yoksa bütçe hesaplanamaz — ve hesaplanmasına gerek de yoktur:
    # köprü çözücüyü çağıramadan NOT_AVAILABLE döner. Import hatası burada
    # yutulup 500'e dönüşmez; kullanıcı açıklayıcı yanıtı alır.
    try:
        from hrma.fea.thermal_axisym import (
            DEFAULT_MAX_DT_HALVINGS,
            DEFAULT_N_STEPS0,
        )
    except ImportError:
        n_steps0 = halvings_istenen = halvings_izinli = None
    else:
        n_steps0 = DEFAULT_N_STEPS0
        halvings_istenen = _fea_thermal_int(data, 'max_halvings',
                                            DEFAULT_MAX_DT_HALVINGS, 0, 12)
        halvings_izinli = 0
        while halvings_izinli < halvings_istenen:
            adim = DEFAULT_N_STEPS0 * 2 ** (halvings_izinli + 1)
            if adim > FEA_THERMAL_MAX_STEPS:
                break
            if (adim + 1) * n_nodes_tahmin > FEA_THERMAL_MAX_HISTORY_VALUES:
                break
            halvings_izinli += 1

    limits = {
        'max_elems': FEA_MAX_ELEMS,
        'max_steps': FEA_THERMAL_MAX_STEPS,
        'max_history_values': FEA_THERMAL_MAX_HISTORY_VALUES,
        'n_axial_requested': n_axial_istenen,
        'n_axial': n_axial,
        'n_radial_requested': n_radial,
        'n_radial_effective': n_radial_fiili,
        'halvings_requested': halvings_istenen,
        'halvings_allowed': halvings_izinli,
        'n_steps0': n_steps0,
        'clamped': (n_axial < n_axial_istenen
                    or (halvings_izinli is not None
                        and halvings_izinli < halvings_istenen)),
        'temperature_history_in_payload': False,
        '_basis': ('two budgets are enforced: the element count (max_elems) '
                   'and the temperature-history size ((n_steps + 1) * '
                   'n_nodes <= max_history_values). Time-step halvings are '
                   'capped so both hold; the mesh is never silently degraded '
                   'and a capped run reports it here, while the solver own '
                   'declaration (time_step.beyan) says whether the time step '
                   'actually converged. The full T(t) field history is not '
                   'serialised - the payload carries the final field and the '
                   'peak-wall-temperature history.'),
    }

    try:
        sonuc = fea_bridge.run_thermal_from_motor(
            motor_results,
            axial_profile=data.get('axial_profile'),
            ambient_temperature_K=data.get('ambient_temperature_K'),
            outer_ambient=data.get('outer_ambient'),
            include_chamber=bool(data.get('include_chamber', False)),
            n_axial=n_axial,
            n_radial=n_radial,
            max_halvings=halvings_izinli,
        )
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': 'Thermal FEA solver failed.',
            'detail': str(e),
        }), 500

    if sonuc.get('status') != fea_bridge.BRIDGE_STATUS_OK:
        # Köprünün redli sonucu AYNEN geçer. 'inputs' bloğu çözücü nesneleri
        # (Material vb.) taşıyabildiği için yalnız beyan zinciri alınır;
        # hiçbir sayı eklenmez, hiçbir alan uydurulmaz.
        payload = {k: v for k, v in sonuc.items() if k != 'inputs'}
        if isinstance(sonuc.get('inputs'), dict):
            payload['inputs_basis'] = sonuc['inputs'].get('_basis')
        payload = sanitize_json_values(payload)
        payload['input_field'] = kaynak_alan
        payload['limits'] = limits
        return jsonify({'status': 'success', 'fea': payload}), 200

    mesh = sonuc['mesh']
    inputs = sonuc['inputs']
    tm = inputs['thermal_material']

    fea = {
        'status': fea_bridge.BRIDGE_STATUS_OK,
        'engine_layout': sonuc.get('engine_layout'),
        'input_field': kaynak_alan,
        'mesh': {
            'nodes': mesh['nodes'],
            'elems': mesh['elems'],
            'node_index_grid': mesh['node_index_grid'],
            'n_nodes': int(mesh['n_nodes']),
            'n_elems': int(mesh['n_elems']),
            'n_axial': int(mesh['n_axial']),
            'n_radial': int(mesh['n_radial']),
            'meta': mesh.get('meta'),
            'coordinate_units': 'm',
            'node_order': ('node index = node_index_grid[i][j]; i is the '
                           'axial station (0..n_axial), j is the through-wall '
                           'layer (0 inner/gas side, n_radial outer surface)'),
        },
        'fields': {
            'temperature_final_K': sonuc['T_final_K'],
            'inner_surface_z_m': sonuc['inner_surface']['z_m'],
            'inner_surface_T_final_K': sonuc['inner_surface']['T_final_K'],
            '_basis': ('nodal temperature field as returned by the solver at '
                       'the end of the burn; the inner-surface arrays are the '
                       'j = 0 nodes of the same field (gas side), read out, '
                       'not recomputed'),
        },
        'history': {
            'times_s': sonuc['times_s'],
            'peak_wall_T_history_K': sonuc['peak_wall_T_history_K'],
            '_basis': ('peak nodal temperature at each time point; this is '
                       'the quantity the solver time-step convergence is '
                       'measured on'),
        },
        'scalars': {
            'peak_wall_T_K': sonuc['peak_wall_T_K'],
            'peak_time_s': sonuc['peak_time_s'],
            'peak_node': sonuc['peak_node'],
            'inner_surface_peak_T_K': sonuc['inner_surface']['peak_T_K'],
            'inner_surface_peak_z_m': sonuc['inner_surface']['peak_z_m'],
            'burn_time_s': inputs['burn_time_s'],
            'ambient_temperature_K': data.get('ambient_temperature_K'),
            'material_key': inputs.get('material_key'),
            'wall_thickness_m': inputs.get('thickness_m'),
            'thermal_conductivity_W_mK': tm.get('thermal_conductivity_W_mK'),
            'specific_heat_J_kgK': tm.get('specific_heat_J_kgK'),
            'density_kg_m3': tm.get('density_kg_m3'),
        },
        'energy': sonuc['energy'],
        'time_step': sonuc['time_step'],
        'material_limits': sonuc['material_limits'],
        'warnings': sonuc['warnings'],
        'limits': limits,
        'meta': sonuc['meta'],
    }
    return jsonify({'status': 'success', 'fea': sanitize_json_values(fea)})


# ===========================================================================
# V2.7 Aşama C — KATI TANE KESİTİ DÜZLEMSEL FEA UCU
# ---------------------------------------------------------------------------
# Kamara/lüle cidarı eksenel simetriktir ve /api/fea/structural onu görür;
# katı yakıt tanesi kesitleri (star/finocyl/slotted) eksenel simetrik
# DEĞİLDİR (docs/mimari/yol-haritasi.md §5). Bu uç, kesit geometrisini
# MOTORUN KENDİ port fonksiyonlarından kuran köprüyü
# (hrma.fea.bridge.run_planar_grain_fea) sürer ve ÇİZİLEBİLİR ham veriyi
# döner (grafik üretmez, karar vermez). Girdi beyaz listesi + eksik zorunlu
# alanda 422 (gimbal/termal-koruma ucu sözleşmesi); Host kapısı tüm
# rotalarda olduğu gibi before_request'te çalışır.
# ===========================================================================

#: İstek gövdesi beyaz listesi. Geometri alan adları motor arayüzünün
#: KENDİ override anahtarlarıdır (bridge.PLANAR_GRAIN_GEOMETRY_KEYS —
#: çeviri katmanı yok; sayılar mm, adetler tam sayı, oran birimsiz).
#: Beyaz liste dışı anahtarlar termal-koruma ucundaki desenle SESSİZCE
#: düşer (TypeError 500'e düşmesin).
_PLANAR_GRAIN_KEYS = (
    'grain_type', 'propellant_type', 'outer_diameter_mm', 'core_diameter_mm',
    'chamber_pressure_bar', 'grain_length_mm',
    'star_points', 'star_radius',
    'fin_count', 'fin_width', 'fin_length', 'finned_length_fraction',
    'slot_count', 'slot_width', 'slot_depth',
)

#: ZORUNLU alanlar — ``run_planar_grain_fea`` imzasındaki varsayılansız
#: argümanların birebir karşılığı (kesit + malzeme + yük; hiçbirinin
#: uydurulabilir varsayılanı yoktur).
_PLANAR_GRAIN_REQUIRED = (
    'grain_type', 'propellant_type', 'outer_diameter_mm', 'core_diameter_mm',
    'chamber_pressure_bar',
)


@app.route('/api/fea/planar-grain', methods=['POST'])
def api_fea_planar_grain():
    """Katı tane kesiti için uçtan uca 2B düzlemsel FEA koşusu.

    Girdi (JSON): ``_PLANAR_GRAIN_KEYS`` beyaz listesi. Kesit çokgeni
    SUNUCUDA motorun kendi port fonksiyonlarıyla kurulur (tek-kaynak
    kuralı); istemci geometri üretmez.

    Çıktı: mesh (düğüm/eleman/indeks ızgarası), von Mises ve maks asal
    gerilme düğüm alanları, port yüzeyi lif gerinimi + kopma uzaması
    karşılaştırması (kabul ölçütü NASA SP-8073 gereği GERİNİM), eleman
    kalite ölçütleri, yakınsama geçmişi ve beyan zinciri.

    Dürüstlük sözleşmesi:
      * Eksik zorunlu alan   -> 422 + ``missing_fields`` (makine-okur).
      * Geçersiz sayı/geometri (negatif web, port >= dış çap, tanınmayan
        tip) -> 400 + modülün kendi gerekçesi (metin yumuşatılmaz).
      * Desteklenmeyen ama meşru tip (wagon_wheel, end_burner) -> HTTP 200
        + köprünün beyanlı NOT_MODELLED paketi (uydurma alan yok).
      * Çözücü hatası -> 500 (boş ama 'başarılı' yanıt üretilmez).
      * Uzun koşu riski -> eleman sayısı FEA_MAX_ELEMS ile sınırlanır ve
        kısıtlama 'limits' bloğunda beyan edilir.
    """
    try:
        data = request.get_json(silent=True) or {}
    except Exception:
        data = {}
    if not isinstance(data, dict):
        data = {}

    params = {k: data[k] for k in _PLANAR_GRAIN_KEYS
              if data.get(k) not in (None, '')}

    # --- ZORUNLU ALAN KAPISI (gimbal/termal-koruma deseni) ----------------
    missing = [k for k in _PLANAR_GRAIN_REQUIRED if k not in params]
    if missing:
        return jsonify({
            'status': 'error',
            'error': 'incomplete_planar_grain_input',
            'message': ('Planar grain FEA cannot run without these inputs; '
                        'they have no default.'),
            'missing_fields': missing,
        }), 422

    from hrma.fea import bridge as fea_bridge
    from hrma.fea.mesh_axisym import (
        DEFAULT_ELEMS_THROUGH_WALL,
        MIN_ELEMS_THROUGH_WALL,
    )
    from hrma.fea.planar_grain import DEFAULT_N_THETA0
    from hrma.fea.structural_axisym import (
        DEFAULT_MAX_REFINE_ROUNDS,
        DEFAULT_REFINE_TOL,
        REFINE_FACTOR,
    )

    def _pozitif_int(alan, varsayilan, alt, ust):
        try:
            sayi = int(data.get(alan))
        except (TypeError, ValueError):
            return int(varsayilan)
        return int(min(max(sayi, alt), ust))

    n_theta0 = _pozitif_int('n_theta0', DEFAULT_N_THETA0, 4, 256)
    n_radial0 = _pozitif_int('n_radial0', DEFAULT_ELEMS_THROUGH_WALL,
                             MIN_ELEMS_THROUGH_WALL, 32)
    rounds_istenen = _pozitif_int('max_rounds', DEFAULT_MAX_REFINE_ROUNDS,
                                  0, 8)

    # Tur kısıtlaması: n_elems(k) = n_theta0 * n_radial0 * F^(2k) <= sınır
    # (yapısal ucun kendi bütçe deseni; mesh sessizce bozulmaz, beyan edilir).
    rounds_izinli = 0
    while rounds_izinli < rounds_istenen:
        sonraki = (n_theta0 * n_radial0
                   * REFINE_FACTOR ** (2 * (rounds_izinli + 1)))
        if sonraki > FEA_MAX_ELEMS:
            break
        rounds_izinli += 1

    limits = {
        'max_elems': FEA_MAX_ELEMS,
        'rounds_requested': rounds_istenen,
        'rounds_allowed': rounds_izinli,
        'refine_factor': REFINE_FACTOR,
        'n_theta0': n_theta0,
        'n_radial0': n_radial0,
        'clamped': rounds_izinli < rounds_istenen,
        '_basis': ('refinement rounds are capped so that the final element '
                   'count stays within max_elems; the mesh is never silently '
                   'degraded, the cap itself is reported here'),
    }

    geometri = {k: params[k]
                for k in fea_bridge.PLANAR_GRAIN_GEOMETRY_KEYS
                if k in params}
    try:
        sonuc = fea_bridge.run_planar_grain_fea(
            grain_type=params['grain_type'],
            propellant_type=params['propellant_type'],
            outer_diameter_mm=params['outer_diameter_mm'],
            core_diameter_mm=params['core_diameter_mm'],
            chamber_pressure_bar=params['chamber_pressure_bar'],
            grain_length_mm=params.get('grain_length_mm'),
            geometry_overrides=geometri,
            tol=DEFAULT_REFINE_TOL,
            n_theta0=n_theta0,
            n_radial0=n_radial0,
            max_rounds=rounds_izinli,
        )
    except ValueError as e:
        # Modülün geçerlilik beyanı (negatif web, tanınmayan tip, geçersiz
        # sayı) — istemci hatasıdır, metin aynen iletilir.
        return jsonify({'status': 'error', 'error': str(e)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'error': 'Planar grain FEA solver failed.',
            'detail': str(e),
        }), 500

    if sonuc.get('status') != fea_bridge.BRIDGE_STATUS_OK:
        # Köprünün beyanlı reddi AYNEN geçer (reason / warning); hiçbir
        # alan eklenmez, hiçbir sayı uydurulmaz.
        payload = sanitize_json_values(sonuc)
        payload['limits'] = limits
        return jsonify({'status': 'success', 'fea': payload}), 200

    mesh = sonuc['mesh']
    ref = sonuc['refinement']
    res = sonuc['result']
    bore = sonuc['bore']

    aspect, scaled_jacobian = _fea_quad_quality(mesh['nodes'], mesh['elems'])
    aspect_bayrak = ~(aspect <= FEA_QUALITY_ASPECT_MAX)
    jacobian_bayrak = ~(scaled_jacobian >= FEA_QUALITY_SCALED_JACOBIAN_MIN)

    quality = {
        '_source': 'hrma.app._fea_quad_quality',
        '_basis': ('Verdict / CUBIT quadrilateral metric definitions (see '
                   '/api/fea/structural for the citation, SAND2007-1751); '
                   'elements outside the ranges are flagged for review, the '
                   'flag is a mesh-quality warning, not a failed result. '
                   'Sharp port corners (star tips, slot roots) concentrate '
                   'distorted elements by construction; the convergence '
                   'declaration is what decides how to read the peak value.'),
        'aspect_ratio': aspect,
        'scaled_jacobian': scaled_jacobian,
        'thresholds': {
            'aspect_ratio_max': FEA_QUALITY_ASPECT_MAX,
            'scaled_jacobian_min': FEA_QUALITY_SCALED_JACOBIAN_MIN,
        },
        'counts': {
            'n_elems': int(mesh['n_elems']),
            'aspect_ratio_flagged': int(np.count_nonzero(aspect_bayrak)),
            'scaled_jacobian_flagged': int(np.count_nonzero(jacobian_bayrak)),
            'flagged': int(np.count_nonzero(aspect_bayrak | jacobian_bayrak)),
        },
        'worst': {
            'aspect_ratio_max': float(np.max(aspect)) if aspect.size else None,
            'scaled_jacobian_min': (float(np.min(scaled_jacobian))
                                    if scaled_jacobian.size else None),
        },
    }

    fea = {
        'status': fea_bridge.BRIDGE_STATUS_OK,
        'grain_type': sonuc.get('grain_type'),
        'propellant_type': sonuc.get('propellant_type'),
        'symmetry_fraction': sonuc.get('symmetry_fraction'),
        'mesh': {
            'nodes': mesh['nodes'],
            'elems': mesh['elems'],
            'node_index_grid': mesh['node_index_grid'],
            'n_nodes': int(mesh['n_nodes']),
            'n_elems': int(mesh['n_elems']),
            'n_theta': int(mesh['n_theta']),
            'n_radial': int(mesh['n_radial']),
            'meta': mesh.get('meta'),
            'coordinate_units': 'm',
            'node_order': ('node index = node_index_grid[i][j]; i is the '
                           'circumferential station (0..n_theta, increasing '
                           'angle), j is the through-web layer (0 port '
                           'surface, n_radial outer/case surface)'),
        },
        'fields': {
            'von_mises_pa': sonuc['von_mises_nodal'],
            'max_principal_pa': sonuc['max_principal_nodal'],
            '_basis': ('nodal fields as returned by the solver (stress '
                       'recovery method declared in meta.cozucu); no safety '
                       'factor from yield is published - the solid grain '
                       'acceptance criterion is STRAIN (NASA SP-8073), see '
                       'the bore block'),
        },
        'bore': {
            'node_ids': bore['node_ids'],
            'strain_nodal': bore['strain_nodal'],
            'max_strain': bore['max_strain'],
            'strain_capability': bore['strain_capability'],
            'strain_margin': bore['strain_margin'],
            '_basis': bore.get('_basis'),
        },
        'scalars': {
            'von_mises_gauss_max_pa': sonuc['von_mises_gauss_max'],
            'von_mises_nodal_max_pa': res.max_von_mises_nodal,
            'max_bore_strain': bore['max_strain'],
            'strain_capability': bore['strain_capability'],
            'strain_margin': bore['strain_margin'],
            'bore_pressure_pa': sonuc['inputs']['yuk']['Pc_Pa'],
            'grain_modulus_pa': sonuc['inputs']['malzeme']['E_Pa'],
            'grain_poisson_ratio': sonuc['inputs']['malzeme']['nu'],
        },
        'quality': quality,
        'convergence': sonuc['convergence'],
        'warnings': sonuc.get('warnings', []),
        'limits': limits,
        'meta': sonuc['meta'],
    }
    return jsonify({'status': 'success', 'fea': sanitize_json_values(fea)})


# ===========================================================================
# DALGA B — v3 CFD LÜLE UCU: POST /api/cfd/nozzle
# ---------------------------------------------------------------------------
# hrma/cfd/ (2B eksenel simetrik Euler çözücüsü + Summerfield ayrılma köprüsü)
# doğrulanmış hâlde depodaydı ama HİÇBİR kullanıcı yüzü onu çağırmıyordu.
# Bu uç zinciri kapatır: yayımlanmış lüle konturu + motor gaz durumu →
# kararlı-hâl çözümü → duvar basıncı → ayrılma hükmü → ÇİZİLEBİLİR ham veri.
# GRAFİK ÜRETMEZ, KARAR VERMEZ (çizim panelin işi).
#
# DÜRÜSTLÜK SÖZLEŞMESİ
#   * Uydurma varsayılan YOK: kontur, P0, T0, gamma, R ve P_ortam istemciden
#     gelir; biri eksikse 422 + hangi alanın NEDEN gerekli olduğu.
#   * Yakınsamayan / ıraksayan koşu 200 döner ama converged=False ve
#     çözücünün kendi beyanıyla döner (çözücü felsefesi); ayrılma hükmü de
#     köprünün kendi 'suspect' güven etiketini taşır.
#   * Çözünürlük SERBEST SAYI DEĞİL, ölçülmüş beyaz listedir: her seviyenin
#     (ni, nj) değeri ve ÖLÇÜLEN en kötü süresi yanıtta beyan edilir.
#   * Sürücü ayarları (tolerans, CFL, rampa) uçta YENİDEN TANIMLANMAZ;
#     çözücünün kendi varsayılanları kullanılır (parametre tutarlılığı).
#     İterasyon TAVANI da uçta yeniden tanımlanmaz — varsayılanı çözücünün
#     DEFAULT_MAX_ITERS'idir; yalnız istemci AÇIKÇA bir bütçe verebilir
#     (max_iterations, bandı beyanlı). Bütçe verilmediğinde aynı ızgarayla
#     doğrudan çağrı, bu ucun ürettiği sonucun BİT-AYNISINI verir (bekçi:
#     tests/test_cfd_endpoint.py).
# ===========================================================================

#: Çözünürlük beyaz listesi: ad → (ni, nj) = (eksenel, radyal) hücre sayısı.
#: SERBEST SAYI ALINMAZ — uzun koşu riski masaüstü tek-worker sunucuda
#: gerçektir ve seviyelerin süresi ÖLÇÜLMÜŞTÜR.
#:
#: ÖLÇÜM — TAZELENDİ (M4 Max, 2026-08-16 akşamı, parti 26; gerçek
#: sample_nozzle_inner_contour konturu, P0=2 MPa, T0=3000 K, γ=1,2, R=350).
#: EN KÖTÜ hâl = bütçe tavanına kadar giden koşu; ölçüm yöntemi tol_res=0 ve
#: settle_tol=0 ile çözücünün erken çıkışı KAPATILARAK tam DEFAULT_MAX_ITERS
#: (20000) iterasyon koşturmaktır (numba yolu doğrudan, NumPy yolu
#: HRMA_CFD_DISABLE_NUMBA=1 ile ayrı süreçte; ikisinde de JIT ısınması
#: ölçümün dışında):
#:      seviye     ni×nj    numba     numpy    (yakınsayan koşu, numba)
#:      coarse     60×12    10,2 s    14,9 s    1,20 s / 2083 iterasyon
#:      standard  120×24    19,9 s    31,8 s    4,91 s / 4505 iterasyon
#: NÖBET DEĞİŞİMİ: bir önceki künye (parti 23) 9,4 / 15,0 ve 15,7 / 30,5
#: diyordu. Aradaki fark ölçüm gürültüsü DEĞİL: parti 25'te çözücüye
#: basınç-tabanlı şok sensörü ve sınırlayıcı tazelemesi, parti 26'da
#: karakteristik giriş sınır koşulu girdi — iterasyon başına iş arttı.
#: Bayat kalan sayı 'standard/numba' idi (15,7 → 19,9 = %27 eksik beyan).
#: Varsayılan 'coarse'tır: ~10-20 sn cüzdanını numba OLMADAN da tutan tek
#: seviye odur (numba isteğe bağlı bağımlılıktır — kernel_backend alanı
#: hangi yolun koştuğunu yanıtta beyan eder). 'standard' istemcinin bilinçli
#: seçimidir ve cüzdanı yalnız numba ile tutar; bu da beyan edilir.
CFD_RESOLUTION_LEVELS = {
    'coarse': (60, 12),
    'standard': (120, 24),
}
CFD_DEFAULT_RESOLUTION = 'coarse'

#: Seviye başına ÖLÇÜLEN en kötü duvar-saati [s] (yukarıdaki ölçüm turu).
#: Panel bu sayıyı kullanıcıya "bu koşu ne kadar sürebilir" diye gösterebilir;
#: uydurma bir ilerleme çubuğu ÜRETİLMEZ (sahte animasyon yasağı).
CFD_RESOLUTION_WORST_CASE_S = {
    'coarse': {'numba': 10.2, 'numpy': 14.9},
    'standard': {'numba': 19.9, 'numpy': 31.8},
}

#: Yanıttaki alan bloğunun hücre tavanı. ÖLÇÜLDÜ (2026-08-16, JSON bayt
#: sayımı): ham 120×24 alan bloğu (2880 hücre × 4 dizi) 232 KB. Tavan 1200
#: hücreye çekilince 'standard' seviyesi eksenel yönde 120→50 kolona
#: inceltilir ve blok 98 KB'ye iner (toplam yanıt 128 KB); 'coarse'
#: (720 hücre, 58 KB) inceltmesiz geçer. İnceltme oranı ve SEÇİLEN İNDEKSLER
#: yanıtta BEYAN edilir — panel hangi hücreleri aldığını bilir.
CFD_FIELD_MAX_CELLS = 1200

#: Kalıntı geçmişi tavanı (nokta). ÖLÇÜLDÜ: yakınsamayan koşunun 20000
#: noktalık ham geçmişi 388 KB — tek başına yanıtın en büyük parçası olurdu
#: (inceltilmiş hâli 11 KB). 400 noktaya inceltilir (0. ve son iterasyon HER
#: ZAMAN içeride); inceltmenin gizleyebileceği iki sayı (son ve en küçük
#: kalıntı) ayrıca TAM değerle yayımlanır.
CFD_RESIDUAL_MAX_POINTS = 400

# ---------------------------------------------------------------------------
# GİRİŞ KOŞULLANDIRMA UYARISI — NÖBET DEĞİŞİMİ (2026-08-16, parti 26)
# ---------------------------------------------------------------------------
# EMEKLİYE AYRILAN SÖZLEŞME: ``CFD_INLET_MACH_ADVISORY = 0.15`` ve ona bağlı
# ``inlet_conditioning.threshold_mach`` / ``inlet_conditioning.advisory``
# alanları. Eşik, çözücünün ESKİ ses-altı giriş sınır koşulunun ölçülmüş
# kırılma noktasıydı (o BC iç hücreden statik basınç dışdeğerleyip Mach'ı
# izantropik p→M bağıntısından çözüyordu; eşleme M→0'da dikleşiyor ve iç
# basınç gürültüsünü giriş debisine büyüterek yansıtıyordu). Tarihçe olarak
# duran ESKİ ölçüm tablosu (2026-08-16 sabahı, aynı kontur ailesi):
#         CR    M_giriş   60×12         120×24
#         1,78   0,357    YAKINSADI     YAKINSADI
#         2,78   0,219    YAKINSADI     YAKINSADI
#         4,00   0,150    YAKINSADI     YAKINSAMADI
#         5,44   0,110    YAKINSAMADI   YAKINSAMADI
#         7,11   0,084    YAKINSAMADI   YAKINSAMADI
#        11,11   0,053    YAKINSAMADI   YAKINSAMADI
# NEDEN EMEKLİ: çözücünün giriş sınır koşulu KARAKTERİSTİK (Riemann
# değişmezi) biçime çevrildi (hrma/cfd/euler_core.py: INLET_BC_NAME =
# 'characteristic_reservoir'). Eşiğin ölçtüğü kusur ORTADAN KALKTI ve eşik
# yanlış ateşlemeye başladı: canlı hibrit motorun giriş Mach'ı 0,043 (eşiğin
# çok altında) olduğu hâlde koşu yakınsıyor. Mach'a bakan bir uyarı artık
# ÖLÇÜLEN hiçbir şeyi bildirmiyordu — yerine geçen sözleşme aşağıdadır.
#
# YENİ SÖZLEŞME: uyarı, giriş Mach'ının değil İTERASYON BÜTÇESİNİN uyarısıdır
# (``budget_advisory``). Kalan gerçek risk yakınsama HIZIDIR: bazı daralma
# oranlarında koşu bütçenin çoğunu yiyor, bir vaka tavana dayanıyor.

#: ÖLÇÜLEN YAKINSAMA TABLOSU (2026-08-16 akşamı, bu depo, M4 Max, numba).
#: Ölçüm UÇ ÜZERİNDEN yapıldı (ürün gerçeği): gerçek örnekleyici konturu
#: (throat 30 mm, exit 75 mm, konik 30°/15°; oda çapı 40→100 mm ile daralma
#: oranı taranarak), P0=2 MPa, T0=3000 K, γ=1,2, R=350, P_ortam=20 kPa,
#: çözücü varsayılanları (bütçe tavanı DEFAULT_MAX_ITERS=20000).
#: 'contraction_ratio' ucun KENDİ ölçtüğü değerdir (yeniden örneklenmiş
#: düğümlerden), o yüzden seviyeye göre binde mertebesinde ayrışır.
#: SINIRLAMA (dürüstlük): tablo TEK BİR KONTUR AİLESİDİR (konik, sabit boğaz
#: ve çıkış çapı, sabit yarı açılar) ve TEK BİR gaz durumudur. Başka bir
#: kontur ailesi (bell/parabolik), başka γ ya da başka boğaz/çıkış oranı bu
#: sayıları taşımaz; blok bunu beyan eder ve HÜKÜM VERMEZ.
CFD_MEASURED_CONVERGENCE = (
    {'resolution': 'coarse', 'contraction_ratio': 1.768,
     'inlet_mach_isentropic': 0.3593, 'iterations': 2169, 'converged': True,
     'residual_last': 9.90e-09, 'mass_balance_rel': 1.75e-10},
    {'resolution': 'coarse', 'contraction_ratio': 2.771,
     'inlet_mach_isentropic': 0.2193, 'iterations': 2083, 'converged': True,
     'residual_last': 2.70e-09, 'mass_balance_rel': 1.49e-10},
    {'resolution': 'coarse', 'contraction_ratio': 3.996,
     'inlet_mach_isentropic': 0.1500, 'iterations': 2949, 'converged': True,
     'residual_last': 3.78e-10, 'mass_balance_rel': 8.36e-11},
    {'resolution': 'coarse', 'contraction_ratio': 5.412,
     'inlet_mach_isentropic': 0.1101, 'iterations': 4131, 'converged': True,
     'residual_last': 5.41e-09, 'mass_balance_rel': 2.00e-10},
    {'resolution': 'coarse', 'contraction_ratio': 7.079,
     'inlet_mach_isentropic': 0.0840, 'iterations': 6149, 'converged': True,
     'residual_last': 9.52e-10, 'mass_balance_rel': 4.54e-11},
    {'resolution': 'coarse', 'contraction_ratio': 11.106,
     'inlet_mach_isentropic': 0.0534, 'iterations': 13563, 'converged': True,
     'residual_last': 2.19e-09, 'mass_balance_rel': 8.74e-10},
    {'resolution': 'standard', 'contraction_ratio': 1.776,
     'inlet_mach_isentropic': 0.3573, 'iterations': 5805, 'converged': True,
     'residual_last': 1.00e-08, 'mass_balance_rel': 1.31e-10},
    {'resolution': 'standard', 'contraction_ratio': 2.771,
     'inlet_mach_isentropic': 0.2193, 'iterations': 4505, 'converged': True,
     'residual_last': 9.99e-09, 'mass_balance_rel': 2.70e-11},
    {'resolution': 'standard', 'contraction_ratio': 3.996,
     'inlet_mach_isentropic': 0.1500, 'iterations': 6774, 'converged': True,
     'residual_last': 9.92e-09, 'mass_balance_rel': 2.60e-10},
    {'resolution': 'standard', 'contraction_ratio': 5.442,
     'inlet_mach_isentropic': 0.1095, 'iterations': 13654, 'converged': True,
     'residual_last': 8.73e-09, 'mass_balance_rel': 2.69e-10},
    {'resolution': 'standard', 'contraction_ratio': 7.095,
     'inlet_mach_isentropic': 0.0838, 'iterations': 20000, 'converged': False,
     'residual_last': 3.41e-06, 'mass_balance_rel': 2.26e-07},
    {'resolution': 'standard', 'contraction_ratio': 11.106,
     'inlet_mach_isentropic': 0.0534, 'iterations': 9345, 'converged': True,
     'residual_last': 4.23e-10, 'mass_balance_rel': 4.83e-10},
)

#: RİSK KESRİ: bütçenin bu kadarını yiyen ÖLÇÜLEN satır "riskli" sayılır.
#: 0,5 uydurma değil, tablonun KENDİ komşuluk ölçümünden gelir: ardışık iki
#: ölçüm noktası arasında iterasyon sayısı 2,2 kata kadar sıçrıyor (coarse
#: 6149 → 13563 = 2,21×; standard 6774 → 13654 = 2,02×). Tablo ~1,3× CR
#: adımlarıyla örneklendiği için İKİ ölçüm noktası ARASINDAKİ bir kontur,
#: hızlı komşusunun iki katını isteyebilir; bütçesinin yarısını çoktan yiyen
#: bir satırın bu çarpana yeri yoktur.
CFD_BUDGET_RISK_FRACTION = 0.5

#: Kullanıcının verebileceği iterasyon bütçesinin ALT sınırı. Çözücünün CFL
#: rampası DEFAULT_RAMP_ITERS=400 iterasyon sürer ve oturma penceresi
#: DEFAULT_SETTLE_WINDOW=200'dür; bunların altında bir bütçe şemayı hedef
#: CFL'ine bile ulaştırmadan koşuyu keser. 500 = rampa + 100 pay; ÜST sınır
#: ayrıca tanımlanmaz, çözücünün kendi DEFAULT_MAX_ITERS'idir (uçta ikinci
#: bir tavan tanımı yazılmaz — parametre tutarlılığı).
CFD_MAX_ITERS_MIN = 500


def _cfd_iteration_budget_band():
    """(alt, üst) iterasyon bütçesi bandı — üst sınır ÇÖZÜCÜDEN gelir."""
    from hrma.cfd.steady import DEFAULT_MAX_ITERS
    return CFD_MAX_ITERS_MIN, int(DEFAULT_MAX_ITERS)


def _cfd_measured_rows(resolution):
    """Bir çözünürlük seviyesinin ölçüm satırları, CR'ye göre sıralı."""
    return sorted((dict(r) for r in CFD_MEASURED_CONVERGENCE
                   if r['resolution'] == resolution),
                  key=lambda r: r['contraction_ratio'])


def _cfd_budget_risk_bands(resolution):
    """ÖLÇÜM TABLOSUNDAN türetilen riskli daralma oranı bantları.

    Kural (uydurma sayı yok, tablo dışında sabit yok):
      1. Riskli satır = bütçe tavanına dayanmış (converged False) YA DA
         ``CFD_BUDGET_RISK_FRACTION`` × bütçe kadar iterasyon istemiş satır.
      2. Ardışık riskli satırlar tek bant olur.
      3. Bandın kenarı, riskli satır ile komşu SAĞLAM satırın GEOMETRİK orta
         noktasıdır (CR çarpımsal bir büyüklüktür ve tablo ~1,3× adımlarla
         örneklenmiştir; aritmetik orta büyük CR'lerde yanlı olurdu).
      4. Riskli satır tablonun ucundaysa o uçta bant AÇIKTIR (None) — orası
         ölçüm DIŞIDIR ve bant üyeliği EKSTRAPOLASYONDUR; blok bunu
         ``extrapolated`` bayrağıyla beyan eder.
    Riskli satır yoksa boş liste döner (uyarı bu kuraldan ATEŞLEYEMEZ).
    """
    _alt, butce = _cfd_iteration_budget_band()
    satirlar = _cfd_measured_rows(resolution)
    riskli = [(not r['converged'])
              or r['iterations'] >= CFD_BUDGET_RISK_FRACTION * butce
              for r in satirlar]
    bantlar = []
    i = 0
    while i < len(satirlar):
        if not riskli[i]:
            i += 1
            continue
        j = i
        while j + 1 < len(satirlar) and riskli[j + 1]:
            j += 1
        cr_alt = (math.sqrt(satirlar[i - 1]['contraction_ratio']
                            * satirlar[i]['contraction_ratio'])
                  if i > 0 else None)
        cr_ust = (math.sqrt(satirlar[j]['contraction_ratio']
                            * satirlar[j + 1]['contraction_ratio'])
                  if j + 1 < len(satirlar) else None)
        bantlar.append({
            'resolution': resolution,
            'cr_min': cr_alt,
            'cr_max': cr_ust,
            'measured_rows': [satirlar[k]['contraction_ratio']
                              for k in range(i, j + 1)],
            'open_below': cr_alt is None,
            'open_above': cr_ust is None,
        })
        i = j + 1
    return bantlar


def _cfd_nearest_measured(resolution, contraction_ratio):
    """CR'ye GEOMETRİK olarak en yakın ölçüm satırı (yoksa None)."""
    satirlar = _cfd_measured_rows(resolution)
    if not satirlar or not (contraction_ratio and contraction_ratio > 0.0):
        return None
    return min(satirlar,
               key=lambda r: abs(math.log(r['contraction_ratio']
                                          / contraction_ratio)))


def _cfd_budget_advisory(resolution, contraction_ratio, max_iters):
    """Koşu öncesi İTERASYON BÜTÇESİ uyarısı — HÜKÜM DEĞİL, koşuyu engellemez.

    Döner: (uyari, gerekceler, bantlar, en_yakin_satir). İki bağımsız gerekçe
    ölçülür ve ADIYLA yayımlanır:
      * ``measured_slow_band``          — CR, ölçülen yavaş bantlardan birinde,
      * ``budget_below_measured_need``  — istenen bütçe, en yakın ölçüm
        satırının GERÇEKTEN harcadığı iterasyondan az.
    Gerekçe listesi boşsa uyarı da yoktur (bool, listenin kendisinden türetilir
    — iki alan asla ayrışamaz).
    """
    bantlar = _cfd_budget_risk_bands(resolution)
    en_yakin = _cfd_nearest_measured(resolution, contraction_ratio)
    gerekceler = []
    if contraction_ratio and contraction_ratio > 0.0:
        for b in bantlar:
            alt_ok = b['cr_min'] is None or contraction_ratio >= b['cr_min']
            ust_ok = b['cr_max'] is None or contraction_ratio <= b['cr_max']
            if alt_ok and ust_ok:
                gerekceler.append('measured_slow_band')
                break
    if en_yakin is not None and max_iters < en_yakin['iterations']:
        gerekceler.append('budget_below_measured_need')
    return bool(gerekceler), gerekceler, bantlar, en_yakin


class _CfdInputError(Exception):
    """İstemci girdisi sözleşmeyi taşımıyor (4xx). Alan adı + gerekçe taşır."""

    def __init__(self, field, message, missing=False):
        super().__init__(message)
        self.field = field
        self.message = message
        self.missing = missing


def _cfd_separation_fn():
    """Ayrılma köprüsünü TAKMA ADLA getirir (sessiz gölgeleme kapısı).

    ``hrma.flow`` da ``assess_separation`` adını dışa verir (yarı-1B sürüm,
    imzası ``(P0, Pa, gamma, ...)``); aynı modülde ikisi birden bulunursa
    biri diğerini SESSİZCE gölgeler. Bu ucun tükettiği sözleşme 2B Euler
    çözümünün duvar basıncıdır, o yüzden köprü tek bir yerde ve takma adla
    alınır. Kimlik denetimi bekçilidir (tests/test_cfd_endpoint.py: dönen
    fonksiyon ``hrma.cfd.separation.assess_separation`` olmalı ve
    ``hrma.flow.assess_separation`` OLMAMALIDIR).
    """
    from hrma.cfd.separation import assess_separation as assess_cfd_separation
    return assess_cfd_separation


def _cfd_pick_contour(data):
    """İstek gövdesinden yayımlanmış lüle konturunu çıkarır → (Nx2 [m], alan).

    Kabul edilen biçimler (motorların YAYIMLADIĞI blokla aynı sözleşme —
    ``results['nozzle_contour'] = {'points': [[z_m, r_m], ...], '_basis': …}``,
    üç motorda da ``nozzle_design.sample_nozzle_inner_contour`` örnekleyicisi;
    parti 25'te wall-profile ekseni de aynı yayına bağlandı):
      * ``nozzle_contour``  — blok (dict, ``points`` taşır) ya da doğrudan
        ``[[z_m, r_m], ...]`` listesi,
      * ``motor_results`` / ``motor`` / ``results`` — motor sonuç sözlüğü;
        içindeki ``nozzle_contour`` bloğu kullanılır (hibrit sayfadaki
        ``{'motor': {...}}`` sarmalı da taranır).
    Kontur UYDURULMAZ: hiçbir alanda bulunamazsa 422 (eksik alan).
    """
    adaylar = [('nozzle_contour', data.get('nozzle_contour'))]
    for alan in ('motor_results', 'motor', 'results'):
        kap = data.get(alan)
        if isinstance(kap, dict):
            adaylar.append((alan + '.nozzle_contour',
                            kap.get('nozzle_contour')))
            ic = kap.get('motor')
            if isinstance(ic, dict):
                adaylar.append((alan + '.motor.nozzle_contour',
                                ic.get('nozzle_contour')))

    ham = None
    kaynak = None
    for alan, deger in adaylar:
        if isinstance(deger, dict):
            noktalar = deger.get('points')
            if noktalar is not None:
                ham, kaynak = noktalar, alan + '.points'
                break
        elif isinstance(deger, (list, tuple)):
            ham, kaynak = deger, alan
            break

    if ham is None:
        raise _CfdInputError(
            'nozzle_contour',
            'The published nozzle contour is required: the CFD grid is built '
            'from it. Send the solver block '
            "results['nozzle_contour'] = {'points': [[z_m, r_m], ...]} "
            '(metres, first point = convergent inlet) either directly as '
            "'nozzle_contour' or inside 'motor_results'/'motor'/'results'. "
            'No contour is invented.',
            missing=True)

    try:
        pts = np.asarray(ham, dtype=float)
    except (TypeError, ValueError):
        raise _CfdInputError(
            kaynak, 'The nozzle contour must be a list of [z_m, r_m] number '
                    'pairs; it could not be read as numbers.')
    if pts.ndim != 2 or pts.shape[1] != 2 or pts.shape[0] < 3:
        raise _CfdInputError(
            kaynak, 'The nozzle contour must be [[z_m, r_m], ...] with at '
                    f'least 3 points; got shape {tuple(pts.shape)}.')
    if not np.all(np.isfinite(pts)):
        raise _CfdInputError(
            kaynak, 'The nozzle contour contains non-finite values (NaN/Inf).')
    if np.any(np.diff(pts[:, 0]) < 0.0):
        raise _CfdInputError(
            kaynak, 'The nozzle contour z coordinate must be non-decreasing '
                    '(sample_nozzle_inner_contour contract: the first point '
                    'is the convergent inlet and z increases toward the '
                    'exit).')
    if pts[-1, 0] <= pts[0, 0]:
        raise _CfdInputError(
            kaynak, 'The nozzle contour has zero axial extent '
                    f'(z stays at {float(pts[0, 0]):g} m).')
    if np.any(pts[:, 1] <= 0.0):
        raise _CfdInputError(
            kaynak, 'The nozzle wall radius must be positive at every '
                    'station (r <= 0 would put the wall on or inside the '
                    'axis).')
    return pts, kaynak


def _cfd_build_grid(points_m, ni, nj):
    """Yayımlanmış kontur (Nx2, METRE) → eksenel simetrik ızgara.

    ÖRNEKLEME SÖZLEŞMESİ: duvar polilinesi ni+1 DÜZGÜN z istasyonuna
    doğrusal ara değerle yeniden örneklenir (``grid_axisym.build_nozzle_grid``
    ile aynı kural; oradaki sürüm milimetre alır, yayımlanan blok metre
    taşıdığı için birim çevirisi yapılmaz ve kontur METRE olarak geçirilir).
    Döner: ``(AxisymGrid, z_nodes [m], r_nodes [m])`` — düğüm dizileri
    çağırana da lazımdır (boğaz istasyonu ve daralma oranı onlardan ölçülür).
    Bekçi bu yardımcıyı doğrudan çağırır — uç ile test aynı ızgarayı kurar,
    ayrılma bloğu bit-tutarlılığı ancak böyle ölçülebilir.
    """
    from hrma.cfd.grid_axisym import build_grid_from_wall
    pts = np.asarray(points_m, dtype=float)
    z_nodes = np.linspace(float(pts[0, 0]), float(pts[-1, 0]), int(ni) + 1)
    r_nodes = np.interp(z_nodes, pts[:, 0], pts[:, 1])
    return build_grid_from_wall(z_nodes, r_nodes, int(nj)), z_nodes, r_nodes


def _cfd_required_float(data, key, aciklama, minimum=None, maximum=None,
                        strict=True):
    """Zorunlu sayısal alan: yoksa 422 (missing), bozuk/bant dışıysa 400."""
    ham = data.get(key)
    if ham is None or ham == '':
        raise _CfdInputError(key, aciklama, missing=True)
    try:
        deger = float(ham)
    except (TypeError, ValueError):
        raise _CfdInputError(key, f"'{key}' must be a number (got {ham!r}).")
    if not math.isfinite(deger):
        raise _CfdInputError(key, f"'{key}' must be finite (got {deger!r}).")
    if minimum is not None:
        if (deger <= minimum) if strict else (deger < minimum):
            raise _CfdInputError(
                key, f"'{key}' must be "
                     f"{'>' if strict else '>='} {minimum:g}; got {deger:g}.")
    if maximum is not None:
        if (deger >= maximum) if strict else (deger > maximum):
            raise _CfdInputError(
                key, f"'{key}' must be "
                     f"{'<' if strict else '<='} {maximum:g}; got {deger:g}.")
    return deger


def _cfd_decimate_indices(n, hedef):
    """[0, n-1] aralığından en fazla `hedef` indeks — uçlar HER ZAMAN içeride."""
    n = int(n)
    hedef = int(hedef)
    if hedef >= n or hedef <= 1:
        return np.arange(n, dtype=int) if hedef >= n else np.array([0, n - 1])
    return np.unique(np.round(np.linspace(0, n - 1, hedef)).astype(int))


@app.route('/api/cfd/nozzle', methods=['POST'])
def api_cfd_nozzle():
    """Lüle iç akışının 2B eksenel simetrik Euler çözümü + ayrılma hükmü.

    Girdi (JSON gövde):
      * ``nozzle_contour``  — ZORUNLU. Motorun yayımladığı blok
        ``{'points': [[z_m, r_m], ...]}`` (METRE, ilk nokta konverjan girişi)
        ya da doğrudan aynı liste. ``motor_results`` / ``motor`` / ``results``
        içinden de okunur.
      * ``P0_Pa``, ``T0_K``  — ZORUNLU. Kamara durma basıncı [Pa] ve
        sıcaklığı [K] (motor çözümünden).
      * ``gamma``, ``R_J_per_kgK`` — ZORUNLU. Kalorik mükemmel gaz sabitleri
        (motor çözümünden; çözücü varsayılan üretmez, uç da üretmez).
      * ``P_ambient_Pa`` — ZORUNLU. Ayrılma ölçütü p_w < k·P_ortam ortam
        basıncına göre tanımlıdır; deniz seviyesi/vakum varsaymak UYDURMA
        sayılır (bkz. hrma/cfd/separation.py).
      * ``Pb_Pa`` — İSTEĞE BAĞLI geri basınç. Verilmezse çıkışta tam
        dışdeğerleme (ses-üstü çıkış) uygulanır ve bu yanıtta beyan edilir.
      * ``resolution`` — İSTEĞE BAĞLI, ``CFD_RESOLUTION_LEVELS`` beyaz
        listesinden ('coarse' | 'standard'); serbest sayı ALINMAZ.
      * ``separation_factor`` — İSTEĞE BAĞLI Summerfield k'sı; verilmezse
        köprünün ithal ettiği varsayılan (bant denetimi köprüdedir).
      * ``max_iterations`` — İSTEĞE BAĞLI iterasyon BÜTÇESİ (tam sayı).
        Bant ``[CFD_MAX_ITERS_MIN, DEFAULT_MAX_ITERS]``; verilmezse
        çözücünün kendi ``DEFAULT_MAX_ITERS``i (uçta ikinci tavan tanımı
        yok). "Yakınsamayan koşu dürüstçe raporlanır" sözleşmesi bu AÇIK
        bütçeye dayanır: koşu tavana dayanırsa ``converged=False`` döner ve
        etkin bütçe ``max_iterations`` + ``max_iterations_basis`` ile
        (kullanıcı mı verdi, varsayılan mı) beyan edilir.

    Yanıt 200: ``{'status': 'success', 'cfd': {...}}`` — yakınsama beyanı,
    kalıntı geçmişi (inceltilmiş, beyanlı), korunum bütçesi, boğaz, kesit
    ortalamaları, duvar basıncı, çizim için inceltilmiş alan bloğu, ayrılma
    bloğu (köprüden AYNEN) ve not_modelled/assumptions beyanları.
    Yakınsamayan ya da ıraksayan koşu da 200 döner: sonuç gizlenmez,
    ``converged=False`` ve çözücünün kendi gerekçesiyle gelir.

    Hata: eksik zorunlu alan 422 + ``missing_fields``; bozuk/bant dışı değer
    400 + ``field``; çözücü çökmesi 500 (boş ama 'başarılı' yanıt üretilmez).
    """
    import time as _time

    t_baslangic = _time.perf_counter()
    try:
        data = request.get_json(silent=True) or {}
    except Exception:
        data = {}
    if not isinstance(data, dict):
        data = {}

    # --- GİRDİ SÖZLEŞMESİ ------------------------------------------------
    # Eksik alanların HEPSİ birden bildirilir (panel tek turda düzeltsin);
    # bozuk/bant dışı değer ilk hatada 400 ile döner.
    eksikler = []
    gerekce = {}
    try:
        try:
            kontur_pts, kontur_alani = _cfd_pick_contour(data)
        except _CfdInputError as e:
            if not e.missing:
                raise
            eksikler.append(e.field)
            gerekce[e.field] = e.message
            kontur_pts, kontur_alani = None, None

        # (alan, gerekçe, alt, üst, katı) — bantlar çözücünün KENDİ
        # bantlarıdır (steady.solve_steady_axisym: P0/T0/R > 0, 1 < γ < 2);
        # burada yalnız ham ValueError'ın 500'e sızması engellenir. P_ortam'ın
        # alt sınırı gevşektir: 0 (vakum) GEÇERLİ girdidir ve köprü orada
        # 'tanımsız' beyanıyla döner — reddedilmez.
        sayisal = {}
        for key, aciklama, alt, ust, kati in (
            ('P0_Pa',
             'Chamber stagnation pressure [Pa] is required: the solver has no '
             'default reservoir state (it comes from the engine solution).',
             0.0, None, True),
            ('T0_K',
             'Chamber stagnation temperature [K] is required: the solver has '
             'no default reservoir state (it comes from the engine solution).',
             0.0, None, True),
            ('gamma',
             'The ratio of specific heats is required: the solver is a '
             'calorically perfect gas model and takes gamma from the engine '
             'solution; no default is invented.',
             1.0, 2.0, True),
            ('R_J_per_kgK',
             'The specific gas constant [J/(kg K)] is required: it comes from '
             'the engine solution (R = R_universal / M); no default is '
             'invented.',
             0.0, None, True),
            ('P_ambient_Pa',
             'Ambient pressure [Pa] is required: the Summerfield separation '
             'criterion p_w < k*P_ambient is defined against it, and assuming '
             'sea level or vacuum would be a fabricated input '
             '(hrma/cfd/separation.py refuses a missing ambient pressure).',
             0.0, None, False),
        ):
            try:
                sayisal[key] = _cfd_required_float(
                    data, key, aciklama, minimum=alt, maximum=ust, strict=kati)
            except _CfdInputError as e:
                if not e.missing:
                    raise
                eksikler.append(e.field)
                gerekce[e.field] = e.message

        if eksikler:
            return jsonify({
                'status': 'error',
                'error': 'incomplete_cfd_nozzle_input',
                'message': ('The nozzle CFD run cannot start without these '
                            'inputs; none of them has a default.'),
                'missing_fields': eksikler,
                'field_reasons': gerekce,
            }), 422

        # İsteğe bağlı geri basınç: yoksa None (ses-üstü dışdeğerleme).
        Pb = None
        if data.get('Pb_Pa') not in (None, ''):
            Pb = _cfd_required_float(
                data, 'Pb_Pa', '', minimum=0.0, strict=True)

        # İsteğe bağlı Summerfield k'sı. Bant SABİTLERİ köprüden İTHAL edilir
        # (ikinci tanım yazılmaz); denetim burada YAPILIR çünkü bant dışı bir
        # k istemci hatasıdır ve 10 saniyelik koşu harcanmadan 400 dönmelidir
        # (köprüye bırakılsaydı hata koşudan SONRA çıkardı).
        from hrma.cfd.separation import (
            SEPARATION_FACTOR_MAX,
            SEPARATION_FACTOR_MIN,
        )
        sep_faktor = None
        if data.get('separation_factor') not in (None, ''):
            sep_faktor = _cfd_required_float(
                data, 'separation_factor', '',
                minimum=SEPARATION_FACTOR_MIN, maximum=SEPARATION_FACTOR_MAX,
                strict=False)

        # İTERASYON BÜTÇESİ (isteğe bağlı, AÇIK sözleşme). Varsayılan
        # çözücünün kendi tavanıdır — uçta ikinci bir tanım yazılmaz, yalnız
        # kullanıcının verebileceği bant burada denetlenir (uzun koşu cüzdanı
        # ucun sorumluluğudur). Tam sayı ISTENİR: JSON tarafı 600.0 gibi
        # sayısal bir tam değeri de kabul eder ama 600.5 reddedilir —
        # sessizce yuvarlamak, kullanıcının yazdığından BAŞKA bir bütçeyle
        # koşmak demek olurdu.
        iter_alt, iter_ust = _cfd_iteration_budget_band()
        etkin_tavan = iter_ust
        tavan_kaynagi = 'default'
        if data.get('max_iterations') not in (None, ''):
            ham_tavan = data.get('max_iterations')
            if isinstance(ham_tavan, bool):
                raise _CfdInputError(
                    'max_iterations',
                    "'max_iterations' must be a whole number of iterations "
                    f'in [{iter_alt}, {iter_ust}]; got {ham_tavan!r}.')
            try:
                sayi_tavan = float(ham_tavan)
            except (TypeError, ValueError):
                raise _CfdInputError(
                    'max_iterations',
                    "'max_iterations' must be a number (got "
                    f'{ham_tavan!r}).')
            if not math.isfinite(sayi_tavan) or sayi_tavan != int(sayi_tavan):
                raise _CfdInputError(
                    'max_iterations',
                    "'max_iterations' must be a WHOLE number of iterations "
                    '(the solver counts iterations, it cannot run half of '
                    f'one); got {ham_tavan!r}.')
            etkin_tavan = int(sayi_tavan)
            if not (iter_alt <= etkin_tavan <= iter_ust):
                raise _CfdInputError(
                    'max_iterations',
                    f"'max_iterations' must be in [{iter_alt}, {iter_ust}]: "
                    f'below {iter_alt} the run is cut before the solver CFL '
                    f'ramp (DEFAULT_RAMP_ITERS=400) completes — and between '
                    f'{iter_alt} and 599 the ramp finishes but a FULL '
                    f'settling window (DEFAULT_SETTLE_WINDOW=200) still may '
                    f'not, which the convergence basis will state — so a '
                    f'"did not converge" answer at tiny budgets says nothing '
                    f'about the flow; above {iter_ust} the run leaves the '
                    f'measured runtime wallet published in '
                    f'grid.levels.worst_case_s. Got {etkin_tavan}.')
            tavan_kaynagi = 'caller'

        # Çözünürlük: beyaz liste, serbest sayı yok.
        seviye = data.get('resolution')
        if seviye in (None, ''):
            seviye = CFD_DEFAULT_RESOLUTION
        seviye = str(seviye)
        if seviye not in CFD_RESOLUTION_LEVELS:
            raise _CfdInputError(
                'resolution',
                f"'resolution' must be one of "
                f"{sorted(CFD_RESOLUTION_LEVELS)} (free grid sizes are not "
                f"accepted: every level's cell count and measured worst-case "
                f"runtime is published in the response); got {seviye!r}.")
        ni, nj = CFD_RESOLUTION_LEVELS[seviye]

        # Izgara: yayımlanan kontur ni+1 düzgün istasyona yeniden örneklenir.
        try:
            grid, z_nodes, r_nodes = _cfd_build_grid(kontur_pts, ni, nj)
        except ValueError as e:
            raise _CfdInputError(
                kontur_alani,
                f'The CFD grid could not be built from the published '
                f'contour: {e}')

        # Boğaz İÇERİDE olmalı: ayrılma ölçütü ıraksak bölgede aranır ve
        # boğaz son istasyondaysa aranacak istasyon KALMAZ (köprü orada
        # ValueError yükseltir). 10 saniyelik koşuyu harcamadan reddedilir.
        i_bogaz = int(np.argmin(r_nodes))
        if not (0 < i_bogaz < len(r_nodes) - 1):
            raise _CfdInputError(
                kontur_alani,
                'The resampled wall has no interior throat: the minimum '
                f'radius sits at station {i_bogaz} of {len(r_nodes) - 1} '
                '(the contour must converge to a throat and then diverge, '
                'otherwise the separation criterion has no divergent section '
                'to search).')
    except _CfdInputError as e:
        return jsonify({
            'status': 'error',
            'error': 'invalid_cfd_nozzle_input',
            'field': e.field,
            'message': e.message,
        }), 400

    # --- KOŞU ------------------------------------------------------------
    # Sürücü ayarları (tolerans, CFL, rampa) VERİLMEZ: çözücünün kendi
    # varsayılanları kullanılır, uçta ikinci bir tanım yazılmaz. TEK istisna
    # iterasyon BÜTÇESİDİR ve o da uçta yeniden TANIMLANMAZ: varsayılanı
    # çözücünün DEFAULT_MAX_ITERS'idir, yalnız kullanıcı açıkça bir bütçe
    # verirse (max_iterations, bant denetimi yukarıda) onun sayısı geçirilir.
    # Böylece bütçe verilmediğinde aynı ızgarayla doğrudan çağrı bu sonucun
    # BİT-AYNISINI üretmeye devam eder (bekçi: tests/test_cfd_endpoint.py::
    # test_cozum_uc_disinda_bit_aynisiyla_uretilebiliyor).
    from hrma.cfd.steady import solve_steady_axisym

    try:
        sonuc = solve_steady_axisym(
            grid,
            P0=sayisal['P0_Pa'], T0=sayisal['T0_K'],
            gamma=sayisal['gamma'], R=sayisal['R_J_per_kgK'], Pb=Pb,
            max_iters=etkin_tavan)
    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'error': 'The axisymmetric Euler solver failed.',
            'detail': str(e),
        }), 500

    arka_uc = sonuc['kernel_backend']

    # --- AYRILMA KÖPRÜSÜ -------------------------------------------------
    # Blok köprüden AYNEN taşınır (tek kaynak). Köprü kendi sözleşmesini
    # ihlal eden bir sonuca (ıraksamış alan, bozuk eksen) hüküm VERMEZ ve
    # ValueError yükseltir; bu red 500'e çevrilmez, BEYAN edilerek geçilir.
    try:
        ayrilma = _cfd_separation_fn()(
            sonuc, sayisal['P_ambient_Pa'], separation_factor=sep_faktor)
    except ValueError as e:
        ayrilma = {
            'applicable': False,
            'bridge_refused': True,
            'not_applicable_reason': str(e),
            '_basis': (
                'hrma.cfd.separation.assess_separation refused to issue a '
                'judgement for this solution and the refusal is reported as '
                'is; no separation verdict is invented in its place. The flow '
                'field above is still returned with its own convergence '
                'declaration.'),
        }
    except Exception as e:
        # ValueError köprünün BEYAN EDİLMİŞ reddidir (yukarıda). Başka bir
        # istisna sözleşme ihlalidir (ör. yanlış fonksiyon ithal edilmiş,
        # imza uyuşmuyor) ve YUTULMAZ — yutulsaydı gölgeleme kusuru "ayrılma
        # hükmü verilemedi" diye sessizce normalleşirdi. 500 döner, ama
        # makine-okur JSON olarak (panel HTML hata sayfası ayrıştırmasın).
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'error': 'The separation bridge raised an unexpected exception.',
            'detail': f'{type(e).__name__}: {e}',
        }), 500

    # --- İNCELTME (beyanlı) ----------------------------------------------
    res_hist = np.asarray(sonuc['residual_history'], dtype=float)
    res_idx = _cfd_decimate_indices(res_hist.size, CFD_RESIDUAL_MAX_POINTS)
    residual = {
        'iteration': res_idx.tolist(),
        'value': res_hist[res_idx].tolist(),
        'n_total': int(res_hist.size),
        'n_returned': int(res_idx.size),
        'decimated': bool(res_idx.size < res_hist.size),
        'last': float(res_hist[-1]) if res_hist.size else None,
        'min': float(np.min(res_hist)) if res_hist.size else None,
        '_basis': (
            f'density L2 residual history, thinned to at most '
            f'{CFD_RESIDUAL_MAX_POINTS} points by uniform index selection '
            f'with the first and last iteration always kept (measured: the '
            f'raw 20000 point history of a non-converging run is 388 KB of '
            f'JSON on its own, the thinned block is 11 KB). The two '
            f'numbers thinning could hide are published exactly: "last" and '
            f'"min". Scale: rho0*a0/L, see convergence_basis.'),
    }

    mach2d = np.asarray(sonuc['fields']['mach'], dtype=float)
    p2d = np.asarray(sonuc['fields']['pressure_Pa'], dtype=float)
    z2d = np.asarray(sonuc['z_centers_m'], dtype=float)
    r2d = np.asarray(sonuc['r_centers_m'], dtype=float)
    # İnceltme YALNIZ eksenel yönde: radyal yön beyaz listede zaten dar
    # (nj <= 24) ve radyal kırpma duvara komşu hücre sırasını seyrekleştirip
    # kontur haritasının duvar kenarını bozar. Eksenel hedef, tavanın radyal
    # sayıya bölümüdür; uçlar (giriş ve çıkış kolonu) her zaman içeride.
    idx_j = np.arange(nj, dtype=int)
    idx_i = _cfd_decimate_indices(ni, max(2, CFD_FIELD_MAX_CELLS // nj))
    kes = np.ix_(idx_i, idx_j)
    field = {
        'z_m': z2d[kes].tolist(),
        'r_m': r2d[kes].tolist(),
        'mach': mach2d[kes].tolist(),
        'pressure_Pa': p2d[kes].tolist(),
        'shape': [int(idx_i.size), int(idx_j.size)],
        'grid_shape': [int(ni), int(nj)],
        'axial_indices': idx_i.tolist(),
        'radial_indices': idx_j.tolist(),
        'n_cells_total': int(ni * nj),
        'n_cells_returned': int(idx_i.size * idx_j.size),
        'decimated': bool(idx_i.size * idx_j.size < ni * nj),
        '_basis': (
            f'cell-centred field block for contour plotting: every entry is a '
            f'[axial][radial] nested list, index 0 on the radial axis is the '
            f'cell next to the symmetry axis and the last index is the cell '
            f'next to the wall. Values are the solver cells as returned (no '
            f'interpolation, no smoothing). If the cell count exceeds '
            f'{CFD_FIELD_MAX_CELLS} the block is thinned by uniform index '
            f'selection with both ends kept, and the kept indices are listed '
            f'here so the client knows exactly which cells it received. '
            f'Thinning is axial only: the radial direction is never sparsened '
            f'so the wall-adjacent cell row stays intact. Measured: the raw '
            f'120x24 block is 232 KB of JSON, the thinned 50x24 block 98 KB; '
            f'the 60x12 "coarse" block (58 KB) needs no thinning.'),
    }

    # --- GİRİŞ KOŞULLANDIRMA + İTERASYON BÜTÇESİ (ölçülmüş, hüküm DEĞİL) --
    # Mach eşikli eski uyarı EMEKLİ (nöbet değişimi künyesi
    # CFD_MEASURED_CONVERGENCE'ın üstünde). Giriş Mach'ı BİLGİ olarak durur —
    # okuyucunun konturu tanıması için — ama hiçbir uyarıyı TETİKLEMEZ.
    from hrma.flow.quasi1d import mach_from_area_ratio
    daralma = max(float((r_nodes[0] / r_nodes[i_bogaz]) ** 2), 1.0)
    try:
        mach_giris = float(mach_from_area_ratio(daralma, sayisal['gamma'],
                                                supersonic=False))
    except Exception:
        mach_giris = None
    butce_uyari, butce_gerekceleri, risk_bantlari, en_yakin_olcum = (
        _cfd_budget_advisory(seviye, daralma, etkin_tavan))
    giris_uyari = {
        'contraction_ratio': daralma,
        'inlet_mach_isentropic': mach_giris,
        'inlet_bc': sonuc['inlet_bc'],
        'inlet_bc_basis': sonuc['inlet_bc_basis'],
        'resolution': seviye,
        'max_iterations': int(etkin_tavan),
        'budget_advisory': bool(butce_uyari),
        'budget_advisory_reasons': list(butce_gerekceleri),
        'budget_advisory_bands': risk_bantlari,
        'budget_risk_fraction': CFD_BUDGET_RISK_FRACTION,
        'nearest_measured': en_yakin_olcum,
        'measured_expectations': [dict(r) for r in CFD_MEASURED_CONVERGENCE],
        '_basis': (
            'Pre-run ITERATION BUDGET advisory. RETIRED (2026-08-16): this '
            'block used to carry an inlet Mach threshold (0.15) measured '
            'against the solver\'s OLD subsonic inlet boundary condition. '
            'That condition extrapolated the interior static pressure and '
            'inverted the isentropic p->M relation, whose slope stiffens as '
            'M->0; high contraction ratios therefore piled ~99% of the '
            'residual into the first columns and did not settle. The solver '
            'inlet is now CHARACTERISTIC (Riemann invariant, see '
            'inlet_bc/inlet_bc_basis), the Mach threshold no longer measures '
            'anything, and it was firing on runs that converge (measured: a '
            'live engine at inlet Mach 0.043 converges). The inlet Mach '
            'number is kept as INFORMATION only (isentropic subsonic '
            'solution of the resampled inlet-to-throat area ratio, '
            'hrma.flow.quasi1d.mach_from_area_ratio) and triggers nothing. '
            'WHAT IS MEASURED NOW: convergence SPEED. Sweep of this endpoint '
            '(2026-08-16, M4 Max, numba, real sampled contour with a 30 mm '
            'throat and a 75 mm exit, conical 30/15 deg, chamber diameter '
            '40->100 mm, P0=2 MPa, T0=3000 K, gamma=1.2, R=350, default '
            'budget 20000) is published verbatim in measured_expectations: '
            'every case converged except standard/CR 7.095, which ran to the '
            'ceiling. Even that run is USABLE, not garbage: its final '
            'residual is 3.4e-06 and its mass imbalance 2.3e-07 (the old '
            'inlet BC left residual 0.33-0.71 and a percent-level mass '
            'imbalance in the same case). The advisory fires when the '
            'contraction ratio falls in a band derived from that table '
            '(budget_advisory_bands: a measured row is "slow" when it did '
            'not converge or spent at least budget_risk_fraction of the '
            'budget; band edges are the geometric midpoints to the nearest '
            'healthy measured rows, and an edge is null where the table '
            'itself ends - membership beyond it is extrapolation), or when '
            'the requested budget is below what the nearest measured row '
            'actually spent (budget_advisory_reasons names which rule '
            'fired). LIMITS, honestly: the table is ONE conical contour '
            'family at ONE gas state and TWO resolutions; a bell contour, '
            'another gamma or another throat-to-exit ratio does not inherit '
            'these numbers. THIS IS NOT A VERDICT: the verdict is the '
            'solver\'s own "converged" flag next to it, this advisory only '
            'states the pre-run expectation with the measurement behind it, '
            'and it never blocks the run.'),
    }

    cfd = {
        'converged': bool(sonuc['converged']),
        'convergence_basis': sonuc['convergence_basis'],
        'iterations': int(sonuc['iterations']),
        'max_iterations': int(etkin_tavan),
        'max_iterations_basis': (
            f'Effective iteration budget {etkin_tavan} '
            + ('(caller supplied "max_iterations")'
               if tavan_kaynagi == 'caller' else
               '(the caller did not supply "max_iterations", so the '
               'solver\'s own DEFAULT_MAX_ITERS is used - the endpoint does '
               'not define a second ceiling)')
            + f'. Accepted band [{iter_alt}, {iter_ust}]: below the lower '
              f'bound the run stops before the CFL ramp and the settling '
              f'window can complete, above the upper bound it leaves the '
              f'measured runtime wallet (grid.levels.worst_case_s). A run '
              f'that reaches this budget returns converged=False with the '
              f'solver\'s own reason; nothing is hidden and no verdict is '
              f'invented.'),
        'max_iterations_source': tavan_kaynagi,
        'limiter_frozen_at_iter': sonuc['limiter_frozen_at_iter'],
        'limiter_freeze_count': int(sonuc['limiter_freeze_count']),
        'residual_history': residual,
        'shock_sensor_columns': int(sonuc['shock_sensor_columns']),
        'shock_sensor_basis': sonuc['shock_sensor_basis'],
        'budget': {
            'mass_flow_in_kg_s': sonuc['mass_flow_in_kg_s'],
            'mass_flow_out_kg_s': sonuc['mass_flow_out_kg_s'],
            'mass_balance_rel': sonuc['mass_balance_rel'],
            'energy_flux_in_W': sonuc['energy_flux_in_W'],
            'energy_flux_out_W': sonuc['energy_flux_out_W'],
            'energy_balance_rel': sonuc['energy_balance_rel'],
            'wall_mass_flux_kg_s': sonuc['wall_mass_flux_kg_s'],
            '_basis': sonuc['budget_basis'],
        },
        'throat': sonuc['throat'],
        'section_average': sonuc['section_average'],
        'wall_pressure': {
            'z_m': sonuc['wall_pressure_z_m'],
            'pressure_Pa': sonuc['wall_pressure_Pa'],
            '_basis': sonuc['wall_pressure_basis'],
        },
        'field': field,
        'separation': ayrilma,
        'inlet_conditioning': giris_uyari,
        'grid': {
            'resolution': seviye,
            'ni': int(ni),
            'nj': int(nj),
            'n_cells': int(ni * nj),
            'z_inlet_m': float(z_nodes[0]),
            'z_exit_m': float(z_nodes[-1]),
            'r_inlet_m': float(r_nodes[0]),
            'r_throat_m': float(r_nodes[i_bogaz]),
            'r_exit_m': float(r_nodes[-1]),
            'levels': {ad: {'ni': a, 'nj': b, 'n_cells': a * b,
                            'worst_case_s': CFD_RESOLUTION_WORST_CASE_S[ad]}
                       for ad, (a, b) in CFD_RESOLUTION_LEVELS.items()},
            'default_resolution': CFD_DEFAULT_RESOLUTION,
            '_basis': (
                'H-type structured axisymmetric grid built from the published '
                'contour: the wall polyline is resampled onto ni+1 uniform '
                'axial stations by linear interpolation and the radial '
                'distribution is uniform from axis to wall (stage 1 of the '
                'CFD design document). Resolution levels are a whitelist, not '
                'a free number; "worst_case_s" is the MEASURED wall clock of '
                'a run that never converges (i.e. runs to the solver '
                'iteration ceiling) on an M4 Max, per kernel backend. '
                'TWO THROAT RADII, ON PURPOSE: r_throat_m here is the '
                'smallest resampled NODE radius and it is what this endpoint '
                'measures the contraction ratio and the pre-flight throat '
                'gate from, whereas the solver\'s own throat.radius_m is the '
                'CELL COLUMN wall radius (the mean of the two node radii '
                'bounding that column). They are close but not equal by '
                'construction; neither is rewritten to match the other.'),
        },
        'inputs': dict(sonuc['inputs'], **{
            'P_ambient_Pa': sayisal['P_ambient_Pa'],
            'separation_factor': sep_faktor,
            'max_iterations': int(etkin_tavan),
            'max_iterations_source': tavan_kaynagi,
            'contour_field': kontur_alani,
            'contour_points': int(kontur_pts.shape[0]),
            'resolution': seviye,
            'back_pressure_basis': (
                'Pb_Pa was supplied, so the outlet applies it wherever the '
                'face-normal Mach number is subsonic.' if Pb is not None else
                'Pb_Pa was NOT supplied, so the outlet is a pure zeroth order '
                'extrapolation. That is the correct treatment for a fully '
                'supersonic exit, but it means an internal shock driven by a '
                'back pressure CANNOT appear in this run: send Pb_Pa to look '
                'for one.'),
        }),
        'kernel_backend': arka_uc,
        'kernel_backend_basis': sonuc['kernel_backend_basis'],
        'not_modelled': sonuc['not_modelled'],
        'assumptions': sonuc['assumptions'],
        'solver_runtime_s': sonuc['runtime_s'],
        'runtime_s': _time.perf_counter() - t_baslangic,
        'runtime_basis': (
            f'solver_runtime_s is the steady driver itself, runtime_s the '
            f'whole request (validation, grid, solve, separation, thinning). '
            f'Measured worst case for this level and backend: '
            f'{CFD_RESOLUTION_WORST_CASE_S[seviye].get(arka_uc)} s.'),
    }
    return jsonify({'status': 'success', 'cfd': sanitize_json_values(cfd)})


@app.route('/api/chemical-database', methods=['GET'])
def get_chemical_database():
    """Get chemical species database information"""
    try:
        validation_results = chemical_db.validate_database()
        all_species = chemical_db.get_all_species_names()
        
        return jsonify({
            'status': 'success',
            'database_info': validation_results,
            'available_species': all_species[:50],  # Return first 50 species
            'total_species': len(all_species)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/chemical-species', methods=['POST'])
def get_chemical_species():
    """Get specific chemical species data"""
    try:
        data = request.json
        species_name = data.get('species_name')
        temperature = data.get('temperature', 2000)  # K
        
        species = chemical_db.get_species(species_name)
        if not species:
            return jsonify({'status': 'error', 'error': 'Species not found'}), 404
        
        # Calculate thermodynamic properties
        cp = chemical_db.calculate_cp(species_name, temperature)
        enthalpy = chemical_db.calculate_enthalpy(species_name, temperature)
        entropy = chemical_db.calculate_entropy(species_name, temperature)
        
        return jsonify({
            'status': 'success',
            'species_data': {
                'name': species.name,
                'formula': species.formula,
                'molecular_weight': species.molecular_weight,
                'phase': species.phase,
                'source': species.source,
                'thermodynamic_properties': {
                    'temperature': temperature,
                    'cp': cp,
                    'enthalpy': enthalpy,
                    'entropy': entropy
                }
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

# v2.5.0 G1 (2026-07-17, karar K5): olu /api/experimental-validation endpoint'i
# KALDIRILDI. Sinifta hic var olmamis metotlari cagiriyordu
# (validate_against_experiments / calculate_confidence_metrics -> AttributeError
# -> 500) ve frontend'de hicbir referansi yoktu. Halefi: G2 dalgasindaki
# korelasyon endpoint'leri (/api/validation/correlation-*) olacak.

@app.route('/api/cfd-analysis', methods=['POST'])
def perform_cfd_analysis():
    """Perform 2D CFD analysis"""
    # Dalga 0 bekçisi (2026-07-14): mevcut çözücü gerçek CFD değil —
    # kütle korunumu yok, 3 iterasyonda ıraksıyor (|u|→7.5e10 m/s),
    # NaN -> 500. Dalga 4A: quasi-1D halef uç noktası yayında —
    # 501 yanıtı artık yönlendirme alanı taşır. Orijinal işleyici korunur.
    return jsonify({
        'error': ('This analysis is being rebuilt on the reduced-order '
                  'physics architecture. Its successor endpoint is live: '
                  'POST /api/flow-analysis (quasi-1D compressible nozzle '
                  'flow with Fast Screening / Engineering fidelity levels).'),
        'status': 'unavailable',
        'successor': '/api/flow-analysis'
    }), 501
    # v2.6.27: 501 sonrasi erisilemez eski govde SOKULDU (teknik borc §4;
    # gecmisi git'te: d36624e oncesi surumler). Halef uclar yasiyor.

@app.route('/api/kinetic-analysis', methods=['POST'])
def perform_kinetic_analysis():
    """Perform nozzle kinetic loss analysis"""
    # Dalga 0 bekçisi (2026-07-14): stiff ODE + explicit RK45 tek istasyonda
    # ~23 dk sürüyor ve tek-worker masaüstü uygulamasını KİLİTLİYOR; bitse
    # bile isp_loss ≡ 0 dönüyordu. Dalga 4A: kademeli kinetik verim halefi
    # yayında — 501 yanıtı yönlendirme alanı taşır. Orijinal işleyici korunur.
    return jsonify({
        'error': ('This analysis is being rebuilt on the reduced-order '
                  'physics architecture. Its successor endpoint is live: '
                  'POST /api/kinetic-efficiency (tiered frozen/shifting '
                  'kinetic-loss model: fast / engineering / high_fidelity).'),
        'status': 'unavailable',
        'successor': '/api/kinetic-efficiency'
    }), 501
    # v2.6.27: 501 sonrasi erisilemez eski govde SOKULDU (teknik borc §4;
    # gecmisi git'te: d36624e oncesi surumler). Halef uclar yasiyor.

@app.route('/api/professional-analysis', methods=['POST'])
def perform_complete_professional_analysis():
    """Perform complete analysis using all modules"""
    # Dalga 0 bekçisi (2026-07-14): bu uç CFD + kinetik çözücüleri birlikte
    # çağırıyor — ikisi de yukarıdaki nedenlerle emekliye ayrıldı (kilitleme
    # + ıraksama riski). v2.4.6: halefler yayında — 501 yanıtı yönlendirme
    # alanı taşır. Orijinal işleyici aşağıda korunur.
    return jsonify({
        'error': ('This analysis is being rebuilt on the reduced-order '
                  'physics architecture. Its successors are live: '
                  'POST /api/flow-analysis, POST /api/kinetic-efficiency '
                  'and the Analysis Deck panels '
                  '(structural/thermal/safety/flow/validation).'),
        'status': 'unavailable',
        'successor': ['/api/flow-analysis', '/api/kinetic-efficiency',
                      'Analysis Deck panels (structural/thermal/safety/flow/validation)']
    }), 501
    # v2.6.27: 501 sonrasi erisilemez eski govde SOKULDU (teknik borc §4;
    # gecmisi git'te: d36624e oncesi surumler). Halef uclar yasiyor.

@app.route('/api/get-fuel-properties', methods=['POST'])
def get_fuel_properties():
    try:
        data = request.json
        fuel_type = data.get('fuel_type', 'htpb')
        temperature = data.get('temperature', 298.15)
        
        print(f"FETCHING NASA CEA DATA: {fuel_type} at {temperature}K")
        
        # Get fuel properties from chemical database
        fuel_mapping = {
            'rp1': 'RP1',
            'lh2': 'H2', 
            'methane': 'CH4',
            'mmh': 'MMH',
            'udmh': 'UDMH',
            'htpb': 'HTPB',
            'paraffin': 'Paraffin'
        }
        
        species_name = fuel_mapping.get(fuel_type, fuel_type.upper())
        species = chemical_db.get_species(species_name)
        
        if species:
            # Calculate properties at requested temperature
            cp = chemical_db.calculate_cp(species_name, temperature)
            enthalpy = chemical_db.calculate_enthalpy(species_name, temperature)
            entropy = chemical_db.calculate_entropy(species_name, temperature)
            
            properties = {
                'density': species.molecular_weight * 10 if species.phase == 'liquid' else species.molecular_weight,
                'enthalpy_formation': species.enthalpy_formation / 1000,  # Convert to kJ/mol
                'formula': species.formula,
                'phase': species.phase,
                'cp': cp / 1000,  # Convert to kJ/mol/K
                'enthalpy': enthalpy / 1000,
                'entropy': entropy / 1000,
                'source': species.source,
                'molecular_weight': species.molecular_weight,
                'temperature': temperature,
                'timestamp': datetime.now().isoformat()
            }
            
            print(f"NASA CEA RESPONSE: {species_name} - MW: {species.molecular_weight}, dHf: {species.enthalpy_formation}")
            
            return jsonify({
                'status': 'success',
                'properties': sanitize_json_values(properties),
                'source': 'NASA CEA Database',
                'real_time': True
            })
        else:
            print(f"Species not found: {species_name}, trying fallback...")
            
            # Fallback properties for common fuels
            fallback_props = get_cached_fuel_properties(fuel_type, temperature)
            
            return jsonify({
                'status': 'success',
                'properties': sanitize_json_values(fallback_props),
                'source': 'Cached Database',
                # Faz 5B / H3-B13: ``species_name`` doğrudan kullanıcı
                # girdisidir. 100 000 karakterlik bir ``fuel_type`` ile bu
                # not 100 049 karakter oluyordu (ölçüldü, HTTP 200 — yani
                # ``_clip_error_body`` kapsamı DIŞINDA). Yankı kaynağında
                # kırpılır ve kırpma açıkça beyan edilir.
                'note': ('Species '
                         + _clip_echo(species_name)
                         + ' not found in NASA CEA, using cached data')
            })
            
    except Exception as e:
        print(f"NASA CEA ERROR: {str(e)}")
        return jsonify({
            'status': 'error', 
            'error': f'NASA CEA Database Error: {str(e)}'
        }), 500


def get_cached_fuel_properties(fuel_type, temperature):
    """Get cached fuel properties"""
    cache_data = {
        'rp1': {
            'density': 810.0,
            'enthalpy_formation': -194.2,  # kJ/mol
            'formula': 'C12H23',
            'phase': 'liquid',
            'heating_value': 43000  # kJ/kg
        },
        'lh2': {
            'density': 71.0,
            'enthalpy_formation': 0.0,
            'formula': 'H2',
            'phase': 'liquid',
            'heating_value': 120000
        },
        'methane': {
            'density': 423.0,
            'enthalpy_formation': -74.6,
            'formula': 'CH4', 
            'phase': 'liquid',
            'heating_value': 50000
        }
    }
    
    props = cache_data.get(fuel_type, cache_data['rp1']).copy()
    props.update({
        'temperature': temperature,
        'source': 'Cached Database',
        'timestamp': datetime.now().isoformat()
    })
    return props

def get_oxidizer_density(oxidizer_type, temperature):
    """Calculate oxidizer density with temperature dependency"""
    base_densities = {
        'lox': (1141.0, 90.15, -4.0),    # (density at Tb, Tb, dρ/dT)
        'n2o4': (1443.0, 261.95, -2.8),
        'n2o': (1220.0, 184.67, -2.5)
    }
    
    if oxidizer_type in base_densities:
        rho_base, t_base, drho_dt = base_densities[oxidizer_type]
        return max(10.0, rho_base + drho_dt * (temperature - t_base))
    return 1141.0  # Default to LOX

def get_oxidizer_viscosity(oxidizer_type, temperature):
    """Calculate oxidizer viscosity"""
    viscosities = {
        'lox': 1.95e-4,
        'n2o4': 4.2e-4, 
        'n2o': 2.8e-4
    }
    return viscosities.get(oxidizer_type, 1.95e-4)

def get_oxidizer_conductivity(oxidizer_type, temperature):
    """Calculate thermal conductivity"""
    conductivities = {
        'lox': 0.15,
        'n2o4': 0.12,
        'n2o': 0.20
    }
    return conductivities.get(oxidizer_type, 0.15)

def get_cached_oxidizer_properties(oxidizer_type, temperature):
    """Get cached oxidizer properties when live data unavailable"""
    
    # Realistic oxidizer properties database with temperature dependency
    cache_data = {
        'lox': {
            'density': get_oxidizer_density('lox', temperature),
            'viscosity': 1.95e-4,
            'heat_capacity': 1.7,
            'thermal_conductivity': 0.15,
            'formula': 'O2',
            'boiling_point': 90.15,
            'critical_temperature': 154.8,
            'molecular_weight': 31.998
        },
        'n2o4': {
            'density': get_oxidizer_density('n2o4', temperature),
            'viscosity': 4.2e-4,
            'heat_capacity': 1.4,
            'thermal_conductivity': 0.12,
            'formula': 'N2O4', 
            'boiling_point': 294.3,
            'critical_temperature': 431.35,
            'molecular_weight': 92.011
        },
        'n2o': {
            'density': get_oxidizer_density('n2o', temperature),
            'viscosity': 2.8e-4,
            'heat_capacity': 2.2,
            'thermal_conductivity': 0.20,
            'formula': 'N2O',
            'boiling_point': 184.67,
            'critical_temperature': 309.57,
            'molecular_weight': 44.013
        }
    }
    
    props = cache_data.get(oxidizer_type, cache_data['lox']).copy()
    props.update({
        'temperature': temperature,
        'source': 'Cached Database',
        'timestamp': datetime.now().isoformat(),
        'note': 'Live NIST data unavailable'
    })
    return props

@app.route('/api/advanced-performance-analysis', methods=['POST'])
def advanced_performance_analysis():
    """Generate advanced performance analysis graphs based on NASA standards"""
    try:
        data = request.json
        analysis_type = data.get('analysis_type', '3d_surface')
        
        if analysis_type == '3d_surface':
            # Chamber Pressure vs Mixture Ratio vs Isp (NASA SP-125)
            # v2.5.2 (Codex bulgusu): yakıt/oksitleyici KİMLİĞİ bu sözlüğe
            # konmuyordu, bu yüzden LOX/RP-1 koşusunda bile denge yüzeyi
            # _resolve_surface_propellant'ın HTPB/N2O referans çiftiyle
            # çözülüyordu. Kimlik ve tarama aralıkları artık aktarılır;
            # verilmezse görselleştirme "referans çift" uyarısını basar.
            engine_data = {
                'base_isp': data.get('base_isp', 300),
                'optimal_of_ratio': data.get('optimal_of_ratio', 3.5),
                'optimal_chamber_pressure': data.get('chamber_pressure', 50),
                'fuel_type': data.get('fuel_type'),
                'fuel_composition': data.get('fuel_composition'),
                'oxidizer_type': data.get('oxidizer_type'),
                'pc_range': data.get('pc_range'),
                'of_range': data.get('of_range'),
                'grid_n': data.get('grid_n'),
            }
            engine_data = {k: v for k, v in engine_data.items() if v is not None}

            plot_json = create_chamber_pressure_mixture_ratio_3d_surface(engine_data)

            return jsonify({
                'status': 'success',
                'plot_data': plot_json,
                'analysis_info': {
                    'title': '3D Performance Surface Analysis',
                    # C2: künye "Liquid-Propellant Rocket Engine Performance"
                    # diyordu; SP-125'in gerçek adı bu değil. Doğrulanmış ad
                    # docs/STANDART_ATIFLARI.md'de (NTRS/ADS 1971NASSP.125.....H).
                    'reference': ('NASA SP-125 — Huzel & Huang, Design of '
                                  'Liquid Propellant Rocket Engines '
                                  '(2nd ed., 1971)'),
                    'description': 'Shows optimum O/F ratio and chamber pressure regions with combustion instability bands'
                }
            })

        elif analysis_type == 'nozzle_mach':
            # Nozzle Mach-Area Ratio Contour
            # (NACA Report 1135, "Equations, Tables, and Charts for
            #  Compressible Flow", Ames Research Staff, 1953; Anderson,
            #  "Modern Compressible Flow", 3. baskı — bkz.
            #  hrma/analysis/nozzle_flow_1d.py:12-26 kaynak listesi)
            #
            # DENETİM DÜZELTMESİ (2026-08-02, C2): bu satır ve aşağıdaki
            # 'reference' alanı konturun kaynağı olarak NASA-STD-5012
            # gösteriyordu. İki ayrı hata vardı: (1) BAŞLIK yanlıştı —
            # belgeye "Pressure Vessels & Pressurized Systems" deniyordu,
            # gerçek adı "Strength and Life Assessment Requirements for
            # Liquid-Fueled Space Propulsion System Engines" (Rev. B, 2016);
            # (2) KONU yanlıştı — o belge motor mukavemeti ve ömür
            # değerlendirmesi hakkındadır, içinde izantropik alan-Mach
            # bağıntısı yoktur, dolayısıyla bu figürün kaynağı olamaz.
            # Doğrulanmış başlıklar: docs/STANDART_ATIFLARI.md; aynı
            # düzeltme formulas.html §14.2, i18n_formulas.js (EN+TR) ve
            # performance_panel.js:7'de yapıldı — kaynak adı üç yerde
            # birebir aynıdır.
            # v2.5.2 (Codex bulgusu): yalnız throat_area / nozzle_length /
            # expansion_ratio aktarılıyordu; gaz hâli (gamma, MW, Tc), oda
            # basıncı, hazne çapı ve ortam basıncı DÜŞÜYORDU. Çözücü kendi
            # 20 bar / gamma 1.20 / 1 atm varsayılanlarına iniyor, oda
            # basıncını değiştirmek grafiği hiç değiştirmiyordu. Aynı ısı
            # akısı dalındaki desen uygulanır: alan varsa geçilir, yoksa
            # hiç konmaz ve figür "assumed" listesinde açıkça yazar.
            cfd_data = {
                'throat_area': data.get('throat_area', 0.001),
                'throat_diameter': data.get('throat_diameter'),
                'exit_diameter': data.get('exit_diameter'),
                'nozzle_length': data.get('nozzle_length', 0.1),
                'expansion_ratio': data.get('expansion_ratio', 16),
                'chamber_diameter': data.get('chamber_diameter'),
                'chamber_pressure': data.get('chamber_pressure'),
                'chamber_temperature': data.get('chamber_temperature'),
                'gamma': data.get('gamma'),
                'molecular_weight': data.get('molecular_weight'),
                # Görselleştirme sözleşmesi: ambient_pressure PASCAL
                # (NozzleFlow1D.from_motor_data ambient_pressure=Pa bekler).
                # Bar cinsinden gönderen çağıranlar için ayrı anahtar.
                'ambient_pressure': (
                    data.get('ambient_pressure')
                    if data.get('ambient_pressure') is not None
                    else (float(data['ambient_pressure_bar']) * 1e5
                          if data.get('ambient_pressure_bar') is not None
                          else None)),
            }
            cfd_data = {k: v for k, v in cfd_data.items() if v is not None}

            plot_json = create_nozzle_mach_area_ratio_contour(cfd_data)
            
            return jsonify({
                'status': 'success',
                'plot_data': plot_json,
                'analysis_info': {
                    'title': 'Nozzle Mach Distribution Analysis',
                    'reference': ('NACA Report 1135, "Equations, Tables, and '
                                  'Charts for Compressible Flow" (Ames '
                                  'Research Staff, 1953); Anderson, "Modern '
                                  'Compressible Flow", 3rd ed.'),
                    'description': 'Visualizes Mach distribution and shock/threshold regions for over/under-expansion detection'
                }
            })
            
        elif analysis_type == 'heat_flux':
            # Wall Heat Flux Waterfall (NASA SP-8124)
            # v2.5.2: panel throat_area / chamber_pressure / expansion_ratio
            # gönderiyordu ama bu sözlüğe konmuyordu, dolayısıyla ısı akısı
            # GERÇEK Bartz hesabına giremeyip "not available" durumuna
            # düşüyordu. Üç alan da geçiriliyor; ek olarak malzeme, gaz
            # özellikleri ve kütle debisi de varsa aktarılır.
            thermal_data = {
                'burn_time': data.get('burn_time', 30),
                'chamber_length': data.get('chamber_length', 0.5),
                'nozzle_length': data.get('nozzle_length', 0.1),
                'base_heat_flux': data.get('base_heat_flux', 2e6),
                'critical_heat_flux': data.get('critical_heat_flux', 4.0),
                'molecular_weight': data.get('molecular_weight'),
                'throat_area': data.get('throat_area'),
                'throat_diameter': data.get('throat_diameter'),
                'chamber_pressure': data.get('chamber_pressure'),
                'chamber_temperature': data.get('chamber_temperature'),
                'expansion_ratio': data.get('expansion_ratio'),
                'chamber_diameter': data.get('chamber_diameter'),
                'mdot_total': data.get('mdot_total'),
                'gamma': data.get('gamma'),
                'material': data.get('material'),
                'c_star': data.get('c_star'),
            }
            thermal_data = {k: v for k, v in thermal_data.items() if v is not None}
            
            plot_json = create_wall_heat_flux_waterfall_plot(thermal_data)
            
            return jsonify({
                'status': 'success',
                'plot_data': plot_json,
                'analysis_info': {
                    'title': 'Wall Heat Flux Waterfall Analysis',
                    # C2: künye "Thermal Design Criteria" diyordu; SP-8124'ün
                    # gerçek adı bu değil (NTRS 78N21211). Doğrulanmış ad
                    # docs/STANDART_ATIFLARI.md'de.
                    'reference': ('NASA SP-8124 — Liquid Rocket Engine '
                                  'Self-Cooled Combustion Chambers (1977)'),
                    'description': 'Gradient colored waterfall showing local heat flux along cooling channels with thermal runaway detection'
                }
            })
            
        else:
            return jsonify({
                'status': 'error',
                'error': f'Unknown analysis type: {analysis_type}',
                'available_types': ['3d_surface', 'nozzle_mach', 'heat_flux']
            }), 400
            
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

# PDF Export Endpoints
def _build_pdf_analysis_sections(motor_data, analysis_results):
    """PDF rapor bölümlerini motorun GERÇEK analiz sonuçlarıyla doldurur.

    Kaynak öncelik sırası (Dalga 2, 2026-07-14):
      motor sonuçları (motor_data.heat_transfer_analysis /
      structural_analysis) > istekle gelen analysis_results alanları.
    Sabit/uydurma değer ÜRETİLMEZ: veri yoksa ilgili alan hiç konmaz;
    pdf_generator eksik alanları 'N/A' olarak basar. (Eski app.js sabit
    SF 4.0/3.0/4.0 dürüstlük sorununun rapor katmanındaki karşılığı.)
    """
    out = dict(analysis_results or {})
    md = motor_data or {}

    def _num(value):
        try:
            f = float(value)
        except (TypeError, ValueError):
            return None
        return f if np.isfinite(f) else None

    # ---- Performans: motor sonucundaki gerçek değerler öncelikli ----
    performance = dict(out.get('performance') or {})
    perf_sources = {
        'thrust': ('thrust',),
        'specific_impulse': ('specific_impulse', 'isp'),
        'chamber_pressure': ('chamber_pressure',),
        'burn_time': ('burn_time',),
        'total_impulse': ('total_impulse',),
        'exit_velocity': ('exit_velocity',),
        'mass_flow_rate': ('mdot_total', 'total_mdot', 'mass_flow_rate'),
    }
    for target, sources in perf_sources.items():
        for source in sources:
            value = _num(md.get(source))
            if value is not None:
                performance[target] = value
                break
    if performance:
        out['performance'] = performance

    # ---- Termal: heat_transfer_analysis (Bartz) gerçek sonuçları ----
    heat = md.get('heat_transfer_analysis') or out.get('heat_transfer_analysis') or {}
    if isinstance(heat, dict) and heat:
        wall = heat.get('wall_analysis') or {}
        gas_side = heat.get('gas_side_analysis') or {}
        cooling = heat.get('cooling_analysis') or {}
        thermal = dict(out.get('thermal') or {})
        heat_flux = _num(gas_side.get('heat_flux'))
        candidates = {
            'max_wall_temp': _num(wall.get('max_temperature')),
            'heat_flux': heat_flux / 1e6 if heat_flux is not None else None,  # W/m^2 -> MW/m^2
            'cooling_req': _num(cooling.get('peak_heat_rate')),  # kW
            'adiabatic_wall_temp': _num(gas_side.get('adiabatic_wall_temperature')),
            'gas_side_coefficient': _num(gas_side.get('gas_side_coefficient')),
        }
        for key, value in candidates.items():
            if value is not None:
                thermal[key] = value
        if thermal:
            out['thermal'] = thermal

    # ---- Yapısal: gerçek SF'ler (sabit 4.0 kalıntısı YOK) ----
    structural_src = md.get('structural_analysis') or out.get('structural_analysis') or {}
    if isinstance(structural_src, dict) and structural_src:
        safety_sub = structural_src.get('safety_analysis') or {}
        chamber = structural_src.get('chamber_analysis') or {}
        structural = dict(out.get('structural') or {})
        candidates = {
            'safety_factor': _num(structural_src.get('safety_factor')),
            'safety_factor_pressure': _num(structural_src.get('safety_factor_pressure')),
            'safety_factor_total': _num(structural_src.get('safety_factor_total')),
            'min_safety_factor': _num(safety_sub.get('minimum_safety_factor')),
            'von_mises_stress_MPa': _num(chamber.get('von_mises_stress')),
            'hoop_stress_MPa': _num(chamber.get('hoop_stress')),
        }
        for key, value in candidates.items():
            if value is not None:
                structural[key] = value
        if safety_sub.get('status'):
            structural['status'] = str(safety_sub['status'])
        if safety_sub.get('risk_level'):
            structural['risk_level'] = str(safety_sub['risk_level'])
        if structural:
            out['structural'] = structural

    # ---- Güvenlik özeti: istekle geldiyse aynen korunur ----
    # (out zaten istekten kopyalandı; 'safety' anahtarına dokunulmaz.)
    return out


@app.route('/api/materials', methods=['GET'])
def get_materials_catalog():
    """Merkezi malzeme kütüphanesini döndürür.

    Sözleşme (v2.5.2): {ok, materials: {key: {name, source, tags, ...}},
    aliases: {...}}. Paneller (static/js/materials_catalog.js) select
    listelerini buradan doldurur; endpoint yoksa hardcoded fallback'e düşer.
    """
    try:
        from hrma.data.materials_db import MATERIALS, ALIASES
        return jsonify(sanitize_json_values({
            'ok': True,
            'materials': MATERIALS,
            'aliases': ALIASES,
        }))
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/propellants', methods=['GET'])
def get_propellants_catalog():
    """Merkezi katı yakıt kataloğunu döndürür.

    Sözleşme (v2.5.2) — /api/materials ile birebir aynı desen:
        {ok: true,
         propellants: {key: {...tüm kayıt alanları}},
         aliases: {alias: canonical_key}}
    Katı sayfası (static/js/propellant_catalog.js) yakıt seçicisini ve
    otomatik dolan özellik alanlarını buradan besler; endpoint yoksa
    sayfa kendi hardcoded fallback listesine düşer.

    Tek doğruluk kaynağı: hrma/data/propellants_db.py — yanma hızı yasası
    olan yakıtlarda (KNDX/KNSB) a-n değerleri merkezi burn_rate_db'den
    türetilir, burada ayrıca yazılmaz.
    """
    try:
        from hrma.data.propellants_db import PROPELLANTS, ALIASES
        return jsonify(sanitize_json_values({
            'ok': True,
            'propellants': PROPELLANTS,
            'aliases': ALIASES,
        }))
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


#: Elektronik tabloların formül başlangıcı saydığı karakterler. Excel ve
#: LibreOffice '=' dışındakileri de bağlama göre formül/komut olarak
#: yorumlayabildiği için dördü birden kaçırılır (CWE-1236).
_FORMULA_LEAD = ('=', '+', '-', '@', '\t', '\r', '\n')


def _looks_numeric(text_value):
    """'-5000' / '+3.2e4' gibi sayı görünen metin mi?

    Formül kaçışının bunlara DOKUNMAMASI gerekir: negatif sayılar '-' ile
    başlar ve körü körüne apostrof eklemek veriyi bozardı.
    """
    try:
        parsed = float(text_value)
    except (TypeError, ValueError):
        return False
    return math.isfinite(parsed)


def _spreadsheet_safe(value):
    """Kullanıcı metnini hücreye yazmadan önce formüle dönüşmekten korur.

    v2.6.26 ÖLÇÜMÜ: '=1+1' ile başlayan bir motor adı XLSX'e yazıldığında
    openpyxl hücreyi ``data_type='f'`` (formül) olarak saklıyordu; dosya
    yeniden açıldığında formül olarak görünüyordu. Başa apostrof koymak
    elektronik tabloların standart 'bu metindir' işaretidir ve hücre metni
    olarak görüntülenirken apostrof gösterilmez.
    """
    text_value = str(value)
    if text_value.startswith(_FORMULA_LEAD) and not _looks_numeric(text_value):
        return "'" + text_value
    return text_value


def _safe_sheet_title(title, idx):
    """Excel'in yasakladığı karakterleri temizler (eskiden HTTP 500 oluyordu)."""
    import re
    cleaned = re.sub(r'[\\/*?:\[\]]', '_', title).strip() or f'Sheet{idx + 1}'
    return cleaned[:31]


# ---------------------------------------------------------------------------
# D8 — XLSX İŞ BÜTÇESİ (2026-08-02)
#
# ÖLÇÜLDÜ: tek bir 23.3 MiB istek 26.3 saniye sürdü ve süreç 2.4 GB RSS'e
# çıktı. 60 000 sütunluk sayfa hiçbir yerde yakalanmıyordu, 200 000
# karakterlik hücreyi openpyxl SESSİZCE 32 767'ye kırpıyordu (bu makinede
# doğrulandı: 200000 karakter yazıldı, geri okunan uzunluk 32767).
#
# Eski kodda iki SESSİZ kırpma daha vardı: ``sheets[:20]`` ve
# ``rows[:100000]``. Kullanıcı eksik bir çalışma kitabı indiriyor ve bunu
# hiçbir yerden öğrenemiyordu. Artık her sınır AÇIK hata döndürür.
#
# Sınırların dayanağı — bu makinede openpyxl ile ölçüldü:
#   500 000 sayısal hücre (250 sütun x 2000 satır) -> 1.2 s, 229 MB tepe RSS
#   500 000 sayısal hücre (10 sütun x 50 000 satır) -> 1.4 s, 263 MB tepe RSS
#   8 000 000 karakter metin                        -> 0.13 s, 59 MB
# Gerçek HRMA çıktılarının en genişi 10 sütun (transient paneli ve katı
# motor itki eğrisi sayfası), en uzunu birkaç bin satır — yani meşru
# kullanım bu tavanların çok altında.
# ---------------------------------------------------------------------------
XLSX_MAX_SHEETS = 20
XLSX_MAX_COLUMNS = 256
XLSX_MAX_ROWS_PER_SHEET = 100_000
XLSX_MAX_TOTAL_CELLS = 500_000
#: Excel'in kendi hücre sınırı. Aşan değeri KIRPMAK yerine reddediyoruz.
XLSX_MAX_CELL_CHARS = 32_767


def _xlsx_budget_error(sheets):
    """İş bütçesi aşıldı mı? Aşıldıysa ``(gövde, http_kodu)`` döner.

    Sessiz kırpma yasak: her ret, hangi sayfanın hangi sınırı hangi değerle
    aştığını söyler. 413 "istek fazla büyük", 422 "istek biçimi geçersiz".
    """
    if len(sheets) > XLSX_MAX_SHEETS:
        return ({'status': 'error', 'error': 'xlsx_budget_exceeded',
                 'limit': 'sheets', 'maximum': XLSX_MAX_SHEETS,
                 'requested': len(sheets),
                 'message': (f'At most {XLSX_MAX_SHEETS} sheets per workbook; '
                             f'{len(sheets)} were supplied.')}, 413)

    total_cells = 0
    for idx, sheet in enumerate(sheets):
        if not isinstance(sheet, dict):
            return ({'status': 'error', 'error': 'invalid_sheet',
                     'sheet_index': idx,
                     'message': 'Each sheet must be an object with '
                                '{name, headers, rows}.'}, 422)
        headers = sheet.get('headers') or []
        rows = sheet.get('rows') or []
        if len(rows) > XLSX_MAX_ROWS_PER_SHEET:
            return ({'status': 'error', 'error': 'xlsx_budget_exceeded',
                     'limit': 'rows_per_sheet', 'sheet_index': idx,
                     'maximum': XLSX_MAX_ROWS_PER_SHEET,
                     'requested': len(rows),
                     'message': (f'Sheet {idx} has {len(rows)} rows; the '
                                 f'limit is {XLSX_MAX_ROWS_PER_SHEET}.')}, 413)

        widest = len(headers)
        for row in rows:
            width = len(row) if isinstance(row, (list, tuple)) else 1
            if width > widest:
                widest = width
        if widest > XLSX_MAX_COLUMNS:
            return ({'status': 'error', 'error': 'xlsx_budget_exceeded',
                     'limit': 'columns', 'sheet_index': idx,
                     'maximum': XLSX_MAX_COLUMNS, 'requested': widest,
                     'message': (f'Sheet {idx} is {widest} columns wide; the '
                                 f'limit is {XLSX_MAX_COLUMNS}.')}, 413)

        total_cells += widest * (len(rows) + (1 if headers else 0))
        if total_cells > XLSX_MAX_TOTAL_CELLS:
            return ({'status': 'error', 'error': 'xlsx_budget_exceeded',
                     'limit': 'total_cells',
                     'maximum': XLSX_MAX_TOTAL_CELLS,
                     'requested': total_cells,
                     'message': (f'The workbook would contain at least '
                                 f'{total_cells} cells; the limit is '
                                 f'{XLSX_MAX_TOTAL_CELLS}.')}, 413)

        def _too_long(values, kind):
            for value in values:
                if isinstance(value, str) and len(value) > XLSX_MAX_CELL_CHARS:
                    return ({'status': 'error', 'error': 'xlsx_cell_too_long',
                             'sheet_index': idx, 'cell_kind': kind,
                             'maximum': XLSX_MAX_CELL_CHARS,
                             'requested': len(value),
                             'message': (f'A {kind} cell holds {len(value)} '
                                         f'characters; Excel stores at most '
                                         f'{XLSX_MAX_CELL_CHARS} and the '
                                         f'value would be truncated without '
                                         f'notice.')}, 422)
            return None

        problem = _too_long(headers, 'header')
        if problem:
            return problem
        for row in rows:
            problem = _too_long(
                row if isinstance(row, (list, tuple)) else [row], 'data')
            if problem:
                return problem
    return None


@app.route('/api/export-xlsx', methods=['POST'])
def export_xlsx():
    """Genel amaçlı Excel (xlsx) dışa aktarma.

    Girdi: {filename: 'name.xlsx', sheets: [{name, headers: [...],
    rows: [[...], ...]}]}. Transient sonuçları, regresyon analizi ve
    genel analiz özeti bu uçtan insan-dostu Excel olarak iner
    (kullanıcı şikayeti: .json indirmesi kullanışsızdı).
    """
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        data = request.json or {}
        sheets = data.get('sheets') or []
        if not sheets:
            return jsonify({'status': 'error', 'error': 'No sheets provided'}), 400
        if not isinstance(sheets, list):
            return jsonify({'status': 'error', 'error': 'invalid_sheets',
                            'message': "'sheets' must be a list."}), 422

        # D8: iş bütçesi. Tek satırlık kırpma yerine açık ret.
        budget = _xlsx_budget_error(sheets)
        if budget:
            body, code = budget
            return jsonify(body), code

        wb = Workbook()
        wb.remove(wb.active)
        header_font = Font(bold=True)
        for idx, sheet in enumerate(sheets):
            raw_title = str(sheet.get('name') or f'Sheet{idx + 1}')[:31]
            title = _safe_sheet_title(raw_title, idx)
            ws = wb.create_sheet(title=title)
            headers = sheet.get('headers') or []
            rows = sheet.get('rows') or []
            if headers:
                ws.append([_spreadsheet_safe(h) for h in headers])
                for cell in ws[1]:
                    cell.font = header_font
            for row in rows:
                ws.append([
                    (float(v) if isinstance(v, (int, float)) and not isinstance(v, bool)
                     else ('' if v is None else _spreadsheet_safe(v)))
                    for v in (row if isinstance(row, (list, tuple)) else [row])
                ])
            # Kolon genişliklerini başlığa göre kabaca ayarla
            for col_idx, h in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(32, len(str(h)) + 4))

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        # D9 (2026-08-02): dosya adı ``safe_name``'den GEÇMİYORDU.
        # ÖLÇÜLDÜ: filename='../../../../etc/passwd.xlsx' isteği HTTP 200
        # döndü ve ad Content-Disposition başlığına aynen girdi. Diğer
        # export uçları (STEP/DXF/ZIP) zaten safe_name kullanıyordu; bu uç
        # atlanmıştı.
        filename = safe_name(data.get('filename') or 'hrma_export.xlsx',
                             default='hrma_export.xlsx')
        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except ImportError:
        return jsonify({
            'status': 'error',
            'error': 'openpyxl is not installed on the server; falling back to CSV is recommended.'
        }), 501
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/export-pdf/<report_type>', methods=['POST'])
def export_pdf_report(report_type):
    """Export motor analysis as PDF report"""
    try:
        from hrma.export.pdf_generator import PDFReportGenerator

        data = request.json
        motor_data = data.get('motor_data', {})
        analysis_results = data.get('analysis_results', {})
        charts = data.get('charts', [])

        # Dalga 2: rapor bölümleri motor sonuçlarındaki GERÇEK analizlerle
        # beslenir (heat_transfer_analysis + structural_analysis + istekle
        # gelen safety özeti). Sabit değer enjekte edilmez.
        analysis_results = _build_pdf_analysis_sections(motor_data, analysis_results)

        pdf_generator = PDFReportGenerator()

        # D5 (2026-08-02): indirme adı ``safe_name``'den GEÇMİYORDU — motor
        # adı ham olarak Content-Disposition'a giriyordu. Depodaki doğru
        # desen aynı dosyada zaten var (app.py:3403 vd., STEP/DXF/ZIP uçları).
        safe_motor_name = safe_name(motor_data.get('motor_name'),
                                    default='unnamed')

        # Generate different types of reports
        if report_type == 'summary':
            pdf_bytes = pdf_generator.generate_quick_summary_report(motor_data, analysis_results)
            filename = f"motor_summary_{safe_motor_name}.pdf"
        elif report_type == 'technical':
            pdf_bytes = pdf_generator.generate_technical_report(motor_data, analysis_results, charts)
            filename = f"motor_technical_{safe_motor_name}.pdf"
        else:
            pdf_bytes = pdf_generator.generate_motor_analysis_report(
                motor_data, analysis_results, charts, 'complete'
            )
            filename = f"motor_complete_{safe_motor_name}.pdf"

        # Return PDF file
        pdf_buffer = io.BytesIO(pdf_bytes)
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'PDF generation failed: {str(e)}'
        }), 500

@app.route('/api/export-chart-pdf', methods=['POST'])
def export_chart_as_pdf():
    """Export individual chart as PDF"""
    try:
        from hrma.export.pdf_generator import PDFReportGenerator
        
        data = request.json
        chart_json = data.get('chart_data', '')
        chart_title = data.get('chart_title', 'Chart')
        motor_name = data.get('motor_name', 'unnamed')
        
        pdf_generator = PDFReportGenerator()
        
        # Convert chart to image
        chart_image = pdf_generator.export_plotly_chart_to_image(chart_json)
        
        if not chart_image:
            return jsonify({
                'status': 'error',
                'error': 'Failed to convert chart to image'
            }), 400
        
        # Create simple PDF with just the chart
        motor_data = {'motor_name': motor_name, 'motor_type': 'analysis'}
        analysis_results = {'chart_title': chart_title}
        
        pdf_bytes = pdf_generator.generate_motor_analysis_report(
            motor_data, analysis_results, [chart_image], 'summary'
        )
        
        pdf_buffer = io.BytesIO(pdf_bytes)
        pdf_buffer.seek(0)
        
        # D5: iki bileşen de ``safe_name``'den geçer — grafik başlığı ve motor
        # adı kullanıcı girdisidir ve doğrudan Content-Disposition'a giriyordu.
        filename = 'chart_{}_{}.pdf'.format(
            safe_name(str(chart_title).lower().replace(' ', '_'),
                      default='chart'),
            safe_name(motor_name, default='unnamed'))

        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'Chart PDF export failed: {str(e)}'
        }), 500

@app.route('/api/detailed-cad/<motor_type>', methods=['POST'])
def generate_detailed_cad(motor_type):
    """Generate detailed engineering CAD visualization"""
    try:
        from hrma.export.cad_visualization import DetailedCADGenerator
        
        data = request.json
        cad_generator = DetailedCADGenerator()
        
        if motor_type == 'liquid':
            result = cad_generator.generate_liquid_motor_cad(data)
        elif motor_type == 'solid':
            result = cad_generator.generate_solid_motor_cad(data)
        else:
            return jsonify({
                'status': 'error',
                'error': f'Unknown motor type: {motor_type}'
            }), 400
        
        return jsonify({
            'status': 'success',
            'cad_data': result['plot_json'],
            'component_details': result['component_details'],
            'dimensions': result.get('dimensions', {}),
            'design_info': {
                'title': f'Engineering CAD: {motor_type.title()} Motor',
                'description': 'Detailed engineering visualization with cross-section view',
                'features': [
                    'External component details',
                    'Internal structure cross-section', 
                    'Injector hole patterns',
                    'Cooling channel layout',
                    'Mounting flanges and sensors'
                ]
            }
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'CAD generation failed: {str(e)}'
        }), 500

# ============================================================================
# Dalga 4A — Akış / Kinetik / Doğrulama / İş kuyruğu uç noktaları (2026-07-14)
# Mimari: docs/ANALIZ_PLATFORM_PLANI.md — sahte CFD/kinetik yerine hızlı
# gerçekçi modeller; UI seviyeleri Fast Screening / Engineering / High-Fidelity.
# ============================================================================

# /api/flow-analysis'in kabul ettiği seviyeler (High-Fidelity kinetik zinciri
# /api/kinetic-efficiency üzerinden yürür; akış modeli quasi-1D kalır)
FLOW_FIDELITY_LEVELS = ('fast', 'engineering')


def _flow_float(data, key, default=None):
    """İstek gövdesinden sayısal alan oku; bozuksa net İngilizce ValueError."""
    value = data.get(key, default)
    if value is None:
        return None
    try:
        value = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"Field '{key}' must be a number, got {value!r}.")
    if not np.isfinite(value):
        raise ValueError(f"Field '{key}' must be finite.")
    return value


def _kinetic_evaluate_job(kwargs, progress_callback=None):
    """job_runner işi: kinetik verim değerlendirmesi (async yol).

    progress_callback runner tarafından enjekte edilir; değerlendirme tek
    parça olduğundan yalnız başla/bitti işaretlenir.
    """
    if progress_callback:
        progress_callback(0.1)
    result = kinetic_efficiency.evaluate(**kwargs)
    if progress_callback:
        progress_callback(1.0)
    return sanitize_json_values(result)


def _run_kinetic_chain(data, chamber_pressure_bar, throat_diameter,
                       nozzle_profile=None):
    """Ortak kinetik verim zinciri (flow-analysis + kinetic-efficiency).

    Returns (result_dict_or_None, note_or_None). Yakıt tanımı yoksa None +
    açıklayıcı not döner; değerlendirme hatası da not olarak raporlanır
    (akış analizi kinetik zincir yüzünden 500'e düşmez).
    """
    of_ratio = data.get('of_ratio')
    if of_ratio is None:
        return None, ("Kinetic-efficiency chain skipped: provide 'of_ratio' "
                      "(and optionally 'fuel_type', 'oxidizer_type') to "
                      "evaluate nozzle kinetic losses.")
    fidelity = str(data.get('kinetic_fidelity',
                            data.get('fidelity', 'engineering')))
    if fidelity not in KINETIC_FIDELITY_LEVELS:
        fidelity = 'engineering'
    fuel_composition = data.get('fuel_composition')
    if not isinstance(fuel_composition, dict) or not fuel_composition:
        fuel_composition = {str(data.get('fuel_type', 'htpb')): 100.0}
    try:
        result = kinetic_efficiency.evaluate(
            fuel_composition=fuel_composition,
            oxidizer_type=str(data.get('oxidizer_type', 'N2O')),
            of_ratio=float(of_ratio),
            chamber_pressure=chamber_pressure_bar,
            fidelity=fidelity,
            characteristic_length=_flow_float(data, 'characteristic_length'),
            throat_diameter=throat_diameter,
            nozzle_profile=nozzle_profile,
        )
        return sanitize_json_values(result), None
    except Exception as exc:
        return None, f"Kinetic-efficiency chain failed: {exc}"


@app.route('/api/flow-analysis', methods=['POST'])
def flow_analysis():
    """Quasi-1D compressible nozzle flow (successor of /api/cfd-analysis).

    Fidelity levels (Wave 4 architecture):
      fast        — isentropic summary: regime classification, CF/thrust,
                    throat state (no station arrays, no Bartz coupling).
      engineering — full 30-60 station arrays (P, M, T, rho, u, wall P),
                    axial Bartz h_g/q coupling, and the kinetic-efficiency
                    chain when the propellant definition is supplied
                    (of_ratio [+ fuel_type/oxidizer_type]).

    Units follow the repo convention: chamber_pressure in bar, temperatures
    in K, diameters in m, ambient_pressure in Pa.
    """
    data = request.get_json(silent=True) or {}

    fidelity = str(data.get('fidelity', 'engineering')).lower()
    if fidelity not in FLOW_FIDELITY_LEVELS:
        return jsonify({
            'status': 'error',
            'error': (f"fidelity must be one of {list(FLOW_FIDELITY_LEVELS)}; "
                      f"got '{fidelity}'. High-fidelity finite-rate kinetics "
                      "is served by POST /api/kinetic-efficiency."),
        }), 400

    try:
        chamber_pressure_bar = _flow_float(data, 'chamber_pressure', 20.0)
        chamber_temperature = _flow_float(data, 'chamber_temperature', 3000.0)
        gamma = _flow_float(data, 'gamma', 1.2)
        molecular_weight = _flow_float(data, 'molecular_weight', 24.0)
        throat_diameter = _flow_float(data, 'throat_diameter', 0.02)
        exit_diameter = _flow_float(data, 'exit_diameter')
        expansion_ratio = _flow_float(data, 'expansion_ratio')
        ambient_pressure = _flow_float(data, 'ambient_pressure', 101325.0)
        n_stations = int(_flow_float(data, 'n_stations', 45))
        separation_factor = _flow_float(data, 'separation_factor', 0.40)
        wall_temperature = _flow_float(data, 'wall_temperature', 800.0)
        # GÖÇ (16 Ağu 2026): varsayılan ARTIK sabit DEĞİL. Alan gövdede
        # yoksa None geçilir; çözücü o zaman ölçülen sınır tabakası kesrini
        # yayımlar (yayımlanamıyorsa 1,5 % sabitine beyanla düşer). Burada
        # sabiti varsayılan olarak geçmek, her isteği "kullanıcı açıkça
        # 0.015 istedi" (source='user') hâline getirir ve göçü SESSİZCE
        # iptal ederdi. Kullanıcı alanı gönderirse üstünlük yine onundur.
        friction_loss_fraction = _flow_float(data, 'friction_loss_fraction')
    except ValueError as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 400

    # Geometri varsayılanı: ne çıkış çapı ne genişleme oranı verilmişse
    # tipik ε=4 (küçük atmosferik motor) ile çalışılır — panel önerileri
    # gerçek motor sonuçlarından doldurur.
    if exit_diameter is None and expansion_ratio is None:
        expansion_ratio = 4.0

    motor_data = {}
    if data.get('nozzle_type'):
        # sample_nozzle_inner_contour konik/bell ayrımını buradan okur
        motor_data['nozzle_angles'] = {'nozzle_type': str(data['nozzle_type'])}
    if data.get('chamber_diameter') is not None:
        try:
            motor_data['chamber_diameter'] = float(data['chamber_diameter'])
        except (TypeError, ValueError):
            pass

    try:
        solver = NozzleFlow1D(
            chamber_pressure=chamber_pressure_bar * 1e5,
            chamber_temperature=chamber_temperature,
            gamma=gamma,
            molecular_weight=molecular_weight,
            throat_diameter=throat_diameter,
            exit_diameter=exit_diameter,
            expansion_ratio=expansion_ratio,
            ambient_pressure=ambient_pressure,
            n_stations=n_stations,
            separation_factor=separation_factor,
            wall_temperature=wall_temperature,
            friction_loss_fraction=friction_loss_fraction,
            motor_data=motor_data,
        )
        flow = solver.solve(include_bartz=(fidelity == 'engineering'))
    except ValueError as exc:
        # Modülün fizik doğrulamaları (Pa >= Pc, gamma bandı...) — net 400
        return jsonify({'status': 'error', 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 500

    kinetic_result = None
    kinetic_note = None
    if fidelity == 'fast':
        # İzantropik özet: istasyon dizileri taşınmaz (hızlı tarama seviyesi)
        flow.pop('stations', None)
        kinetic_note = ("Fast Screening level: station arrays and the "
                        "kinetic-efficiency chain require the Engineering "
                        "fidelity level.")
    else:
        nozzle_profile = None
        stations = flow.get('stations') or {}
        if stations.get('x_mm'):
            # Yüksek doğruluk kinetiği için quasi-1D profil (mm -> m)
            nozzle_profile = {
                'x': [v / 1000.0 for v in stations['x_mm']],
                'T': stations['temperature_K'],
                'P': stations['pressure_Pa'],
                'u': stations['velocity_m_s'],
            }
        kinetic_result, kinetic_note = _run_kinetic_chain(
            data, chamber_pressure_bar, throat_diameter,
            nozzle_profile=nozzle_profile)

    response = {
        'status': 'success',
        'fidelity': fidelity,
        'fidelity_levels': list(FLOW_FIDELITY_LEVELS),
        'flow': sanitize_json_values(flow),
        'kinetic_efficiency': kinetic_result,
    }
    if kinetic_note:
        response['kinetic_note'] = kinetic_note
    return jsonify(response)


@app.route('/api/kinetic-efficiency', methods=['POST'])
def kinetic_efficiency_analysis():
    """Tiered nozzle kinetic-efficiency analysis (successor of
    /api/kinetic-analysis).

    Fidelity: 'fast' (equilibrium reference), 'engineering' (JANNAF-style
    Damköhler correlation), 'high_fidelity' (Cantera finite-rate along a
    nozzle T(x), P(x) profile; graceful fallback to engineering — the
    'fidelity_used' field always reports what actually ran).

    Special modes:
      {'probe': true}  — capability probe only: reports whether the
                         high-fidelity path is available (no heavy work).
      {'async': true}  — queue the evaluation on the job runner; returns
                         202 with a job id to poll via GET /api/jobs/<id>.
    """
    data = request.get_json(silent=True) or {}

    # --- Yetenek sondası: Cantera + reaksiyonlu mekanizma var mı? ---
    # Panel, High-Fidelity seçeneğini fidelity_used alanından tespit eder.
    if data.get('probe'):
        available = False
        detail = 'Cantera is not installed'
        if KINETIC_CANTERA_AVAILABLE:
            try:
                # Modülün kendi mekanizma çözücüsü (önbellekli); reaksiyonsuz
                # termo-dosyalar (nasa_gas.yaml) elenir.
                kinetic_efficiency._get_kinetics_gas()
                available = True
                detail = (f"Cantera mechanism "
                          f"'{kinetic_efficiency._kin_mech_name}' ready")
            except Exception as exc:
                detail = str(exc)
        levels = ['fast', 'engineering'] + (['high_fidelity'] if available
                                            else [])
        return jsonify({
            'status': 'success',
            'probe': True,
            'fidelity_requested': 'high_fidelity',
            'fidelity_used': 'high_fidelity' if available else 'engineering',
            'cantera_available': bool(KINETIC_CANTERA_AVAILABLE),
            'fidelity_levels': levels,
            'detail': detail,
        })

    fidelity = str(data.get('fidelity', 'engineering'))
    if fidelity not in KINETIC_FIDELITY_LEVELS:
        return jsonify({
            'status': 'error',
            'error': (f"fidelity must be one of "
                      f"{list(KINETIC_FIDELITY_LEVELS)}; got '{fidelity}'."),
        }), 400

    try:
        of_ratio = _flow_float(data, 'of_ratio')
        chamber_pressure_bar = _flow_float(data, 'chamber_pressure', 20.0)
        characteristic_length = _flow_float(data, 'characteristic_length')
        throat_diameter = _flow_float(data, 'throat_diameter')
    except ValueError as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 400
    if of_ratio is None:
        return jsonify({
            'status': 'error',
            'error': ("Field 'of_ratio' is required (oxidizer-to-fuel mass "
                      "ratio, e.g. 6.0 for N2O/HTPB)."),
        }), 400

    fuel_composition = data.get('fuel_composition')
    if not isinstance(fuel_composition, dict) or not fuel_composition:
        fuel_composition = {str(data.get('fuel_type', 'htpb')): 100.0}

    kwargs = {
        'fuel_composition': fuel_composition,
        'oxidizer_type': str(data.get('oxidizer_type', 'N2O')),
        'of_ratio': of_ratio,
        'chamber_pressure': chamber_pressure_bar,
        'fidelity': fidelity,
        'characteristic_length': characteristic_length,
        'throat_diameter': throat_diameter,
        'nozzle_profile': data.get('nozzle_profile'),
    }

    if data.get('async'):
        # Uzun sürebilecek yol (Cantera BDF) iş kuyruğuna atılır; istemci
        # GET /api/jobs/<id> ile yoklar (job_runner sözleşmesi).
        job_id = job_runner.submit(_kinetic_evaluate_job, kwargs)
        return jsonify({
            'status': 'queued',
            'job_id': job_id,
            'poll_url': f'/api/jobs/{job_id}',
        }), 202

    try:
        result = kinetic_efficiency.evaluate(**kwargs)
    except ValueError as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 400
    except Exception as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 500

    payload = {'status': 'success'}
    payload.update(sanitize_json_values(result))
    return jsonify(payload)


@app.route('/api/jobs/<job_id>', methods=['GET'])
def get_job_status(job_id):
    """Poll a queued analysis job (job_runner contract).

    States: queued | running | done | error. A finished job carries
    'result'; a failed one carries 'error'. Records expire after the
    runner TTL (default 1 h) and then return 404.
    """
    try:
        status = job_runner.status(job_id)
    except KeyError:
        return jsonify({
            'status': 'error',
            'error': f'Unknown or expired job id: {job_id}',
        }), 404
    return jsonify({'status': 'success', 'job': sanitize_json_values(status)})


@app.route('/api/validation/upload-csv', methods=['POST'])
def validation_upload_csv():
    """Parse a user static-fire thrust CSV and (optionally) compare it with
    the HRMA prediction.

    Accepts either:
      - a plain-text body (text/csv, text/plain): parse only, or
      - JSON {'csv_text': str, 'predicted_curve': {'time': [...],
        'thrust': [...]}} — parse + quantitative comparison (total impulse,
        peak/mean thrust, NFPA 1125 burn time, RMSE/NRMSE, English
        assessment).
    """
    predicted_curve = None
    csv_text = None
    if request.is_json:
        data = request.get_json(silent=True) or {}
        csv_text = data.get('csv_text')
        predicted_curve = data.get('predicted_curve')
    else:
        csv_text = request.get_data(as_text=True)

    if not isinstance(csv_text, str) or not csv_text.strip():
        return jsonify({
            'status': 'error',
            'error': ("No CSV content provided. Send the file as a "
                      "text/csv body, or as JSON {'csv_text': '...'} with "
                      "an optional 'predicted_curve' {'time', 'thrust'}."),
        }), 400

    try:
        parsed = parse_thrust_csv(csv_text)
    except ValueError as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 400

    parsed_out = {
        'time': parsed['time'].tolist(),
        'thrust': parsed['thrust'].tolist(),
        'n_points': parsed['n_points'],
        'warnings': parsed['warnings'],
    }

    comparison = None
    if predicted_curve is not None:
        try:
            comparison = sanitize_json_values(
                compare_thrust_curves(parsed, predicted_curve))
        except ValueError as exc:
            # CSV çözüldü ama karşılaştırma anlamsız (örtüşme yok vb.):
            # panel çözümlemeyi yine gösterebilsin diye parsed eklenir.
            return jsonify({
                'status': 'error',
                'error': str(exc),
                'parsed': parsed_out,
            }), 400

    return jsonify({
        'status': 'success',
        'parsed': parsed_out,
        'comparison': comparison,
    })


# ---------------------------------------------------------------------------
# v2.5.0 G3 — Belirsizlik nicelemesi (UQ) endpoint'i
# ---------------------------------------------------------------------------
# Kontrat (G3 API dalgası): POST /api/uncertainty-analysis
#   fast / engineering  -> senkron 'ok' gövdesi
#   high_fidelity       -> job_runner'a kuyruklanır (202 + job_id; sonuç
#                          GET /api/jobs/<id> sözleşmesiyle, job.result =
#                          aynı 'ok' gövdesi)
# Motor print gürültüsü endpoint kapsamında os.devnull'a yönlendirilir.

_UQ_MOTOR_TYPES = ('hybrid', 'solid', 'liquid')


class _UQAnalysisError(RuntimeError):
    """UQ koşusu status='error' döndürdü (örnek #0 tutarlılık kırılması
    dahil) — sessiz düşme yok, 500 + mesajla yukarı taşınır."""


def _uq_contract_body(motor_type, level, result):
    """run_uncertainty sonucunu G3 API kontrat gövdesine çevirir.

    Kontrat alanlarına ek olarak şeffaflık alanları taşınır (sampler,
    uq_version, inputs_used künyeleri, yöntem notu) — kontratın üst kümesi.
    'cv' yüzde cinsindendir (mevcut solid MC cv_percent geleneği); aynı değer
    'cv_percent' adıyla da yankılanır ki birim belirsizliği kalmasın.
    """
    outputs = {}
    for key, block in result['outputs'].items():
        outputs[key] = {
            'nominal': block['nominal'],
            'mean': block['mean'],
            'std': block['std'],
            'cv': block['cv_percent'],
            'cv_percent': block['cv_percent'],
            'p5': block['p5'],
            'p25': block['p25'],
            'p50': block['p50'],
            'p75': block['p75'],
            'p95': block['p95'],
            'histogram': block['histogram'],
        }
    # sensitivity sözlüğü hem çıktı-başına SATIR LİSTESİ hem de skaler
    # meta alanları taşır ('method_note', v2.6.2'de eklenen 'noise_floor').
    # Skalerleri satır listesi sanıp döngüye sokmak TypeError üretiyordu;
    # bu yüzden liste olmayan her değer meta kabul edilip aynen taşınır.
    sensitivity = {}
    sensitivity_meta = {}
    for key, rows in result['sensitivity'].items():
        if not isinstance(rows, (list, tuple)):
            sensitivity_meta[key] = rows
            continue
        sensitivity[key] = [
            {'param': row['param'], 'rho': row['spearman'],
             # Gürültü tabanı: |rho| bu değerin altındaysa duyarlılık
             # örnekleme gürültüsünden ayırt edilemez (v2.6.2 fizik denetimi).
             **({'noise_floor': row['noise_floor']}
                if isinstance(row, dict) and 'noise_floor' in row else {})}
            for row in rows
        ]
    body = {
        'status': 'ok',
        'motor_type': motor_type,
        'level': level,
        'n_samples': result['n_samples'],
        'failed_samples': result['failed_samples'],
        'seed': result['seed'],
        'timing_s': result['timing']['wall_s'],
        'mean_shift_percent': result['consistency']['mean_shift_percent'],
        'outputs': outputs,
        'sensitivity': sensitivity,
        'sensitivity_method_note': result['sensitivity'].get('method_note'),
        # Skaler meta alanları (ör. noise_floor): |rho| bu eşiğin altındaysa
        # duyarlılık örnekleme gürültüsünden ayırt edilemez. Kullanıcı bunu
        # görmeden sıralamaya anlam yükleyemez.
        'sensitivity_meta': sensitivity_meta,
        'sampler': result['sampler'],
        'uq_version': result['uq_version'],
        'inputs_used': result['inputs_used'],
        'consistency_note': result['consistency'].get('note'),
    }
    if result.get('warning'):
        body['warning'] = result['warning']
    return body


def _run_uq_analysis(motor_type, level, seed, inputs, overrides,
                     n_samples=None, progress_callback=None):
    """Senkron ve job yolunun ortak çekirdeği.

    Raises:
        ValueError: girdi/dağılım doğrulama hatası (endpoint 400'e çevirir).
        _UQAnalysisError: koşu status='error' bitirdi (endpoint 500).
    """
    from hrma.analysis import uncertainty as _uq
    from hrma.analysis import uq_adapters as _uqa

    distributions = _uqa.build_distributions(motor_type, overrides)
    if n_samples is None:
        n = _uq.LEVEL_BUDGETS[level]
    else:
        n = max(50, min(int(n_samples), 10000))  # spec 7.1 kırpması
    track = (level == 'high_fidelity')  # spec: yalnız High-Fidelity O/F izler

    cb = None
    if progress_callback is not None:
        def cb(done, total):
            progress_callback(done / max(total, 1))

    with open(os.devnull, 'w') as devnull, \
            contextlib.redirect_stdout(devnull):
        factory = _uqa.make_factory(motor_type, inputs,
                                    track_performance=track)
        result = _uq.run_uncertainty(
            factory, distributions, n_samples=n, seed=seed,
            progress_callback=cb)

    if result.get('status') != 'success':
        raise _UQAnalysisError(
            result.get('error') or 'uncertainty analysis failed')
    return sanitize_json_values(_uq_contract_body(motor_type, level, result))


def _uncertainty_job(payload, progress_callback=None):
    """job_runner işi: high_fidelity UQ koşusu (job.result = 'ok' gövdesi)."""
    return _run_uq_analysis(
        payload['motor_type'], payload['level'], payload['seed'],
        payload['inputs'], payload['overrides'], payload.get('n_samples'),
        progress_callback=progress_callback)


@app.route('/api/uncertainty-analysis', methods=['POST'])
def uncertainty_analysis():
    """Monte Carlo / LHS uncertainty analysis (G3 contract).

    Request: {motor_type, level, seed?, inputs, distribution_overrides?,
    n_samples?}. fast/engineering run synchronously; high_fidelity is queued
    on the job runner (202 + job_id, poll GET /api/jobs/<id>; the finished
    job's result is the same 'ok' body).
    """
    data = request.get_json(silent=True) or {}

    motor_type = data.get('motor_type')
    if motor_type not in _UQ_MOTOR_TYPES:
        return jsonify({
            'status': 'error',
            'error': (f"motor_type must be one of {list(_UQ_MOTOR_TYPES)}; "
                      f"got {motor_type!r}."),
        }), 400

    from hrma.analysis.uncertainty import LEVEL_BUDGETS as _LEVELS
    level = data.get('level')
    if level not in _LEVELS:
        return jsonify({
            'status': 'error',
            'error': (f"level must be one of {sorted(_LEVELS)}; "
                      f"got {level!r}."),
        }), 400

    try:
        seed = int(data.get('seed', 42))
    except (TypeError, ValueError):
        return jsonify({'status': 'error',
                        'error': "Field 'seed' must be an integer."}), 400

    inputs = data.get('inputs')
    if inputs is None:
        inputs = {}
    if not isinstance(inputs, dict):
        return jsonify({'status': 'error',
                        'error': "Field 'inputs' must be an object with the "
                                 "motor form fields."}), 400

    overrides = data.get('distribution_overrides')
    n_samples = data.get('n_samples')
    if n_samples is not None:
        try:
            n_samples = int(n_samples)
        except (TypeError, ValueError):
            return jsonify({'status': 'error',
                            'error': "Field 'n_samples' must be an "
                                     "integer."}), 400

    # Dağılım kümesi erken doğrulanır: bozuk override job kuyruğuna girmeden
    # net bir 400 ile dönsün (job yolunda hata ancak poll'da görünürdü).
    try:
        from hrma.analysis import uq_adapters as _uqa
        _uqa.build_distributions(motor_type, overrides)
    except ValueError as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 400

    if level == 'high_fidelity':
        job_id = job_runner.submit(_uncertainty_job, {
            'motor_type': motor_type,
            'level': level,
            'seed': seed,
            'inputs': inputs,
            'overrides': overrides,
            'n_samples': n_samples,
        })
        return jsonify({
            'status': 'queued',
            'job_id': job_id,
            'poll_url': f'/api/jobs/{job_id}',
        }), 202

    try:
        body = _run_uq_analysis(motor_type, level, seed, inputs, overrides,
                                n_samples)
    except ValueError as exc:
        return jsonify({'status': 'error', 'error': str(exc)}), 400
    except _UQAnalysisError as exc:
        # Örnek #0 tutarlılık kırılması dahil: sessiz düşme yok (spec 7.3)
        return jsonify({'status': 'error', 'error': str(exc)}), 500
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(exc)}), 500
    return jsonify(body)


# ---------------------------------------------------------------------------
# v2.5.0 G3 — Otomatik korelasyon raporu endpoint'i
# ---------------------------------------------------------------------------
# GET  /api/correlation-report          -> önbellekli özet (ilk çağrı koşar)
# POST /api/correlation-report {refresh:true} -> önbelleği yok say, yeniden koş
# Önbellek anahtarı: deney DB içerik hash'i (correlation_runner.db_content_hash)

_CORRELATION_CACHE = {}

# Korelasyon koşusu tek seferde BİR kez çalışsın (v2.6.2).
#
# Koşu soğukta ~2 dakika sürüyor (Cantera denge çözümleri; docstring "~15-25 s"
# diyordu ama DB 136 -> 209 kayda çıkınca güncellenmemişti). Flask geliştirme
# sunucusu çok iş parçacıklı olduğu için, önbellek soğukken gelen iki eşzamanlı
# istek AYNI ağır işi iki kez başlatıyordu: CPU ikiye katlanıyor, ikisi de
# önbelleğe yazıyor ve kullanıcı iki katı bekliyordu. Kilit, ikinci isteğin
# birincinin sonucunu beklemesini sağlar (sonra önbellekten anında döner).
_CORRELATION_LOCK = threading.Lock()


def _correlation_report_body(refresh=False):
    """Korelasyon rapor gövdesini kurar (modül-içi {db_hash: gövde} önbelleği).

    Koşu ~15-25 s sürer (Cantera denge çözümleri); ilk çağrı senkron kabul
    edilir, sonrakiler cached=true ile anında döner. DB içeriği değişince
    hash değişir ve önbellek kendiliğinden ıskalar.
    """
    from hrma.validation import correlation_runner as _cr
    from hrma.validation.experiment_db import (load_records,
                                               records_for_statistics)
    from hrma.validation.status_report import correlation_cells

    records = load_records()
    stat_records = sorted(records_for_statistics(records),
                          key=lambda r: r.get('test_id', ''))
    db_hash = _cr.db_content_hash(stat_records)

    if not refresh and db_hash in _CORRELATION_CACHE:
        body = dict(_CORRELATION_CACHE[db_hash])
        body['cached'] = True
        return body

    with open(os.devnull, 'w') as devnull, \
            contextlib.redirect_stdout(devnull):
        result = _cr.run_correlation(records=records)

    skipped_scores = {}
    for rr in result['records']:
        for score in rr.get('scores', {}).values():
            status = score.get('status')
            if status and status != 'scored':
                skipped_scores[status] = skipped_scores.get(status, 0) + 1

    body = sanitize_json_values({
        'status': 'ok',
        'db_hash': result['db_content_hash'],
        'cached': False,
        'generated_s': result['timing']['total_s'],
        'record_counts': {
            'total': result['n_records'],
            'scored': result['status_counts'].get('ok', 0),
            'insufficient_inputs': result['status_counts'].get(
                'insufficient_inputs', 0),
            'not_supported': result['status_counts'].get('not_supported', 0),
            'runner_error': result['status_counts'].get('runner_error', 0),
        },
        'cells': correlation_cells(result['statistics']),
        'skipped_summary': {
            'status_counts': result['status_counts'],
            'not_supported': result['not_supported'],
            'insufficient_inputs': result['insufficient_inputs'],
            'runner_errors': result['runner_errors'],
            'skipped_score_counts': dict(sorted(skipped_scores.items())),
        },
        'markdown': _cr.to_markdown(result),
    })

    # Tek girdilik önbellek: DB değişince eski sonuç bellekte birikmesin
    _CORRELATION_CACHE.clear()
    _CORRELATION_CACHE[db_hash] = body
    return dict(body)


@app.route('/api/correlation-report', methods=['GET', 'POST'])
def correlation_report():
    """Real-experiment correlation report (G3 contract).

    GET returns the cached report when the experiment DB is unchanged
    (cached=true); the first call runs the full correlation synchronously.
    POST with {"refresh": true} ignores the cache and re-runs.
    """
    refresh = False
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        refresh = bool(data.get('refresh'))
    try:
        # Kilit: eşzamanlı istekler ağır koşuyu tekrarlamaz. İlk istek koşar,
        # ikincisi bekler ve kilidi aldığında önbellekten anında döner.
        with _CORRELATION_LOCK:
            body = _correlation_report_body(refresh=refresh)
    except Exception as exc:
        traceback.print_exc()
        return jsonify({'status': 'error', 'error': str(exc)}), 500
    return jsonify(body)


if __name__ == '__main__':
    # Gerçek giriş noktası hrma/run.py (waitress, 8080); bu blok yalnız
    # geliştirme içindir. debug VARSAYILAN OLARAK KAPALIDIR: Flask debug modu
    # Werkzeug'un interaktif hata ayıklayıcısını açar ve localhost'ta bile
    # rastgele kod çalıştırılmasına izin verebilir (2026-07-24 denetim bulgusu).
    # Açıkça istemek için HRMA_DEBUG=1 ortam değişkeni kullanılır.
    debug = os.environ.get('HRMA_DEBUG', '') in ('1', 'true', 'True')
    print("Starting Motor Analysis on port 8080..." + (" [debug]" if debug else ""))
    app.run(debug=debug, port=8080, host='127.0.0.1')